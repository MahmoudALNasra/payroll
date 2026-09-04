"""A standalone local Payroll Application using ttkbootstrap with AES Encryption."""

import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
import sys
import os

# PyInstaller --windowed safe stream fallbacks to prevent NoneType attribute crashes
class _SafeStream:
    def write(self, *args, **kwargs): pass
    def flush(self, *args, **kwargs): pass
if sys.stdout is None:
    sys.stdout = _SafeStream()
if sys.stderr is None:
    sys.stderr = _SafeStream()

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

class SafeTkProxy:
    """Safeguards Tkinter against Tcl crashes from non-BMP Unicode characters (> 0xFFFF).
    On Windows, Tcl 8.6 is built with 16-bit characters and raises TclError when any character
    with codepoint > 0xFFFF is passed. This proxy intercepts Tk calls and safely sanitizes
    strings, mapping common emojis to safe representations and stripping unsupported codepoints.
    """
    _EMOJI_MAP = {
        0x1F680: "[Update]",
        0x1F441: "Show",
        0x1F648: "Hide",
        0x1F48E: "❖",
        0x1F504: "⟳",
        0x1F4E5: "↓",
        0x1F4E4: "↑",
        0x1F5D1: "X",
        0x1F4CA: "■",
        0x1F4C1: "[Folder]",
        0x1F4C2: "[Folder]",
        0x1F4C4: "[File]",
        0x1F4BE: "[Save]",
        0x1F511: "*",
        0x1F512: "[Lock]",
        0x1F513: "[Unlock]",
        0x1F4C5: "[Date]",
        0x1F465: "[Users]",
        0x1F4B8: "[$]",
        0x1F3ED: "[Factory]",
        0x1F389: "!",
        0x1F9F9: "[Clear]",
        0x1F7E2: "●",
        0x1F534: "●",
        0x1F7E1: "●",
        0x1F4E6: "■",
        0x1F4CB: "❖",
        0x1F3AF: "*",
        0x1F488: "[Shop]",
        0x1F487: "[Stylist]",
        0x1F4BB: "[PC]",
        0x1F4D6: "[Guide]",
        0x1F4CD: "*",
        0x1F3F7: "*",
        0x1F4B3: "[$]",
        0x1F5C4: "[DB]",
        0x1F50D: "[Search]",
        0x1F4F4: "[Offline]",
    }

    def __init__(self, real_tk):
        self._real_tk = real_tk

    @classmethod
    def _sanitize(cls, arg):
        if isinstance(arg, str):
            if not any(ord(c) > 0xFFFF or (0xFE00 <= ord(c) <= 0xFE0F) for c in arg):
                return arg
            res = []
            for c in arg:
                cp = ord(c)
                if cp > 0xFFFF:
                    res.append(cls._EMOJI_MAP.get(cp, ""))
                elif 0xFE00 <= cp <= 0xFE0F:
                    continue
                else:
                    res.append(c)
            return "".join(res)
        elif isinstance(arg, (tuple, list)):
            return type(arg)(cls._sanitize(x) for x in arg)
        elif isinstance(arg, dict):
            return {k: cls._sanitize(v) for k, v in arg.items()}
        return arg

    def call(self, *args):
        return self._real_tk.call(*[self._sanitize(a) for a in args])

    def eval(self, *args):
        return self._real_tk.eval(*[self._sanitize(a) for a in args])

    def setvar(self, name, value):
        return self._real_tk.setvar(name, self._sanitize(value))

    def globalsetvar(self, name, value):
        return self._real_tk.globalsetvar(name, self._sanitize(value))

    def __getattr__(self, name):
        return getattr(self._real_tk, name)

import csv
import hashlib
import tempfile
import os
import atexit
import platform
import uuid
import shutil
import subprocess
import time
import json
import gzip
import threading
from datetime import datetime, timedelta
from contextlib import contextmanager

try:
    import ttkbootstrap as tb
    from ttkbootstrap.constants import *
    from cryptography.fernet import Fernet
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    import base64
    try:
        from ttkbootstrap.widgets import ToolTip
    except ImportError:
        try:
            from ttkbootstrap.tooltip import ToolTip
        except ImportError:
            class ToolTip:
                def __init__(self, widget, text, **kwargs):
                    self.widget = widget
                    self.text = text
                    self.tip_window = None
                    self.widget.bind("<Enter>", self.show_tip)
                    self.widget.bind("<Leave>", self.hide_tip)
                def show_tip(self, event=None):
                    if self.tip_window or not self.text:
                        return
                    x = self.widget.winfo_rootx() + 20
                    y = self.widget.winfo_rooty() + 25
                    self.tip_window = tw = tk.Toplevel(self.widget)
                    tw.wm_overrideredirect(True)
                    tw.wm_geometry(f"+{x}+{y}")
                    label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                                    background="#333333", foreground="#ffffff", relief=tk.SOLID, borderwidth=1,
                                    font=("Segoe UI", 10, "normal"))
                    label.pack(ipadx=6, ipady=3)
                def hide_tip(self, event=None):
                    tw = self.tip_window
                    self.tip_window = None
                    if tw:
                        tw.destroy()
    HAS_DEPS = True
except ImportError:
    HAS_DEPS = False

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# standard tk constants fallbacks to avoid NameErrors
CENTER = tk.CENTER
TOP = tk.TOP
BOTTOM = tk.BOTTOM
LEFT = tk.LEFT
RIGHT = tk.RIGHT
BOTH = tk.BOTH
X = tk.X
Y = tk.Y
E = tk.E
W = tk.W
N = tk.N
S = tk.S
NE = tk.NE
NW = tk.NW
SE = tk.SE
SW = tk.SW
NS = tk.NS
EW = tk.EW
NSEW = tk.NSEW
VERTICAL = tk.VERTICAL
HORIZONTAL = tk.HORIZONTAL


# --- TRANSLATIONS ---
TRANSLATIONS = {
    "Cycle": "الدورة",
    "Cycle:": "الدورة:",
    "Cycles": "دورات",
    "All Cycles": "جميع الدورات",
    "Recent Cycles": "الدورات الأخيرة",
    "Browse by Year": "تصفح حسب السنة",
    "Cycles Selected": "دورات محددة",
    "Target Cycle:": "الدورة المستهدفة:",
    "Detected Period:": "الفترة المكتشفة:",
    "Tip Included": "يشمل البقشيش",
    "Tip Included:": "يشمل البقشيش:",
    "1. Select Target Pay Period & Cycle": "1. اختر فترة الدورة المستهدفة",
    "2. Match Column Headers": "2. مطابقة أعمدة الملف",
    "Import Sales Data": "استيراد بيانات المبيعات",
    "Match Column Headers:": "مطابقة أعمدة الملف:",
    "Current Cycle": "الدورة الحالية",
    "Pay Cycles": "دورات الرواتب",
    "Browse Cycles": "تصفح الدورات",
    "Table Columns": "أعمدة الجدول",
    "Customize Table Columns": "تخصيص أعمدة الجدول",
    "Select Columns to Display": "حدد الأعمدة المراد عرضها",
    "Save & Apply": "حفظ وتطبيق",
    "Reset to Default": "إعادة التعيين للافتراضي",
    "Columns Updated": "تم تحديث الأعمدة",
    "Table columns display updated successfully.": "تم تحديث عرض أعمدة الجدول بنجاح.",
    "Cycle Selected": "دورة محددة",
    "Select All": "تحديد الكل",
    "Clear Selection": "مسح التحديد",
    "Year:": "السنة:",
    "Revenue Tab": "تبويب الإيرادات",
    "Employees": "الموظفين",
    "Data Entry": "إدخال البيانات",
    "Filters": "تصفية",
    "From:": "من:",
    "To:": "إلى:",
    "Employee:": "الموظف:",
    "Actions": "الإجراءات",
    "💾 Export CSV": "💾 تصدير CSV",
    "✏️ Edit": "✏️ تعديل",
    "🗑️ Delete": "🗑️ حذف",
    "📅 Quick Periods": "📅 فترات سريعة",
    "Today": "اليوم",
    "This Week": "هذا الأسبوع",
    "This Month": "هذا الشهر",
    "All Time": "كل الوقت",
    "Record ID": "معرف السجل",
    "Service Sales": "مبيعات الخدمة",
    "Service Add-on Sales": "مبيعات الخدمات الإضافية",
    "Tips": "البقشيش",
    "Date": "التاريخ",
    "Name": "الاسم",
    "Revenue": "الإيرادات",
    "Hour Rate": "سعر الساعة",
    "Percentage": "النسبة",
    "Hours": "الساعات",
    "Calculation": "الحساب",
    "Payment Amount": "مبلغ الدفع",
    "Payment Type": "نوع الدفع",
    "Notes": "ملاحظات",
    "Written Up": "مكتوب",
    "Rev:": "إيرادات:",
    "Hrs:": "ساعات:",
    "Calc:": "حساب:",
    "Paid:": "مدفوع:",
    "=== TOTALS ===": "=== الإجمالي ===",
    "👥 Employee Management": "👥 إدارة الموظفين",
    "🔑 Change App Password": "🔑 تغيير كلمة المرور",
    "History & Backups": "السجل والنسخ الاحتياطي",
    "Activity & Backups": "النشاط والنسخ الاحتياطي",
    "Activity log": "سجل النشاط",
    "Filter by user:": "تصفية حسب المستخدم:",
    "All users": "كل المستخدمين",
    "Cloud backups": "النسخ الاحتياطي السحابي",
    "Two backups are saved in Supabase every day (morning and afternoon).": "يتم حفظ نسختين احتياطيتين في Supabase كل يوم (صباحاً ومساءً).",
    "📥 Backup now to cloud": "📥 نسخ احتياطي إلى السحابة الآن",
    "Restore selected backup": "استعادة النسخة المحددة",
    "Database Backups": "النسخ الاحتياطي لقاعدة البيانات",
    "📥 Generate Immediate Backup Copy": "📥 إنشاء نسخة احتياطية فورية",
    "Version Reversion History (Supabase Mode Only)": "سجل إرجاع الإصدارات (وضع Supabase فقط)",
    "🔄 Refresh Logs": "🔄 تحديث السجل",
    "⏪ Revert selected revision": "⏪ استعادة التعديل المحدد",
    "Date": "التاريخ",
    "User": "المستخدم",
    "Table": "الجدول",
    "Action": "الإجراء",
    "Refresh": "تحديث",
    "+ Add New Employee": "+ إضافة موظف",
    "✏️ Edit Selected": "✏️ تعديل المحدد",
    "🗑️ Delete Selected": "🗑️ حذف المحدد",
    "📊 View Performance": "📊 عرض الأداء",
    "ID": "المعرف",
    "✨ Add New Payroll Record": "✨ إضافة سجل رواتب جديد",
    "Employee Name:": "اسم الموظف:",
    "Payment Type:": "نوع الدفع:",
    "Payment Amount:": "مبلغ الدفع:",
    "Revenue:": "الإيرادات:",
    "Hours:": "الساعات:",
    "Notes:": "ملاحظات:",
    "Written Up:": "مكتوب:",
    "Calculate & Save to Database": "حساب وحفظ",
    "All": "الكل",
    "Logged in as": "تم تسجيل الدخول كـ",
    "Employee Folders": "ملفات الموظفين",
    "📁 Employee Folders": "📁 ملفات الموظفين",
    "First Name:": "الاسم الأول:",
    "Last Name:": "الاسم الأخير:",
    "First Name:*": "الاسم الأول:*",
    "Phone:": "رقم الهاتف:",
    "Email:": "البريد الإلكتروني:",
    "First Name": "الاسم الأول",
    "Last Name": "الاسم الأخير",
    "Phone": "الهاتف",
    "Email": "البريد",
    "Files Count": "عدد الملفات",
    "Documents": "المستندات",
    "Employee details and documents": "تفاصيل ومستندات الموظف",
    "Upload Document": "رفع مستند",
    "Open Folder": "فتح المجلد",
    "Open Selected": "فتح المحدد",
    "Select File": "اختر ملف",
    "First Name is required.": "الاسم الأول مطلوب.",
    "✏️ Edit Info": "✏️ تعديل المعلومات",
    "Edit Employee": "تعديل بيانات الموظف",
    "Add Employee": "إضافة موظف",
    "Save": "حفظ",
    "Save Employee": "حفظ الموظف",
    "Edit Info": "تعديل المعلومات",
    "SSN:": "الرقم الوطني:",
    "Address:": "العنوان:",
    "Start Date:": "تاريخ البدء:",
    "End Date:": "تاريخ الانتهاء:",
    "CV:": "السيرة الذاتية:",
    "ID Photo:": "صورة الهوية:",
    "Personal Photo:": "الصورة الشخصية:",
    "Browse": "تصفح",
    "💸 Expense Reports": "💸 تقارير المصروفات",
    "Expense Reports": "تقارير المصروفات",
    "Category": "الفئة",
    "Amount": "المبلغ",
    "Description": "الوصف",
    "Status": "الحالة",
    "Category:": "الفئة:",
    "Amount:": "المبلغ:",
    "Calculated Pay": "الراتب المحسوب",
    "Expense Amount": "مبلغ المصروف",
    "Groceries": "البقالة",
    "Description:": "الوصف:",
    "Document / Receipt:": "المستند / الإيصال:",
    "Upload Document": "رفع مستند",
    "Preview Document": "معاينة المستند",
    "Clear Document": "مسح المستند",
    "No document attached": "لا يوجد مستند مرفق",
    "Document Preview": "معاينة المستند",
    "Open with default app": "فتح بالتطبيق الافتراضي",
    "Click filename to preview": "انقر على اسم الملف للمعاينة",
    "Print": "طباعة",
    "Add files": "إضافة ملفات",
    "Status:": "الحالة:",
    "+ Add Expense": "+ إضافة مصروف",
    "Add Expense": "إضافة مصروف",
    "Edit Expense": "تعديل المصروف",
    "Delete Expense": "حذف المصروف",
    "General/None": "عام / لا يوجد",
    "Total Filtered": "إجمالي المصفى",
    "Salary Payment": "دفع الرواتب",
    "Salary includes tip?": "هل يشمل الراتب البقشيش؟",
    "Tip given to employee:": "البقشيش المعطى للموظف:",
    "How much of the tip did you give this employee?": "كم من البقشيش أعطيت لهذا الموظف؟",
    "Enter how much tip you gave the employee.": "أدخل مبلغ البقشيش الذي أعطيته للموظف.",
    "Amazon Order": "طلب أمازون",
    "Expense Distribution": "توزيع المصروفات",
    "No expense data to display in chart": "لا توجد مصاريف لعرضها في المخطط",
    "Cash Envelope Received": "ظرف نقدي مستلم",
    "Total Revenue": "إجمالي الإيرادات",
    "Total Expenses": "إجمالي المصروفات",
    "Net Income": "صافي الدخل",
    "Integration: Vagaro Pull": "مزامنة: سحب فاجارو",
    "Vagaro Integration Pull": "سحب البيانات من فاجارو",
    "Pull Data": "سحب البيانات",
    "Close Window": "إغلاق النافذة",
    "Employee Revenue": "إيرادات الموظف",
    "Expenses vs. Revenue": "المصروفات مقابل الإيرادات",
    "No financial data to display in chart": "لا توجد بيانات مالية لعرضها في المخطط",
    "Vagaro Sync:": "مزامنة فاجارو:",
    "Employees": "الموظفين",
    "Revenue": "الإيرادات",
    "Phone": "الهاتف",
    "Email": "البريد الإلكتروني",
    "Expenses": "المصروفات",
    "None": "بلا",
    "Employees, Revenue": "الموظفين، الإيرادات",
    "Phone Number": "رقم الهاتف",
    "Email Address": "البريد الإلكتروني",
    "Synced Date": "تاريخ المزامنة",
    "Pull Timestamp": "وقت السحب",
    "Sync History Log": "سجل تاريخ المزامنة",
    "Import Excel Sales": "استيراد مبيعات Excel",
    "📁 Shop Files": "📁 ملفات المحل",
    "Choose a shop location": "اختر موقع المحل",
    "Select which shop you want to open files for.": "اختر المحل الذي تريد فتح ملفاته.",
    "Shop document archive": "أرشيف مستندات المحل",
    "← Back to locations": "← العودة للمواقع",
    "+ Add Document": "+ إضافة مستند",
    "Document name": "اسم المستند",
    "Document date": "تاريخ المستند",
    "No shop documents yet.\nAdd a scanned lease, rent receipt, or other record.": "لا توجد مستندات بعد.\nأضف عقد إيجار أو إيصال أو أي سجل آخر.",
    "Save Document": "حفظ المستند",
    "Cancel": "إلغاء",
    "Open / Preview": "فتح / معاينة",
    "Delete Document": "حذف المستند",
    "Attached file:": "الملف المرفق:",
    "Choose file…": "اختر ملف…",
    "A name and file are required.": "الاسم والملف مطلوبان.",
    "Shop document saved.": "تم حفظ مستند المحل.",
    "Open a location to manage leases, rent, permits, and other shop records.": "افتح موقعاً لإدارة عقود الإيجار والإيصالات والتراخيص وسجلات المحل.",
    "documents": "مستندات",
    "Open archive →": "فتح الأرشيف ←",
    "← Back to archive": "← العودة للأرشيف",
    "Keep rent leases, permits, and scanned shop papers in one place.": "احفظ عقود الإيجار والتراخيص والمستندات الممسوحة في مكان واحد.",
    "Select a document first.": "اختر مستنداً أولاً.",
    "Delete this shop document and its file?": "حذف مستند المحل وملفه؟",
    "No file chosen yet": "لم يتم اختيار ملف بعد",
    "Enter a name, date, short description, then attach the scanned file.": "أدخل الاسم والتاريخ ووصفاً قصيراً ثم أرفق الملف الممسوح.",
    "Tip: scan the paper first, then upload the PDF or photo here.": "نصيحة: امسح الورقة أولاً ثم ارفع ملف PDF أو صورة هنا.",
    "No locations configured. Add locations in Settings first.": "لا توجد مواقع. أضف المواقع من الإعدادات أولاً.",
    "Document file was not found.": "لم يتم العثور على ملف المستند.",
    "File": "الملف",
    "No exclude": "بدون استبعاد",
    "Exclude:": "استبعاد:",
    "excluded": "مستبعد",
    "🧰 Tools": "🧰 أدوات",
    "Import & Tools": "الاستيراد والأدوات",
    "Commissions": "العمولات",
    "Service sales commissions": "عمولات مبيعات الخدمات",
    "Product sales commissions": "عمولات مبيعات المنتجات",
    "From sales ()":"منالمبيعات()",
    "To sales ()":"إلىالمبيعات()",
    "From sales () → Tosales()  =  commission %. Dollar signs are optional.": "من المبيعات ()←إلىالمبيعات() = نسبة العمولة %. علامة $ اختيارية.",
    "Select All": "تحديد الكل",
    "Service Sales Calculations": "حساب مبيعات الخدمات",
    "Total Calculation": "الحساب الإجمالي",
    "Dynamic commission tiers": "شرائح العمولة الديناميكية",
    "From sales": "من المبيعات",
    "To sales": "إلى المبيعات",
    "Commission %": "نسبة العمولة %",
    "👥 Users": "👥 المستخدمون",
    "Add User": "إضافة مستخدم",
    "New username": "اسم مستخدم جديد",
    "Username added. Default password is admin until they change it.": "تمت إضافة المستخدم. كلمة المرور الافتراضية admin حتى يغيرها.",
    "Cannot delete the admin account.": "لا يمكن حذف حساب المسؤول.",
    "Amount is required.": "المبلغ مطلوب.",
    "Delete Month Envelopes": "حذف مغلفات الشهر",
    "This will permanently delete all cash envelopes for this month.": "سيتم حذف كل مغلفات هذا الشهر نهائياً.",
    "Preparing your workspace…": "جاري تجهيز مساحة العمل…",
    "Uploading local changes…": "جاري رفع التغييرات المحلية…",
    "Downloading latest records…": "جاري تنزيل أحدث السجلات…",
    "Almost ready…": "أوشك على الانتهاء…",
    "Saving…": "جاري الحفظ…",
    "Please wait, do not click…": "يرجى الانتظار، لا تضغط…",
    "🔑 Account": "🔑 الحساب",
    "Change username": "تغيير اسم المستخدم",
    "Current username": "اسم المستخدم الحالي",
    "New username:": "اسم المستخدم الجديد:",
    "Update Username": "تحديث اسم المستخدم",
    "Change Username": "تغيير اسم المستخدم",
    "Rename Selected": "إعادة تسمية المحدد",
    "Username updated. Use the new name at login.": "تم تحديث اسم المستخدم. استخدم الاسم الجديد عند تسجيل الدخول.",

    "Import Excel sales files into payroll records.": "استيراد ملفات مبيعات Excel إلى سجلات الرواتب.",
    "All expenses": "كل المصروفات",
    "categories": "فئات",
    "None": "لا شيء",
    "Select Excel Sales File": "اختر ملف مبيعات Excel",
    "Excel Column Mapper": "مخطط أعمدة Excel",
    "Match Excel Headers:": "تطابق أعمدة Excel:",
    "Employee Column:": "عمود الموظف:",
    "Service Sales Column:": "عمود مبيعات الخدمات:",
    "Product Sales Column:": "عمود مبيعات المنتجات:",
    "Tip Column:": "عمود الإكرامية (Tips):",
    "Import Data": "استيراد البيانات",
    "📊 P&L / Financials": "📊 الأرباح والخسائر / المالية",
    "Dashboard Filters": "تصفية لوحة التحكم",
    "Include Earnings": "تضمين الأرباح",
    "Include Expenses": "تضمين المصروفات",
    "Show All Employees": "عرض جميع الموظفين",
    "Total Earnings": "إجمالي الأرباح",
    "Net Profit": "صافي الأرباح",
    "Summary by Employee / Shop": "ملخص حسب الموظف / المحل",
    "Detailed Transaction Log": "سجل المعاملات المفصل",
    "Use Tiered Payout:": "استخدام الدفع المتدرج:",
    "Use Tiered Payout": "استخدام الدفع المتدرج",
    "dynamic commission tiers + product sales split": "فئات عمولة ديناميكية + حصة مبيعات المنتجات",
    "Type": "النوع",
    "Earning": "ربح",
    "Expense": "مصروف",
    "📅 Cash Calendar": "📅 تقويم النقدية",
    "All Employees": "جميع الموظفين",
    "Select Action": "اختر الإجراء",
    "✏️ Edit Employee": "✏️ تعديل الموظف",
    "📊 Performance Report": "📊 تقرير الأداء",
    "Assignee": "المفوض",
    "Assignee:": "المفوض:",
    "Received From": "مستلم من",
    "Received From:": "مستلم من:",
    "Other Earnings / Envelopes": "أرباح أخرى / أظرف",
    "Include:": "تضمين:",
    "Hour Rate ():":"أجرالساعة():",
    "Percentage (%):": "النسبة (%):",
    "Write Up Reason:": "سبب الإنذار/المخالفة:",
    "🟢 💈 Shop Earnings": "🟢 💈 تبويب الأرباح",
    "🔵 💇‍♂️ Barbers / Stylists": "🔵 الحلاقين / المصممين",
    "🟠 ✍️ Manual Ledger": "🟠 السجل اليدوي",
    "🟣 📊 P&L / Financials": "🟣 الأرباح والخسائر / المالية",
    "🔴 📅 Cash Calendar": "🔴 تقويم النقدية",
    "Auto": "تلقائي",
    "Field Info": "معلومات الحقل",
    "Table Info": "معلومات الجدول",
    "✅ Approve Selected": "✅ قبول المحدد",
    "❌ Set Pending": "❌ قيد الانتظار",
    "Sync Cloud": "مزامنة السحابة",
    "Syncing…": "جارِ المزامنة…",
    "Connecting…": "جارِ الاتصال…",
    "Connected": "متصل",
    "Synced": "تمت المزامنة",
    "Offline": "غير متصل",
    "pending": "معلق",
    "Sync Cloud Now": "مزامنة السحابة الآن",
    "Cloud Synchronization": "مزامنة السحابة",
    "Revenue Data Export": "تصدير بيانات الإيرادات",
    "Export Revenue to CSV / Excel": "تصدير الإيرادات إلى CSV / Excel",
    "+Add Rate": "+إضافة النسبة/الأجر",
    "+Add Rate for": "+إضافة النسبة لـ",
    "Missing Rate": "النسبة غير محددة"
}

# --- APP CONFIGURATION ---
APP_TITLE = "Highend Payroll App - Custom Made ✂"
APP_LOGO_TITLE = "★ HIGHEND PAYROLL ★"
APP_GEOMETRY = "1250x900"
APP_THEME = "darkly"

# Versioning Policy:
# - Format: MAJOR.MINOR.PATCH (e.g., 2.5.3)
# - Every commit: Increment PATCH (2.5.1 -> 2.5.2 -> 2.5.3 -> ...)
# - Big change / major feature / overhaul: Increment MINOR (e.g., 2.6.0, 2.7.0) or MAJOR (3.0.0)
APP_VERSION = "2.5.3"
APP_BUILD_DATE = "2026-09-04"
DEFAULT_UPDATE_SERVER_URL = "https://raw.githubusercontent.com/MahmoudALNasra/payroll/main/main.py"
DEFAULT_GITHUB_RAW_URL = DEFAULT_UPDATE_SERVER_URL

def get_default_app_dir():
    import platform
    if platform.system() == "Windows":
        base_dir = os.environ.get("APPDATA") or os.path.expanduser("~")
        app_dir = os.path.join(base_dir, "PayrollProData")
    elif platform.system() == "Darwin":
        base_dir = os.path.expanduser("~/Library/Application Support")
        app_dir = os.path.join(base_dir, "PayrollProData")
    else:
        base_dir = os.path.expanduser("~")
        app_dir = os.path.join(base_dir, ".payrollprodata")
    return app_dir

def get_updates_dir():
    d = os.path.join(get_default_app_dir(), "updates")
    try:
        os.makedirs(d, exist_ok=True)
    except Exception:
        pass
    return d

def get_updates_script_path():
    return os.path.join(get_updates_dir(), "payroll_app.py")

def get_safe_mode_flag_path():
    return os.path.join(get_updates_dir(), "safe_mode.flag")

def get_last_crash_log_path():
    return os.path.join(get_updates_dir(), "last_crash.log")

def _auto_discover_existing_database():
    """If the active app data directory has no database, scan adjacent folders for previous payroll_data.enc."""
    default_dir = get_default_app_dir()
    target_db = os.path.join(default_dir, "payroll_data.enc")
    if os.path.exists(target_db) and os.path.getsize(target_db) > 1000:
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(script_dir, "payroll_data.enc"),
        os.path.join(os.getcwd(), "payroll_data.enc"),
        os.path.join(os.path.expanduser("~"), "Downloads", "pythonapp", "payroll_data.enc"),
        os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "app", "payroll_data.enc"),
        os.path.join(os.path.expanduser("~"), "Desktop", "app", "payroll_data.enc"),
        os.path.join(os.path.expanduser("~"), "AppData", "Roaming", "HighendPayroll", "payroll_data.enc"),
        os.path.join(os.path.expanduser("~"), "Library", "Application Support", "HighendPayroll", "payroll_data.enc"),
    ]
    for c in candidates:
        if os.path.exists(c) and os.path.abspath(c) != os.path.abspath(target_db):
            try:
                if os.path.getsize(c) > 500:
                    import shutil
                    os.makedirs(default_dir, exist_ok=True)
                    shutil.copy2(c, target_db)
                    break
            except Exception:
                pass

def get_app_dir():
    default_dir = get_default_app_dir()
    os.makedirs(default_dir, exist_ok=True)
    _auto_discover_existing_database()
    
    config_file = os.path.join(default_dir, "location_config.json")
    if os.path.exists(config_file):
        try:
            import json
            with open(config_file, "r", encoding="utf-8") as f:
                config_data = json.load(f)
            custom_path = config_data.get("custom_db_directory")
            if custom_path:
                custom_path = os.path.expanduser(custom_path)
                os.makedirs(custom_path, exist_ok=True)
                return custom_path
        except Exception:
            pass
            
    return default_dir

def get_db_config():
    default_dir = get_default_app_dir()
    config_file = os.path.join(default_dir, "location_config.json")
    if os.path.exists(config_file):
        try:
            import json
            with open(config_file, "r", encoding="utf-8") as f:
                res = json.load(f)
                if isinstance(res, dict):
                    return res
        except Exception:
            pass
    return {}

def get_db_mode():
    return get_db_config().get("mode", "local")


def save_last_selected_username(username):
    try:
        cfg = get_db_config()
        cfg["last_selected_username"] = username
        config_file = os.path.join(get_default_app_dir(), "location_config.json")
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4)
    except Exception:
        pass

def load_ui_column_preferences():
    pref_file = os.path.join(get_default_app_dir(), "ui_column_preferences.json")
    if os.path.isfile(pref_file):
        try:
            with open(pref_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_ui_column_preferences(prefs):
    try:
        pref_file = os.path.join(get_default_app_dir(), "ui_column_preferences.json")
        os.makedirs(os.path.dirname(pref_file), exist_ok=True)
        with open(pref_file, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=2)
    except Exception:
        pass

def get_saved_column_widths(table_key):
    prefs = load_ui_column_preferences()
    return prefs.get(table_key, {})

def save_table_column_widths(table_key, tree, columns):
    if tree is None:
        return
    try:
        if not bool(tree.winfo_exists()):
            return
        widths = {}
        for col in columns:
            try:
                w = int(tree.column(col, "width"))
                if w >= 0:
                    widths[str(col)] = w
            except Exception:
                pass
        if widths:
            prefs = load_ui_column_preferences()
            prefs[table_key] = widths
            save_ui_column_preferences(prefs)
    except Exception:
        pass

ALL_CALENDAR_COLUMNS = [
    ("Date", "Record Date (e.g. 2026-08-20)"),
    ("Cycle", "Pay Cycle & Range (e.g. August - Cycle 2)"),
    ("Name", "Employee / Barber Name"),
    ("Location", "Branch / Location"),
    ("Service Sales", "Gross Service Revenue ($)"),
    ("Service Sales Calculations", "Service Commission Payout ($)"),
    ("Service Add-on Sales", "Add-on Sales ($)"),
    ("Product Sales", "Product Retail Sales ($)"),
    ("Tip", "Tips Received ($)"),
    ("Hour Rate", "Hourly Base Rate ($/hr)"),
    ("Percentage", "Commission Payout Rate (%)"),
    ("Hours", "Hours Worked"),
    ("Total Calculation", "Total Payout Calculation ($)"),
    ("Notes", "Record Notes"),
    ("Written Up", "Written Up Status / Violations"),
]

def get_calendar_hidden_columns():
    prefs = load_ui_column_preferences()
    hidden_list = prefs.get("calendar_hidden_columns")
    if hidden_list is None:
        return {"Written Up", "Record ID"}
    return set(hidden_list)

def save_calendar_hidden_columns(hidden_set):
    prefs = load_ui_column_preferences()
    prefs["calendar_hidden_columns"] = list(hidden_set)
    save_ui_column_preferences(prefs)

CURRENT_SESSION_USER = "admin"
SUPABASE_SALT = b"\x80\xa7\xbf\xcc\xa3\x12\xcc\x81\xf2\x93\xb4\x37\x13\xc3\xb4\x3a"

def init_supabase_cipher():
    global CIPHER_SUITE, SALT
    if CIPHER_SUITE is None:
        SALT = SUPABASE_SALT
        CIPHER_SUITE = get_cipher(DEFAULT_ENCRYPTION_PASSWORD, SUPABASE_SALT)

def encrypt_val(val):
    """Non-deterministic Fernet encryption (prefix enc:)."""
    if val is None:
        return None
    if isinstance(val, str) and (val.startswith("enc:") or val.startswith("denc:")):
        return val
    init_supabase_cipher()
    val_str = str(val)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        val_str = f"num:{val}"
    else:
        val_str = f"str:{val}"
    return "enc:" + CIPHER_SUITE.encrypt(val_str.encode()).decode()


def _det_aes_key():
    init_supabase_cipher()
    return hashlib.sha256(DEFAULT_ENCRYPTION_PASSWORD.encode("utf-8") + SUPABASE_SALT).digest()


def encrypt_val_deterministic(val):
    """
    Deterministic AES-GCM encryption (prefix denc:).
    Same plaintext -> same ciphertext, so WHERE/UNIQUE lookups still work.
    """
    if val is None:
        return None
    if isinstance(val, str) and (val.startswith("enc:") or val.startswith("denc:")):
        return val
    try:
        from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    except ImportError:
        return encrypt_val(val)

    if isinstance(val, (int, float)) and not isinstance(val, bool):
        payload = f"num:{val}"
    else:
        payload = f"str:{val}"
    key = _det_aes_key()
    nonce = hashlib.sha256(key + payload.encode("utf-8")).digest()[:12]
    ct = AESGCM(key).encrypt(nonce, payload.encode("utf-8"), None)
    return "denc:" + base64.urlsafe_b64encode(nonce + ct).decode("ascii")


def plain_label(val):
    """Decrypt a stored text cell and return a stripped string."""
    if val is None:
        return ""
    v = decrypt_val(val)
    return str(v).strip() if v is not None else ""


INCOME_EXPENSE_CATEGORIES = {"Cash Envelope Received", "Employee Revenue"}
SALARY_EXPENSE_CATEGORIES = {"Salary Payment", "Salary Payment (Tip)"}
EXPENSE_EDIT_SELECT = (
    "SELECT expense_date, category, amount, employee_id, status, description, "
    "payment_type, location, assignee_id, document_path, is_tip, tip_given, cycle_key "
    "FROM expenses WHERE id=?"
)


def is_income_expense_category(cat):
    return plain_label(cat) in INCOME_EXPENSE_CATEGORIES


def is_envelope_category(cat):
    p = plain_label(cat)
    return p in ("Cash Envelope Received", "ظرف نقدي مستلم")


def normalize_iso_date(val):
    """Return YYYY-MM-DD or empty string."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    s2 = s.replace(",", " ").replace(".", "/")
    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%m-%d-%Y",
        "%b %d %Y",
        "%B %d %Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(s2[:32].strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def decrypt_val(val):
    if not isinstance(val, str):
        return val
    try:
        if val.startswith("denc:"):
            from cryptography.hazmat.primitives.ciphers.aead import AESGCM
            raw = base64.urlsafe_b64decode(val[5:].encode("ascii"))
            nonce, ct = raw[:12], raw[12:]
            plain = AESGCM(_det_aes_key()).decrypt(nonce, ct, None).decode("utf-8")
            if plain.startswith("num:"):
                s = plain[4:]
                return float(s) if "." in s else int(s)
            if plain.startswith("str:"):
                return plain[4:]
            return plain
        if val.startswith("enc:"):
            init_supabase_cipher()
            decrypted_bytes = CIPHER_SUITE.decrypt(val[4:].encode())
            decrypted_str = decrypted_bytes.decode()
            if decrypted_str.startswith("num:"):
                s = decrypted_str[4:]
                return float(s) if '.' in s else int(s)
            elif decrypted_str.startswith("str:"):
                return decrypted_str[4:]
    except Exception:
        pass
    return val


def to_float(val, default=0.0):
    """Coerce DB / decrypted values to float for payroll math (never return str)."""
    if val is None or val == "":
        return float(default)
    if isinstance(val, bool):
        return float(default)
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except (TypeError, ValueError):
            return float(default)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return float(default)
        # Decrypt ciphertext if a caller passed a raw encrypted cell
        if s.startswith("denc:") or s.startswith("enc:"):
            try:
                val = decrypt_val(s)
                return to_float(val, default)
            except Exception:
                return float(default)
        s = s.replace(",", "").replace("$", "").replace("%", "").strip()
        try:
            return float(s)
        except (TypeError, ValueError):
            return float(default)
    try:
        return float(val)
    except (TypeError, ValueError):
        return float(default)


def load_payout_tiers(kind="service"):
    """Return [(from_sales, to_sales, percent), ...] sorted by from_sales."""
    kind = (kind or "service").strip().lower() or "service"
    tiers = []
    try:
        conn = sqlite3.connect(TEMP_DB_PATH)
        cur = conn.cursor()
        try:
            cur.execute(
                "SELECT from_sales, to_sales, percentage FROM payout_tiers "
                "WHERE LOWER(COALESCE(kind, 'service')) = ? ORDER BY from_sales ASC",
                (kind,),
            )
        except Exception:
            if kind != "service":
                conn.close()
                return []
            cur.execute(
                "SELECT from_sales, to_sales, percentage FROM payout_tiers ORDER BY from_sales ASC"
            )
        for row in cur.fetchall() or []:
            a = to_float(row[0], 0.0)
            b = to_float(row[1], 9999999.0)
            p = to_float(row[2], 0.0)
            tiers.append((a, b, p))
        conn.close()
    except Exception:
        tiers = []
    if not tiers and kind == "service":
        tiers = [
            (0, 2500, 35),
            (2500.01, 3500, 38),
            (3500.01, 4500, 40),
            (4500.01, 6000, 45),
            (6000.01, 7500, 48),
            (7500.01, 9999999, 50),
        ]
    if not tiers and kind == "product":
        tiers = [
            (0, 149.99, 0),
            (150, 250, 15),
            (250.01, 9999999, 20),
        ]
    return tiers


def service_percent_for_sales(total_service):
    total = to_float(total_service, 0.0)
    chosen = 0.0
    for from_s, to_s, perc in load_payout_tiers("service"):
        if total >= from_s:
            chosen = perc / 100.0
            if total <= to_s:
                break
    return chosen


def product_percent_for_sales(total_prod):
    total = to_float(total_prod, 0.0)
    chosen = 0.0
    matched = False
    for from_s, to_s, perc in load_payout_tiers("product"):
        if total >= from_s:
            chosen = perc / 100.0
            matched = True
            if total <= to_s:
                break
    if matched:
        return chosen
    if total < 150.0:
        return 0.0
    if total <= 250.0:
        return 0.15
    return 0.20

# Deterministic = safe for WHERE / UNIQUE equality.
# Random Fernet = notes/free text only.
# Numeric money/rate fields are NOT encrypted (caused type errors across the UI).
DET_ENCRYPT_COLS = {
    'name', 'first_name', 'last_name', 'username', 'user_name', 'phone', 'email', 'ssn', 'address',
    'category', 'payment_type', 'location', 'status', 'written_up', 'pulled_date', 'is_tip',
    'title',
}
RAND_ENCRYPT_COLS = {
    'notes', 'written_up_desc', 'description', 'old_data', 'new_data', 'password',
    'summary', 'details',
}
ALL_ENCRYPT_COLS = DET_ENCRYPT_COLS | RAND_ENCRYPT_COLS

# Formerly encrypted numeric columns — kept as TEXT-compatible, stored as plaintext numbers.
PLAIN_NUMERIC_COLS = {
    "employees": ["hour_rate", "percentage"],
    "payroll_records": [
        "revenue", "hours", "calculation", "payment_amount", "product_sales",
        "tip", "service_addon_sales", "hour_rate", "percentage",
    ],
    "expenses": ["amount", "tip_given"],
}

# Real/double columns that must become TEXT before any legacy ciphertext can be rewritten.
NUMERIC_TO_TEXT = dict(PLAIN_NUMERIC_COLS)


def _encrypt_for_col(col, val):
    if val is None:
        return None
    col = (col or "").lower()
    if col in DET_ENCRYPT_COLS:
        return encrypt_val_deterministic(val)
    if col in RAND_ENCRYPT_COLS:
        # passwords are already sha256 hex — still wrap for at-rest privacy
        return encrypt_val(val)
    return val


def encrypt_params(query, params):
    """Encrypt bound params for Supabase inserts/updates/where equality filters."""
    if not params:
        return params
    if get_db_mode() != "supabase" or is_supabase_offline():
        return params

    import re
    param_list = list(params)
    query_upper = query.upper()

    def _encrypt_eq_placeholders(sql_chunk, start_param_idx=0):
        """Walk col/op/? placeholders in order; encrypt only equality on DET cols."""
        idx = start_param_idx
        for m in re.finditer(
            r"(?:(\w+)\.)?(\w+)\s*(=|!=|<>|>=|<=|>|<|LIKE)\s*\?",
            sql_chunk,
            flags=re.IGNORECASE,
        ):
            if idx >= len(param_list):
                break
            col = (m.group(2) or "").lower()
            op = (m.group(3) or "").strip()
            if op == "=" and col in DET_ENCRYPT_COLS:
                param_list[idx] = _encrypt_for_col(col, param_list[idx])
            idx += 1
        return idx

    if "INSERT INTO" in query_upper:
        start_idx = query.find('(')
        end_idx = query.find(')')
        if start_idx != -1 and end_idx != -1:
            cols_str = query[start_idx + 1:end_idx]
            cols = [c.strip().lower() for c in cols_str.split(',')]
            for i, col in enumerate(cols):
                if col in ALL_ENCRYPT_COLS and i < len(param_list):
                    param_list[i] = _encrypt_for_col(col, param_list[i])

    elif "UPDATE " in query_upper and " SET " in query_upper:
        set_idx = query_upper.find(' SET ')
        where_idx = query_upper.find(' WHERE ')
        if set_idx != -1:
            set_str = query[set_idx + 5: where_idx] if where_idx != -1 else query[set_idx + 5:]
            assignments = set_str.split(',')
            cols = [a.split('=')[0].strip().lower() for a in assignments if '=' in a]
            for i, col in enumerate(cols):
                if col in ALL_ENCRYPT_COLS and i < len(param_list):
                    param_list[i] = _encrypt_for_col(col, param_list[i])
            if where_idx != -1:
                _encrypt_eq_placeholders(query[where_idx:], start_param_idx=len(cols))

    else:
        # SELECT / DELETE: map each ? to its column in order.
        # Important: date filters use >= / <= — must not shift equality indexes.
        _encrypt_eq_placeholders(query)

    return tuple(param_list)

_SUPABASE_LOCK = threading.RLock()

class PostgresConnectionProxy:
    def __init__(self, pg_conn, shared=False):
        self.conn = pg_conn
        self._shared = shared

    def cursor(self):
        return PostgresCursorProxy(self.conn.cursor(), self.conn, shared=self._shared)

    def commit(self):
        with _SUPABASE_LOCK:
            try:
                self.conn.commit()
            except Exception as e:
                if "failed transaction block" in str(e).lower() or "aborted" in str(e).lower():
                    try:
                        self.conn.rollback()
                    except Exception:
                        pass
                raise

    def rollback(self):
        with _SUPABASE_LOCK:
            self.conn.rollback()

    def close(self):
        # Shared connection stays open for the app lifetime.
        if self._shared:
            with _SUPABASE_LOCK:
                try:
                    self.conn.commit()
                except Exception:
                    pass
            return
        with _SUPABASE_LOCK:
            self.conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self.conn.rollback()
        else:
            self.conn.commit()
        if not self._shared:
            self.conn.close()

class PostgresCursorProxy:
    def __init__(self, pg_cursor, pg_conn, shared=True):
        self.cursor = pg_cursor
        self.conn = pg_conn
        self._shared = shared
        self._lastrowid = None
        self._description = None
        self._history_table = None
        self._history_row_id = None
        self._history_action = None
        self._old_data_history = None
        self._savepoint_seq = 0
        self._pending_rows = None

    @property
    def description(self):
        return self._description or self.cursor.description

    @property
    def rowcount(self):
        return self.cursor.rowcount

    @property
    def lastrowid(self):
        return self._lastrowid

    def _fetch_row_json(self, table, row_id):
        try:
            sub_cur = self.conn.cursor()
            sub_cur.execute(f"SELECT * FROM {table} WHERE id = %s", (row_id,))
            row = sub_cur.fetchone()
            desc = sub_cur.description
            sub_cur.close()
            if row and desc:
                cols = [col[0] for col in desc]
                row_dict = dict(zip(cols, row))
                import json
                return json.dumps(row_dict)
        except Exception:
            pass
        return None

    def execute(self, query, params=None):
        with _SUPABASE_LOCK:
            try:
                return self._execute_unlocked(query, params)
            except Exception as e:
                if self._shared and (_is_dead_pg_error(e) or _is_connectivity_error(e)):
                    try:
                        get_shared_supabase_conn(force_reconnect=True)
                        proxy = get_shared_supabase_conn()
                        self.conn = proxy.conn
                        self.cursor = self.conn.cursor()
                        return self._execute_unlocked(query, params)
                    except Exception as e2:
                        if _is_connectivity_error(e2) or _is_connectivity_error(e):
                            enter_supabase_offline_mode(str(e2))
                        raise
                raise

    def _execute_unlocked(self, query, params=None):
        import re
        self._pending_rows = None
        
        # Track history details before mutation
        query_upper = query.upper()
        is_history_table = "database_history_log" in query.lower()
        self._history_table = None
        self._history_row_id = None
        self._history_action = None
        self._old_data_history = None
        
        if get_db_mode() == "supabase" and not is_history_table:
            # 1. DELETE FROM table WHERE id = ?
            m_del = re.match(r"DELETE\s+FROM\s+(\w+)\s+WHERE\s+id\s*=\s*\?", query, re.IGNORECASE)
            if m_del:
                self._history_table = m_del.group(1).lower()
                self._history_row_id = params[0] if params else None
                self._history_action = "DELETE"
                if self._history_row_id:
                    self._old_data_history = self._fetch_row_json(self._history_table, self._history_row_id)
            
            # 2. UPDATE table SET ... WHERE id = ?
            m_upd = re.match(r"UPDATE\s+(\w+)\s+SET.*WHERE\s+id\s*=\s*\?", query, re.IGNORECASE)
            if m_upd:
                self._history_table = m_upd.group(1).lower()
                self._history_row_id = params[-1] if params else None
                self._history_action = "UPDATE"
                if self._history_row_id:
                    self._old_data_history = self._fetch_row_json(self._history_table, self._history_row_id)
            
            # 3. INSERT INTO table
            if "INSERT INTO" in query_upper:
                tbl_match = re.search(r"INSERT\s+INTO\s+(\w+)", query, re.IGNORECASE)
                if tbl_match:
                    tbl_name = tbl_match.group(1).lower()
                    if tbl_name in ["employees", "payroll_records", "expenses", "shop_documents"]:
                        self._history_table = tbl_name
                        self._history_action = "INSERT"

        # Encrypt sensitive params
        encrypted_p = encrypt_params(query, params)
        translated_query = query.replace('?', '%s')
        
        if "PRAGMA table_info" in query:
            m = re.findall(r"PRAGMA\s+table_info\((\w+)\)", query, re.IGNORECASE)
            if m:
                tbl = m[0]
                sql = f"""
                SELECT 0 as cid, column_name as name, data_type as type, 
                    (is_nullable = 'NO')::integer as notnull, 
                    column_default as dflt_value, 
                    0 as pk
                FROM information_schema.columns 
                WHERE table_name = '{tbl}'
                ORDER BY ordinal_position
                """
                self.cursor.execute(sql)
                self._pending_rows = list(self.cursor.fetchall() or [])
                self._description = [
                    ('cid', None, None, None, None, None, None),
                    ('name', None, None, None, None, None, None),
                    ('type', None, None, None, None, None, None),
                    ('notnull', None, None, None, None, None, None),
                    ('dflt_value', None, None, None, None, None, None),
                    ('pk', None, None, None, None, None, None)
                ]
                return
                
        if "INSERT OR IGNORE" in query:
            translated_query = translated_query.replace("INSERT OR IGNORE INTO", "INSERT INTO")
            tbl_match = re.search(r"INSERT\s+INTO\s+(\w+)", translated_query, re.IGNORECASE)
            if tbl_match:
                tbl = tbl_match.group(1).lower()
                if tbl == "users":
                    translated_query += " ON CONFLICT (username) DO NOTHING"
                elif tbl in ["config_locations", "config_categories", "config_payments", "config_languages"]:
                    translated_query += " ON CONFLICT (name) DO NOTHING"

        if "AUTOINCREMENT" in translated_query.upper():
            translated_query = re.sub(
                r"\bINTEGER\s+PRIMARY\s+KEY\s+AUTOINCREMENT\b", 
                "SERIAL PRIMARY KEY", 
                translated_query, 
                flags=re.IGNORECASE
            )

        is_insert = translated_query.strip().upper().startswith("INSERT INTO")
        ret_val = None
        
        if is_insert and "RETURNING" not in translated_query.upper():
            words = translated_query.split()
            if len(words) > 2:
                tbl_name = words[2].lower().strip("()")
                if tbl_name not in ["users", "config_locations", "config_categories", "config_payments", "config_languages", "cash_month_locks"]:
                    translated_query += " RETURNING id"
                    ret_val = True

        # Savepoints for any mutating statement so one failure cannot abort
        # the shared connection (Postgres "commands ignored until end of transaction").
        q0 = translated_query.lstrip().upper()
        needs_savepoint = not (
            q0.startswith("SELECT")
            or q0.startswith("WITH")
            or q0.startswith("PRAGMA")
            or q0.startswith("CREATE")
            or q0.startswith("ALTER")
            or q0.startswith("DROP")
            or q0.startswith("SET")
            or q0.startswith("SHOW")
        )
        self._pending_rows = []

        def _buffer_results():
            try:
                rows = self.cursor.fetchall()
                self._pending_rows = list(rows) if rows is not None else []
            except Exception:
                self._pending_rows = []
            if ret_val and self._pending_rows:
                self._lastrowid = self._pending_rows[0][0]
                self._pending_rows = []

        if needs_savepoint:
            self._savepoint_seq += 1
            sp_name = f"sp_payroll_{self._savepoint_seq}"
            try:
                self.cursor.execute(f"SAVEPOINT {sp_name}")
                try:
                    self.cursor.execute(translated_query, encrypted_p or ())
                    _buffer_results()
                    self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}")
                except Exception as inner_e:
                    try:
                        self.cursor.execute(f"ROLLBACK TO SAVEPOINT {sp_name}")
                        self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}")
                    except Exception:
                        try:
                            self.conn.rollback()
                        except Exception:
                            pass

                    err_txt = str(inner_e)
                    needs_type_fix = (
                        "22P02" in err_txt
                        or "invalid input syntax for type real" in err_txt.lower()
                        or "invalid input syntax for type double" in err_txt.lower()
                        or "invalid input syntax for type numeric" in err_txt.lower()
                    ) and ("denc:" in err_txt or "enc:" in err_txt)

                    if needs_type_fix and get_db_mode() == "supabase" and not is_supabase_offline():
                        try:
                            ensure_numeric_columns_are_text()
                            self.cursor.execute(f"SAVEPOINT {sp_name}_r")
                            self.cursor.execute(translated_query, encrypted_p or ())
                            _buffer_results()
                            self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}_r")
                            # recovered after migrating column types
                        except Exception as retry_e:
                            try:
                                self.cursor.execute(f"ROLLBACK TO SAVEPOINT {sp_name}_r")
                                self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}_r")
                            except Exception:
                                try:
                                    self.conn.rollback()
                                except Exception:
                                    pass
                            # Last resort: write plaintext numbers if TEXT migration is blocked
                            try:
                                self.cursor.execute(f"SAVEPOINT {sp_name}_p")
                                self.cursor.execute(translated_query, params or ())
                                _buffer_results()
                                self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}_p")
                            except Exception:
                                try:
                                    self.cursor.execute(f"ROLLBACK TO SAVEPOINT {sp_name}_p")
                                    self.cursor.execute(f"RELEASE SAVEPOINT {sp_name}_p")
                                except Exception:
                                    try:
                                        self.conn.rollback()
                                    except Exception:
                                        pass
                                if _is_connectivity_error(inner_e):
                                    pass
                                raise inner_e
                    else:
                        raise
            finally:
                pass
        else:
            try:
                self.cursor.execute(translated_query, encrypted_p or ())
                # Always buffer under the lock so another thread cannot interleave
                # between execute() and fetchall() on the shared socket.
                _buffer_results()
                try:
                    self._description = self.cursor.description
                except Exception:
                    pass
            except Exception as inner_e:
                raise
        if needs_savepoint:
            self._description = None

        # Log change history (skipped during bulk init for speed)
        if (
            SUPABASE_HISTORY_ENABLED
            and get_db_mode() == "supabase"
            and self._history_table
        ):
            try:
                new_data = None
                row_id_to_log = self._history_row_id or self._lastrowid
                if row_id_to_log:
                    new_data = self._fetch_row_json(self._history_table, row_id_to_log)
                
                sub_cur = self.conn.cursor()
                sql = """
                INSERT INTO database_history_log (change_timestamp, user_name, table_name, record_id, action_type, old_data, new_data)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                from datetime import datetime
                ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                sub_cur.execute(sql, (ts, CURRENT_SESSION_USER, self._history_table, row_id_to_log, self._history_action, self._old_data_history, new_data))
                sub_cur.close()
            except Exception:
                pass

    def executemany(self, query, seq_of_params):
        for params in seq_of_params:
            self.execute(query, params)

    def fetchone(self):
        with _SUPABASE_LOCK:
            if self._pending_rows is not None:
                if not self._pending_rows:
                    return None
                res = self._pending_rows.pop(0)
            else:
                try:
                    res = self.cursor.fetchone()
                except Exception as e:
                    if _is_dead_pg_error(e):
                        try:
                            get_shared_supabase_conn(force_reconnect=True)
                        except Exception:
                            pass
                    return None
            if res:
                res_list = list(res)
                for idx, val in enumerate(res_list):
                    # Fast path: only decrypt ciphertext cells
                    if isinstance(val, str) and (val.startswith("denc:") or val.startswith("enc:")):
                        res_list[idx] = decrypt_val(val)
                return tuple(res_list)
            return None

    def fetchall(self):
        with _SUPABASE_LOCK:
            if self._pending_rows is not None:
                rows = self._pending_rows
                self._pending_rows = []
            else:
                try:
                    rows = self.cursor.fetchall() or []
                except Exception as e:
                    if _is_dead_pg_error(e):
                        try:
                            get_shared_supabase_conn(force_reconnect=True)
                        except Exception:
                            pass
                    rows = []
            decrypted_rows = []
            for row in rows:
                row_list = list(row)
                for idx, val in enumerate(row_list):
                    if isinstance(val, str) and (val.startswith("denc:") or val.startswith("enc:")):
                        row_list[idx] = decrypt_val(val)
                decrypted_rows.append(tuple(row_list))
            return decrypted_rows
if not hasattr(sqlite3, "_real_connect"):
    sqlite3._real_connect = sqlite3.connect
_original_sqlite3_connect = sqlite3._real_connect
_SUPABASE_PG_CONN = None
SUPABASE_HISTORY_ENABLED = True

def _is_dead_pg_error(exc):
    msg = str(exc).lower()
    return (
        ("nonetype" in msg and "bytearray" in msg)
        or ("nonetype" in msg and "write" in msg)
        or ("nonetype" in msg and "read" in msg)
        or "connection is closed" in msg
        or "server closed the connection" in msg
        or "connection already closed" in msg
        or "broken pipe" in msg
        or "sock.read" in msg
        or "interfaceerror" in msg
        or "operationalerror" in msg
        or "connection reset" in msg
        or "timed out" in msg
        or "timeout" in msg
    )

def _is_pg_conn_alive(conn):
    if conn is None:
        return False
    try:
        c = getattr(conn, "_c", None)
        if c is None or getattr(c, "_sock", None) is None:
            return False
        cur = conn.cursor()
        cur.execute("SELECT 1")
        cur.fetchone()
        cur.close()
        conn.commit()
        return True
    except Exception:
        return False

def _clean_supabase_config(config):
    """Auto-clean and parse host, full URI, port, user, and database."""
    cfg = dict(config or {})
    host = str(cfg.get("supabase_host") or "").strip()
    port = str(cfg.get("supabase_port") or "5432").strip()
    user = str(cfg.get("supabase_user") or "postgres").strip()
    password = str(cfg.get("supabase_password") or "").strip()
    database = str(cfg.get("supabase_database") or "postgres").strip()

    # Handle full connection URI if pasted in host
    if host.startswith("postgres://") or host.startswith("postgresql://"):
        try:
            import urllib.parse
            u = urllib.parse.urlparse(host)
            if u.hostname:
                host = u.hostname
            if u.port:
                port = str(u.port)
            if u.username:
                user = urllib.parse.unquote(u.username)
            if u.password:
                password = urllib.parse.unquote(u.password)
            if u.path and len(u.path) > 1:
                database = u.path.lstrip("/")
        except Exception:
            pass

    # Strip http:// or https://
    if host.startswith("http://"):
        host = host[7:]
    elif host.startswith("https://"):
        host = host[8:]

    # Strip trailing path / slashes
    if "/" in host:
        parts = host.split("/", 1)
        host = parts[0]
        if not database or database == "postgres":
            database = parts[1].strip() or "postgres"

    # Strip trailing port in host (e.g. host:5432 or host:6543)
    if ":" in host:
        h_parts = host.split(":")
        host = h_parts[0]
        port = h_parts[1]

    # Auto-fix missing db. prefix for Supabase database hosts
    if host.endswith(".supabase.co") and not host.startswith("db."):
        host = "db." + host
    elif "." not in host and len(host) >= 12 and not host.startswith("db."):
        # Bare project reference pasted
        host = f"db.{host}.supabase.co"

    # Clean port
    try:
        port = int(port)
    except Exception:
        port = 5432

    cfg["supabase_host"] = host
    cfg["supabase_port"] = port
    cfg["supabase_user"] = user
    cfg["supabase_password"] = password
    cfg["supabase_database"] = database
    return cfg

def _open_supabase_pg_conn(timeout=15):
    import pg8000.dbapi
    raw_config = get_db_config()
    config = _clean_supabase_config(raw_config)
    host = config.get("supabase_host")
    port = int(config.get("supabase_port", 5432))
    user = config.get("supabase_user", "postgres")
    password = config.get("supabase_password")
    database = config.get("supabase_database", "postgres")

    try:
        conn = pg8000.dbapi.connect(
            host=host,
            port=port,
            user=user,
            password=password,
            database=database,
            timeout=timeout,
        )
        try:
            conn.commit()
        except Exception:
            pass
        return conn
    except Exception as first_err:
        err_msg = str(first_err).lower()
        is_timeout_or_network = any(k in err_msg for k in ["timeout", "handshake", "timed out", "source_address is none", "connection refused", "network unreachable", "10060", "10061"])

        # Extract project reference
        ref = None
        if host:
            clean_h = host.replace("db.", "").strip()
            if ".supabase.co" in clean_h:
                ref = clean_h.split(".supabase.co")[0].strip()
        if not ref and user and "." in user:
            ref = user.split(".", 1)[1].strip()

        if is_timeout_or_network and ref:
            # Try connection pooler (IPv4 compatible) across candidate regions
            candidate_regions = ["us-east-2", "us-east-1", "us-west-1", "us-west-2", "eu-central-1", "eu-west-1", "ap-southeast-1"]
            pooler_user = f"postgres.{ref}"
            for reg in candidate_regions:
                pooler_host = f"aws-0-{reg}.pooler.supabase.com"
                for pooler_port in [6543, 5432]:
                    try:
                        conn = pg8000.dbapi.connect(
                            host=pooler_host,
                            port=pooler_port,
                            user=pooler_user,
                            password=password,
                            database=database,
                            timeout=min(timeout, 8),
                        )
                        try:
                            conn.commit()
                        except Exception:
                            pass
                        # Auto-update local config with working IPv4 pooler
                        try:
                            cfg_path = os.path.join(get_default_app_dir(), "location_config.json")
                            if os.path.exists(cfg_path):
                                with open(cfg_path, "r", encoding="utf-8") as f:
                                    cur_cfg = json.load(f)
                                cur_cfg["supabase_host"] = pooler_host
                                cur_cfg["supabase_port"] = str(pooler_port)
                                cur_cfg["supabase_user"] = pooler_user
                                with open(cfg_path, "w", encoding="utf-8") as f:
                                    json.dump(cur_cfg, f, indent=4)
                        except Exception:
                            pass
                        return conn
                    except Exception as pooler_err:
                        p_msg = str(pooler_err).lower()
                        if "password authentication failed" in p_msg:
                            raise pooler_err
                        continue
        raise first_err

def get_shared_supabase_conn(force_reconnect=False, timeout=15):
    """Reuse one Postgres connection for the whole app session, auto-reconnecting if dead."""
    global _SUPABASE_PG_CONN
    if force_reconnect and _SUPABASE_PG_CONN is not None:
        with _SUPABASE_LOCK:
            if _SUPABASE_PG_CONN is not None:
                try:
                    _SUPABASE_PG_CONN.close()
                except Exception:
                    pass
                _SUPABASE_PG_CONN = None
    if _SUPABASE_PG_CONN is None or not _is_pg_conn_alive(_SUPABASE_PG_CONN):
        new_conn = _open_supabase_pg_conn(timeout=timeout)
        with _SUPABASE_LOCK:
            if _SUPABASE_PG_CONN is not None:
                try:
                    _SUPABASE_PG_CONN.close()
                except Exception:
                    pass
            _SUPABASE_PG_CONN = new_conn
    return PostgresConnectionProxy(_SUPABASE_PG_CONN, shared=True)

def close_shared_supabase_conn():
    global _SUPABASE_PG_CONN
    with _SUPABASE_LOCK:
        if _SUPABASE_PG_CONN is not None:
            try:
                _SUPABASE_PG_CONN.close()
            except Exception:
                pass
            _SUPABASE_PG_CONN = None


# --- Offline / reconnect support for Supabase mode ---
# When the cloud is unreachable, the app keeps working against a local encrypted
# cache and queues mutations. On reconnect, the queue is flushed to Postgres.
_SUPABASE_OFFLINE = False
OFFLINE_TEMP_DB_PATH = None
_OFFLINE_CACHE_SALT = None
_OFFLINE_CACHE_CIPHER = None
CLOUD_CACHE_FILE = os.path.join(get_app_dir(), "payroll_cloud_cache.enc")
OFFLINE_SYNC_TABLES = (
    "users",
    "employees",
    "config_locations",
    "config_categories",
    "config_payments",
    "payroll_records",
    "expenses",
    "shop_documents",
    "vagaro_pull_logs",
    "payout_tiers",
    "cash_month_locks",
    "user_action_log",
    "database_history_log",
)
_LOCAL_FIRST = False
_SYNC_IN_PROGRESS = False
_SYNC_LOCK = threading.Lock()
_persist_timer = None
_push_timer = None
_persist_lock = threading.Lock()
_LAST_SYNC_ERROR = ""
_LOCAL_PROTECTED_EXPENSE_IDS = set()
_SKIP_EXPENSE_PULL_UNTIL = 0.0


def save_local_daily_snapshot():
    """Create a local encrypted backup snapshot for today so records are always preserved on disk."""
    try:
        app_dir = get_app_dir()
        backups_dir = os.path.join(app_dir, "local_daily_backups")
        os.makedirs(backups_dir, exist_ok=True)
        today_str = datetime.now().strftime("%Y-%m-%d")
        daily_backup_file = os.path.join(backups_dir, f"payroll_backup_{today_str}.enc")

        if os.path.exists(CLOUD_CACHE_FILE):
            import shutil
            shutil.copy2(CLOUD_CACHE_FILE, daily_backup_file)

        # Keep latest 60 daily backups on disk
        existing = sorted(
            [f for f in os.listdir(backups_dir) if f.startswith("payroll_backup_") and f.endswith(".enc")]
        )
        if len(existing) > 60:
            for old_f in existing[:-60]:
                try:
                    os.remove(os.path.join(backups_dir, old_f))
                except Exception:
                    pass
    except Exception:
        pass


def protect_local_expense_id(row_id):
    """Keep a just-saved expense/envelope from being wiped by a cloud pull."""
    global _SKIP_EXPENSE_PULL_UNTIL
    try:
        if row_id is not None:
            _LOCAL_PROTECTED_EXPENSE_IDS.add(int(row_id))
    except Exception:
        pass
    _SKIP_EXPENSE_PULL_UNTIL = time.time() + 60.0


def is_supabase_offline():
    return get_db_mode() == "supabase" and _SUPABASE_OFFLINE


def _is_connectivity_error(exc):
    if _is_dead_pg_error(exc):
        return True
    msg = str(exc).lower()
    needles = (
        "failed to connect",
        "could not connect",
        "can't create a connection",
        "timed out",
        "timeout",
        "handshake",
        "source_address is none",
        "network is unreachable",
        "unreachable",
        "no route to host",
        "connection refused",
        "connection reset",
        "broken pipe",
        "server closed the connection",
        "getaddrinfo",
        "name or service not known",
        "temporarily unavailable",
        "ssl syscall",
        "eof occurred in violation",
        "nodename nor servname",
        "can't connect",
        "connection aborted",
        "winsock",
    )
    return any(n in msg for n in needles)


def _sync_log(msg):
    try:
        path = os.path.join(get_app_dir(), "sync_log.txt")
        with open(path, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {msg}\n")
    except Exception:
        pass


def _persist_offline_cache():
    """Encrypt the offline working SQLite file to disk and update daily local snapshot."""
    global _OFFLINE_CACHE_CIPHER, _OFFLINE_CACHE_SALT
    path = OFFLINE_TEMP_DB_PATH
    if not path or not os.path.exists(path):
        return
    if _OFFLINE_CACHE_CIPHER is None or _OFFLINE_CACHE_SALT is None:
        return
    try:
        try:
            ck = _original_sqlite3_connect(path, timeout=5)
            try:
                ck.execute("PRAGMA wal_checkpoint(PASSIVE)")
            finally:
                ck.close()
        except Exception:
            pass
        with open(path, "rb") as f:
            data = f.read()
        encrypted = _OFFLINE_CACHE_CIPHER.encrypt(data)
        with open(CLOUD_CACHE_FILE, "wb") as f:
            f.write(_OFFLINE_CACHE_SALT + encrypted)
        # Save timestamped daily snapshot on disk
        save_local_daily_snapshot()
    except Exception:
        pass


def _schedule_persist_offline_cache(delay=1.25):
    """Encrypt the cache in the background so Save stays instant."""
    global _persist_timer
    def _run():
        global _persist_timer
        with _persist_lock:
            _persist_timer = None
        _persist_offline_cache()
    with _persist_lock:
        if _persist_timer is not None:
            try:
                _persist_timer.cancel()
            except Exception:
                pass
        _persist_timer = threading.Timer(delay, _run)
        _persist_timer.daemon = True
        _persist_timer.start()


def schedule_cloud_push(delay=0.08):
    """Upload queued local edits without pulling (keeps the UI snappy)."""
    global _push_timer
    if get_db_mode() != "supabase":
        return
    def _run():
        global _push_timer, _LAST_SYNC_ERROR
        _push_timer = None
        if get_db_mode() != "supabase":
            return
        acquired = _SYNC_LOCK.acquire(blocking=False)
        if not acquired:
            if offline_pending_count():
                schedule_cloud_push(2.0)
            return
        try:
            ok, msg, n = flush_offline_queue_to_cloud()
            if not ok:
                _LAST_SYNC_ERROR = msg or "upload failed"
                _sync_log(f"push failed: {msg}")
                if offline_pending_count():
                    schedule_cloud_push(5.0)
            else:
                if n:
                    _sync_log(f"pushed {n} change(s)")
                if not offline_pending_count():
                    _LAST_SYNC_ERROR = ""
        except Exception as e:
            _LAST_SYNC_ERROR = str(e)
            _sync_log(f"push exception: {e}")
            if offline_pending_count():
                schedule_cloud_push(5.0)
        finally:
            try:
                _SYNC_LOCK.release()
            except Exception:
                pass
    try:
        if _push_timer is not None:
            _push_timer.cancel()
    except Exception:
        pass
    _push_timer = threading.Timer(delay, _run)
    _push_timer.daemon = True
    _push_timer.start()


def cloud_sync_status_label(ok=True):
    """Short status for the live-sync label."""
    pending = offline_pending_count()
    err = (_LAST_SYNC_ERROR or "").strip()
    if pending:
        extra = f" — {err[:24]}…" if err else ""
        return f"☁️ {pending} change(s) uploading{extra}"
    if err:
        short = err if len(err) < 48 else err[:45] + "…"
        return f"☁️ Sync issue — {short}"
    return f"☁️ Synced {datetime.now().strftime('%H:%M:%S')}"


def _close_offline_temp(remove_file=True):
    global OFFLINE_TEMP_DB_PATH
    path = OFFLINE_TEMP_DB_PATH
    OFFLINE_TEMP_DB_PATH = None
    if remove_file and path and os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass


def ensure_offline_cache_open():
    """Decrypt or create the local cloud-cache SQLite and return its temp path."""
    global OFFLINE_TEMP_DB_PATH, _OFFLINE_CACHE_SALT, _OFFLINE_CACHE_CIPHER
    if OFFLINE_TEMP_DB_PATH and os.path.exists(OFFLINE_TEMP_DB_PATH):
        return OFFLINE_TEMP_DB_PATH

    password = DEFAULT_ENCRYPTION_PASSWORD
    decrypted = b""
    if os.path.exists(CLOUD_CACHE_FILE):
        try:
            with open(CLOUD_CACHE_FILE, "rb") as f:
                content = f.read()
            if len(content) >= 17:
                _OFFLINE_CACHE_SALT = content[:16]
                _OFFLINE_CACHE_CIPHER = get_cipher(password, _OFFLINE_CACHE_SALT)
                decrypted = _OFFLINE_CACHE_CIPHER.decrypt(content[16:])
        except Exception:
            decrypted = b""
            _OFFLINE_CACHE_SALT = None
            _OFFLINE_CACHE_CIPHER = None

    if _OFFLINE_CACHE_SALT is None or _OFFLINE_CACHE_CIPHER is None:
        _OFFLINE_CACHE_SALT = os.urandom(16)
        _OFFLINE_CACHE_CIPHER = get_cipher(password, _OFFLINE_CACHE_SALT)

    fd, path = tempfile.mkstemp(suffix="_cloud_cache.db")
    os.close(fd)
    if decrypted:
        with open(path, "wb") as f:
            f.write(decrypted)

    OFFLINE_TEMP_DB_PATH = path
    lite = _original_sqlite3_connect(path, timeout=15)
    try:
        lite.execute("PRAGMA journal_mode=WAL")
        lite.execute("PRAGMA synchronous=NORMAL")
        lite.execute("PRAGMA busy_timeout=8000")
        cur = lite.cursor()
        # Build local (plaintext) schema — temporarily pretend we're not on live PG.
        _init_offline_schema(cur)
        lite.commit()
    finally:
        lite.close()
    return path


def _init_offline_schema(cursor):
    """Schema for the offline cache (plaintext SQLite, REAL money columns)."""
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS offline_sync_queue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payload TEXT NOT NULL,
            created_at TEXT
        )
        """
    )
    # Reuse app schema builder with local-style types by forcing non-supabase branch.
    # _init_db_schema checks get_db_mode(); while offline flag is set it should use REAL.
    _init_db_schema(cursor, seed=False)


def enter_supabase_offline_mode(reason=""):
    """Switch working DB from Postgres to the local encrypted cache."""
    global _SUPABASE_OFFLINE, TEMP_DB_PATH
    if get_db_mode() != "supabase":
        return
    close_shared_supabase_conn()
    path = ensure_offline_cache_open()
    _SUPABASE_OFFLINE = True
    TEMP_DB_PATH = path
    try:
        print(f"[offline] Working locally. {reason}".strip())
    except Exception:
        pass


def leave_supabase_offline_mode():
    """Switch working DB back to the shared Postgres connection."""
    global _SUPABASE_OFFLINE, TEMP_DB_PATH
    _persist_offline_cache()
    _SUPABASE_OFFLINE = False
    if _LOCAL_FIRST:
        # Stay on the local cache for UI speed; cloud is only used by background sync.
        path = ensure_offline_cache_open()
        TEMP_DB_PATH = path
        return
    TEMP_DB_PATH = SUPABASE_DB_SENTINEL


def using_local_cache():
    return bool(OFFLINE_TEMP_DB_PATH and TEMP_DB_PATH == OFFLINE_TEMP_DB_PATH)


def enable_local_first_mode():
    """Use the local SQLite cache for all UI reads/writes; cloud syncs in the background."""
    global TEMP_DB_PATH, _LOCAL_FIRST
    _LOCAL_FIRST = True
    path = ensure_offline_cache_open()
    TEMP_DB_PATH = path
    return path


def sync_local_cache_with_cloud(progress_cb=None, backfill=False, init_schema=False):
    """Flush local edits to Supabase, then download the latest copy into the local cache."""
    global _SYNC_IN_PROGRESS, _SUPABASE_OFFLINE, _LAST_SYNC_ERROR
    if get_db_mode() != "supabase":
        if progress_cb:
            progress_cb("Ready")
        return True, "local"
    if not _SYNC_LOCK.acquire(blocking=False):
        return False, "busy"
    _SYNC_IN_PROGRESS = True
    try:
        enable_local_first_mode()
        if progress_cb:
            progress_cb("Connecting to cloud…")
        try:
            get_shared_supabase_conn(force_reconnect=False)
        except Exception:
            get_shared_supabase_conn(force_reconnect=True)
        _SUPABASE_OFFLINE = False
        if init_schema:
            try:
                db_conn = _open_supabase_pg_conn(timeout=10)
                try:
                    ensure_all_supabase_tables(db_conn)
                finally:
                    try:
                        db_conn.close()
                    except Exception:
                        pass
            except Exception:
                pass
        if progress_cb:
            progress_cb("Uploading local changes…")
        ok, msg, _flushed = flush_offline_queue_to_cloud()
        if not ok:
            if _is_connectivity_error(Exception(msg or "")):
                enter_supabase_offline_mode(msg)
                enable_local_first_mode()
                return False, msg
            _LAST_SYNC_ERROR = msg or "upload failed"
        if backfill:
            try:
                if progress_cb:
                    progress_cb("Checking for unsynced local records…")
                if backfill_local_rows_missing_from_cloud():
                    flush_offline_queue_to_cloud()
            except Exception:
                pass
        if progress_cb:
            progress_cb("Downloading latest records…")
        refresh_offline_cache_from_cloud()
        if offline_pending_count():
            flush_offline_queue_to_cloud()
        if progress_cb:
            progress_cb("Almost ready…")
        enable_local_first_mode()
        return True, "ok"
    except Exception as e:
        enter_supabase_offline_mode(str(e))
        enable_local_first_mode()
        return False, str(e)
    finally:
        _SYNC_IN_PROGRESS = False
        try:
            _SYNC_LOCK.release()
        except Exception:
            pass


def _queue_offline_op(payload, raw_conn=None):
    """Queue a mutation for Supabase. Use raw_conn (same SQLite connection) when
    still inside an open write transaction — a second connection would deadlock."""
    own = raw_conn is None
    if own:
        path = ensure_offline_cache_open()
        raw_conn = _original_sqlite3_connect(path, timeout=15)
        try:
            raw_conn.execute("PRAGMA busy_timeout=8000")
        except Exception:
            pass
    try:
        cur = raw_conn.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS offline_sync_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, payload TEXT NOT NULL, created_at TEXT)"
        )
        cur.execute(
            "INSERT INTO offline_sync_queue (payload, created_at) VALUES (?, ?)",
            (
                json.dumps(payload, default=str),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        if own:
            raw_conn.commit()
    finally:
        if own:
            try:
                raw_conn.close()
            except Exception:
                pass


def offline_pending_count():
    if not is_supabase_offline() and get_db_mode() != "supabase":
        return 0
    try:
        path = ensure_offline_cache_open()
        conn = _original_sqlite3_connect(path)
        try:
            cur = conn.cursor()
            cur.execute(
                "CREATE TABLE IF NOT EXISTS offline_sync_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, payload TEXT NOT NULL, created_at TEXT)"
            )
            cur.execute("SELECT COUNT(*) FROM offline_sync_queue")
            row = cur.fetchone()
            return int(row[0] or 0) if row else 0
        finally:
            conn.close()
    except Exception:
        return 0


_ACTION_LOG_SKIP_TABLES = {
    "offline_sync_queue",
    "sqlite_sequence",
    "user_action_log",
    "database_history_log",
    "cloud_backups",
}
_LAST_BACKUP_SLOT = None
CLOUD_BACKUP_KEEP_DAYS = 30
BACKUP_DUMP_TABLES = (
    "users",
    "employees",
    "config_locations",
    "config_categories",
    "config_payments",
    "payroll_records",
    "expenses",
    "shop_documents",
    "vagaro_pull_logs",
    "payout_tiers",
    "cash_month_locks",
    "user_action_log",
    "database_history_log",
)


def _session_user_name():
    try:
        return plain_label(CURRENT_SESSION_USER) or "unknown"
    except Exception:
        return "unknown"


def _friendly_user_action(action, table=None, row=None, record_id=None):
    row = row or {}
    t = str(table or "").lower()
    op = str(action or "").lower()
    cat = plain_label(row.get("category") or "")
    name = plain_label(row.get("name") or row.get("username") or "")
    amt = row.get("amount")
    if amt is None:
        amt = row.get("revenue")
    money = ""
    try:
        if amt is not None and str(amt) != "":
            money = f" ${to_float(amt, 0.0):,.2f}"
    except Exception:
        money = ""
    verbs = {"insert": "Added", "update": "Updated", "delete": "Deleted", "upsert": "Updated"}
    verb = verbs.get(op, op.replace("_", " ").title())
    if op == "login":
        return "Logged in"
    if op == "backup":
        extra = plain_label(row.get("summary") or "")
        return extra or "Saved cloud backup"
    if op == "rename_user":
        return f"Changed username to {name or record_id or ''}".strip()
    if t == "expenses":
        label = cat or "expense"
        return f"{verb} {label}{money}".strip()
    if t == "payroll_records":
        return f"{verb} payroll record{money}".strip()
    if t == "employees":
        return f"{verb} employee {name}".strip()
    if t == "users":
        return f"{verb} login user {name or record_id or ''}".strip()
    if t == "shop_documents":
        title = plain_label(row.get("title") or "")
        return f"{verb} shop document {title}".strip()
    if t.startswith("config_"):
        return f"{verb} setting {name or record_id or t}".strip()
    if t:
        return f"{verb} {t.replace('_', ' ')}".strip()
    return verb or "Action"


def log_user_action(
    action,
    table=None,
    record_id=None,
    row=None,
    extra_summary=None,
    raw_conn=None,
):
    """Record one user action locally and queue it for Supabase."""
    table_l = str(table or "").lower()
    if table_l in _ACTION_LOG_SKIP_TABLES:
        return
    row = dict(row or {})
    summary = extra_summary or _friendly_user_action(action, table, row, record_id)
    details = ""
    try:
        details = json.dumps(row, default=str)[:2500]
    except Exception:
        details = str(row)[:2500]
    uid = str(uuid.uuid4())
    rec = {
        "log_uid": uid,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "user_name": _session_user_name(),
        "action": str(action or ""),
        "table_name": table_l,
        "record_id": "" if record_id is None else str(record_id),
        "summary": summary,
        "details": details,
    }
    own = raw_conn is None
    if own:
        try:
            if get_db_mode() == "supabase":
                path = ensure_offline_cache_open()
            else:
                path = TEMP_DB_PATH
            if not path or path == SUPABASE_DB_SENTINEL:
                return
            raw_conn = _original_sqlite3_connect(path, timeout=15)
            try:
                raw_conn.execute("PRAGMA busy_timeout=8000")
            except Exception:
                pass
        except Exception:
            return
    try:
        cur = raw_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_action_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_uid TEXT UNIQUE,
                created_at TEXT,
                user_name TEXT,
                action TEXT,
                table_name TEXT,
                record_id TEXT,
                summary TEXT,
                details TEXT
            )
            """
        )
        cur.execute(
            """
            INSERT INTO user_action_log
                (log_uid, created_at, user_name, action, table_name, record_id, summary, details)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                rec["log_uid"],
                rec["created_at"],
                rec["user_name"],
                rec["action"],
                rec["table_name"],
                rec["record_id"],
                rec["summary"],
                rec["details"],
            ),
        )
        if get_db_mode() == "supabase":
            _queue_offline_op({"op": "insert", "table": "user_action_log", "row": rec}, raw_conn)
        if own:
            raw_conn.commit()
            if get_db_mode() == "supabase":
                try:
                    schedule_cloud_push(0.2)
                except Exception:
                    pass
    except Exception:
        if own:
            try:
                raw_conn.rollback()
            except Exception:
                pass
    finally:
        if own:
            try:
                raw_conn.close()
            except Exception:
                pass


def _ensure_audit_backup_schema(pg_cur):
    try:
        pg_cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_action_log (
                id SERIAL PRIMARY KEY,
                log_uid TEXT UNIQUE,
                created_at TEXT,
                user_name TEXT,
                action TEXT,
                table_name TEXT,
                record_id TEXT,
                summary TEXT,
                details TEXT
            )
            """
        )
    except Exception:
        pass
    try:
        pg_cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cloud_backups (
                id SERIAL PRIMARY KEY,
                slot_key TEXT UNIQUE,
                backup_date TEXT,
                slot TEXT,
                created_at TEXT,
                created_by TEXT,
                payload TEXT,
                size_bytes INTEGER
            )
            """
        )
    except Exception:
        pass


def current_cloud_backup_slot():
    now = datetime.now()
    slot = "am" if now.hour < 12 else "pm"
    day = now.strftime("%Y-%m-%d")
    return day, slot, f"{day}_{slot}"


def _build_cloud_backup_payload():
    path = ensure_offline_cache_open() if get_db_mode() == "supabase" else TEMP_DB_PATH
    if not path or path == SUPABASE_DB_SENTINEL:
        raise RuntimeError("No local database to back up")
    conn = _original_sqlite3_connect(path, timeout=20)
    snapshot = {
        "version": 1,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "created_by": _session_user_name(),
        "tables": {},
    }
    try:
        cur = conn.cursor()
        for tbl in BACKUP_DUMP_TABLES:
            try:
                cur.execute(f"PRAGMA table_info({tbl})")
                cols = [r[1] for r in (cur.fetchall() or []) if r and r[1]]
                if not cols:
                    continue
                cur.execute(f"SELECT * FROM {tbl}")
                rows = []
                for row in cur.fetchall() or []:
                    rec = {}
                    for i, c in enumerate(cols):
                        val = row[i] if i < len(row) else None
                        if isinstance(val, bytes):
                            val = {"__bytes__": base64.b64encode(val).decode("ascii")}
                        rec[c] = val
                    rows.append(rec)
                snapshot["tables"][tbl] = {"columns": cols, "rows": rows}
            except Exception:
                continue
    finally:
        conn.close()
    raw = json.dumps(snapshot, default=str).encode("utf-8")
    gz = gzip.compress(raw)
    token = base64.b64encode(gz).decode("ascii")
    return encrypt_val(token), len(gz)


def _decode_cloud_backup_payload(payload):
    if payload is None:
        return None
    try:
        token = decrypt_val(payload) if isinstance(payload, str) else payload
        if isinstance(token, dict):
            return token
        raw = token
        if isinstance(raw, str):
            try:
                gz = base64.b64decode(raw.encode("ascii"))
                data = gzip.decompress(gz)
                return json.loads(data.decode("utf-8"))
            except Exception:
                try:
                    return json.loads(raw)
                except Exception:
                    pass
        elif isinstance(raw, (bytes, bytearray)):
            try:
                data = gzip.decompress(raw)
                return json.loads(data.decode("utf-8"))
            except Exception:
                try:
                    return json.loads(raw.decode("utf-8"))
                except Exception:
                    pass
    except Exception:
        pass
    return None
def get_device_identifier():
    import platform
    import socket
    try:
        host = socket.gethostname() or platform.node() or "Device"
    except Exception:
        host = "Device"
    user = os.environ.get("USERNAME") or os.environ.get("USER") or "user"
    return f"{user}@{host}"


def get_local_backups_dir():
    d = os.path.join(get_default_app_dir(), "Local_Backups")
    os.makedirs(d, exist_ok=True)
    return d


def create_local_backup(slot_key=None, slot=None, backup_date=None):
    """Saves an encrypted snapshot backup on this machine's local disk."""
    try:
        b_dir = get_local_backups_dir()
        if not slot_key:
            backup_date, slot, slot_key = current_cloud_backup_slot()
        
        # 1. Copy encrypted database if available
        enc_file = os.path.join(get_default_app_dir(), "payroll_data.enc")
        if os.path.isfile(enc_file):
            import shutil
            dest_enc = os.path.join(b_dir, f"backup_{slot_key}.enc")
            try:
                shutil.copy2(enc_file, dest_enc)
            except Exception:
                pass

        # 2. Save encrypted JSON snapshot
        try:
            payload, size = _build_cloud_backup_payload()
            json_file = os.path.join(b_dir, f"snapshot_{slot_key}.json.gz")
            with open(json_file, "w", encoding="utf-8") as f:
                f.write(payload)
        except Exception:
            pass

        # Rotate: keep newest 40 local backups
        try:
            files = sorted(
                [os.path.join(b_dir, f) for f in os.listdir(b_dir) if f.startswith("backup_") or f.startswith("snapshot_")],
                key=os.path.getmtime,
            )
            if len(files) > 40:
                for old_f in files[:-40]:
                    try:
                        os.remove(old_f)
                    except Exception:
                        pass
        except Exception:
            pass
        return True, slot_key
    except Exception as e:
        return False, str(e)


def create_cloud_backup(slot_key=None, slot=None, backup_date=None, kind="auto"):
    """Save a snapshot locally on this PC AND upload it to Supabase cloud_backups."""
    if not slot_key:
        backup_date, slot, slot_key = current_cloud_backup_slot()
        if kind == "manual":
            slot_key = f"manual_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            slot = "manual"
            backup_date = datetime.now().strftime("%Y-%m-%d")

    # Step 1: ALWAYS create local backup on this PC
    ok_local, msg_local = create_local_backup(slot_key=slot_key, slot=slot, backup_date=backup_date)

    # Step 2: Upload to Supabase if connected
    if get_db_mode() != "supabase" or is_supabase_offline():
        return ok_local, f"Saved locally on this device ({msg_local})"

    try:
        payload, size_bytes = _build_cloud_backup_payload()
        pg = get_shared_supabase_conn()
        cur = pg.cursor()
        _ensure_audit_backup_schema(cur)
        cur.execute("SELECT slot_key FROM cloud_backups WHERE slot_key = %s", (slot_key,))
        exists = cur.fetchone()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        who = f"{_session_user_name()} ({get_device_identifier()})"
        if exists:
            cur.execute(
                """
                UPDATE cloud_backups
                SET created_at=%s, created_by=%s, payload=%s, size_bytes=%s, backup_date=%s, slot=%s
                WHERE slot_key=%s
                """,
                (now, who, payload, int(size_bytes), backup_date, slot, slot_key),
            )
        else:
            cur.execute(
                """
                INSERT INTO cloud_backups
                    (slot_key, backup_date, slot, created_at, created_by, payload, size_bytes)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                (slot_key, backup_date, slot, now, who, payload, int(size_bytes)),
            )
        cutoff = (datetime.now() - timedelta(days=CLOUD_BACKUP_KEEP_DAYS)).strftime("%Y-%m-%d")
        try:
            cur.execute(
                "DELETE FROM cloud_backups WHERE backup_date < %s AND slot IN ('am', 'pm')",
                (cutoff,),
            )
        except Exception:
            pass
        pg.commit()
        label = "morning" if slot == "am" else "afternoon" if slot == "pm" else "manual"
        try:
            log_user_action(
                "backup",
                extra_summary=f"Saved {label} backup locally and uploaded to cloud",
                row={"summary": f"Saved {label} backup locally and uploaded to cloud", "slot_key": slot_key},
            )
        except Exception:
            pass
        return True, slot_key
    except Exception as e:
        return ok_local, f"Saved locally on this device. Cloud sync: {e}"


def maybe_run_scheduled_cloud_backup():
    """Daily morning/afternoon backup for each device linked to DB: saves local + cloud."""
    global _LAST_BACKUP_SLOT
    backup_date, slot, slot_key = current_cloud_backup_slot()
    device_key = f"{slot_key}_{get_device_identifier()}"
    if _LAST_BACKUP_SLOT == device_key:
        return True, "cached"
    ok, msg = create_cloud_backup(
        slot_key=slot_key, slot=slot, backup_date=backup_date, kind="auto"
    )
    if ok:
        _LAST_BACKUP_SLOT = device_key
    return ok, msg


def list_local_backups(limit=30):
    """List backups saved locally on this machine."""
    b_dir = get_local_backups_dir()
    if not os.path.exists(b_dir):
        return []
    items = []
    seen = set()
    for f in os.listdir(b_dir):
        if (f.startswith("backup_") or f.startswith("snapshot_")) and (f.endswith(".enc") or f.endswith(".json.gz")):
            base_key = f.replace("backup_", "").replace("snapshot_", "").replace(".enc", "").replace(".json.gz", "")
            if base_key in seen:
                continue
            seen.add(base_key)
            full_p = os.path.join(b_dir, f)
            mtime = os.path.getmtime(full_p)
            dt_str = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            sz = os.path.getsize(full_p)
            slot = "AM" if "_am" in base_key.lower() else ("PM" if "_pm" in base_key.lower() else "Manual")
            items.append({
                "slot_key": f"local::{base_key}",
                "backup_date": dt_str.split()[0],
                "slot": f"💻 Local ({slot})",
                "created_at": dt_str,
                "created_by": f"This PC ({get_device_identifier()})",
                "size_bytes": sz,
                "is_local": True
            })
    items.sort(key=lambda x: x["created_at"], reverse=True)
    return items[:limit]


def list_cloud_backups(limit=40):
    """List cloud backups from Supabase."""
    if get_db_mode() != "supabase" or is_supabase_offline():
        return []
    try:
        pg = get_shared_supabase_conn()
        cur = pg.cursor()
        _ensure_audit_backup_schema(cur)
        cur.execute(
            """
            SELECT slot_key, backup_date, slot, created_at, created_by, size_bytes
            FROM cloud_backups
            ORDER BY created_at DESC
            LIMIT %s
            """,
            (int(limit),),
        )
        rows = []
        for r in cur.fetchall() or []:
            slot_name = r[2] or ""
            slot_label = "AM" if slot_name == "am" else ("PM" if slot_name == "pm" else "Manual")
            rows.append(
                {
                    "slot_key": r[0],
                    "backup_date": r[1],
                    "slot": f"☁️ Cloud ({slot_label})",
                    "created_at": r[3],
                    "created_by": plain_label(r[4]),
                    "size_bytes": r[5],
                    "is_local": False
                }
            )
        return rows
    except Exception:
        return []


def list_all_backups(limit=50):
    """Combines cloud backups and local device backups into a single chronological list."""
    cloud_list = list_cloud_backups(limit=limit)
    local_list = list_local_backups(limit=limit)
    combined = cloud_list + local_list
    combined.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return combined[:limit]


def restore_snapshot_dict(snapshot, source_name="backup"):
    """Restores database tables from a decoded snapshot dictionary."""
    tables = (snapshot or {}).get("tables") or {}
    if not tables:
        return False, "No tables found in snapshot."
    path = ensure_offline_cache_open()
    lite = _original_sqlite3_connect(path, timeout=30)
    try:
        lite.execute("PRAGMA busy_timeout=8000")
        lcur = lite.cursor()
        _init_offline_schema(lcur)
        for tbl, pack in tables.items():
            if str(tbl).lower() in ("offline_sync_queue", "sqlite_sequence"):
                continue
            cols = list((pack or {}).get("columns") or [])
            rows = list((pack or {}).get("rows") or [])
            if not cols:
                continue
            lcur.execute(f"PRAGMA table_info({tbl})")
            local_cols = [r[1] for r in (lcur.fetchall() or []) if r and r[1]]
            use_cols = [c for c in cols if c in local_cols]
            if not use_cols:
                continue
            lcur.execute(f"DELETE FROM {tbl}")
            placeholders = ", ".join(["?"] * len(use_cols))
            col_list = ", ".join(use_cols)
            for rec in rows:
                vals = []
                for c in use_cols:
                    val = rec.get(c)
                    if isinstance(val, dict) and "__bytes__" in val:
                        val = base64.b64decode(val["__bytes__"])
                    vals.append(val)
                lcur.execute(
                    f"INSERT INTO {tbl} ({col_list}) VALUES ({placeholders})",
                    vals,
                )
        lite.commit()
    finally:
        lite.close()
    try:
        _persist_offline_cache()
    except Exception:
        pass
    if get_db_mode() == "supabase" and not is_supabase_offline():
        try:
            backfill_local_rows_missing_from_cloud()
            flush_offline_queue_to_cloud()
        except Exception:
            pass
    try:
        log_user_action("backup", extra_summary=f"Restored backup from {source_name}")
    except Exception:
        pass
    return True, "ok"


def restore_cloud_backup(slot_key):
    """Restores data from either a local device snapshot or a Supabase cloud backup."""
    if str(slot_key).startswith("local::"):
        raw_key = str(slot_key)[7:]
        b_dir = get_local_backups_dir()
        json_file = os.path.join(b_dir, f"snapshot_{raw_key}.json.gz")
        payload = None
        if os.path.isfile(json_file):
            with open(json_file, "r", encoding="utf-8") as f:
                payload = f.read()
        if not payload:
            enc_file = os.path.join(b_dir, f"backup_{raw_key}.enc")
            if os.path.isfile(enc_file):
                active_enc = os.path.join(get_default_app_dir(), "payroll_data.enc")
                import shutil
                shutil.copy2(enc_file, active_enc)
                load_database()
                return True, "Restored from local encrypted file."
            return False, "Local backup file not found."
    else:
        if get_db_mode() != "supabase" or is_supabase_offline():
            return False, "Cloud is offline"
        pg = get_shared_supabase_conn()
        cur = pg.cursor()
        cur.execute("SELECT payload FROM cloud_backups WHERE slot_key = %s", (slot_key,))
        row = cur.fetchone()
        if not row or not row[0]:
            return False, "Backup not found in cloud."
        payload = row[0]

    snapshot = _decode_cloud_backup_payload(payload)
    if not snapshot:
        return False, "Could not decode backup payload."
    return restore_snapshot_dict(snapshot, source_name=f"slot:{slot_key}")


# --- IN-APP CLOUD AUTO-UPDATER & FAIL-SAFE CRASH GUARD ---

def restart_app():
    """Cleanly restart the current application process."""
    try:
        clean_env = os.environ.copy()
        clean_env.pop("_DYNAMIC_UPDATE_RUNNING", None)
        clean_env.pop("_DYNAMIC_UPDATE_ACTIVE", None)
        clean_env.pop("_RUNNING_SAFE_MODE_FALLBACK", None)
        clean_env.pop("_MEIPASS2", None)
        clean_env.pop("_MEIPASS", None)
        clean_env.pop("_PYI_APPLICATION_HOME_DIR", None)
        clean_env.pop("_PYI_PARENT_PROCESS_LEVEL", None)
        # Purge any temporary PyInstaller _MEI paths from PATH so new instance extracts cleanly
        if "PATH" in clean_env:
            raw_paths = clean_env["PATH"].split(os.pathsep)
            purged = [p for p in raw_paths if "_MEI" not in p]
            clean_env["PATH"] = os.pathsep.join(purged)
        if platform.system() == "Windows":
            app_dir = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(sys.argv[0] or "."))
            if getattr(sys, "frozen", False):
                run_target = f'""{sys.executable}""'
                if len(sys.argv) > 1:
                    run_target += " " + " ".join(f'""{a}""' for a in sys.argv[1:])
            else:
                script_path = os.path.abspath(sys.argv[0]) if sys.argv else "payroll_app.py"
                run_target = f'""{sys.executable}"" ""{script_path}""'
                if len(sys.argv) > 1:
                    run_target += " " + " ".join(f'""{a}""' for a in sys.argv[1:])

            # Release current working directory lock from any PyInstaller _MEI folder
            try:
                os.chdir(app_dir)
            except Exception:
                pass

            # Launch completely silently via native Windows wscript without flashing any console or terminal window
            try:
                vbs_path = os.path.join(tempfile.gettempdir(), f"payroll_restart_{os.getpid()}.vbs")
                vbs_code = (
                    "WScript.Sleep 2500\n"
                    "Set sh = CreateObject(\"WScript.Shell\")\n"
                    "Set env = sh.Environment(\"Process\")\n"
                    "On Error Resume Next\n"
                    "env.Remove(\"_MEIPASS2\")\n"
                    "env.Remove(\"_MEIPASS\")\n"
                    "env.Remove(\"_PYI_APPLICATION_HOME_DIR\")\n"
                    "env.Remove(\"_PYI_PARENT_PROCESS_LEVEL\")\n"
                    "On Error GoTo 0\n"
                    f"sh.CurrentDirectory = \"{app_dir}\"\n"
                    f"sh.Run \"{run_target}\", 1, False\n"
                    "Set fso = CreateObject(\"Scripting.FileSystemObject\")\n"
                    "On Error Resume Next\n"
                    "fso.DeleteFile WScript.ScriptFullName\n"
                )
                with open(vbs_path, "w", encoding="utf-8") as vf:
                    vf.write(vbs_code)
                subprocess.Popen(["wscript.exe", vbs_path], cwd=app_dir, close_fds=True)
            except Exception as wscript_err:
                messagebox.showinfo("Restart Needed", "Please close and reopen the application to complete the update.")
                return
        elif platform.system() == "Darwin":
            exe = sys.executable
            if getattr(sys, "frozen", False):
                if ".app/Contents/MacOS" in exe:
                    app_bundle = exe[:exe.rfind(".app/Contents/MacOS") + 4]
                    subprocess.Popen(["open", "-n", app_bundle], env=clean_env)
                else:
                    subprocess.Popen([exe] + sys.argv[1:], env=clean_env)
            else:
                subprocess.Popen([sys.executable] + sys.argv, env=clean_env)
        else:
            if getattr(sys, "frozen", False):
                subprocess.Popen([sys.executable] + sys.argv[1:], env=clean_env)
            else:
                subprocess.Popen([sys.executable] + sys.argv, env=clean_env)
    except Exception as e:
        messagebox.showinfo("Restart Needed", f"Please close and reopen the app manually to complete the update.\n\n({e})")
        return
    sys.exit(0)


def get_custom_update_server_url():
    """Retrieve the configured update server URL or return the default."""
    try:
        conn = sqlite3.connect(TEMP_DB_PATH)
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)")
        cur.execute("SELECT value FROM app_settings WHERE key IN ('update_server_url', 'github_raw_url') ORDER BY CASE WHEN key='update_server_url' THEN 0 ELSE 1 END")
        r = cur.fetchone()
        conn.close()
        if r and r[0] and str(r[0]).strip():
            return str(r[0]).strip()
    except Exception:
        pass
    return DEFAULT_UPDATE_SERVER_URL

get_custom_github_raw_url = get_custom_update_server_url


def set_custom_update_server_url(url):
    """Save a custom update server URL to the settings database."""
    try:
        conn = sqlite3.connect(TEMP_DB_PATH)
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)")
        cur.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('update_server_url', ?)", (url.strip(),))
        conn.commit()
        conn.close()
    except Exception:
        pass

set_custom_github_raw_url = set_custom_update_server_url


def get_update_auth_token():
    """Retrieve private cloud update access token if configured (from Supabase cloud_config or local DB)."""
    try:
        conn = sqlite3.connect(TEMP_DB_PATH)
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)")
        cur.execute("SELECT value FROM app_settings WHERE key = 'update_auth_token'")
        r = cur.fetchone()
        conn.close()
        if r and r[0] and str(r[0]).strip():
            return str(r[0]).strip()
    except Exception:
        pass
    if get_db_mode() == "supabase" and not is_supabase_offline():
        try:
            pg = get_shared_supabase_conn()
            cur = pg.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS cloud_config (key TEXT PRIMARY KEY, value TEXT)")
            cur.execute("SELECT value FROM cloud_config WHERE key = 'update_auth_token'")
            row = cur.fetchone()
            if row and row[0] and str(row[0]).strip():
                return str(row[0]).strip()
        except Exception:
            pass
    return os.environ.get("UPDATE_AUTH_TOKEN") or ""


def set_update_auth_token(token_str, sync_to_cloud=True):
    """Save the private update access token locally and optionally sync to Supabase."""
    tok = str(token_str or "").strip()
    try:
        conn = sqlite3.connect(TEMP_DB_PATH)
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)")
        cur.execute("INSERT OR REPLACE INTO app_settings (key, value) VALUES ('update_auth_token', ?)", (tok,))
        conn.commit()
        conn.close()
    except Exception:
        pass
    if sync_to_cloud and get_db_mode() == "supabase" and not is_supabase_offline():
        try:
            pg = get_shared_supabase_conn()
            cur = pg.cursor()
            cur.execute("CREATE TABLE IF NOT EXISTS cloud_config (key TEXT PRIMARY KEY, value TEXT)")
            cur.execute("INSERT INTO cloud_config (key, value) VALUES ('update_auth_token', %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value", (tok,))
            pg.commit()
        except Exception:
            pass


def _parse_version_tuple(v_str):
    """
    Parses a version string into a comparable tuple of integers.
    Examples:
        '2.5.0' -> (2, 5, 0)
        'v2.5.1' -> (2, 5, 1)
        '2.5' -> (2, 5, 0)
    """
    if not v_str:
        return (0, 0, 0)
    import re
    clean = re.sub(r'^[^\d]*', '', str(v_str).strip())
    nums = []
    for chunk in clean.split('.'):
        m = re.match(r'^(\d+)', chunk)
        if m:
            nums.append(int(m.group(1)))
        else:
            break
    while len(nums) < 3:
        nums.append(0)
    return tuple(nums)


def get_active_code_info():
    """Returns metadata about the currently running code engine."""
    is_safe_mode = os.environ.get("_RUNNING_SAFE_MODE_FALLBACK") == "1" or os.path.isfile(get_safe_mode_flag_path())
    is_dynamic = (os.environ.get("_DYNAMIC_UPDATE_ACTIVE") == "1" or os.environ.get("_DYNAMIC_UPDATE_RUNNING") == "1") and not is_safe_mode
    
    current_file = None
    if is_dynamic and os.path.isfile(get_updates_script_path()):
        current_file = get_updates_script_path()
    else:
        current_file = os.path.abspath(__file__)
        
    code_hash = ""
    norm_hash = ""
    file_size = 0
    version = APP_VERSION
    build_date = APP_BUILD_DATE
    last_modified = ""

    # Check version_meta.json if dynamic
    if is_dynamic:
        meta_file = os.path.join(get_updates_dir(), "version_meta.json")
        if os.path.isfile(meta_file):
            try:
                with open(meta_file, "r", encoding="utf-8") as mf:
                    meta = json.load(mf)
                    if meta.get("version"):
                        version = meta.get("version")
                    if meta.get("build_date"):
                        build_date = meta.get("build_date")
                    if meta.get("hash"):
                        code_hash = meta.get("hash")
                    if meta.get("norm_hash"):
                        norm_hash = meta.get("norm_hash")
            except Exception:
                pass

    try:
        if current_file and os.path.isfile(current_file):
            with open(current_file, "rb") as f:
                content = f.read()
            if not code_hash:
                code_hash = hashlib.sha256(content).hexdigest()
            if not norm_hash:
                norm_hash = hashlib.sha256(content.replace(b"\r\n", b"\n").strip()).hexdigest()
            file_size = len(content)
            mtime = os.path.getmtime(current_file)
            last_modified = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            import re
            m = re.search(r'APP_VERSION\s*=\s*["\']([^"\']+)["\']', content.decode("utf-8", errors="ignore"))
            if m:
                version = m.group(1)
            m_d = re.search(r'APP_BUILD_DATE\s*=\s*["\']([^"\']+)["\']', content.decode("utf-8", errors="ignore"))
            if m_d:
                build_date = m_d.group(1)
    except Exception:
        pass

    return {
        "version": version,
        "build_date": build_date,
        "hash": code_hash,
        "norm_hash": norm_hash,
        "is_dynamic": is_dynamic,
        "is_safe_mode": is_safe_mode,
        "file_size": file_size,
        "last_modified": last_modified,
        "current_file": current_file,
    }


def check_for_cloud_update():
    """
    Checks the cloud server for new updates.
    Returns (status: str, data: dict)
    status: 'update_available' | 'up_to_date' | 'error'
    """
    url = get_custom_update_server_url()
    import time, urllib.request
    sep = "&" if "?" in url else "?"
    req_url = f"{url}{sep}_cb={int(time.time())}"
    headers = {
        "User-Agent": "PayrollApp-Updater/2.5",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
    }
    token = get_update_auth_token()
    if token:
        headers["Authorization"] = f"token {token}"
    req = urllib.request.Request(req_url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw_bytes = resp.read()
            remote_code = raw_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        err_str = str(e)
        if "404" in err_str:
            clean_err = "Update server resource not found (404)."
        elif "403" in err_str:
            clean_err = "Access to update server was denied (403)."
        elif "timed out" in err_str.lower():
            clean_err = "Connection to update server timed out. Please check your network."
        else:
            clean_err = f"Could not connect to update service: {err_str}"
        return "error", {"error": clean_err}

    remote_hash = hashlib.sha256(raw_bytes).hexdigest()
    remote_norm_hash = hashlib.sha256(raw_bytes.replace(b"\r\n", b"\n").strip()).hexdigest()

    local_info = get_active_code_info()
    local_hash = local_info.get("hash") or ""
    local_norm_hash = local_info.get("norm_hash") or ""
    local_version = local_info.get("version") or APP_VERSION
    local_build_date = local_info.get("build_date") or APP_BUILD_DATE

    import re
    m_ver = re.search(r'APP_VERSION\s*=\s*["\']([^"\']+)["\']', remote_code)
    remote_version = m_ver.group(1) if m_ver else local_version

    m_date = re.search(r'APP_BUILD_DATE\s*=\s*["\']([^"\']+)["\']', remote_code)
    remote_build_date = m_date.group(1) if m_date else ""

    remote_tuple = _parse_version_tuple(remote_version)
    local_tuple = _parse_version_tuple(local_version)

    meta_hash = ""
    meta_norm_hash = ""
    meta_file = os.path.join(get_updates_dir(), "version_meta.json")
    if os.path.isfile(meta_file):
        try:
            with open(meta_file, "r", encoding="utf-8") as mf:
                meta_data = json.load(mf)
                meta_hash = meta_data.get("hash") or ""
                meta_norm_hash = meta_data.get("norm_hash") or ""
        except Exception:
            pass

    # Check if this exact update is already downloaded and saved on disk
    upd_path = get_updates_script_path()
    disk_is_current = False
    if os.path.isfile(upd_path):
        try:
            with open(upd_path, "rb") as uf:
                disk_bytes = uf.read()
            disk_norm = hashlib.sha256(disk_bytes.replace(b"\r\n", b"\n").strip()).hexdigest()
            if disk_norm == remote_norm_hash:
                disk_is_current = True
        except Exception:
            pass

    is_update_available = False

    if disk_is_current:
        if local_info.get("is_dynamic"):
            is_update_available = False
        else:
            data = {
                "remote_code": remote_code,
                "remote_hash": remote_hash,
                "remote_version": remote_version,
                "remote_build_date": remote_build_date,
                "size_bytes": len(raw_bytes),
                "local_info": local_info,
            }
            return "installed_pending_restart", data
    elif remote_tuple > local_tuple:
        is_update_available = True
    elif remote_tuple < local_tuple:
        is_update_available = False
    else:
        # Same semantic version
        if meta_hash and (remote_hash == meta_hash or (meta_norm_hash and remote_norm_hash == meta_norm_hash)):
            is_update_available = False
        elif local_norm_hash and remote_norm_hash == local_norm_hash:
            is_update_available = False
        elif local_hash and remote_hash == local_hash:
            is_update_available = False
        else:
            upd_path = get_updates_script_path()
            if os.path.isfile(upd_path):
                try:
                    with open(upd_path, "rb") as uf:
                        disk_bytes = uf.read()
                    disk_norm = hashlib.sha256(disk_bytes.replace(b"\r\n", b"\n").strip()).hexdigest()
                    if disk_norm == remote_norm_hash:
                        is_update_available = False
                    else:
                        if remote_build_date and local_build_date and remote_build_date < local_build_date:
                            is_update_available = False
                        else:
                            is_update_available = True
                except Exception:
                    is_update_available = False
            else:
                # Factory binary (no dynamic update file on disk yet)
                if remote_build_date and local_build_date and remote_build_date < local_build_date:
                    is_update_available = False
                else:
                    # Remote code differs from local factory code, update is available
                    is_update_available = True
    
    try:
        log_user_action(
            "app_update_check",
            extra_summary=f"Checked for software updates: {'Update available' if is_update_available else 'Already up to date'}",
            row={"remote_version": remote_version, "remote_hash": remote_hash[:10]}
        )
    except Exception:
        pass

    disp_version = remote_version
    if remote_version == local_version and is_update_available:
        disp_version = f"{remote_version} (Rev {remote_hash[:8]})"

    data = {
        "remote_code": remote_code,
        "remote_hash": remote_hash,
        "remote_version": disp_version,
        "remote_build_date": remote_build_date,
        "size_bytes": len(raw_bytes),
        "local_info": local_info,
    }
    if is_update_available:
        return "update_available", data
    return "up_to_date", data

check_for_github_update = check_for_cloud_update


def install_cloud_update(code_str, remote_hash=""):
    """
    Validates syntax, backs up previous version, writes update to disk,
    logs action to Supabase, and clears the safe-mode flag.
    Returns (ok: bool, msg: str)
    """
    if not code_str or len(code_str.strip()) < 500:
        return False, "Downloaded code is empty or truncated."

    # 1. Syntax Check
    try:
        compile(code_str, "payroll_app.py", "exec")
    except SyntaxError as syn_err:
        err_msg = f"Syntax error at line {syn_err.lineno}: {syn_err.msg}"
        try:
            log_user_action(
                "update_syntax_error",
                extra_summary=f"Update failed syntax validation: {err_msg[:80]}",
                row={"error": str(syn_err), "line": syn_err.lineno, "hash": remote_hash[:10]}
            )
        except Exception:
            pass
        return False, f"Downloaded update cannot be installed because it contains syntax errors:\n\n{err_msg}"

    updates_dir = get_updates_dir()
    try:
        os.makedirs(updates_dir, exist_ok=True)
    except Exception:
        pass
    current_update = get_updates_script_path()
    bak_file = os.path.join(updates_dir, "payroll_app.py.bak")

    # 2. Backup existing update
    try:
        if os.path.isfile(current_update):
            shutil.copy2(current_update, bak_file)
            archive_dir = os.path.join(updates_dir, "backups")
            os.makedirs(archive_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy2(current_update, os.path.join(archive_dir, f"payroll_app_{ts}.py"))
            old_baks = sorted([os.path.join(archive_dir, f) for f in os.listdir(archive_dir) if f.endswith(".py")], key=os.path.getmtime)
            if len(old_baks) > 10:
                for ob in old_baks[:-10]:
                    try:
                        os.remove(ob)
                    except Exception:
                        pass
    except Exception:
        pass

    # 3. Write new code
    try:
        with open(current_update, "w", encoding="utf-8", newline="\n") as f:
            f.write(code_str)
    except Exception as e:
        return False, f"Failed to write update file: {e}"

    # 4. Remove safe mode flag if any
    safe_flag = get_safe_mode_flag_path()
    if os.path.isfile(safe_flag):
        try:
            os.remove(safe_flag)
        except Exception:
            pass

    # 5. Write metadata
    try:
        import re
        m = re.search(r'APP_VERSION\s*=\s*["\']([^"\']+)["\']', code_str)
        new_ver = m.group(1) if m else "Latest"
        m_d = re.search(r'APP_BUILD_DATE\s*=\s*["\']([^"\']+)["\']', code_str)
        new_date = m_d.group(1) if m_d else ""
        norm_h = hashlib.sha256(code_str.replace("\r\n", "\n").strip().encode("utf-8")).hexdigest()
        meta_file = os.path.join(updates_dir, "version_meta.json")
        with open(meta_file, "w", encoding="utf-8") as f:
            json.dump({
                "version": new_ver,
                "build_date": new_date,
                "installed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "hash": remote_hash,
                "norm_hash": norm_h,
                "user": _session_user_name(),
                "machine": get_device_identifier(),
            }, f, indent=2)
    except Exception:
        new_ver = "Latest"

    # 6. Log to Supabase
    try:
        log_user_action(
            "app_update",
            extra_summary=f"Installed software update v{new_ver}",
            row={"version": new_ver, "size_bytes": len(code_str.encode("utf-8"))}
        )
    except Exception:
        pass

    return True, f"Version v{new_ver} installed successfully!"

install_github_update = install_cloud_update


def rollback_cloud_update(target="bak"):
    """
    Rolls back update:
    target='bak': restores payroll_app.py.bak
    target='factory': moves payroll_app.py to .disabled to run the built-in factory version.
    """
    updates_dir = get_updates_dir()
    current_update = get_updates_script_path()
    bak_file = os.path.join(updates_dir, "payroll_app.py.bak")
    safe_flag = get_safe_mode_flag_path()

    if target == "bak":
        if not os.path.isfile(bak_file):
            return False, "No previous update backup (.bak) found."
        try:
            shutil.copy2(bak_file, current_update)
            if os.path.isfile(safe_flag):
                os.remove(safe_flag)
            log_user_action("app_rollback", extra_summary="Rolled back to previous update (.bak)")
            return True, "Successfully restored previous update version."
        except Exception as e:
            return False, f"Failed to restore backup: {e}"
    elif target == "factory":
        try:
            if os.path.isfile(current_update):
                dis_path = os.path.join(updates_dir, f"payroll_app_{datetime.now().strftime('%Y%m%d_%H%M%S')}.disabled")
                shutil.move(current_update, dis_path)
            if os.path.isfile(safe_flag):
                os.remove(safe_flag)
            meta_file = os.path.join(updates_dir, "version_meta.json")
            if os.path.isfile(meta_file):
                try:
                    os.remove(meta_file)
                except Exception:
                    pass
            log_user_action("app_rollback", extra_summary="Reverted to built-in factory version")
            return True, "Reverted to factory built-in version."
        except Exception as e:
            return False, f"Failed to revert to factory version: {e}"
    return False, "Unknown rollback target."

rollback_github_update = rollback_cloud_update


def _check_and_run_dynamic_update():
    """
    Host bootstrap loader:
    If a downloaded update exists at <app_dir>/updates/payroll_app.py,
    attempt to run it dynamically via runpy.
    If it crashes on launch, catch the crash, record details in last_crash.log,
    flag safe_mode, log to Supabase, display a recovery popup, and allow the built-in code to run!
    """
    if os.environ.get("_DYNAMIC_UPDATE_RUNNING") == "1":
        return False
        
    update_file = get_updates_script_path()
    if not os.path.isfile(update_file):
        return False
        
    safe_flag = get_safe_mode_flag_path()
    if os.path.isfile(safe_flag):
        try:
            # If the downloaded update is newer than the crash flag, clear the stale flag
            if os.path.getmtime(update_file) > os.path.getmtime(safe_flag):
                os.remove(safe_flag)
            else:
                os.environ["_RUNNING_SAFE_MODE_FALLBACK"] = "1"
                return False
        except Exception:
            os.environ["_RUNNING_SAFE_MODE_FALLBACK"] = "1"
            return False

    try:
        with open(update_file, "r", encoding="utf-8") as f:
            code_str = f.read()
        if len(code_str.strip()) < 500:
            return False
        compile(code_str, update_file, "exec")
    except Exception as e:
        try:
            with open(safe_flag, "w", encoding="utf-8") as sf:
                sf.write(f"Syntax error in downloaded update: {e}")
        except Exception:
            pass
        os.environ["_RUNNING_SAFE_MODE_FALLBACK"] = "1"
        return False

    # Execute dynamic update
    os.environ["_DYNAMIC_UPDATE_RUNNING"] = "1"
    os.environ["_DYNAMIC_UPDATE_ACTIVE"] = "1"
    import runpy
    import traceback
    try:
        runpy.run_path(update_file, run_name="__main__")
        sys.exit(0)
    except SystemExit as se:
        sys.exit(se.code if se.code is not None else 0)
    except Exception as exc:
        tb_str = traceback.format_exc()
        try:
            with open(get_last_crash_log_path(), "w", encoding="utf-8") as lf:
                lf.write(f"Crash Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                lf.write(f"Error: {exc}\n\nTraceback:\n{tb_str}\n")
        except Exception:
            pass
        try:
            with open(safe_flag, "w", encoding="utf-8") as sf:
                sf.write(f"Update crashed on boot: {exc}\n{tb_str[:500]}")
        except Exception:
            pass
        try:
            log_user_action(
                "update_crash",
                extra_summary=f"Update failed to boot: {str(exc)[:80]}",
                row={"error": str(exc), "traceback": tb_str[:2000]}
            )
        except Exception:
            pass
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showwarning(
                "Safe Mode Activated ⚠️",
                f"The recently downloaded update encountered an error during startup:\n\n{exc}\n\n"
                f"The application has safely recovered and fallen back to the built-in version.\n"
                f"You can continue working, check for a fixed update, or roll back in Settings ➔ App Updates.",
            )
            root.destroy()
        except Exception:
            pass
        os.environ["_RUNNING_SAFE_MODE_FALLBACK"] = "1"
        os.environ.pop("_DYNAMIC_UPDATE_RUNNING", None)
        return False



def _merge_action_logs_into_local(lcur, use_cols, packed_rows):
    if "log_uid" not in use_cols:
        return
    uid_idx = use_cols.index("log_uid")
    existing = set()
    try:
        lcur.execute("SELECT log_uid FROM user_action_log")
        for r in lcur.fetchall() or []:
            if r and r[0]:
                existing.add(plain_label(r[0]))
    except Exception:
        existing = set()
    insert_cols = [c for c in use_cols if c != "id"]
    if not insert_cols:
        return
    col_list = ", ".join(insert_cols)
    placeholders = ", ".join(["?"] * len(insert_cols))
    for row in packed_rows or []:
        row = list(row)
        uid = plain_label(row[uid_idx]) if uid_idx < len(row) else ""
        if uid and uid in existing:
            continue
        vals = []
        for c in insert_cols:
            vals.append(row[use_cols.index(c)])
        try:
            lcur.execute(
                f"INSERT INTO user_action_log ({col_list}) VALUES ({placeholders})",
                vals,
            )
            if uid:
                existing.add(uid)
        except Exception:
            continue


def _pull_user_action_logs(pg_cur, lcur):
    try:
        pg_cur.execute("SELECT * FROM user_action_log")
        rows = pg_cur.fetchall() or []
        desc = pg_cur.description
        if not desc:
            return
        cloud_cols = [d[0] for d in desc]
        lcur.execute("PRAGMA table_info(user_action_log)")
        local_cols = [r[1] for r in (lcur.fetchall() or []) if r and r[1]]
        use_cols = [c for c in cloud_cols if c in local_cols]
        if not use_cols:
            return
        indexes = [cloud_cols.index(c) for c in use_cols]
        packed = [tuple(row[i] for i in indexes) for row in rows]
        _merge_action_logs_into_local(lcur, use_cols, packed)
    except Exception:
        pass


class OfflineTrackingConnection:
    """SQLite connection wrapper that queues mutations for later cloud flush."""

    def __init__(self, conn):
        self._conn = conn

    def cursor(self):
        return OfflineTrackingCursor(self._conn.cursor(), self._conn)

    def commit(self):
        self._conn.commit()
        if get_db_mode() == "supabase":
            _schedule_persist_offline_cache()
            schedule_cloud_push()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        try:
            self._conn.commit()
        except Exception:
            pass
        try:
            self._conn.close()
        except Exception:
            pass

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            try:
                self._conn.rollback()
            except Exception:
                pass
        else:
            try:
                self._conn.commit()
            except Exception:
                pass
        self.close()


class OfflineTrackingCursor:
    def __init__(self, cursor, conn):
        self._cur = cursor
        self._conn = conn
        self._lastrowid = None

    @property
    def description(self):
        return self._cur.description

    @property
    def rowcount(self):
        return self._cur.rowcount

    @property
    def lastrowid(self):
        if self._lastrowid is not None:
            return self._lastrowid
        return self._cur.lastrowid

    def _enqueue(self, payload):
        _queue_offline_op(payload, self._conn)

    def _log_change(self, op, table, row=None, record_id=None):
        if not table:
            return
        try:
            log_user_action(
                op,
                table=table,
                record_id=record_id,
                row=row,
                raw_conn=self._conn,
            )
        except Exception:
            pass

    def _fetch_row_dict(self, table, row_id):
        try:
            self._cur.execute(f"PRAGMA table_info({table})")
            cols = [r[1] for r in self._cur.fetchall() or []]
            if not cols:
                return None
            self._cur.execute(
                f"SELECT * FROM {table} WHERE id = ?", (row_id,)
            )
            row = self._cur.fetchone()
            if not row:
                return None
            return {c: row[i] for i, c in enumerate(cols)}
        except Exception:
            return None

    def _fetch_named_row(self, table, key_col, key_val):
        try:
            self._cur.execute(f"PRAGMA table_info({table})")
            cols = [r[1] for r in self._cur.fetchall() or []]
            if not cols or key_col not in cols:
                return None
            self._cur.execute(
                f"SELECT * FROM {table} WHERE {key_col} = ?", (key_val,)
            )
            row = self._cur.fetchone()
            if not row:
                return None
            return {c: row[i] for i, c in enumerate(cols)}
        except Exception:
            return None

    def execute(self, query, params=None):
        import re

        params = () if params is None else tuple(params)
        self._cur.execute(query, params)
        try:
            if (query or "").strip().upper().startswith("INSERT"):
                self._lastrowid = self._cur.lastrowid
        except Exception:
            pass
        q = (query or "").strip()
        qu = q.upper()
        if "OFFLINE_SYNC_QUEUE" in qu:
            return self
        try:
            if qu.startswith("INSERT"):
                m = re.search(
                    r"INSERT\s+(?:OR\s+\w+\s+)?INTO\s+(\w+)", q, re.IGNORECASE
                )
                if not m:
                    return self
                table = m.group(1)
                if table.lower() in (
                    "offline_sync_queue",
                    "sqlite_sequence",
                    "user_action_log",
                    "database_history_log",
                ):
                    return self
                if table.lower() in (
                    "config_locations",
                    "config_categories",
                    "config_payments",
                ):
                    # name primary key
                    if params:
                        row = self._fetch_named_row(table, "name", params[0])
                        if row:
                            self._enqueue(
                                {"op": "upsert_key", "table": table, "key": "name", "row": row}
                            )
                            self._log_change("insert", table, row=row, record_id=params[0])
                    return self
                if table.lower() == "users" and params:
                    row = self._fetch_named_row(table, "username", params[0])
                    if row:
                        self._enqueue(
                            {
                                "op": "upsert_key",
                                "table": table,
                                "key": "username",
                                "row": row,
                            }
                        )
                        self._log_change("insert", table, row=row, record_id=params[0])
                    return self
                lid = self._lastrowid or self._cur.lastrowid
                if lid:
                    row = self._fetch_row_dict(table, lid)
                    if row:
                        self._enqueue({"op": "insert", "table": table, "row": row})
                        self._log_change("insert", table, row=row, record_id=lid)
                    if table.lower() == "expenses":
                        protect_local_expense_id(lid)
            elif qu.startswith("UPDATE"):
                m = re.match(
                    r"UPDATE\s+(\w+)\s+SET.*WHERE\s+id\s*=\s*\?",
                    q,
                    re.IGNORECASE | re.DOTALL,
                )
                if m and params:
                    table = m.group(1)
                    row = self._fetch_row_dict(table, params[-1])
                    if row:
                        self._enqueue({"op": "upsert", "table": table, "row": row})
                        self._log_change("update", table, row=row, record_id=params[-1])
                    if str(table).lower() == "expenses":
                        protect_local_expense_id(params[-1])
                else:
                    m2 = re.match(
                        r"UPDATE\s+(\w+)\s+SET.*WHERE\s+(\w+)\s*=\s*\?",
                        q,
                        re.IGNORECASE | re.DOTALL,
                    )
                    if m2 and params:
                        table, key = m2.group(1), m2.group(2)
                        row = self._fetch_named_row(table, key, params[-1])
                        if row:
                            self._enqueue(
                                {
                                    "op": "upsert_key",
                                    "table": table,
                                    "key": key,
                                    "row": row,
                                }
                            )
                            self._log_change("update", table, row=row, record_id=params[-1])
                    else:
                        self._enqueue(
                            {"op": "sql", "sql": q, "params": list(params)}
                        )
                        self._log_change("update", None, row={"sql": q[:120]})
            elif qu.startswith("DELETE"):
                m = re.match(
                    r"DELETE\s+FROM\s+(\w+)\s+WHERE\s+id\s*=\s*\?",
                    q,
                    re.IGNORECASE,
                )
                if m and params:
                    self._enqueue(
                        {"op": "delete", "table": m.group(1), "id": params[0]}
                    )
                    self._log_change("delete", m.group(1), record_id=params[0])
                else:
                    m2 = re.match(
                        r"DELETE\s+FROM\s+(\w+)\s+WHERE\s+(\w+)\s*=\s*\?",
                        q,
                        re.IGNORECASE,
                    )
                    if m2 and params:
                        self._enqueue(
                            {
                                "op": "delete_key",
                                "table": m2.group(1),
                                "key": m2.group(2),
                                "value": params[0],
                            }
                        )
                        self._log_change("delete", m2.group(1), record_id=params[0])
                    else:
                        self._enqueue(
                            {"op": "sql", "sql": q, "params": list(params)}
                        )
                        self._log_change("delete", None, row={"sql": q[:120]})
        except Exception as e:
            try:
                global _LAST_SYNC_ERROR
                _LAST_SYNC_ERROR = f"queue: {e}"
            except Exception:
                pass
        return self

    def executemany(self, query, seq_of_params):
        for params in seq_of_params:
            self.execute(query, params)

    def fetchone(self):
        return self._cur.fetchone()

    def fetchall(self):
        return self._cur.fetchall()

    def close(self):
        try:
            self._cur.close()
        except Exception:
            pass


def _row_without_bad_column(row, exc):
    """Drop a column named in a Postgres 'does not exist' error so the row can retry."""
    import re

    if not row:
        return None
    msg = str(exc or "")
    names = re.findall(r'column "?([A-Za-z_][A-Za-z0-9_]*)"?', msg, re.I)
    if not names:
        m = re.search(r'([A-Za-z_][A-Za-z0-9_]*) does not exist', msg, re.I)
        if m:
            names = [m.group(1)]
    keys = {str(k).lower(): k for k in row}
    changed = False
    out = dict(row)
    for n in names:
        k = keys.get(str(n).lower())
        if k is not None:
            out.pop(k, None)
            changed = True
    return out if changed else None


def _pg_insert(pg_cur, table, row):
    cols = [c for c in row.keys()]
    col_list = ", ".join(cols)
    placeholders = ", ".join(["?"] * len(cols))
    vals = [row[c] for c in cols]
    table_l = str(table or "").lower()
    if table_l == "user_action_log":
        pg_cur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) ON CONFLICT (log_uid) DO NOTHING", vals)
    elif table_l in ("config_locations", "config_categories", "config_payments", "config_languages"):
        pg_cur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) ON CONFLICT (name) DO NOTHING", vals)
    elif table_l == "users":
        pg_cur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) ON CONFLICT (username) DO NOTHING", vals)
    elif table_l == "cash_month_locks":
        pg_cur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders}) ON CONFLICT (year_month) DO NOTHING", vals)
    else:
        pg_cur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", vals)


def _pg_inserted_id(pg_cur, row):
    if row and row.get("id") is not None:
        return row.get("id")
    lastid = getattr(pg_cur, "_lastrowid", None)
    if lastid is not None:
        return lastid
    return None


def _cloud_insert_row(pg_cur, table, row):
    """Insert a new row. Never overwrite an existing cloud id (avoids wiping other PCs).

    Returns the cloud id used. For payroll_records, an id clash with the *same*
    logical row is treated as already uploaded — we do not insert a second copy.
    """
    if not row:
        return None
    row = dict(row)
    last_err = None
    popped_id = False
    table_l = str(table or "").lower()
    for _ in range(16):
        try:
            _pg_insert(pg_cur, table, row)
            return _pg_inserted_id(pg_cur, row)
        except Exception as e:
            last_err = e
            if _is_duplicate_key_error(e):
                if table_l in ("user_action_log", "config_locations", "config_categories", "config_payments", "config_languages", "users", "cash_month_locks"):
                    return row.get("id")
            stripped = _row_without_bad_column(row, e)
            if stripped is not None:
                row = stripped
                continue
            if not popped_id and row.get("id") is not None and _is_duplicate_key_error(e):
                if table_l == "payroll_records":
                    try:
                        pg_cur.execute(
                            "SELECT * FROM payroll_records WHERE id = ?",
                            (row["id"],),
                        )
                        found = pg_cur.fetchone()
                        desc = pg_cur.description
                        if found and desc:
                            cols = [d[0] for d in desc]
                            existing = {cols[i]: found[i] for i in range(min(len(cols), len(found)))}
                            if _payroll_identity(existing) == _payroll_identity(row):
                                return row["id"]
                    except Exception:
                        return row.get("id")
                row = dict(row)
                row.pop("id", None)
                popped_id = True
                if not row:
                    return None
                continue
            if not popped_id and "id" in row:
                row = dict(row)
                row.pop("id", None)
                popped_id = True
                if not row:
                    return None
                continue
            raise
    if last_err:
        if _is_duplicate_key_error(last_err) and table_l == "user_action_log":
            return row.get("id")
        raise last_err
    return _pg_inserted_id(pg_cur, row)


def _cloud_upsert_row(pg_cur, table, row):
    if not row:
        return
    cols = [c for c in row.keys()]
    if "id" in row and row["id"] is not None:
        pg_cur.execute(f"SELECT id FROM {table} WHERE id = ?", (row["id"],))
        exists = pg_cur.fetchone()
        if exists:
            set_cols = [c for c in cols if c != "id"]
            if not set_cols:
                return
            sets = ", ".join(f"{c}=?" for c in set_cols)
            vals = [row[c] for c in set_cols] + [row["id"]]
            pg_cur.execute(f"UPDATE {table} SET {sets} WHERE id = ?", vals)
            return
    col_list = ", ".join(cols)
    placeholders = ", ".join(["?"] * len(cols))
    vals = [row[c] for c in cols]
    try:
        pg_cur.execute(
            f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", vals
        )
    except Exception:
        if "id" in row and row["id"] is not None:
            set_cols = [c for c in cols if c != "id"]
            if set_cols:
                sets = ", ".join(f"{c}=?" for c in set_cols)
                vals = [row[c] for c in set_cols] + [row["id"]]
                pg_cur.execute(f"UPDATE {table} SET {sets} WHERE id = ?", vals)


def _cloud_upsert_by_key(pg_cur, table, key, row):
    if not row or key not in row:
        return
    cols = list(row.keys())
    pg_cur.execute(f"SELECT {key} FROM {table} WHERE {key} = ?", (row[key],))
    if pg_cur.fetchone():
        set_cols = [c for c in cols if c != key]
        if not set_cols:
            return
        sets = ", ".join(f"{c}=?" for c in set_cols)
        vals = [row[c] for c in set_cols] + [row[key]]
        pg_cur.execute(f"UPDATE {table} SET {sets} WHERE {key} = ?", vals)
    else:
        col_list = ", ".join(cols)
        placeholders = ", ".join(["?"] * len(cols))
        vals = [row[c] for c in cols]
        pg_cur.execute(
            f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", vals
        )


def _pg_columns_for_table(pg_cur, table, cache):
    key = str(table or "").lower()
    if not key:
        return set()
    if key in cache:
        return cache[key]
    cols = set()
    try:
        pg_cur.execute(
            "SELECT column_name FROM information_schema.columns "
            "WHERE table_schema = 'public' AND table_name = ?",
            (key,),
        )
        for row in pg_cur.fetchall() or []:
            if row and row[0]:
                cols.add(str(row[0]).lower())
    except Exception:
        cols = set()
    cache[key] = cols
    return cols


def _filter_row_for_postgres(pg_cur, table, row, cache):
    if not row:
        return {}
    allowed = _pg_columns_for_table(pg_cur, table, cache)
    if not allowed:
        try:
            pg_cur.execute(f"PRAGMA table_info({table})")
            for r in pg_cur.fetchall() or []:
                if r and len(r) > 1 and r[1]:
                    allowed.add(str(r[1]).lower())
            cache[str(table or "").lower()] = set(allowed)
        except Exception:
            pass
    if not allowed:
        return dict(row)
    return {k: v for k, v in row.items() if str(k).lower() in allowed}


def _is_duplicate_key_error(exc):
    msg = str(exc).lower()
    return (
        "duplicate key" in msg
        or "unique constraint" in msg
        or "already exists" in msg
        or "23505" in msg
    )


def _is_unrecoverable_queue_error(exc):
    """Schema/SQL errors that will never succeed — drop the queue item so later rows can upload."""
    if _is_connectivity_error(exc):
        return False
    if _is_duplicate_key_error(exc):
        return True
    msg = str(exc).lower()
    needles = (
        "does not exist",
        "undefined column",
        "undefinedcolumn",
        "undefined table",
        "undefinedtable",
        "syntax error",
        "invalid input syntax",
        "not-null constraint",
        "null value in column",
        "42703",
        "42p01",
        "23502",
    )
    return any(n in msg for n in needles)


def _apply_offline_payload(pg_cur, payload, col_cache):
    op = payload.get("op")
    table = payload.get("table")
    row = payload.get("row") or {}
    if op in ("insert", "upsert", "upsert_key", "rename_key"):
        row = _filter_row_for_postgres(pg_cur, table, row, col_cache)
        if op != "rename_key" and not row:
            raise ValueError(f"no matching columns for {table}")
        if str(table or "").lower() == "expenses":
            row = _remap_shop_employee_id(pg_cur, row)
    last_err = None
    for _ in range(12):
        try:
            remap = None
            orig_id = row.get("id") if op == "insert" else None
            if op == "insert":
                new_id = _cloud_insert_row(pg_cur, table, row)
                try:
                    if (
                        orig_id is not None
                        and new_id is not None
                        and int(orig_id) != int(new_id)
                        and str(table or "").lower() in ("payroll_records", "expenses")
                    ):
                        remap = (str(table).lower(), int(orig_id), int(new_id))
                except Exception:
                    remap = None
            elif op == "upsert":
                _cloud_upsert_row(pg_cur, table, row)
            elif op == "upsert_key":
                _cloud_upsert_by_key(pg_cur, table, payload.get("key"), row)
            elif op == "rename_key":
                key = payload.get("key") or "username"
                old = payload.get("old")
                if not table or not key or old is None:
                    raise ValueError("rename_key missing table/key/old")
                pg_cur.execute(
                    f"DELETE FROM {table} WHERE {key} = ?",
                    (old,),
                )
                if row:
                    _cloud_upsert_by_key(pg_cur, table, key, row)
            elif op == "delete":
                pg_cur.execute(
                    f"DELETE FROM {table} WHERE id = ?",
                    (payload.get("id"),),
                )
            elif op == "delete_key":
                pg_cur.execute(
                    f"DELETE FROM {table} WHERE {payload.get('key')} = ?",
                    (payload.get("value"),),
                )
            elif op == "sql":
                pg_cur.execute(payload.get("sql") or "", payload.get("params") or [])
            else:
                raise ValueError(f"unknown offline op {op!r}")
            return remap
        except Exception as e:
            last_err = e
            if op in ("insert", "upsert", "upsert_key", "rename_key"):
                stripped = _row_without_bad_column(row, e)
                if stripped is not None:
                    row = stripped
                    continue
                msg = str(e).lower()
                if "foreign key" in msg or "23503" in msg:
                    row = dict(row)
                    cleared = False
                    for fk in ("employee_id", "assignee_id"):
                        if row.get(fk) is not None:
                            row[fk] = None
                            cleared = True
                            break
                    if cleared:
                        continue
            raise
    if last_err:
        raise last_err


def flush_offline_queue_to_cloud():
    """Push queued offline mutations to Supabase. Returns (ok, message, flushed_count)."""
    global SUPABASE_HISTORY_ENABLED, _LAST_SYNC_ERROR
    if get_db_mode() != "supabase":
        return True, "not supabase", 0

    path = ensure_offline_cache_open()
    lite = _original_sqlite3_connect(path)
    try:
        cur = lite.cursor()
        cur.execute(
            "CREATE TABLE IF NOT EXISTS offline_sync_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, payload TEXT NOT NULL, created_at TEXT)"
        )
        cur.execute("SELECT id, payload FROM offline_sync_queue ORDER BY id ASC")
        rows = cur.fetchall() or []
    finally:
        lite.close()

    if not rows:
        return True, "nothing pending", 0

    history_prev = SUPABASE_HISTORY_ENABLED
    SUPABASE_HISTORY_ENABLED = False
    flushed = 0
    applied_ids = []
    remaps = []
    stop_reason = None
    try:
        try:
            pg_proxy = get_shared_supabase_conn(force_reconnect=False)
        except Exception:
            pg_proxy = get_shared_supabase_conn(force_reconnect=True)
        pg_cur = pg_proxy.cursor()
        col_cache = {}
        for qid, payload_raw in rows:
            try:
                payload = json.loads(payload_raw)
            except Exception:
                applied_ids.append(qid)
                continue
            try:
                remap = _apply_offline_payload(pg_cur, payload, col_cache)
                applied_ids.append(qid)
                flushed += 1
                if remap:
                    remaps.append(remap)
            except Exception as e:
                if _is_connectivity_error(e):
                    stop_reason = str(e)
                    _sync_log(f"flush connectivity stop #{qid}: {e}")
                    break
                if _is_duplicate_key_error(e):
                    applied_ids.append(qid)
                    flushed += 1
                    continue
                # Keep this queue item for retry, but still try later items so
                # one bad row cannot block every other expense from uploading.
                stop_reason = f"Failed applying offline change #{qid}: {e}"
                _LAST_SYNC_ERROR = stop_reason
                _sync_log(stop_reason)
                continue

        committed = False
        try:
            pg_proxy.commit()
            committed = True
        except Exception as e:
            if not stop_reason:
                stop_reason = str(e)
            try:
                pg_proxy.rollback()
            except Exception:
                pass

        # Fix sequences for tables with id
        if committed:
            try:
                raw = pg_proxy.conn.cursor()
                for tbl in ("employees", "payroll_records", "expenses", "shop_documents", "payout_tiers", "vagaro_pull_logs"):
                    try:
                        raw.execute(
                            f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), "
                            f"COALESCE((SELECT MAX(id) FROM {tbl}), 1))"
                        )
                    except Exception:
                        pass
                raw.close()
                pg_proxy.commit()
            except Exception:
                pass

        if committed and (applied_ids or remaps):
            lite = _original_sqlite3_connect(path)
            try:
                cur = lite.cursor()
                for qid in applied_ids:
                    cur.execute("DELETE FROM offline_sync_queue WHERE id = ?", (qid,))
                for table, old_id, new_id in remaps:
                    try:
                        cur.execute(f"SELECT id FROM {table} WHERE id = ?", (new_id,))
                        if cur.fetchone():
                            cur.execute(f"DELETE FROM {table} WHERE id = ?", (old_id,))
                        else:
                            cur.execute(
                                f"UPDATE {table} SET id = ? WHERE id = ?",
                                (new_id, old_id),
                            )
                        _repoint_queued_row_id(cur, table, old_id, new_id)
                    except Exception:
                        pass
                lite.commit()
            finally:
                lite.close()
            _persist_offline_cache()

        if stop_reason:
            return False, stop_reason, flushed
        return True, f"flushed {flushed}", flushed
    except Exception as e:
        return False, str(e), flushed
    finally:
        SUPABASE_HISTORY_ENABLED = history_prev

def _apply_queue_payload_to_sqlite(lcur, payload):
    """Re-apply a queued mutation to local SQLite after a cloud pull."""
    op = (payload or {}).get("op")
    table = (payload or {}).get("table")
    if op in ("insert", "upsert"):
        row = payload.get("row") or {}
        if not table or not row:
            return
        cols = list(row.keys())
        col_list = ", ".join(cols)
        placeholders = ", ".join(["?"] * len(cols))
        vals = [row[c] for c in cols]
        if "id" in row and row["id"] is not None:
            lcur.execute(f"DELETE FROM {table} WHERE id = ?", (row["id"],))
        lcur.execute(f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})", vals)
    elif op == "upsert_key":
        table = payload.get("table")
        key = payload.get("key")
        row = payload.get("row") or {}
        if not table or not key or key not in row:
            return
        cols = list(row.keys())
        col_list = ", ".join(cols)
        placeholders = ", ".join(["?"] * len(cols))
        lcur.execute(f"DELETE FROM {table} WHERE {key} = ?", (row[key],))
        lcur.execute(
            f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})",
            [row[c] for c in cols],
        )
    elif op == "delete" and table:
        lcur.execute(f"DELETE FROM {table} WHERE id = ?", (payload.get("id"),))
    elif op == "delete_key" and table:
        lcur.execute(
            f"DELETE FROM {table} WHERE {payload.get('key')} = ?",
            (payload.get("value"),),
        )
    elif op == "rename_key" and table:
        key = payload.get("key")
        old = payload.get("old")
        row = payload.get("row") or {}
        if key and old is not None:
            lcur.execute(f"DELETE FROM {table} WHERE {key} = ?", (old,))
        if key and row.get(key) is not None:
            lcur.execute(f"DELETE FROM {table} WHERE {key} = ?", (row[key],))
        if row:
            cols = list(row.keys())
            col_list = ", ".join(cols)
            placeholders = ", ".join(["?"] * len(cols))
            lcur.execute(
                f"INSERT INTO {table} ({col_list}) VALUES ({placeholders})",
                [row[c] for c in cols],
            )
    elif op == "sql":
        lcur.execute(payload.get("sql") or "", payload.get("params") or [])


def _queued_row_ids_for_table(lcur, table):
    ids = set()
    try:
        lcur.execute("SELECT payload FROM offline_sync_queue")
        for (raw,) in lcur.fetchall() or []:
            try:
                payload = json.loads(raw)
            except Exception:
                continue
            if str(payload.get("table") or "").lower() != str(table).lower():
                continue
            row = payload.get("row") or {}
            if row.get("id") is not None:
                try:
                    ids.add(int(row["id"]))
                except Exception:
                    pass
            if payload.get("id") is not None:
                try:
                    ids.add(int(payload["id"]))
                except Exception:
                    pass
    except Exception:
        pass
    return ids


def _expense_identity(rec):
    return (
        normalize_iso_date(rec.get("expense_date")),
        plain_label(rec.get("category")).lower(),
        plain_label(rec.get("location")).lower(),
        round(to_float(rec.get("amount"), 0.0), 2),
        plain_label(rec.get("description")).lower(),
    )


def _money_key(val):
    return round(to_float(decrypt_val(val) if val is not None else 0, 0.0), 2)


def _payroll_identity(rec):
    """Logical fingerprint so the same shop-earnings row is not uploaded twice."""
    if not rec:
        return None
    try:
        emp = rec.get("employee_id")
        emp_i = int(emp) if emp is not None and str(emp).strip() != "" else 0
    except Exception:
        emp_i = 0
    dt_raw = rec.get("record_date")
    dt_plain = plain_label(dt_raw) if dt_raw is not None else ""
    dt_s = normalize_iso_date(dt_plain) or str(dt_plain or "")[:10]
    return (
        emp_i,
        dt_s,
        _money_key(rec.get("revenue")),
        _money_key(rec.get("service_addon_sales")),
        _money_key(rec.get("hours")),
        _money_key(rec.get("calculation")),
        _money_key(rec.get("tip")),
        _money_key(rec.get("product_sales")),
        plain_label(rec.get("notes") or "").strip().lower(),
        plain_label(rec.get("location") or "").strip().lower(),
    )


def _dedupe_payroll_table(cur):
    """Keep one payroll_records row per identity; delete extra copies. Returns count removed."""
    try:
        cur.execute("SELECT * FROM payroll_records")
        rows = cur.fetchall() or []
        desc = cur.description
        if not desc or not rows:
            return 0
        cols = [d[0] for d in desc]
    except Exception:
        return 0
    groups = {}
    for row in rows:
        rec = {cols[i]: row[i] for i in range(min(len(cols), len(row)))}
        key = _payroll_identity(rec)
        try:
            rid = int(rec.get("id"))
        except Exception:
            continue
        groups.setdefault(key, []).append(rid)
    dropped = 0
    for ids in groups.values():
        uniq = sorted(set(ids))
        if len(uniq) < 2:
            continue
        for drop in uniq[1:]:
            try:
                cur.execute("DELETE FROM payroll_records WHERE id = ?", (drop,))
                dropped += 1
            except Exception:
                pass
    return dropped


def _remap_shop_employee_id(pg_cur, row):
    """Point envelope rows at the cloud Shop employee so FK inserts succeed."""
    if not row or not is_envelope_category(row.get("category")):
        return row
    try:
        pg_cur.execute("SELECT id, name FROM employees")
        for eid, name in pg_cur.fetchall() or []:
            if plain_label(name).lower() == "shop":
                row = dict(row)
                row["employee_id"] = eid
                return row
    except Exception:
        pass
    return row


def backfill_local_rows_missing_from_cloud():
    """Queue local rows that exist only on this PC (never uploaded because the queue failed)."""
    if get_db_mode() != "supabase":
        return 0
    path = ensure_offline_cache_open()
    lite = _original_sqlite3_connect(path, timeout=15)
    queued = 0
    try:
        try:
            lite.execute("PRAGMA busy_timeout=8000")
        except Exception:
            pass
        lcur = lite.cursor()
        lcur.execute(
            "CREATE TABLE IF NOT EXISTS offline_sync_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, payload TEXT NOT NULL, created_at TEXT)"
        )
        try:
            pg_cur = get_shared_supabase_conn().cursor()
        except Exception:
            return 0
        cloud_expense_ids = set()
        cloud_envelope_keys = set()
        cloud_payroll_keys = set()
        try:
            pg_cur.execute(
                "SELECT id, expense_date, category, location, amount, description FROM expenses"
            )
            for r in pg_cur.fetchall() or []:
                rec = {
                    "id": r[0],
                    "expense_date": r[1],
                    "category": r[2],
                    "location": r[3],
                    "amount": r[4],
                    "description": r[5],
                }
                try:
                    cloud_expense_ids.add(int(r[0]))
                except Exception:
                    pass
                if is_envelope_category(rec.get("category")):
                    cloud_envelope_keys.add(_expense_identity(rec))
        except Exception:
            cloud_expense_ids = set()
            cloud_envelope_keys = set()
        try:
            pg_cur.execute("SELECT * FROM payroll_records")
            prow = pg_cur.fetchall() or []
            pdesc = pg_cur.description
            pcols = [d[0] for d in pdesc] if pdesc else []
            for r in prow:
                prec = {pcols[i]: r[i] for i in range(min(len(pcols), len(r)))}
                cloud_payroll_keys.add(_payroll_identity(prec))
        except Exception:
            cloud_payroll_keys = set()

        for table in ("employees", "payroll_records", "expenses", "shop_documents"):
            queued_ids = _queued_row_ids_for_table(lcur, table)
            try:
                pg_cur.execute(f"SELECT id FROM {table}")
                cloud_ids = set()
                for r in pg_cur.fetchall() or []:
                    if r and r[0] is not None:
                        try:
                            cloud_ids.add(int(r[0]))
                        except Exception:
                            pass
            except Exception:
                continue
            if table == "expenses":
                cloud_ids = set(cloud_expense_ids)
            try:
                lcur.execute(f"PRAGMA table_info({table})")
                cols = [r[1] for r in lcur.fetchall() or [] if r and r[1]]
                if not cols:
                    continue
                lcur.execute(f"SELECT * FROM {table}")
                local_rows = lcur.fetchall() or []
            except Exception:
                continue
            for row in local_rows:
                rec = {c: row[i] for i, c in enumerate(cols)}
                rid = rec.get("id")
                if rid is None:
                    continue
                try:
                    rid_i = int(rid)
                except Exception:
                    continue
                if rid_i in queued_ids:
                    continue
                if table == "payroll_records":
                    if _payroll_identity(rec) in cloud_payroll_keys:
                        continue
                    payload_row = dict(rec)
                    if rid_i in cloud_ids:
                        payload_row.pop("id", None)
                    _queue_offline_op({"op": "insert", "table": table, "row": payload_row}, lite)
                    queued_ids.add(rid_i)
                    queued += 1
                    continue
                if table == "expenses" and is_envelope_category(rec.get("category")):
                    if _expense_identity(rec) in cloud_envelope_keys:
                        continue
                    payload_row = dict(rec)
                    if rid_i in cloud_ids:
                        payload_row.pop("id", None)
                    _queue_offline_op({"op": "insert", "table": table, "row": payload_row}, lite)
                    queued_ids.add(rid_i)
                    queued += 1
                    continue
                if rid_i in cloud_ids:
                    continue
                _queue_offline_op({"op": "insert", "table": table, "row": rec}, lite)
                queued_ids.add(rid_i)
                queued += 1
        try:
            lcur.execute(
                """
                CREATE TABLE IF NOT EXISTS user_action_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    log_uid TEXT UNIQUE,
                    created_at TEXT,
                    user_name TEXT,
                    action TEXT,
                    table_name TEXT,
                    record_id TEXT,
                    summary TEXT,
                    details TEXT
                )
                """
            )
            cloud_uids = set()
            try:
                pg_cur.execute("SELECT log_uid FROM user_action_log")
                for r in pg_cur.fetchall() or []:
                    if r and r[0]:
                        cloud_uids.add(plain_label(r[0]))
            except Exception:
                cloud_uids = set()
            lcur.execute(
                "SELECT log_uid, created_at, user_name, action, table_name, record_id, summary, details FROM user_action_log"
            )
            for row in lcur.fetchall() or []:
                uid = plain_label(row[0]) if row and row[0] else ""
                if not uid or uid in cloud_uids:
                    continue
                rec = {
                    "log_uid": uid,
                    "created_at": row[1],
                    "user_name": row[2],
                    "action": row[3],
                    "table_name": row[4],
                    "record_id": row[5],
                    "summary": row[6],
                    "details": row[7],
                }
                _queue_offline_op({"op": "insert", "table": "user_action_log", "row": rec}, lite)
                queued += 1
        except Exception:
            pass
        if queued:
            lite.commit()
        return queued
    except Exception:
        try:
            lite.rollback()
        except Exception:
            pass
        return queued
    finally:
        try:
            lite.close()
        except Exception:
            pass


def _row_category(use_cols, row):
    try:
        if "category" not in use_cols:
            return ""
        return row[use_cols.index("category")]
    except Exception:
        return ""


def _row_id_value(use_cols, row):
    try:
        if "id" not in use_cols:
            return None
        return int(row[use_cols.index("id")])
    except Exception:
        return None


def _next_sqlite_id(lcur, table, extra_taken=None):
    taken = set(extra_taken or ())
    try:
        lcur.execute(f"SELECT COALESCE(MAX(id), 0) FROM {table}")
        n = int((lcur.fetchone() or [0])[0] or 0)
    except Exception:
        n = 0
    n += 1
    while n in taken:
        n += 1
    return n


def _active_protected_expense_ids():
    """Just-saved local expense ids, only while the short protect window is open."""
    global _LOCAL_PROTECTED_EXPENSE_IDS, _SKIP_EXPENSE_PULL_UNTIL
    if time.time() >= _SKIP_EXPENSE_PULL_UNTIL:
        _LOCAL_PROTECTED_EXPENSE_IDS = set()
        return set()
    return set(_LOCAL_PROTECTED_EXPENSE_IDS)


def _repoint_queued_row_id(lcur, table, old_id, new_id):
    """Keep the upload queue pointing at a row after its local id was moved."""
    if old_id is None or new_id is None or old_id == new_id:
        return
    try:
        lcur.execute("SELECT id, payload FROM offline_sync_queue")
        rows = lcur.fetchall() or []
    except Exception:
        return
    for qid, raw in rows:
        try:
            payload = json.loads(raw)
        except Exception:
            continue
        if str(payload.get("table") or "").lower() != str(table).lower():
            continue
        changed = False
        row = payload.get("row")
        if isinstance(row, dict) and row.get("id") is not None:
            try:
                if int(row["id"]) == int(old_id):
                    row = dict(row)
                    row["id"] = int(new_id)
                    payload["row"] = row
                    changed = True
            except Exception:
                pass
        if payload.get("id") is not None:
            try:
                if int(payload["id"]) == int(old_id):
                    payload["id"] = int(new_id)
                    changed = True
            except Exception:
                pass
        if changed:
            try:
                lcur.execute(
                    "UPDATE offline_sync_queue SET payload = ? WHERE id = ?",
                    (json.dumps(payload, default=str), qid),
                )
            except Exception:
                pass


def _merge_cloud_rows_into_local(lcur, tbl, use_cols, packed_rows):
    """Upsert cloud rows without wiping local envelopes or unsynced records."""
    if "id" not in use_cols:
        lcur.execute(f"DELETE FROM {tbl}")
        if packed_rows:
            col_list = ", ".join(use_cols)
            placeholders = ", ".join(["?"] * len(use_cols))
            lcur.executemany(
                f"INSERT INTO {tbl} ({col_list}) VALUES ({placeholders})",
                packed_rows,
            )
        return
    col_list = ", ".join(use_cols)
    placeholders = ", ".join(["?"] * len(use_cols))
    id_idx = use_cols.index("id")
    queued_ids = _queued_row_ids_for_table(lcur, tbl)

    envelope_snaps = []
    local_envelope_ids = set()
    if tbl == "expenses" and "category" in use_cols:
        try:
            lcur.execute(f"SELECT {col_list} FROM {tbl}")
            for row in lcur.fetchall() or []:
                if is_envelope_category(_row_category(use_cols, row)):
                    envelope_snaps.append(list(row))
                    rid = _row_id_value(use_cols, row)
                    if rid is not None:
                        local_envelope_ids.add(rid)
        except Exception:
            envelope_snaps = []

    cloud_ids = set()
    cloud_envelope_ids = set()
    for row in packed_rows or []:
        rid = _row_id_value(use_cols, row)
        if rid is not None:
            cloud_ids.add(rid)
        if tbl == "expenses" and is_envelope_category(_row_category(use_cols, row)):
            if rid is not None:
                cloud_envelope_ids.add(rid)

    taken_ids = set(cloud_ids) | set(local_envelope_ids)

    # 1. Clean up local records deleted from cloud (except envelopes & un-pushed local queue)
    if cloud_ids:
        try:
            lcur.execute(f"SELECT id FROM {tbl}")
            local_all_ids = {r[0] for r in (lcur.fetchall() or []) if r and r[0] is not None}
            del_ids = local_all_ids - cloud_ids - queued_ids - local_envelope_ids
            for did in del_ids:
                lcur.execute(f"DELETE FROM {tbl} WHERE id = ?", (did,))
        except Exception:
            pass

    for row in packed_rows or []:
        row = list(row)
        rid = _row_id_value(use_cols, row)
        cloud_is_envelope = tbl == "expenses" and is_envelope_category(
            _row_category(use_cols, row)
        )
        if (
            tbl == "expenses"
            and rid is not None
            and rid in local_envelope_ids
            and not cloud_is_envelope
        ):
            new_id = _next_sqlite_id(lcur, tbl, taken_ids)
            try:
                lcur.execute(
                    f"UPDATE {tbl} SET id = ? WHERE id = ?",
                    (new_id, rid),
                )
                _repoint_queued_row_id(lcur, tbl, rid, new_id)
                local_envelope_ids.discard(rid)
                local_envelope_ids.add(new_id)
                taken_ids.add(new_id)
                if rid in queued_ids:
                    queued_ids.discard(rid)
                    queued_ids.add(new_id)
                for snap in envelope_snaps:
                    if _row_id_value(use_cols, snap) == rid:
                        snap[id_idx] = new_id
            except Exception:
                pass
        if rid is not None and rid in queued_ids:
            continue
        lcur.execute(
            f"INSERT OR REPLACE INTO {tbl} ({col_list}) VALUES ({placeholders})",
            tuple(row),
        )

    for snap in envelope_snaps:
        sid = _row_id_value(use_cols, snap)
        if sid is not None and sid in cloud_envelope_ids:
            continue
        if sid is not None and sid in cloud_ids:
            continue
        lcur.execute(
            f"INSERT OR REPLACE INTO {tbl} ({col_list}) VALUES ({placeholders})",
            tuple(snap),
        )
    if tbl == "payroll_records":
        try:
            _dedupe_payroll_table(lcur)
        except Exception:
            pass


def refresh_offline_cache_from_cloud():
    """Copy decrypted cloud tables into the local offline cache (for future offline use)."""
    if get_db_mode() != "supabase" or is_supabase_offline():
        return False
    path = ensure_offline_cache_open()
    try:
        pg_proxy = get_shared_supabase_conn()
        pg_cur = pg_proxy.cursor()
        lite = _original_sqlite3_connect(path, timeout=30)
        try:
            lite.execute("PRAGMA busy_timeout=8000")
            lcur = lite.cursor()
            _init_offline_schema(lcur)
            lcur.execute(
                "CREATE TABLE IF NOT EXISTS offline_sync_queue (id INTEGER PRIMARY KEY AUTOINCREMENT, payload TEXT NOT NULL, created_at TEXT)"
            )
            try:
                dropped = _dedupe_payroll_table(pg_cur)
                if dropped:
                    pg_proxy.commit()
            except Exception:
                try:
                    pg_proxy.rollback()
                except Exception:
                    pass
            for tbl in OFFLINE_SYNC_TABLES:
                try:
                    pg_cur.execute(f"SELECT * FROM {tbl}")
                    rows = pg_cur.fetchall() or []
                    desc = pg_cur.description
                    if not desc:
                        continue
                    cloud_cols = [d[0] for d in desc]
                    lcur.execute(f"PRAGMA table_info({tbl})")
                    local_cols = [r[1] for r in (lcur.fetchall() or []) if r and r[1]]
                    use_cols = [c for c in cloud_cols if c in local_cols]
                    if not use_cols:
                        continue
                    indexes = [cloud_cols.index(c) for c in use_cols]
                    packed = [
                        tuple(row[i] for i in indexes)
                        for row in rows
                    ]
                    # Never replace a table that has local data with an empty cloud result
                    # (a failed/blank SELECT would delete every envelope/expense).
                    if not packed:
                        try:
                            lcur.execute(f"SELECT COUNT(*) FROM {tbl}")
                            local_n = int((lcur.fetchone() or [0])[0] or 0)
                        except Exception:
                            local_n = 0
                        if local_n > 0:
                            continue
                    sp = f"sp_pull_{tbl}"
                    lcur.execute(f"SAVEPOINT {sp}")
                    try:
                        if tbl in ("expenses", "payroll_records"):
                            _merge_cloud_rows_into_local(lcur, tbl, use_cols, packed)
                        else:
                            lcur.execute(f"DELETE FROM {tbl}")
                            if packed:
                                col_list = ", ".join(use_cols)
                                placeholders = ", ".join(["?"] * len(use_cols))
                                lcur.executemany(
                                    f"INSERT INTO {tbl} ({col_list}) VALUES ({placeholders})",
                                    packed,
                                )
                        if "id" in local_cols:
                            try:
                                lcur.execute("DELETE FROM sqlite_sequence WHERE name=?", (tbl,))
                                lcur.execute(
                                    "INSERT INTO sqlite_sequence(name, seq) "
                                    f"SELECT ?, COALESCE(MAX(id), 0) FROM {tbl}",
                                    (tbl,),
                                )
                            except Exception:
                                pass
                        lcur.execute(f"RELEASE SAVEPOINT {sp}")
                    except Exception:
                        try:
                            lcur.execute(f"ROLLBACK TO SAVEPOINT {sp}")
                            lcur.execute(f"RELEASE SAVEPOINT {sp}")
                        except Exception:
                            pass
                except Exception:
                    continue
            try:
                _pull_user_action_logs(pg_cur, lcur)
            except Exception:
                pass
            try:
                lcur.execute("SELECT payload FROM offline_sync_queue ORDER BY id ASC")
                pending = lcur.fetchall() or []
                for (payload_raw,) in pending:
                    try:
                        payload = json.loads(payload_raw)
                        # expenses/payroll are merged (not wiped). Replaying those
                        # queue items would DELETE the other PC's just-pulled row
                        # and put a local envelope/expense back on the same id.
                        if str(payload.get("table") or "").lower() in (
                            "expenses",
                            "payroll_records",
                            "user_action_log",
                        ):
                            continue
                        _apply_queue_payload_to_sqlite(lcur, payload)
                    except Exception:
                        continue
            except Exception:
                pass
            lite.commit()
        finally:
            lite.close()
        _schedule_persist_offline_cache(0.5)
        return True
    except Exception:
        return False


def try_reconnect_supabase():
    """
    If offline, attempt to reach Supabase, flush the queue, and go back online.
    Returns (went_online: bool, status_text: str)
    """
    if get_db_mode() != "supabase":
        return False, ""
    pending = offline_pending_count()
    try:
        # Prefer the existing shared socket; only force-reconnect if the probe fails.
        try:
            conn = get_shared_supabase_conn(force_reconnect=False)
            cur = conn.cursor()
            cur.execute("SELECT 1")
            cur.fetchone()
        except Exception:
            conn = get_shared_supabase_conn(force_reconnect=True)
            cur = conn.cursor()
            cur.execute("SELECT 1")
            cur.fetchone()
    except Exception:
        return False, f"📴 Offline — changes saved locally ({pending} pending)"

    if is_supabase_offline() or pending:
        ok, msg, flushed = flush_offline_queue_to_cloud()
        if not ok:
            if _is_connectivity_error(Exception(msg or "")):
                enter_supabase_offline_mode(msg)
                enable_local_first_mode()
                return False, f"📴 Offline — sync paused ({msg})"
            global _LAST_SYNC_ERROR
            _LAST_SYNC_ERROR = msg or "upload failed"

    enable_local_first_mode()
    pending_after = offline_pending_count()
    if pending_after:
        return True, f"☁️ Online — {pending_after} changes still pending"
    return True, f"☁️ Synced {datetime.now().strftime('%H:%M:%S')}"


def cloud_data_fingerprint():
    """Change detector so live sync can skip full UI reloads when nothing changed."""
    try:
        path = OFFLINE_TEMP_DB_PATH or TEMP_DB_PATH
        if not path or path == SUPABASE_DB_SENTINEL:
            return None
        conn = _original_sqlite3_connect(path, timeout=8)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*), COALESCE(MAX(id), 0), COALESCE(SUM(id), 0) FROM payroll_records"
        )
        p = cur.fetchone() or (0, 0, 0)
        cur.execute(
            "SELECT COUNT(*), COALESCE(MAX(id), 0), COALESCE(SUM(id), 0), "
            "COALESCE(SUM(CAST(amount AS REAL)), 0) FROM expenses"
        )
        e = cur.fetchone() or (0, 0, 0, 0)
        sig = 0
        try:
            cur.execute(
                "SELECT id, amount, expense_date, category, status, ifnull(description,'') "
                "FROM expenses"
            )
            for row in cur.fetchall() or []:
                sig = (
                    sig
                    + hash(
                        (
                            row[0],
                            str(row[1]),
                            str(row[2]),
                            plain_label(row[3]),
                            plain_label(row[4]),
                            str(row[5])[:40],
                        )
                    )
                ) & 0x7FFFFFFF
        except Exception:
            pass
        cur.execute("SELECT COUNT(*), COALESCE(MAX(id), 0) FROM employees")
        em = cur.fetchone() or (0, 0)
        try:
            cur.execute("SELECT COUNT(*) FROM cash_month_locks")
            locks = cur.fetchone() or (0,)
        except Exception:
            locks = (0,)
        try:
            cur.execute("SELECT COUNT(*), COALESCE(MAX(id), 0) FROM shop_documents")
            docs = cur.fetchone() or (0, 0)
        except Exception:
            docs = (0, 0)
        try:
            cur.execute("SELECT COUNT(*), COALESCE(MAX(id), 0) FROM payout_tiers")
            tiers = cur.fetchone() or (0, 0)
        except Exception:
            tiers = (0, 0)
        try:
            cur.execute("SELECT COUNT(*) FROM config_locations")
            locs = cur.fetchone() or (0,)
        except Exception:
            locs = (0,)
        try:
            cur.execute("SELECT COUNT(*) FROM config_categories")
            cats = cur.fetchone() or (0,)
        except Exception:
            cats = (0,)
        try:
            cur.execute("SELECT COUNT(*) FROM config_payments")
            pmts = cur.fetchone() or (0,)
        except Exception:
            pmts = (0,)
        conn.close()
        return (
            int(p[0] or 0),
            int(p[1] or 0),
            int(p[2] or 0),
            int(e[0] or 0),
            int(e[1] or 0),
            int(e[2] or 0),
            round(to_float(e[3], 0.0), 2),
            int(sig),
            int(em[0] or 0),
            int(em[1] or 0),
            int(locks[0] or 0),
            int(docs[0] or 0),
            int(docs[1] or 0),
            int(tiers[0] or 0),
            int(locs[0] or 0),
            int(cats[0] or 0),
            int(pmts[0] or 0),
        )
    except Exception:
        return None


def _same_db_path(a, b):
    if not a or not b:
        return False
    if a == b:
        return True
    try:
        return os.path.normcase(os.path.abspath(str(a))) == os.path.normcase(
            os.path.abspath(str(b))
        )
    except Exception:
        return False


def db_connect(database, *args, **kwargs):
    # Route the app's working DB handle to Supabase when configured.
    # Do not pass factory=None through to sqlite — that raises
    # TypeError: 'NoneType' object is not callable.
    if database is None:
        if get_db_mode() == "supabase":
            database = SUPABASE_DB_SENTINEL
        else:
            database = TEMP_DB_PATH or ensure_offline_cache_open()
    if get_db_mode() == "supabase":
        is_sentinel = database == SUPABASE_DB_SENTINEL or (
            TEMP_DB_PATH == SUPABASE_DB_SENTINEL and database == TEMP_DB_PATH
        )
        use_local_cache = (not is_sentinel) and (
            _SUPABASE_OFFLINE
            or _same_db_path(database, OFFLINE_TEMP_DB_PATH)
            or (
                TEMP_DB_PATH
                and TEMP_DB_PATH != SUPABASE_DB_SENTINEL
                and _same_db_path(database, TEMP_DB_PATH)
            )
        )
        if use_local_cache:
            path = ensure_offline_cache_open()
            conn = _original_sqlite3_connect(path, timeout=15)
            try:
                conn.execute("PRAGMA journal_mode=WAL")
                conn.execute("PRAGMA synchronous=NORMAL")
                conn.execute("PRAGMA busy_timeout=8000")
            except Exception:
                pass
            return OfflineTrackingConnection(conn)
        if is_sentinel:
            try:
                return get_shared_supabase_conn()
            except Exception as e:
                try:
                    return get_shared_supabase_conn(force_reconnect=True)
                except Exception as e2:
                    if _is_connectivity_error(e2) or _is_connectivity_error(e):
                        enter_supabase_offline_mode(str(e2))
                        path = ensure_offline_cache_open()
                        conn = _original_sqlite3_connect(path, timeout=15)
                        try:
                            conn.execute("PRAGMA busy_timeout=8000")
                        except Exception:
                            pass
                        return OfflineTrackingConnection(conn)
                    raise sqlite3.OperationalError(
                        f"Failed to connect to Supabase: {str(e)}"
                    )
    if database is None:
        database = ensure_offline_cache_open()
    return _original_sqlite3_connect(database, *args, **kwargs)

sqlite3.connect = db_connect

DB_FILE = os.path.join(get_app_dir(), "payroll_data.enc")
EMPLOYEE_FOLDERS_DIR = os.path.join(get_app_dir(), "Employee_Folders")
os.makedirs(EMPLOYEE_FOLDERS_DIR, exist_ok=True)


def get_expense_docs_dir():
    docs_dir = os.path.join(get_app_dir(), "Expense_Documents")
    os.makedirs(docs_dir, exist_ok=True)
    return docs_dir


# Keep a module alias for older call sites; prefer get_expense_docs_dir() at runtime.
EXPENSE_DOCS_DIR = get_expense_docs_dir()


def resolve_local_doc_path(filepath):
    """Resolve cross-platform and cross-device document paths (Windows <-> Mac <-> Linux)."""
    if not filepath:
        return ""
    sp = str(filepath).strip()
    if not sp:
        return ""
    
    # 1. Direct path check
    if os.path.isfile(sp):
        return os.path.abspath(sp)
    
    norm = sp.replace("\\", "/")
    app_dir = get_app_dir()
    
    # 2. Check for known subfolders: Expense_Documents, Shop_Files, Employee_Folders
    for marker in ("Expense_Documents", "Shop_Files", "Employee_Folders"):
        if marker in norm:
            rel = norm[norm.index(marker):]
            candidate = os.path.join(app_dir, *rel.split("/"))
            if os.path.isfile(candidate):
                return os.path.abspath(candidate)
            return os.path.abspath(candidate)
            
    # 3. Check by basename
    base = os.path.basename(norm)
    for folder_name in ("Expense_Documents", "Shop_Files", "Employee_Folders"):
        candidate = os.path.join(app_dir, folder_name, base)
        if os.path.isfile(candidate):
            return os.path.abspath(candidate)
            
    return os.path.join(get_expense_docs_dir(), base)


def ensure_document_file_available(filepath):
    """Ensure the file is on local disk. If missing on this PC, fetch it from Supabase cloud_file_storage."""
    if not filepath:
        return ""
    local_path = resolve_local_doc_path(filepath)
    if os.path.isfile(local_path) and os.path.getsize(local_path) > 0:
        return local_path

    # If not found locally and cloud mode is active, try to fetch from Supabase
    if get_db_mode() == "supabase" and not is_supabase_offline():
        base_name = os.path.basename(str(filepath).replace("\\", "/"))
        stem, ext = os.path.splitext(base_name)
        lookup_keys = [base_name, f"{stem}.jpg", f"{stem}.jpeg", f"{stem}.png", f"{stem}.pdf", stem]
        placeholders = ", ".join(["%s"] * len(lookup_keys))
        try:
            db_conn = _open_supabase_pg_conn(timeout=15)
            try:
                cur = db_conn.cursor()
                cur.execute(
                    f"SELECT file_data FROM cloud_file_storage WHERE file_name IN ({placeholders}) OR file_key IN ({placeholders}) LIMIT 1",
                    tuple(lookup_keys + lookup_keys)
                )
                row = cur.fetchone()
                if row and row[0]:
                    raw_bytes = row[0]
                    if isinstance(raw_bytes, str):
                        import base64
                        try:
                            raw_bytes = base64.b64decode(raw_bytes)
                        except Exception:
                            raw_bytes = raw_bytes.encode("utf-8")
                    elif isinstance(raw_bytes, memoryview):
                        raw_bytes = raw_bytes.tobytes()
                    os.makedirs(os.path.dirname(local_path), exist_ok=True)
                    with open(local_path, "wb") as f:
                        f.write(raw_bytes)
                    return local_path
            finally:
                try:
                    db_conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    return local_path if os.path.isfile(local_path) else None


def _push_file_to_cloud_storage(file_path):
    """Save a compressed file copy into cloud_file_storage reliably."""
    if not file_path or not os.path.isfile(file_path):
        return
    if get_db_mode() != "supabase" or is_supabase_offline():
        return

    def _bg():
        try:
            with open(file_path, "rb") as f:
                data = f.read()
            import base64
            token = base64.b64encode(data).decode("ascii")
            base_name = os.path.basename(file_path)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db_conn = _open_supabase_pg_conn(timeout=15)
            try:
                cur = db_conn.cursor()
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS cloud_file_storage (
                        file_key TEXT PRIMARY KEY,
                        file_name TEXT,
                        file_data TEXT,
                        file_size INTEGER,
                        uploaded_at TEXT
                    )
                    """
                )
                db_conn.commit()
                cur.execute(
                    """
                    INSERT INTO cloud_file_storage (file_key, file_name, file_data, file_size, uploaded_at)
                    VALUES (%s, %s, %s, %s, %s)
                    ON CONFLICT (file_key) DO UPDATE SET file_data = EXCLUDED.file_data, uploaded_at = EXCLUDED.uploaded_at
                    """,
                    (base_name, base_name, token, len(data), now)
                )
                db_conn.commit()
            finally:
                try:
                    db_conn.close()
                except Exception:
                    pass
        except Exception:
            pass

    import threading
    threading.Thread(target=_bg, daemon=True).start()


def open_path_with_default_app(filepath):
    """Open a file with the OS default application, auto-downloading from cloud if missing."""
    actual_path = ensure_document_file_available(filepath)
    if not actual_path or not os.path.isfile(actual_path):
        raise FileNotFoundError(f"Document was not found: {filepath}")
    filepath = actual_path
    system = platform.system()
    if system == "Windows":
        os.startfile(filepath)
    elif system == "Darwin":
        subprocess.call(["open", filepath])
    else:
        subprocess.call(["xdg-open", filepath])


def print_path_with_default_app(filepath):
    """Send a file to the OS print handler."""
    if not filepath or not os.path.isfile(filepath):
        raise FileNotFoundError(filepath or "No file")
    system = platform.system()
    if system == "Windows":
        os.startfile(filepath, "print")
    elif system == "Darwin":
        subprocess.call(["lpr", filepath])
    else:
        subprocess.call(["lpr", filepath])


def parse_expense_documents(raw):
    """Return a list of document paths from DB (supports legacy single path)."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    if s.startswith("["):
        try:
            data = json.loads(s)
            if isinstance(data, list):
                return [str(p).strip() for p in data if str(p).strip()]
        except Exception:
            pass
    return [s]


def serialize_expense_documents(paths):
    cleaned = []
    seen = set()
    for p in paths or []:
        sp = str(p).strip()
        if not sp or sp in seen:
            continue
        seen.add(sp)
        cleaned.append(sp)
    if not cleaned:
        return None
    return json.dumps(cleaned)


def optimize_and_save_file(src_path, dest_folder, dest_filename=None, max_dim=1280, quality=80):
    """
    Copy a file to dest_folder. If it's an image (.jpg, .png, .jpeg, .webp, .bmp, .heic),
    automatically resize down to max_dim and compress as optimized JPEG (quality=80)
    to save 95%+ storage space while maintaining visual clarity.
    Non-image files (PDFs, docs) are safely copied.
    """
    if not src_path:
        return ""
    src_path = os.path.abspath(os.path.expanduser(str(src_path).strip()))
    if not os.path.isfile(src_path):
        return src_path

    os.makedirs(dest_folder, exist_ok=True)
    base_name = os.path.basename(src_path)
    stem, ext = os.path.splitext(base_name)
    ext_lower = ext.lower()

    if ext_lower in ('.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tiff'):
        out_name = dest_filename or f"{stem}.jpg"
        if not out_name.lower().endswith(".jpg") and not out_name.lower().endswith(".jpeg"):
            out_name = f"{os.path.splitext(out_name)[0]}.jpg"
        dest_path = os.path.join(dest_folder, out_name)
        try:
            from PIL import Image, ImageOps
            with Image.open(src_path) as img:
                try:
                    img = ImageOps.exif_transpose(img)
                except Exception:
                    pass
                if img.mode in ("RGBA", "P", "LA"):
                    img = img.convert("RGB")
                elif img.mode != "RGB":
                    img = img.convert("RGB")
                w, h = img.size
                if max(w, h) > max_dim:
                    resample = getattr(Image, "Resampling", Image).LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS
                    img.thumbnail((max_dim, max_dim), resample)
                img.save(dest_path, "JPEG", quality=quality, optimize=True)
                _push_file_to_cloud_storage(dest_path)
                return dest_path
        except Exception:
            pass

    # Fallback / Non-image copy
    out_name = dest_filename or base_name
    dest_path = os.path.join(dest_folder, out_name)
    if os.path.abspath(src_path) != os.path.abspath(dest_path):
        try:
            import shutil
            shutil.copy2(src_path, dest_path)
        except Exception:
            return src_path
    _push_file_to_cloud_storage(dest_path)
    return dest_path


def store_expense_document(src_path, expense_id):
    """Copy and compress an uploaded receipt/document into the app Expense_Documents folder."""
    if not src_path:
        return ""
    src_path = os.path.abspath(os.path.expanduser(str(src_path).strip()))
    if not os.path.isfile(src_path):
        return src_path
    docs_root = os.path.abspath(get_expense_docs_dir())
    if src_path.startswith(docs_root + os.sep) or src_path == docs_root:
        return src_path
    safe_base = "".join(c if (c.isalnum() or c in "._-") else "_" for c in os.path.basename(src_path))
    dest_name = f"expense_{expense_id or 'new'}_{int(time.time())}_{safe_base}"
    return optimize_and_save_file(src_path, docs_root, dest_filename=dest_name)


def store_expense_documents(src_paths, expense_id):
    """Copy many uploaded docs; returns list of stored paths."""
    out = []
    for i, src in enumerate(src_paths or []):
        # slight unique suffix when many files share a second timestamp
        stored = store_expense_document(src, f"{expense_id}_{i}" if expense_id is not None else f"new_{i}")
        if stored:
            out.append(stored)
    return out


def delete_expense_document_file(path):
    """Remove a stored expense doc file if it lives in Expense_Documents."""
    if not path:
        return
    try:
        abs_path = os.path.abspath(str(path))
        docs_root = os.path.abspath(get_expense_docs_dir())
        if abs_path.startswith(docs_root + os.sep) and os.path.isfile(abs_path):
            os.remove(abs_path)
    except Exception:
        pass


def get_shop_files_dir(location=None):
    root = os.path.join(get_app_dir(), "Shop_Files")
    os.makedirs(root, exist_ok=True)
    if not location:
        return root
    safe = "".join(c if (c.isalnum() or c in "._- ") else "_" for c in str(location).strip()).strip().replace(" ", "_")
    path = os.path.join(root, safe or "General")
    os.makedirs(path, exist_ok=True)
    return path


def store_shop_document_file(src_path, location, doc_id=None):
    if not src_path:
        return ""
    src_path = os.path.abspath(os.path.expanduser(str(src_path).strip()))
    if not os.path.isfile(src_path):
        return src_path
    dest_dir = get_shop_files_dir(location)
    if src_path.startswith(os.path.abspath(dest_dir) + os.sep):
        return src_path
    safe_base = "".join(c if (c.isalnum() or c in "._-") else "_" for c in os.path.basename(src_path))
    dest_name = f"shop_{doc_id or 'new'}_{int(time.time())}_{safe_base}"
    return optimize_and_save_file(src_path, dest_dir, dest_filename=dest_name)


def delete_shop_document_file(path):
    if not path:
        return
    try:
        abs_path = os.path.abspath(str(path))
        root = os.path.abspath(get_shop_files_dir())
        if abs_path.startswith(root + os.sep) and os.path.isfile(abs_path):
            os.remove(abs_path)
    except Exception:
        pass

def parse_period_dates(period_str):
    """Robustly parse start and end dates from spreadsheet period headers and strings."""
    if not period_str:
        return None
    s = str(period_str).strip()
    import re
    
    # 1. Check if period contains (MM/DD - MM/DD) or (MM/DD → MM/DD)
    m_paren = re.search(r'\((\d{1,2}/\d{1,2})\s*[-–—→to]+\s*(\d{1,2}/\d{1,2})\)', s)
    if m_paren:
        m_yr = re.search(r'\b(20\d\d)\b', s)
        yr = int(m_yr.group(1)) if m_yr else datetime.today().year
        try:
            d1_parts = [int(x) for x in m_paren.group(1).split('/')]
            d2_parts = [int(x) for x in m_paren.group(2).split('/')]
            d1 = datetime(yr, d1_parts[0], d1_parts[1])
            yr2 = yr + 1 if (d1_parts[0] == 12 and d2_parts[0] == 1) else yr
            d2 = datetime(yr2, d2_parts[0], d2_parts[1])
            return d1, d2
        except Exception:
            pass

    # 2. Check for ISO date range: YYYY-MM-DD to/through/-/– YYYY-MM-DD
    m_iso = re.search(r'(\d{4}-\d{2}-\d{2})\s*(?:[-–—to~/]|through|\.\.)\s*(\d{4}-\d{2}-\d{2})', s)
    if m_iso:
        try:
            d1 = datetime.strptime(m_iso.group(1), '%Y-%m-%d')
            d2 = datetime.strptime(m_iso.group(2), '%Y-%m-%d')
            return d1, d2
        except Exception:
            pass

    # 3. Check for MM/DD/YYYY to/through/-/– MM/DD/YYYY
    m_slash = re.search(r'(\d{1,2}/\d{1,2}/\d{2,4})\s*(?:[-–—to~/]|through|\.\.)\s*(\d{1,2}/\d{1,2}/\d{2,4})', s)
    if m_slash:
        fmt1 = '%m/%d/%Y' if len(m_slash.group(1).split('/')[-1]) == 4 else '%m/%d/%y'
        fmt2 = '%m/%d/%Y' if len(m_slash.group(2).split('/')[-1]) == 4 else '%m/%d/%y'
        try:
            d1 = datetime.strptime(m_slash.group(1), fmt1)
            d2 = datetime.strptime(m_slash.group(2), fmt2)
            return d1, d2
        except Exception:
            pass

    # 4. Check for 'Month DD, YYYY - Month DD, YYYY' or 'Month DD - Month DD, YYYY' or 'Month DD - DD, YYYY'
    m_words = re.search(r'([A-Za-z]+)\s+(\d{1,2})(?:st|nd|rd|th)?(?:,\s*(\d{4}))?\s*(?:[-–—]|to|through)\s*(?:([A-Za-z]+)\s+)?(\d{1,2})(?:st|nd|rd|th)?,\s*(\d{4})', s)
    if m_words:
        m1_name, d1_num, yr1_num, m2_name, d2_num, yr2_num = m_words.groups()
        yr1 = int(yr1_num) if yr1_num else int(yr2_num)
        yr2 = int(yr2_num)
        m2_str = m2_name if m2_name else m1_name
        for m_fmt in ('%B', '%b'):
            try:
                m1_dt = datetime.strptime(m1_name, m_fmt)
                m2_dt = datetime.strptime(m2_str, m_fmt)
                d1 = datetime(yr1, m1_dt.month, int(d1_num))
                d2 = datetime(yr2, m2_dt.month, int(d2_num))
                return d1, d2
            except Exception:
                pass

    # 5. Clean prefix 'Period:' and delimiters
    clean_s = re.sub(r'^(?:Period|Pay\s*Period|Cycle|Dates?)\s*:\s*', '', s, flags=re.IGNORECASE).strip()
    delims = [r'\s+to\s+', r'\s+through\s+', r'\s*–\s*', r'\s*—\s*', r'\s*-\s*', r'\s*\.\.\s*']
    for d_pat in delims:
        parts = re.split(d_pat, clean_s, flags=re.IGNORECASE)
        if len(parts) == 2:
            p1 = parts[0].strip().strip('()')
            p2 = parts[1].strip().strip('()')
            parsed_d1 = None
            parsed_d2 = None
            for fmt in ('%b %d,%Y', '%b %d, %Y', '%B %d,%Y', '%B %d, %Y', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
                if not parsed_d1:
                    try:
                        parsed_d1 = datetime.strptime(p1, fmt)
                    except ValueError:
                        pass
                if not parsed_d2:
                    try:
                        parsed_d2 = datetime.strptime(p2, fmt)
                    except ValueError:
                        pass
            if parsed_d1 and parsed_d2:
                return parsed_d1, parsed_d2

    return None

# --- PAY-CYCLE MODEL -------------------------------------------------------
# The shop runs two pay cycles per month, anchored on the 3rd and the 17th.
#   Cycle 1 of month M : M-03 .. M-16
#   Cycle 2 of month M : M-17 .. (M+1)-02  (so the 1st/2nd of a month belong to
#                                           the PREVIOUS month's Cycle 2, and the
#                                           31st belongs to that month's Cycle 2)
# A cycle is identified by a canonical, sortable key "YYYY-MM-H" where H is 1 or 2
# and MM is the anchor month. Labels look like "August - Cycle 1".
CYCLE_FIRST_ANCHOR = 3    # Cycle 1 starts on the 3rd
CYCLE_SECOND_ANCHOR = 17  # Cycle 2 starts on the 17th

def _parse_any_date(date_str):
    """Best-effort parse of a date string into a datetime, or None."""
    if date_str is None:
        return None
    if isinstance(date_str, datetime):
        return date_str
    s = str(date_str).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%b %d, %Y', '%B %d, %Y', '%b %d,%Y', '%B %d,%Y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d')
    except Exception:
        return None

# Continuous 14-day (bi-weekly) gapless payroll cycle cadence.
# Anchor cycle: 2026-08-03 (Monday) to 2026-08-16 (Sunday)
CYCLE_ANCHOR_DT = datetime(2026, 8, 3)

def cycle_for_date(date_str):
    """Return the canonical cycle_key (YYYY-MM-DD start date) for the 14-day cycle that owns the given date."""
    dt = _parse_any_date(date_str)
    if dt is None:
        return None
    d_only = datetime(dt.year, dt.month, dt.day)
    diff = (d_only - CYCLE_ANCHOR_DT).days
    idx = diff // 14
    c_start = CYCLE_ANCHOR_DT + timedelta(days=idx * 14)
    return c_start.strftime("%Y-%m-%d")

def parse_cycle_key(cycle_key):
    """Return (start_datetime, end_datetime) from a cycle_key (supports YYYY-MM-DD or legacy YYYY-MM-H)."""
    if not cycle_key:
        return None
    ck = str(cycle_key).strip()
    # YYYY-MM-DD format (10 chars)
    if len(ck) == 10 and ck[4] == '-' and ck[7] == '-':
        try:
            start = datetime.strptime(ck, "%Y-%m-%d")
            end = start + timedelta(days=13)
            return start, end
        except Exception:
            return None
    # Legacy YYYY-MM-H fallback
    parts = ck.split('-')
    if len(parts) == 3:
        try:
            y, m, h = int(parts[0]), int(parts[1]), int(parts[2])
            approx_d = 3 if h == 1 else 17
            canonical_ck = cycle_for_date(f"{y:04d}-{m:02d}-{approx_d:02d}")
            if canonical_ck:
                start = datetime.strptime(canonical_ck, "%Y-%m-%d")
                end = start + timedelta(days=13)
                return start, end
        except Exception:
            pass
    return None

def cycle_bounds(cycle_key):
    """Return (start_date, end_date) as 'YYYY-MM-DD' strings for a cycle_key."""
    parsed = parse_cycle_key(cycle_key)
    if parsed is None:
        return None
    start, end = parsed
    return start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')

def cycles_for_year(year):
    """All 14-day cycle_keys for a calendar year (gapless continuous bi-weekly schedule)."""
    jan1 = datetime(int(year), 1, 1)
    diff = (jan1 - CYCLE_ANCHOR_DT).days
    idx = diff // 14
    c_start = CYCLE_ANCHOR_DT + timedelta(days=idx * 14)
    
    cycles = []
    curr = c_start
    while curr.year <= int(year):
        cend = curr + timedelta(days=13)
        if curr.year == int(year) or cend.year == int(year):
            cycles.append(curr.strftime('%Y-%m-%d'))
        curr += timedelta(days=14)
    return cycles

def cycle_short_label(cycle_key):
    """Short label for cards e.g. 'Aug 03 - 16' or 'Aug 31 - Sep 13'."""
    import calendar
    parsed = parse_cycle_key(cycle_key)
    if parsed is None:
        return str(cycle_key or "")
    start, end = parsed
    s_m = calendar.month_abbr[start.month]
    e_m = calendar.month_abbr[end.month]
    if start.month == end.month:
        return f"{s_m} {start.day:02d} - {end.day:02d}"
    return f"{s_m} {start.day:02d} - {e_m} {end.day:02d}"

def cycle_label(cycle_key):
    """Full human label e.g. 'August 2026 - Cycle 1 (08/03 - 08/16)'."""
    import calendar
    b = cycle_bounds(cycle_key)
    if not b:
        return str(cycle_key or "")
    s = datetime.strptime(b[0], '%Y-%m-%d')
    e = datetime.strptime(b[1], '%Y-%m-%d')
    
    all_year = cycles_for_year(s.year)
    same_month_cycles = [c for c in all_year if datetime.strptime(c, '%Y-%m-%d').month == s.month]
    try:
        c_num = same_month_cycles.index(cycle_key) + 1
    except Exception:
        c_num = 1
    
    m_name = calendar.month_name[s.month]
    s_dt = s.strftime('%m/%d')
    e_dt = e.strftime('%m/%d')
    return f"{m_name} {s.year} - Cycle {c_num} ({s_dt} - {e_dt})"

def cycle_label_with_year(cycle_key):
    """Full human label e.g. 'August 2026 - Cycle 1 (08/03 - 08/16)'."""
    return cycle_label(cycle_key)

def last_completed_cycle_for_date(date_val=None):
    """Return the cycle_key for the most recently completed 14-day payroll cycle.
    If the cycle containing date_val is still in progress, the last completed cycle
    is the 14-day cycle immediately preceding it."""
    if date_val is None:
        dt = datetime.today()
    elif isinstance(date_val, str):
        dt = _parse_any_date(date_val) or datetime.today()
    else:
        dt = date_val
    cur_ck = cycle_for_date(dt.strftime("%Y-%m-%d"))
    if not cur_ck:
        return cur_ck
    parsed = parse_cycle_key(cur_ck)
    if not parsed:
        return cur_ck
    cur_start, _ = parsed
    last_ended_start = cur_start - timedelta(days=14)
    return last_ended_start.strftime("%Y-%m-%d")

def get_earliest_entry_date(default=None):
    """Return the earliest record_date or expense_date from the database as 'YYYY-MM-DD'."""
    if default is None:
        default = (datetime.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    try:
        if not TEMP_DB_PATH or not os.path.exists(TEMP_DB_PATH):
            return default
        conn = sqlite3.connect(TEMP_DB_PATH)
        c = conn.cursor()
        c.execute("""
            SELECT MIN(d) FROM (
                SELECT MIN(record_date) AS d FROM payroll_records WHERE record_date IS NOT NULL AND TRIM(record_date) != ''
                UNION ALL
                SELECT MIN(expense_date) AS d FROM expenses WHERE expense_date IS NOT NULL AND TRIM(expense_date) != ''
            )
        """)
        row = c.fetchone()
        conn.close()
        if row and row[0]:
            iso = normalize_iso_date(str(row[0]).strip())
            if iso:
                return iso
            return str(row[0]).strip()
    except Exception:
        pass
    return default


def get_formatted_cycle_choices(active_cycle_key=None, start_from_june_2026=True):
    """
    Generate pay cycle choices starting from June 2026 onwards for easy selection.
    Highlights/marks the active cycle with '✓ ' so users immediately know it's selected.
    Returns: (display_choices_list, cycle_key_map_dict, active_display_label)
    """
    cycle_key_map = {}
    display_choices = []
    
    cur_year = datetime.today().year
    years_to_scan = range(2026, max(cur_year + 2, 2028))
    
    all_candidate_cycles = []
    for yr in years_to_scan:
        for ck in cycles_for_year(yr):
            bounds = cycle_bounds(ck)
            if start_from_june_2026:
                # Filter out cycles that end before June 1, 2026
                if bounds and bounds[1] < "2026-06-01":
                    continue
            all_candidate_cycles.append(ck)
            
    # If active_cycle_key is before June 2026, include it at the top so existing record works
    if active_cycle_key and active_cycle_key not in all_candidate_cycles:
        all_candidate_cycles.insert(0, active_cycle_key)
        
    active_display = None
    for ck in all_candidate_cycles:
        raw_lbl = cycle_label_with_year(ck)
        no_yr_lbl = cycle_label(ck)
        
        is_active = (ck == active_cycle_key)
        disp_lbl = f"✓ {raw_lbl}" if is_active else f"   {raw_lbl}"
        
        display_choices.append(disp_lbl)
        cycle_key_map[disp_lbl] = ck
        cycle_key_map[raw_lbl] = ck
        cycle_key_map[no_yr_lbl] = ck
        cycle_key_map[f"✓ {raw_lbl}"] = ck
        cycle_key_map[f"   {raw_lbl}"] = ck
        cycle_key_map[ck] = ck
        
        if is_active:
            active_display = disp_lbl
            
    if not active_display and display_choices:
        active_display = display_choices[0]
        
    return display_choices, cycle_key_map, active_display

def add_cycles(cycle_key, delta):
    """Return the cycle_key offset by `delta` 14-day cycles (can be negative)."""
    parsed = parse_cycle_key(cycle_key)
    if parsed is None:
        return None
    start, _ = parsed
    n_start = start + timedelta(days=int(delta) * 14)
    return n_start.strftime('%Y-%m-%d')

DEFAULT_ENCRYPTION_PASSWORD = "default_encryption_password"
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin"

# --- INTEGRATION CONFIGURATION ---
# Replace with your actual active API credentials to query Vagaro services live
VAGARO_API_KEY = "YOUR_VAGARO_API_KEY_HERE"
VAGARO_CLIENT_ID = "YOUR_VAGARO_CLIENT_ID_HERE"
VAGARO_CLIENT_SECRET = "YOUR_VAGARO_CLIENT_SECRET_HERE"
VAGARO_API_ENDPOINT = "https://api.vagaro.com/v1/revenue"

# --- LICENSE / MACHINE LOCK CONFIGURATION ---
# Set to "ANY" to disable machine locking (good for your own testing).
# Set to "" (empty string) to make the app show a popup with the PC's unique ID on startup.
# Set to a specific ID (e.g. "A1B2C3D4E5F67890") to lock the app strictly to that client's PC.
AUTHORIZED_MACHINE_ID = "ANY"

# While logged in on Supabase, pull the other PC's changes this often (ms).
LIVE_SYNC_INTERVAL_MS = 30000  # Sync every 30 seconds
# How often to refresh the local offline cache while online (seconds).
OFFLINE_CACHE_PULL_SEC = 30
# Minimum seconds between forced UI reloads when cloud data is unchanged.
LIVE_SYNC_MIN_UI_REFRESH_SEC = 10

# --- INTERNAL GLOBALS ---
# Sentinel used as the sqlite3.connect() "database" argument in Supabase mode
# so the monkey-patched db_connect() can route to Postgres.
SUPABASE_DB_SENTINEL = "supabase://payroll"
TEMP_DB_PATH = None
CIPHER_SUITE = None
SALT = None

def get_cipher(password, salt):
    """Generates an AES key from the user's password and salt."""
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
    )
    key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
    return Fernet(key)

def commit_and_save(conn):
    """Commits to the temp SQLite database, then encrypts it to the permanent file."""
    if get_db_mode() == "supabase":
        conn.commit()
        if is_supabase_offline() or using_local_cache():
            _schedule_persist_offline_cache()
            schedule_cloud_push(0.05)
        return
    conn.commit()
    
    # Read the decrypted temporary database
    with open(TEMP_DB_PATH, "rb") as f:
        data = f.read()
        
    # Encrypt the raw bytes
    encrypted = CIPHER_SUITE.encrypt(data)
    
    # Write the Salt + Encrypted Data to the portable file
    with open(DB_FILE, "wb") as f:
        f.write(SALT + encrypted)

def cleanup():
    """Securely deletes the temporary decrypted database when the app closes and releases directory locks."""
    try:
        if getattr(sys, "frozen", False):
            app_dir = os.path.dirname(sys.executable)
            if os.path.isdir(app_dir):
                os.chdir(app_dir)
    except Exception:
        pass
    if get_db_mode() == "supabase":
        close_shared_supabase_conn()
        _persist_offline_cache()
        _close_offline_temp(remove_file=True)
        return
    if TEMP_DB_PATH and os.path.exists(TEMP_DB_PATH):
        try:
            os.remove(TEMP_DB_PATH)
        except:
            pass
            
atexit.register(cleanup)


def _existing_columns(cursor, table_name):
    """Return lowercase column names for a table (Postgres or SQLite)."""
    cols = set()
    try:
        cursor.execute(f"PRAGMA table_info({table_name})")
        rows = cursor.fetchall() or []
        if rows:
            for row in rows:
                if len(row) > 1 and row[1]:
                    cols.add(str(row[1]).lower())
            if cols:
                return cols
    except Exception:
        pass
    try:
        if get_db_mode() == "supabase" and not is_supabase_offline():
            cursor.execute(
                "SELECT column_name FROM information_schema.columns "
                "WHERE table_schema = 'public' AND table_name = ?",
                (table_name,),
            )
            for row in cursor.fetchall() or []:
                if row and row[0]:
                    cols.add(str(row[0]).lower())
    except Exception:
        pass
    return cols


def _add_missing_columns(cursor, table_name, columns):
    """columns: list of (name, sql_type_with_optional_default) e.g. ('tip', 'REAL')"""
    existing = _existing_columns(cursor, table_name)
    for col_name, col_type in columns:
        if col_name.lower() in existing:
            continue
        try:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}")
        except Exception:
            pass


def _repair_employee_names_and_shop(cursor):
    """Decode previously encrypted employee names and ensure a single Shop row."""
    needs_name_repair = False
    if get_db_mode() == "supabase" and not is_supabase_offline() and hasattr(cursor, "conn"):
        try:
            probe = cursor.conn.cursor()
            probe.execute("SELECT name FROM employees LIMIT 20")
            for (stored_name,) in probe.fetchall() or []:
                if isinstance(stored_name, str) and stored_name.startswith("enc:"):
                    needs_name_repair = True
                    break
            if not needs_name_repair:
                probe.execute("SELECT name FROM config_locations LIMIT 20")
                for (stored_name,) in probe.fetchall() or []:
                    if isinstance(stored_name, str) and stored_name.startswith("enc:"):
                        needs_name_repair = True
                        break
            probe.close()
        except Exception:
            needs_name_repair = True

    try:
        cursor.execute(
            "SELECT id, name, first_name, last_name FROM employees ORDER BY id ASC"
        )
        rows = cursor.fetchall() or []
    except Exception:
        return

    shop_ids = []
    normalized = []
    for row in rows:
        if not row:
            continue
        eid = row[0]
        name = decrypt_val(row[1]) if row[1] is not None else None
        first_name = decrypt_val(row[2]) if len(row) > 2 and row[2] is not None else ""
        last_name = decrypt_val(row[3]) if len(row) > 3 and row[3] is not None else ""
        normalized.append((eid, name, first_name, last_name))
        if name == "Shop":
            shop_ids.append(eid)

    # Fast path: already repaired and exactly one Shop.
    if not needs_name_repair and len(shop_ids) == 1:
        return
    if not needs_name_repair and len(shop_ids) == 0:
        try:
            cursor.execute(
                "INSERT INTO employees (name, first_name, last_name, hour_rate, percentage, use_tiered_payout) "
                "VALUES (?, ?, ?, 0.0, 0.0, 0)",
                ("Shop", "Shop", ""),
            )
        except Exception:
            pass
        return

    # Prefer raw SQL + savepoints on Supabase so UNIQUE conflicts cannot abort init_db.
    raw = getattr(cursor, "conn", None)
    if get_db_mode() == "supabase" and not is_supabase_offline() and raw is not None:
        ctrl = raw.cursor()
        try:
            ctrl.execute("SAVEPOINT sp_repair_names")
            if len(shop_ids) > 1:
                keep_id = min(shop_ids)
                for sid in shop_ids:
                    if sid == keep_id:
                        continue
                    ctrl.execute(
                        "UPDATE payroll_records SET employee_id = %s WHERE employee_id = %s",
                        (keep_id, sid),
                    )
                    ctrl.execute(
                        "UPDATE expenses SET employee_id = %s WHERE employee_id = %s",
                        (keep_id, sid),
                    )
                    ctrl.execute("DELETE FROM employees WHERE id = %s", (sid,))
                normalized = [r for r in normalized if r[0] == keep_id or r[0] not in shop_ids]

            for eid, name, first_name, last_name in normalized:
                ctrl.execute(
                    "UPDATE employees SET name = %s, first_name = %s, last_name = %s WHERE id = %s",
                    (
                        encrypt_val_deterministic(name),
                        encrypt_val_deterministic(first_name or ""),
                        encrypt_val_deterministic(last_name or ""),
                        eid,
                    ),
                )

            if not shop_ids:
                ctrl.execute(
                    "INSERT INTO employees (name, first_name, last_name, hour_rate, percentage, use_tiered_payout) "
                    "VALUES (%s, %s, %s, %s, %s, 0)",
                    (
                        encrypt_val_deterministic("Shop"),
                        encrypt_val_deterministic("Shop"),
                        encrypt_val_deterministic(""),
                        encrypt_val_deterministic(0.0),
                        encrypt_val_deterministic(0.0),
                    ),
                )

            for table in ("config_locations", "config_categories", "config_payments"):
                ctrl.execute(f"SELECT name FROM {table}")
                raw_rows = ctrl.fetchall() or []
                seen = set()
                for (stored_name,) in raw_rows:
                    plain = decrypt_val(stored_name)
                    if plain in seen or (
                        isinstance(stored_name, str)
                        and not stored_name.startswith("denc:")
                        and not stored_name.startswith("enc:")
                    ) or (isinstance(stored_name, str) and stored_name.startswith("enc:")):
                        ctrl.execute(f"DELETE FROM {table} WHERE name = %s", (stored_name,))
                    if plain not in seen:
                        seen.add(plain)
                        ctrl.execute(
                            f"INSERT INTO {table} (name) VALUES (%s) ON CONFLICT DO NOTHING",
                            (encrypt_val_deterministic(plain),),
                        )

            ctrl.execute("RELEASE SAVEPOINT sp_repair_names")
        except Exception:
            try:
                ctrl.execute("ROLLBACK TO SAVEPOINT sp_repair_names")
                ctrl.execute("RELEASE SAVEPOINT sp_repair_names")
            except Exception:
                pass
        finally:
            try:
                ctrl.close()
            except Exception:
                pass
        return

    # Local SQLite path
    if len(shop_ids) > 1:
        keep_id = min(shop_ids)
        for sid in shop_ids:
            if sid == keep_id:
                continue
            try:
                cursor.execute(
                    "UPDATE payroll_records SET employee_id = ? WHERE employee_id = ?",
                    (keep_id, sid),
                )
                cursor.execute(
                    "UPDATE expenses SET employee_id = ? WHERE employee_id = ?",
                    (keep_id, sid),
                )
                cursor.execute("DELETE FROM employees WHERE id = ?", (sid,))
            except Exception:
                pass
        normalized = [r for r in normalized if r[0] == keep_id or r[0] not in shop_ids]

    for eid, name, first_name, last_name in normalized:
        try:
            cursor.execute(
                "UPDATE employees SET name = ?, first_name = ?, last_name = ? WHERE id = ?",
                (name, first_name or "", last_name or "", eid),
            )
        except Exception:
            pass

    if not shop_ids:
        try:
            cursor.execute(
                "INSERT INTO employees (name, first_name, last_name, hour_rate, percentage, use_tiered_payout) "
                "VALUES (?, ?, ?, 0.0, 0.0, 0)",
                ("Shop", "Shop", ""),
            )
        except Exception:
            pass


def _table_count(cursor, table_name):
    try:
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        row = cursor.fetchone()
        return int(row[0]) if row and row[0] is not None else 0
    except Exception:
        return 0


def _cloud_is_empty(cursor):
    """True when cloud/local DB has no real app data yet."""
    try:
        cursor.execute("SELECT COUNT(*) FROM employees WHERE name != 'Shop' AND first_name != 'Shop'")
        emp_cnt = cursor.fetchone()[0]
    except Exception:
        emp_cnt = 0
    return _table_count(cursor, "users") <= 3 and emp_cnt == 0


def _seed_defaults_if_empty(cursor):
    """Seed login users / configs / Shop only for a brand-new empty database."""
    if _table_count(cursor, "users") == 0:
        hashed_default = hashlib.sha256("admin".encode()).hexdigest()
        for uname in ("admin", "moe", "ziad"):
            try:
                cursor.execute(
                    "INSERT OR IGNORE INTO users (username, password) VALUES (?, ?)",
                    (uname, hashed_default),
                )
            except Exception:
                pass

    if not _cloud_is_empty(cursor):
        return

    try:
        cursor.execute("SELECT COUNT(*) FROM config_locations")
        if cursor.fetchone()[0] == 0:
            cursor.executemany(
                "INSERT INTO config_locations (name) VALUES (?)",
                [("Shavano Park",), ("Stone Oak",)],
            )
    except Exception:
        pass

    try:
        cursor.execute("SELECT COUNT(*) FROM config_categories")
        if cursor.fetchone()[0] == 0:
            defaults = [
                ("Travel",), ("Equipment",), ("Office Supplies",), ("Meals",),
                ("Software",), ("Salary Payment",), ("Amazon Order",), ("Groceries",),
                ("Other",), ("Cash Envelope Received",),
            ]
            cursor.executemany("INSERT INTO config_categories (name) VALUES (?)", defaults)
    except Exception:
        pass

    try:
        cursor.execute("SELECT COUNT(*) FROM config_payments")
        if cursor.fetchone()[0] == 0:
            cursor.executemany(
                "INSERT INTO config_payments (name) VALUES (?)",
                [("Cash",), ("W2",), ("Cheque",), ("Gift Card",)],
            )
    except Exception:
        pass

    try:
        cursor.execute("SELECT id FROM employees WHERE name = ?", ("Shop",))
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO employees (name, first_name, last_name, hour_rate, percentage, use_tiered_payout) "
                "VALUES (?, ?, ?, 0.0, 0.0, 0)",
                ("Shop", "Shop", ""),
            )
    except Exception:
        pass


def cleanup_supabase_duplicates():
    """Remove duplicate Shop/config rows and wipe noisy history on the shared cloud DB."""
    if get_db_mode() != "supabase":
        return

    global TEMP_DB_PATH, SUPABASE_HISTORY_ENABLED
    if TEMP_DB_PATH != SUPABASE_DB_SENTINEL:
        TEMP_DB_PATH = SUPABASE_DB_SENTINEL
        init_supabase_cipher()

    history_prev = SUPABASE_HISTORY_ENABLED
    SUPABASE_HISTORY_ENABLED = False
    try:
        db_conn = _open_supabase_pg_conn(timeout=10)
        try:
            raw = db_conn.cursor()
            try:
                raw.execute("DELETE FROM database_history_log")
                db_conn.commit()
            except Exception:
                try:
                    db_conn.rollback()
                except Exception:
                    pass

            # Extra safety: collapse duplicate employee names (keep lowest id).
            try:
                raw.execute("SELECT id, name FROM employees ORDER BY id")
                rows = raw.fetchall() or []
                seen = {}
                for eid, name in rows:
                    plain = decrypt_val(name)
                    if plain in seen:
                        keep = seen[plain]
                        try:
                            raw.execute(
                                "UPDATE payroll_records SET employee_id = %s WHERE employee_id = %s",
                                (keep, eid),
                            )
                            raw.execute(
                                "UPDATE expenses SET employee_id = %s WHERE employee_id = %s",
                                (keep, eid),
                            )
                            raw.execute("DELETE FROM employees WHERE id = %s", (eid,))
                        except Exception:
                            pass
                    else:
                        seen[plain] = eid
                        if plain != name:
                            try:
                                raw.execute(
                                    "UPDATE employees SET name = %s WHERE id = %s",
                                    (plain, eid),
                                )
                            except Exception:
                                pass
                db_conn.commit()
            except Exception:
                try:
                    db_conn.rollback()
                except Exception:
                    pass

            try:
                raw.close()
            except Exception:
                pass
        finally:
            try:
                db_conn.close()
            except Exception:
                pass
    except Exception:
        pass
    finally:
        SUPABASE_HISTORY_ENABLED = history_prev


def _open_local_sqlite_readonly():
    """Decrypt the local payroll_data.enc into a temp SQLite DB and return (path, conn)."""
    if not os.path.exists(DB_FILE):
        raise FileNotFoundError(f"Local database not found: {DB_FILE}")

    with open(DB_FILE, "rb") as f:
        content = f.read()
    if len(content) < 17:
        raise ValueError("Local database file is empty or corrupt.")

    salt = content[:16]
    encrypted = content[16:]
    cipher = get_cipher(DEFAULT_ENCRYPTION_PASSWORD, salt)
    decrypted = cipher.decrypt(encrypted)

    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    with open(path, "wb") as f:
        f.write(decrypted)

    # Use the real sqlite connect, not the Supabase monkeypatch.
    lite = _original_sqlite3_connect(path)
    return path, lite


def ensure_all_supabase_tables(db_conn):
    """Ensure every required Postgres table and column exists on Supabase with native PostgreSQL types."""
    ddls = [
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS employees (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE,
            first_name TEXT,
            last_name TEXT,
            phone TEXT,
            email TEXT,
            hour_rate TEXT,
            percentage TEXT,
            ssn TEXT,
            address TEXT,
            start_date TEXT,
            end_date TEXT,
            cv_path TEXT,
            id_photo_path TEXT,
            personal_photo_path TEXT,
            vagaro_id TEXT,
            use_tiered_payout INTEGER DEFAULT 0
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS payroll_records (
            id SERIAL PRIMARY KEY,
            employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
            record_date TEXT,
            payment_amount TEXT,
            payment_type TEXT,
            revenue TEXT,
            hours TEXT,
            calculation TEXT,
            notes TEXT,
            written_up TEXT,
            location TEXT,
            product_sales TEXT,
            tip TEXT,
            written_up_desc TEXT,
            service_addon_sales TEXT DEFAULT '0.0',
            hour_rate TEXT,
            percentage TEXT,
            cycle_key TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS expenses (
            id SERIAL PRIMARY KEY,
            expense_date TEXT,
            category TEXT,
            amount TEXT,
            description TEXT,
            employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
            status TEXT,
            payment_type TEXT,
            location TEXT,
            is_tip TEXT DEFAULT 'No',
            assignee_id INTEGER,
            document_path TEXT,
            tip_given TEXT DEFAULT '0',
            cycle_key TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS shop_documents (
            id SERIAL PRIMARY KEY,
            location TEXT,
            title TEXT,
            doc_date TEXT,
            description TEXT,
            file_path TEXT,
            created_at TEXT,
            category TEXT,
            date_entered TEXT,
            notes TEXT,
            file_size INTEGER,
            file_type TEXT,
            filename TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS payout_tiers (
            id SERIAL PRIMARY KEY,
            from_sales TEXT,
            to_sales TEXT,
            percentage TEXT,
            kind TEXT DEFAULT 'service'
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS cash_month_locks (
            year_month TEXT PRIMARY KEY,
            locked_by TEXT,
            locked_at TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS vagaro_pull_logs (
            pulled_date TEXT PRIMARY KEY,
            pull_timestamp TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS database_history_log (
            id SERIAL PRIMARY KEY,
            change_timestamp TEXT,
            user_name TEXT,
            table_name TEXT,
            record_id INTEGER,
            action_type TEXT,
            old_data TEXT,
            new_data TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS user_action_log (
            id SERIAL PRIMARY KEY,
            log_uid TEXT UNIQUE,
            created_at TEXT,
            user_name TEXT,
            action TEXT,
            table_name TEXT,
            record_id TEXT,
            summary TEXT,
            details TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS cloud_backups (
            id SERIAL PRIMARY KEY,
            slot_key TEXT UNIQUE,
            backup_date TEXT,
            slot TEXT,
            created_at TEXT,
            created_by TEXT,
            payload TEXT,
            size_bytes INTEGER
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS config_locations (
            name TEXT PRIMARY KEY
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS config_categories (
            name TEXT PRIMARY KEY
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS config_payments (
            name TEXT PRIMARY KEY
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS cloud_file_storage (
            file_key TEXT PRIMARY KEY,
            file_name TEXT,
            file_data TEXT,
            file_size INTEGER,
            uploaded_at TEXT
        )
        """
    ]
    cur = db_conn.cursor()
    for ddl in ddls:
        try:
            cur.execute(ddl)
            db_conn.commit()
        except Exception:
            try:
                db_conn.rollback()
            except Exception:
                pass

    col_migrations = [
        ("employees", "first_name", "TEXT"),
        ("employees", "last_name", "TEXT"),
        ("employees", "phone", "TEXT"),
        ("employees", "email", "TEXT"),
        ("employees", "ssn", "TEXT"),
        ("employees", "address", "TEXT"),
        ("employees", "start_date", "TEXT"),
        ("employees", "end_date", "TEXT"),
        ("employees", "cv_path", "TEXT"),
        ("employees", "id_photo_path", "TEXT"),
        ("employees", "personal_photo_path", "TEXT"),
        ("employees", "vagaro_id", "TEXT"),
        ("employees", "use_tiered_payout", "INTEGER DEFAULT 0"),
        ("payroll_records", "location", "TEXT"),
        ("payroll_records", "product_sales", "TEXT"),
        ("payroll_records", "tip", "TEXT"),
        ("payroll_records", "written_up_desc", "TEXT"),
        ("payroll_records", "service_addon_sales", "TEXT DEFAULT '0.0'"),
        ("payroll_records", "hour_rate", "TEXT"),
        ("payroll_records", "percentage", "TEXT"),
        ("payroll_records", "cycle_key", "TEXT"),
        ("expenses", "payment_type", "TEXT"),
        ("expenses", "location", "TEXT"),
        ("expenses", "is_tip", "TEXT DEFAULT 'No'"),
        ("expenses", "assignee_id", "INTEGER"),
        ("expenses", "document_path", "TEXT"),
        ("expenses", "tip_given", "TEXT DEFAULT '0'"),
        ("expenses", "cycle_key", "TEXT"),
        ("payout_tiers", "kind", "TEXT DEFAULT 'service'"),
        ("shop_documents", "category", "TEXT"),
        ("shop_documents", "date_entered", "TEXT"),
        ("shop_documents", "notes", "TEXT"),
        ("shop_documents", "file_size", "INTEGER"),
        ("shop_documents", "file_type", "TEXT"),
        ("shop_documents", "filename", "TEXT"),
    ]
    for tbl, col, col_type in col_migrations:
        try:
            cur.execute(f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS {col} {col_type}")
            db_conn.commit()
        except Exception:
            try:
                db_conn.rollback()
            except Exception:
                pass


def sync_all_local_files_to_cloud(progress_cb=None):
    """Scan Expense_Documents, Shop_Files, and Employee_Folders and upload all documents to cloud_file_storage in Supabase."""
    if get_db_mode() != "supabase" or is_supabase_offline():
        return False, "Cloud is offline"
    
    app_dir = get_app_dir()
    folders_to_scan = [
        os.path.join(app_dir, "Expense_Documents"),
        os.path.join(app_dir, "Shop_Files"),
        os.path.join(app_dir, "Employee_Folders"),
    ]
    
    all_files = []
    for fld in folders_to_scan:
        if os.path.exists(fld):
            for root, _, files in os.walk(fld):
                for fname in files:
                    full_p = os.path.join(root, fname)
                    if os.path.isfile(full_p) and not fname.startswith(".") and not fname.endswith(".db"):
                        all_files.append(full_p)
                        
    if not all_files:
        return True, "No local files found to sync"
        
    try:
        db_conn = _open_supabase_pg_conn(timeout=15)
        cur = db_conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cloud_file_storage (
                file_key TEXT PRIMARY KEY,
                file_name TEXT,
                file_data TEXT,
                file_size INTEGER,
                uploaded_at TEXT
            )
            """
        )
        db_conn.commit()
        
        total = len(all_files)
        for idx, fpath in enumerate(all_files):
            if progress_cb:
                progress_cb(f"Syncing file {idx + 1}/{total}: {os.path.basename(fpath)}")
            try:
                with open(fpath, "rb") as f:
                    data = f.read()
                import base64
                token = base64.b64encode(data).decode("ascii")
                base_name = os.path.basename(fpath)
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cur.execute(
                    """
                    INSERT INTO cloud_file_storage (file_key, file_name, file_data, file_size, uploaded_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT (file_key) DO UPDATE SET file_data = EXCLUDED.file_data, uploaded_at = EXCLUDED.uploaded_at
                    """,
                    (base_name, base_name, token, len(data), now)
                )
                db_conn.commit()
            except Exception:
                try:
                    db_conn.rollback()
                except Exception:
                    pass
        try:
            db_conn.close()
        except Exception:
            pass
        return True, f"Successfully synced {total} files to cloud"
    except Exception as e:
        return False, str(e)


def get_supabase_storage_usage():
    """
    Calculate current storage used in Supabase (Database size + Cloud file storage bytes).
    Returns (used_bytes, total_bytes_1gb, percent_used, formatted_string)
    """
    total_bytes = 1024 * 1024 * 1024  # 1 GB in bytes
    if get_db_mode() != "supabase" or is_supabase_offline():
        return 0, total_bytes, 0.0, "Offline / Local Mode"
        
    try:
        db_conn = _open_supabase_pg_conn(timeout=10)
        try:
            cur = db_conn.cursor()
            
            # 1. Database table & index size
            try:
                cur.execute("SELECT pg_database_size(current_database());")
                r = cur.fetchone()
                db_size = r[0] if r and r[0] else 0
            except Exception:
                db_size = 0
                
            # 2. Uploaded documents & receipts size
            try:
                cur.execute("SELECT COALESCE(SUM(file_size), 0) FROM cloud_file_storage;")
                r = cur.fetchone()
                files_size = r[0] if r and r[0] else 0
            except Exception:
                files_size = 0
                
            total_used = int(db_size) + int(files_size)
            pct = min(100.0, (total_used / float(total_bytes)) * 100.0)
            
            used_mb = total_used / (1024.0 * 1024.0)
            total_mb = 1024.0
            
            status_text = f"{used_mb:.1f} MB / {total_mb:,.0f} MB ({pct:.1f}% of 1 GB Free Tier)"
            return total_used, total_bytes, pct, status_text
        finally:
            try:
                db_conn.close()
            except Exception:
                pass
    except Exception as e:
        return 0, total_bytes, 0.0, f"Error: {e}"


def upload_local_database_to_supabase(progress_cb=None):
    """
    First-sync: replace cloud tables with the contents of the local encrypted DB.
    Guarantees that the local database is NEVER modified, erased, or corrupted.
    """
    # 1. ALWAYS create an immutable local backup before cloud operations
    try:
        safety_key = f"pre_cloud_safety_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        create_local_backup(slot_key=safety_key)
    except Exception:
        pass

    if progress_cb:
        progress_cb("Reading local database (Safe Read-Only)...")

    global TEMP_DB_PATH, SUPABASE_HISTORY_ENABLED
    TEMP_DB_PATH = SUPABASE_DB_SENTINEL
    init_supabase_cipher()

    lite_path, lite_conn = _open_local_sqlite_readonly()
    history_prev = SUPABASE_HISTORY_ENABLED
    SUPABASE_HISTORY_ENABLED = False

    try:
        if progress_cb:
            progress_cb("Connecting to Supabase...")
        db_conn = _open_supabase_pg_conn(timeout=15)
        
        # 1. Ensure all native tables and columns exist in Postgres
        if progress_cb:
            progress_cb("Creating cloud tables...")
        ensure_all_supabase_tables(db_conn)
        ensure_numeric_columns_are_text()
        try:
            db_cur = db_conn.cursor()

            tables_to_clear = [
                "user_action_log",
                "database_history_log",
                "payroll_records",
                "expenses",
                "shop_documents",
                "payout_tiers",
                "cash_month_locks",
                "vagaro_pull_logs",
                "employees",
                "users",
                "config_locations",
                "config_categories",
                "config_payments",
                "cloud_backups",
            ]

            if progress_cb:
                progress_cb("Clearing cloud tables...")
            for tbl in tables_to_clear:
                try:
                    db_cur.execute(f"DELETE FROM {tbl}")
                    db_conn.commit()
                except Exception:
                    try:
                        db_conn.rollback()
                    except Exception:
                        pass

            copy_order = [
                "users",
                "employees",
                "config_locations",
                "config_categories",
                "config_payments",
                "payroll_records",
                "expenses",
                "shop_documents",
                "payout_tiers",
                "cash_month_locks",
                "vagaro_pull_logs",
                "user_action_log",
                "database_history_log",
            ]

            lite_cur = lite_conn.cursor()
            for tbl in copy_order:
                if progress_cb:
                    progress_cb(f"Uploading {tbl}...")
                try:
                    lite_cur.execute(f"PRAGMA table_info({tbl})")
                    cols = [r[1] for r in lite_cur.fetchall()]
                    if not cols:
                        continue
                    lite_cur.execute(f"SELECT * FROM {tbl}")
                    rows = lite_cur.fetchall()
                    if not rows:
                        continue

                    # Query cloud columns with reconnect retry
                    cloud_cols = set()
                    for _attempt in range(3):
                        try:
                            db_cur.execute(
                                "SELECT column_name FROM information_schema.columns "
                                "WHERE table_schema = 'public' AND table_name = %s",
                                (tbl,),
                            )
                            cloud_cols = {r[0] for r in (db_cur.fetchall() or [])}
                            break
                        except Exception as e:
                            if _is_dead_pg_error(e):
                                try:
                                    db_conn.close()
                                except Exception:
                                    pass
                                db_conn = _open_supabase_pg_conn()
                                db_cur = db_conn.cursor()
                            else:
                                raise

                    use_cols = [c for c in cols if c in cloud_cols]
                    if not use_cols:
                        continue
                    col_indexes = [cols.index(c) for c in use_cols]
                    col_list = ", ".join(use_cols)
                    placeholders = ", ".join(["%s"] * len(use_cols))
                    insert_sql = f"INSERT INTO {tbl} ({col_list}) VALUES ({placeholders})"
                    for row in rows:
                        cleaned = []
                        for idx, col_name in zip(col_indexes, use_cols):
                            val = row[idx]
                            # Normalize any legacy ciphertext, then encrypt for cloud at-rest storage
                            if isinstance(val, str) and (val.startswith("enc:") or val.startswith("denc:")):
                                val = decrypt_val(val)
                            if col_name.lower() in ALL_ENCRYPT_COLS:
                                cleaned.append(_encrypt_for_col(col_name, val))
                            else:
                                cleaned.append(val)
                        for _ins_attempt in range(3):
                            try:
                                db_cur.execute(insert_sql, cleaned)
                                break
                            except Exception as ins_e:
                                if _is_dead_pg_error(ins_e):
                                    try:
                                        db_conn.close()
                                    except Exception:
                                        pass
                                    db_conn = _open_supabase_pg_conn()
                                    db_cur = db_conn.cursor()
                                else:
                                    raise
                    if "id" in use_cols:
                        try:
                            db_cur.execute(
                                f"SELECT setval(pg_get_serial_sequence('{tbl}', 'id'), "
                                f"COALESCE((SELECT MAX(id) FROM {tbl}), 1))"
                            )
                        except Exception:
                            pass
                    try:
                        db_conn.commit()
                    except Exception:
                        pass
                except Exception as e:
                    raise RuntimeError(f"Failed uploading table {tbl}: {e}") from e
        finally:
            try:
                db_cur.close()
            except Exception:
                pass
            try:
                db_conn.close()
            except Exception:
                pass
        if progress_cb:
            progress_cb("Syncing documents & photos to cloud...")
        try:
            sync_all_local_files_to_cloud(progress_cb=progress_cb)
        except Exception:
            pass
        if progress_cb:
            progress_cb("Finalizing upload...")
        try:
            cleanup_supabase_duplicates()
        except Exception:
            pass
        if progress_cb:
            progress_cb("Done.")
    finally:
        SUPABASE_HISTORY_ENABLED = history_prev
        try:
            lite_conn.close()
        except Exception:
            pass
        try:
            os.remove(lite_path)
        except Exception:
            pass


def ensure_numeric_columns_are_text():
    """Postgres REAL columns cannot store denc:/enc: strings — convert them to TEXT."""
    if get_db_mode() != "supabase" or is_supabase_offline():
        return
    conn = get_shared_supabase_conn()
    raw = conn.conn.cursor()
    try:
        for table, cols in NUMERIC_TO_TEXT.items():
            for col in cols:
                try:
                    raw.execute("SAVEPOINT sp_coltype")
                    raw.execute(
                        "SELECT data_type, udt_name FROM information_schema.columns "
                        "WHERE table_schema='public' AND table_name=%s AND column_name=%s",
                        (table, col),
                    )
                    row = raw.fetchone()
                    if not row:
                        raw.execute("RELEASE SAVEPOINT sp_coltype")
                        continue
                    dtype = (row[0] or "").lower()
                    udt = (row[1] or "").lower() if len(row) > 1 else ""
                    if dtype in ("text", "character varying", "varchar", "character") or udt in (
                        "text",
                        "varchar",
                        "bpchar",
                    ):
                        raw.execute("RELEASE SAVEPOINT sp_coltype")
                        continue
                    # Force TEXT so encrypted denc:/enc: values can be stored.
                    raw.execute(
                        f'ALTER TABLE {table} ALTER COLUMN {col} TYPE TEXT '
                        f"USING TRIM(BOTH FROM ({col})::text)"
                    )
                    raw.execute("RELEASE SAVEPOINT sp_coltype")
                    conn.commit()
                except Exception:
                    try:
                        raw.execute("ROLLBACK TO SAVEPOINT sp_coltype")
                        raw.execute("RELEASE SAVEPOINT sp_coltype")
                    except Exception:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                    # Fallback ALTER without TRIM
                    try:
                        raw.execute("SAVEPOINT sp_coltype2")
                        raw.execute(
                            f"ALTER TABLE {table} ALTER COLUMN {col} TYPE TEXT "
                            f"USING ({col})::text"
                        )
                        raw.execute("RELEASE SAVEPOINT sp_coltype2")
                        conn.commit()
                    except Exception:
                        try:
                            raw.execute("ROLLBACK TO SAVEPOINT sp_coltype2")
                            raw.execute("RELEASE SAVEPOINT sp_coltype2")
                        except Exception:
                            try:
                                conn.rollback()
                            except Exception:
                                pass
    finally:
        try:
            raw.close()
        except Exception:
            pass
        conn.close()


def decrypt_plain_numeric_columns():
    """Rewrite legacy denc:/enc: money/rate values to plaintext numbers (no longer encrypted)."""
    if get_db_mode() != "supabase" or is_supabase_offline():
        return
    ensure_numeric_columns_are_text()
    conn = get_shared_supabase_conn()
    raw = conn.conn.cursor()
    try:
        for table, cols in PLAIN_NUMERIC_COLS.items():
            for col in cols:
                try:
                    raw.execute("SAVEPOINT sp_dec_num")
                    raw.execute(
                        "SELECT column_name FROM information_schema.columns "
                        "WHERE table_schema='public' AND table_name=%s AND column_name=%s",
                        (table, col),
                    )
                    if not raw.fetchone():
                        raw.execute("RELEASE SAVEPOINT sp_dec_num")
                        continue
                    raw.execute(f"SELECT id, {col} FROM {table}")
                    rows = raw.fetchall() or []
                    for row_id, val in rows:
                        if not isinstance(val, str):
                            continue
                        if not (val.startswith("denc:") or val.startswith("enc:")):
                            continue
                        plain = decrypt_val(val)
                        num = to_float(plain, 0.0)
                        raw.execute(
                            f"UPDATE {table} SET {col} = %s WHERE id = %s",
                            (str(num), row_id),
                        )
                    raw.execute("RELEASE SAVEPOINT sp_dec_num")
                    conn.commit()
                except Exception:
                    try:
                        raw.execute("ROLLBACK TO SAVEPOINT sp_dec_num")
                        raw.execute("RELEASE SAVEPOINT sp_dec_num")
                    except Exception:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
    finally:
        try:
            raw.close()
        except Exception:
            pass
        conn.close()


def migrate_supabase_encryption():
    """Encrypt text fields at rest; keep money/rate columns as plaintext numbers."""
    if get_db_mode() != "supabase":
        return
    ensure_numeric_columns_are_text()
    try:
        decrypt_plain_numeric_columns()
    except Exception:
        pass
    conn = get_shared_supabase_conn()
    raw = conn.conn.cursor()
    try:
        plans = [
            ("employees", "id", [
                "name", "first_name", "last_name", "phone", "email", "ssn", "address",
                "vagaro_id",
            ]),
            ("payroll_records", "id", [
                "location", "payment_type", "notes",
                "written_up", "written_up_desc",
            ]),
            ("expenses", "id", [
                "category", "description", "status", "payment_type", "location", "is_tip",
            ]),
            ("shop_documents", "id", [
                "location", "title", "description",
            ]),
            ("config_locations", "name", ["name"]),
            ("config_categories", "name", ["name"]),
            ("config_payments", "name", ["name"]),
            ("users", "username", ["username", "password"]),
            ("vagaro_pull_logs", "pulled_date", ["pulled_date", "pull_timestamp"]),
        ]
        for table, key_col, cols in plans:
            sp = f"sp_enc_{table}"
            try:
                raw.execute(f"SAVEPOINT {sp}")
                raw.execute(
                    "SELECT column_name FROM information_schema.columns "
                    "WHERE table_schema='public' AND table_name=%s",
                    (table,),
                )
                existing = {r[0] for r in (raw.fetchall() or [])}
                use_cols = [c for c in cols if c in existing]
                if key_col not in existing or not use_cols:
                    raw.execute(f"RELEASE SAVEPOINT {sp}")
                    continue
                select_cols = [key_col] + [c for c in use_cols if c != key_col]
                raw.execute(f"SELECT {', '.join(select_cols)} FROM {table}")
                rows = raw.fetchall() or []
                for row in rows:
                    key_val = row[0]
                    updates = []
                    vals = []
                    for i, col in enumerate(select_cols):
                        val = row[i]
                        if val is None:
                            continue
                        if isinstance(val, str) and val.startswith("denc:"):
                            continue
                        if isinstance(val, str) and val.startswith("enc:"):
                            if col in DET_ENCRYPT_COLS:
                                plain = decrypt_val(val)
                                updates.append(f"{col} = %s")
                                vals.append(_encrypt_for_col(col, plain))
                            continue
                        if col in ALL_ENCRYPT_COLS:
                            updates.append(f"{col} = %s")
                            vals.append(_encrypt_for_col(col, val))
                    if not updates:
                        continue
                    vals.append(key_val)
                    raw.execute(
                        f"UPDATE {table} SET {', '.join(updates)} WHERE {key_col} = %s",
                        vals,
                    )
                raw.execute(f"RELEASE SAVEPOINT {sp}")
                conn.commit()
            except Exception:
                try:
                    raw.execute(f"ROLLBACK TO SAVEPOINT {sp}")
                    raw.execute(f"RELEASE SAVEPOINT {sp}")
                except Exception:
                    pass
    finally:
        try:
            raw.close()
        except Exception:
            pass
        conn.close()


def init_db(defer_heavy_migrations=False):
    """Ensures schema exists. Seeds defaults only when the DB is completely empty."""
    global SUPABASE_HISTORY_ENABLED
    history_prev = SUPABASE_HISTORY_ENABLED
    SUPABASE_HISTORY_ENABLED = False
    path = TEMP_DB_PATH or ensure_offline_cache_open()
    conn = _original_sqlite3_connect(path, timeout=10)
    cursor = conn.cursor()
    try:
        _init_db_schema(cursor, seed=True)
        conn.commit()
    finally:
        SUPABASE_HISTORY_ENABLED = history_prev
        conn.close()


_CLOUD_MAINTENANCE_DONE = False


def run_deferred_cloud_maintenance():
    """Heavy Supabase work deferred until after login so startup stays fast.

    Uses a private socket under the shared lock so the UI connection is not
    interleaved / corrupted by background SELECT/UPDATE traffic.
    """
    global _CLOUD_MAINTENANCE_DONE, _SUPABASE_PG_CONN
    if _CLOUD_MAINTENANCE_DONE:
        return
    if get_db_mode() != "supabase" or is_supabase_offline():
        return
    with _SUPABASE_LOCK:
        private = None
        old = _SUPABASE_PG_CONN
        try:
            private = _open_supabase_pg_conn()
            _SUPABASE_PG_CONN = private
            try:
                migrate_supabase_encryption()
            except Exception:
                pass
            try:
                refresh_offline_cache_from_cloud()
            except Exception:
                pass
            _CLOUD_MAINTENANCE_DONE = True
        except Exception:
            pass
        finally:
            if private is not None:
                try:
                    private.close()
                except Exception:
                    pass
            _SUPABASE_PG_CONN = old


def _init_db_schema(cursor, seed=True):
    """Create / migrate schema. Seed only when empty so a 2nd PC just loads cloud data."""
    # On Supabase (online), money/rate columns are TEXT so encrypted ciphertext can be stored.
    num = "TEXT" if get_db_mode() == "supabase" and not is_supabase_offline() else "REAL"

    def _commit_step():
        try:
            if hasattr(cursor, "conn") and hasattr(cursor.conn, "commit"):
                cursor.conn.commit()
            elif hasattr(cursor, "connection") and hasattr(cursor.connection, "commit"):
                cursor.connection.commit()
        except Exception:
            pass

    # 0. Create database history log table
    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS database_history_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                change_timestamp TEXT,
                user_name TEXT,
                table_name TEXT,
                record_id INTEGER,
                action_type TEXT,
                old_data TEXT,
                new_data TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_action_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                log_uid TEXT UNIQUE,
                created_at TEXT,
                user_name TEXT,
                action TEXT,
                table_name TEXT,
                record_id TEXT,
                summary TEXT,
                details TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cloud_backups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                slot_key TEXT UNIQUE,
                backup_date TEXT,
                slot TEXT,
                created_at TEXT,
                created_by TEXT,
                payload TEXT,
                size_bytes INTEGER
            )
        ''')
        _commit_step()
    except Exception:
        pass

    # 1. Create Users table
    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                password TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass
    
    # 2. Create Employees
    try:
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                first_name TEXT,
                last_name TEXT,
                phone TEXT,
                email TEXT,
                hour_rate {num},
                percentage {num},
                use_tiered_payout INTEGER DEFAULT 0
            )
        ''')
        _commit_step()
    except Exception:
        pass
    
    # 3. Create Payroll Records
    try:
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS payroll_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                record_date TEXT,
                payment_amount {num},
                payment_type TEXT,
                revenue {num},
                hours {num},
                calculation {num},
                notes TEXT,
                written_up TEXT,
                FOREIGN KEY(employee_id) REFERENCES employees(id)
            )
        ''')
        _commit_step()
    except Exception:
        pass

    # Create Expenses table
    try:
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                expense_date TEXT,
                category TEXT,
                amount {num},
                description TEXT,
                employee_id INTEGER,
                status TEXT,
                payment_type TEXT,
                FOREIGN KEY(employee_id) REFERENCES employees(id)
            )
        ''')
        _commit_step()
    except Exception:
        pass

    # Create Vagaro Pull logs table
    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS vagaro_pull_logs (
                pulled_date TEXT PRIMARY KEY,
                pull_timestamp TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass

    # 4. Add only missing employee columns (avoids dozens of failing ALTER round-trips)
    _add_missing_columns(cursor, "employees", [
        ("first_name", "TEXT"), ("last_name", "TEXT"), ("phone", "TEXT"), ("email", "TEXT"),
        ("ssn", "TEXT"), ("address", "TEXT"), ("start_date", "TEXT"), ("end_date", "TEXT"),
        ("cv_path", "TEXT"), ("id_photo_path", "TEXT"), ("personal_photo_path", "TEXT"),
        ("vagaro_id", "TEXT"), ("use_tiered_payout", "INTEGER DEFAULT 0"),
    ])
    _commit_step()
            
    try:
        if get_db_mode() != "supabase":
            cursor.execute("UPDATE employees SET first_name = name, last_name = '', phone='', email='' WHERE first_name IS NULL OR first_name = ''")
            _commit_step()
    except Exception:
        pass

    # 5. Add only missing expense / payroll columns
    _add_missing_columns(cursor, "expenses", [
        ("payment_type", "TEXT"),
        ("location", "TEXT"),
        ("is_tip", "TEXT DEFAULT 'No'"),
        ("assignee_id", "INTEGER"),
        ("document_path", "TEXT"),
        ("tip_given", f"{num} DEFAULT 0"),
        ("cycle_key", "TEXT"),
    ])
    _commit_step()

    _add_missing_columns(cursor, "payroll_records", [
        ("location", "TEXT"),
        ("product_sales", "REAL"),
        ("tip", "REAL"),
        ("written_up_desc", "TEXT"),
        ("service_addon_sales", "REAL DEFAULT 0.0"),
        ("hour_rate", "REAL"),
        ("percentage", "REAL"),
        ("cycle_key", "TEXT"),
    ])
    _commit_step()

    # Backfill missing cycle_key values for historical records
    try:
        cursor.execute("SELECT id, record_date, notes FROM payroll_records WHERE cycle_key IS NULL OR TRIM(cycle_key)=''")
        for _r_id, _r_date, _r_notes in cursor.fetchall() or []:
            _ck = None
            if _r_notes:
                import re as _re_bf
                _m = _re_bf.search(r'Period:\s*([^|)]+)', str(_r_notes))
                if _m:
                    _pd = parse_period_dates(_m.group(1).strip())
                    if _pd:
                        _ck = cycle_for_date(_pd[0])
            if not _ck and _r_date:
                _ck = cycle_for_date(_r_date)
            if _ck:
                cursor.execute("UPDATE payroll_records SET cycle_key=? WHERE id=?", (_ck, _r_id))
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute("SELECT id, expense_date FROM expenses WHERE cycle_key IS NULL OR TRIM(cycle_key)=''")
        for _e_id, _e_date in cursor.fetchall() or []:
            if _e_date:
                _ck = cycle_for_date(_e_date)
                if _ck:
                    cursor.execute("UPDATE expenses SET cycle_key=? WHERE id=?", (_ck, _e_id))
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute("CREATE TABLE IF NOT EXISTS config_locations (name TEXT PRIMARY KEY)")
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute("CREATE TABLE IF NOT EXISTS config_categories (name TEXT PRIMARY KEY)")
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute("CREATE TABLE IF NOT EXISTS config_payments (name TEXT PRIMARY KEY)")
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cash_month_locks (
                year_month TEXT PRIMARY KEY,
                locked_by TEXT,
                locked_at TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass

    try:
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS payout_tiers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                from_sales {num},
                to_sales {num},
                percentage {num},
                kind TEXT DEFAULT 'service'
            )
        ''')
        _commit_step()
    except Exception:
        pass

    _add_missing_columns(cursor, "payout_tiers", [("kind", "TEXT DEFAULT 'service'")])
    _commit_step()
    try:
        cursor.execute("UPDATE payout_tiers SET kind='service' WHERE kind IS NULL OR TRIM(kind)=''")
        _commit_step()
    except Exception:
        pass

    def _seed_payout_kind(kind, defaults):
        try:
            cursor.execute(
                "SELECT COUNT(*) FROM payout_tiers WHERE LOWER(COALESCE(kind, 'service')) = ?",
                (kind,),
            )
            row = cursor.fetchone()
            if row and int(row[0] or 0) > 0:
                return
        except Exception:
            if kind != "service":
                return
            try:
                cursor.execute("SELECT COUNT(*) FROM payout_tiers")
                row = cursor.fetchone()
                if row and int(row[0] or 0) > 0:
                    return
            except Exception:
                return
        for a, b, p in defaults:
            try:
                cursor.execute(
                    "INSERT INTO payout_tiers (from_sales, to_sales, percentage, kind) VALUES (?, ?, ?, ?)",
                    (a, b, p, kind),
                )
            except Exception:
                pass

    _seed_payout_kind("service", [
        (0, 2500, 35),
        (2500.01, 3500, 38),
        (3500.01, 4500, 40),
        (4500.01, 6000, 45),
        (6000.01, 7500, 48),
        (7500.01, 9999999, 50),
    ])
    _seed_payout_kind("product", [
        (0, 149.99, 0),
        (150, 250, 15),
        (250.01, 9999999, 20),
    ])

    try:
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS shop_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                location TEXT,
                title TEXT,
                doc_date TEXT,
                description TEXT,
                file_path TEXT,
                created_at TEXT
            )
        ''')
        _commit_step()
    except Exception:
        pass

    # Repair encrypted-name leftovers / Shop dupes, then seed ONLY if empty.
    # Skip expensive repair on warm Supabase DBs that already have employees —
    # startup only needs schema + login users to be ready.
    skip_heavy_repair = False
    if get_db_mode() == "supabase" and not is_supabase_offline():
        try:
            cursor.execute("SELECT 1 FROM employees LIMIT 1")
            if cursor.fetchone():
                skip_heavy_repair = True
        except Exception:
            skip_heavy_repair = False
    # Enable Row Level Security (RLS) on all Supabase tables to resolve Supabase security advisory
    if get_db_mode() == "supabase" and not is_supabase_offline():
        all_pg_tables = [
            "users", "employees", "config_locations", "config_categories",
            "config_payments", "payroll_records", "expenses", "shop_documents",
            "payout_tiers", "cash_month_locks", "vagaro_pull_logs", "cloud_backups",
            "offline_sync_queue", "database_history_log"
        ]
        for tbl in all_pg_tables:
            try:
                cursor.execute(f"ALTER TABLE {tbl} ENABLE ROW LEVEL SECURITY")
            except Exception:
                pass

    if seed:
        _seed_defaults_if_empty(cursor)


if HAS_DEPS:
    
    class PayrollApp(tk.Tk):
        def __init__(self):
            super().__init__()
            try:
                self.tk = SafeTkProxy(self.tk)
            except Exception:
                pass
            self.style = tb.Style(theme=APP_THEME)
            self.title(APP_TITLE)
            
            # Start zoomed/maximized automatically (ideal for 13" MacBook keeping dock and top menu bar visible)
            try:
                self.wm_attributes("-zoomed", True)
            except Exception:
                try:
                    self.state('zoomed')
                except Exception:
                    self.geometry("1100x700")
            
            # Check Tcl/Tk version on macOS to warn if running the buggy 8.5 system version
            if platform.system() == "Darwin":
                try:
                    tcl_version = self.tk.eval('info patchlevel')
                    if tcl_version.startswith('8.5'):
                        messagebox.showwarning(
                            "Deprecated Tk Version",
                            f"Your macOS is using an old Tcl/Tk version ({tcl_version}) which is known to render window contents as completely blank.\n\n"
                            "To fix this, please download and install Python from the official python.org website, or run:\n"
                            "brew install tcl-tk",
                            parent=self
                        )
                except Exception:
                    pass

            self.style.configure('Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('primary.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('secondary.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('warning.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('info.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('success.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('danger.Treeview', rowheight=34, font=('Segoe UI', 10))
            self.style.configure('Treeview.Heading', font=('Segoe UI', 11, 'bold'), padding=6)
            self.style.configure('TButton', font=('Segoe UI', 11, 'bold'))
            self.style.configure('TNotebook.Tab', font=('Segoe UI', 12, 'bold'), padding=[15, 10])

            # Initialize Vagaro Sync Variables
            self.sync_employees_var = tk.BooleanVar(value=True)
            self.sync_revenue_var = tk.BooleanVar(value=True)
            self.sync_phone_var = tk.BooleanVar(value=False)
            self.sync_email_var = tk.BooleanVar(value=False)
            self.sync_expenses_var = tk.BooleanVar(value=False)

            # Inactivity tracker variables
            self.is_logged_in = False
            self.last_activity_time = time.time()
            
            self.bind_all("<Key>", self.reset_idle_timer)
            self.bind_all("<Button>", self.reset_idle_timer)
            self.bind_all("<Motion>", self.reset_idle_timer)
            self.bind_all("<MouseWheel>", self.reset_idle_timer)
            self.bind_all("<Button-4>", self.reset_idle_timer)
            self.bind_all("<Button-5>", self.reset_idle_timer)
            
            self.after(5000, self.check_idle_time)

            def _handle_tk_callback_error(exc_type, exc_val, exc_tb):
                if issubclass(exc_type, (tk.TclError,)):
                    msg = str(exc_val).lower()
                    if "can't delete tcl command" in msg or "application has been destroyed" in msg:
                        return  # Benign Python 3.14 / Windows Tkinter teardown race
                try:
                    import traceback
                    tb_str = "".join(traceback.format_exception(exc_type, exc_val, exc_tb))
                    with open(get_last_crash_log_path(), "a", encoding="utf-8") as f:
                        f.write(f"\n--- Tk Callback Error ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ---\n{tb_str}\n")
                except Exception:
                    pass
                if sys.stderr is not None:
                    try:
                        import traceback
                        traceback.print_exception(exc_type, exc_val, exc_tb)
                    except Exception:
                        pass

            self.report_callback_exception = _handle_tk_callback_error

            self.show_startup_splash()

        def show_startup_splash(self):
            self.clear_window()
            self._startup_offline_warning = None
            container = tb.Frame(self, padding=40)
            container.place(relx=0.5, rely=0.5, anchor=CENTER)
            
            lbl_b = tb.Label(container, text="★ Highend Payroll App ★", font=("Segoe UI", 22, "bold"), bootstyle="success")
            lbl_b.pack(pady=(0, 5))
            lbl_c = tb.Label(container, text="Custom Made ✂️", font=("Segoe UI", 12, "italic"), bootstyle="secondary")
            lbl_c.pack(pady=(0, 25))
            
            progress = tb.Progressbar(container, orient=tk.HORIZONTAL, length=380, mode="indeterminate", bootstyle="success")
            progress.pack(pady=10)
            progress.start(15)
            self._startup_progress = progress
            
            self.lbl_startup_status = tb.Label(
                container,
                text="Connecting to database…",
                font=("Segoe UI", 10, "italic"),
                bootstyle="secondary",
            )
            self.lbl_startup_status.pack()
            self.update_idletasks()
            # Run DB unlock/init off the UI thread so the window stays responsive.
            self.after(30, self._begin_startup_bootstrap)

        def _begin_startup_bootstrap(self):
            import threading

            def work():
                err = None
                try:
                    self.after(0, lambda: self._set_startup_status("Connecting to secure database…"))
                    self.unlock_database_silently()
                    self.after(0, lambda: self._set_startup_status("Preparing workspace…"))
                    init_db(defer_heavy_migrations=True)
                except Exception as e:
                    err = e
                self.after(0, lambda: self._finish_startup_bootstrap(err))

            threading.Thread(target=work, daemon=True).start()

        def _set_startup_status(self, text):
            lbl = getattr(self, "lbl_startup_status", None)
            if lbl is not None:
                try:
                    lbl.config(text=text)
                except Exception:
                    pass

        def _finish_startup_bootstrap(self, err):
            prog = getattr(self, "_startup_progress", None)
            if prog is not None:
                try:
                    prog.stop()
                except Exception:
                    pass
            if err is not None:
                try:
                    messagebox.showerror("Startup Error", f"Failed to open database:\n{err}", parent=self)
                except Exception:
                    pass
            warn = getattr(self, "_startup_offline_warning", None)
            if warn:
                try:
                    messagebox.showwarning("Working Offline", warn, parent=self)
                except Exception:
                    pass
                self._startup_offline_warning = None
            self._set_startup_status("Ready")
            self.show_login_page()
            
        def get_db_locations(self):
            cache = getattr(self, "_cache_locations", None)
            if cache is not None:
                return list(cache)
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM config_locations ORDER BY name ASC")
                res = [r[0] for r in cursor.fetchall()]
                conn.close()
                self._cache_locations = res
                return list(res)
            except Exception:
                return ["Shavano Park", "Stone Oak"]

        def get_db_categories(self):
            cache = getattr(self, "_cache_categories", None)
            if cache is not None:
                return list(cache)
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM config_categories ORDER BY name ASC")
                res = []
                for r in cursor.fetchall() or []:
                    name = plain_label(r[0] if r else None)
                    if name:
                        res.append(name)
                conn.close()
                self._cache_categories = res
                return list(res)
            except Exception:
                return ["Travel", "Equipment", "Office Supplies", "Meals", "Software", "Salary Payment", "Amazon Order", "Groceries", "Other", "Cash Envelope Received"]

        def get_shop_employee_id(self):
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute("SELECT id, name FROM employees")
                rows = cur.fetchall() or []
                conn.close()
                for emp_id, name in rows:
                    if plain_label(name).lower() == "shop":
                        return emp_id
            except Exception:
                pass
            return None

        def get_db_payments(self):
            cache = getattr(self, "_cache_payments", None)
            if cache is not None:
                return list(cache)
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM config_payments ORDER BY name ASC")
                res = [r[0] for r in cursor.fetchall()]
                conn.close()
                self._cache_payments = res
                return list(res)
            except Exception:
                return ["Cash", "W2", "Cheque", "Gift Card"]

        def invalidate_config_caches(self):
            self._cache_locations = None
            self._cache_categories = None
            self._cache_payments = None
            self._cache_payroll_cols = None
            self._last_sync_fingerprint = None
            if hasattr(self, "rebuild_fin_expense_cat_menu"):
                try:
                    self.rebuild_fin_expense_cat_menu()
                except Exception:
                    pass
            
        def unlock_database_silently(self):
            global TEMP_DB_PATH, CIPHER_SUITE, SALT
            
            if get_db_mode() == "supabase":
                init_supabase_cipher()
                # Work from the local cache immediately for instantaneous startup.
                enable_local_first_mode()
                return

            password = DEFAULT_ENCRYPTION_PASSWORD
            
            if os.path.exists(DB_FILE):
                with open(DB_FILE, "rb") as f:
                    content = f.read()
                SALT = content[:16]
                encrypted_data = content[16:]
                CIPHER_SUITE = get_cipher(password, SALT)
                
                try:
                    decrypted = CIPHER_SUITE.decrypt(encrypted_data)
                except Exception:
                    messagebox.showerror("Access Denied", "Failed to decrypt database. It may be corrupted or encrypted with a different key.")
                    sys.exit(1)
            else:
                # New Database
                SALT = os.urandom(16)
                CIPHER_SUITE = get_cipher(password, SALT)
                decrypted = b""
                
            # Create isolated temporary file for sqlite to use
            fd, TEMP_DB_PATH = tempfile.mkstemp(suffix=".db")
            os.close(fd)
            
            if decrypted:
                with open(TEMP_DB_PATH, "wb") as f:
                    f.write(decrypted)

        def clear_window(self):
            """Safely tear down the main UI. Ignores Tcl destroy races (DateEntry/ToolTip/busy)."""
            self.stop_filter_pollers()
            try:
                self.stop_live_sync()
            except Exception:
                pass

            for attr in ("_fin_refresh_after_id", "_busy_anim_id"):
                aid = getattr(self, attr, None)
                if aid is not None:
                    try:
                        self.after_cancel(aid)
                    except Exception:
                        pass
                    setattr(self, attr, None)

            try:
                self.hide_busy(force=True)
            except Exception:
                pass

            # Destroy Menubutton menus BEFORE parents — avoids "can't delete Tcl command"
            for attr in ("fin_exclude_menu", "fin_exp_cat_menu"):
                menu = getattr(self, attr, None)
                setattr(self, attr, None)
                if menu is None:
                    continue
                try:
                    menu.unpost()
                except Exception:
                    pass
                try:
                    menu.destroy()
                except Exception:
                    pass

            # Drop refs early so in-flight pollers/callbacks can't touch half-dead widgets
            for attr in (
                "fin_exclude_menu",
                "fin_exp_cat_menu",
                "btn_fin_exclude",
                "btn_fin_exp_cats",
                "cal_name_filter",
                "fin_emp_filter",
                "cbo_employee",
                "tree_calendar",
                "tree_names",
                "tree_fin_details",
                "tree_expenses",
                "tab_financials",
                "tab_calendar",
                "tab_names",
                "tab_data_entry",
                "tab_cash_cal",
                "notebook",
                "cal_from_date",
                "cal_to_date",
                "fin_from_date",
                "fin_to_date",
            ):
                if hasattr(self, attr):
                    setattr(self, attr, None)

            for attr in ("shop_files_win", "expenses_win", "folders_win"):
                win = getattr(self, attr, None)
                setattr(self, attr, None)
                if win is None:
                    continue
                try:
                    if win.winfo_exists():
                        try:
                            win.grab_release()
                        except Exception:
                            pass
                        win.destroy()
                except Exception:
                    pass

            for widget in list(self.winfo_children()):
                try:
                    if isinstance(widget, tk.Toplevel):
                        try:
                            widget.grab_release()
                        except Exception:
                            pass
                        widget.destroy()
                except Exception:
                    pass

            try:
                self.update_idletasks()
            except Exception:
                pass

            # Retry destroy — ttkbootstrap DateEntry/Notebook can fail once on Py3.14/Windows
            for _ in range(3):
                children = list(self.winfo_children())
                if not children:
                    break
                for widget in children:
                    try:
                        widget.destroy()
                    except Exception:
                        try:
                            self.tk.call("destroy", widget._w)
                        except Exception:
                            pass
                try:
                    self.update_idletasks()
                except Exception:
                    pass

        def _widget_alive(self, widget):
            """True if widget exists and is still a live Tk object (not None / destroyed)."""
            if widget is None:
                return False
            try:
                return bool(widget.winfo_exists())
            except Exception:
                return False

        def reset_idle_timer(self, event=None):
            self.last_activity_time = time.time()

        def _auto_save_all_pending_edits(self):
            """Safely commit any active cell entry or open dialog before auto-logout."""
            # 1. Unfocus active widget so any FocusOut/cell validation triggers
            try:
                focused = self.focus_get()
                if focused:
                    try:
                        focused.event_generate("<FocusOut>")
                    except Exception:
                        pass
            except Exception:
                pass

            # 2. Automatically save any open dialogs that registered a save function
            try:
                for widget in list(self.winfo_children()):
                    if isinstance(widget, tk.Toplevel) and widget.winfo_exists():
                        save_fn = getattr(widget, "_save_fn", None)
                        if callable(save_fn):
                            try:
                                save_fn()
                            except Exception:
                                pass
            except Exception:
                pass

            # 3. Check specific known popups like envelope / cash details
            try:
                pop = getattr(self, "_envelope_popup", None)
                if self._widget_alive(pop):
                    save_fn = getattr(pop, "_save_fn", None)
                    if callable(save_fn):
                        try:
                            save_fn()
                        except Exception:
                            pass
            except Exception:
                pass

        def check_idle_time(self):
            if not self.winfo_exists():
                return
            try:
                if getattr(self, 'is_logged_in', False):
                    elapsed = time.time() - self.last_activity_time
                    if elapsed >= 180:  # 3 minutes of inactivity
                        self.is_logged_in = False
                        self._auto_save_all_pending_edits()
                        self.logout()
                        try:
                            messagebox.showinfo(
                                "Auto-Saved & Logged Out",
                                "You have been safely logged out due to inactivity.\n\nAll your unsaved fields and open forms were automatically saved.",
                                parent=self
                            )
                        except Exception:
                            pass
            except Exception:
                pass
            finally:
                if self.winfo_exists():
                    self.after(5000, self.check_idle_time)

        def logout(self):
            self.is_logged_in = False
            self.stop_live_sync()
            self.hide_busy(force=True)
            for widget in list(self.winfo_children()):
                if isinstance(widget, tk.Toplevel):
                    try:
                        widget.destroy()
                    except Exception:
                        pass
            self.clear_window()
            self.show_login_page()

        def show_busy(self, message="Please wait…"):
            """Lightweight floating loader (does not cover the whole window). Cancelable."""
            if not self.winfo_exists():
                return
            self._busy_cancelled = False
            self._busy_depth = getattr(self, "_busy_depth", 0) + 1
            msg = message or "Please wait…"
            if getattr(self, "_busy_msg_var", None) is None:
                self._busy_msg_var = tk.StringVar(value=msg)
            else:
                self._busy_msg_var.set(msg)

            panel = getattr(self, "_busy_panel", None)
            if panel is not None and panel.winfo_exists():
                try:
                    self.update_idletasks()
                except Exception:
                    pass
                return

            # Floating card — background UI stays visible
            panel = tk.Frame(
                self,
                bg="#1e1e1e",
                padx=22,
                pady=18,
                highlightbackground="#4CAF50",
                highlightthickness=2,
            )
            panel.place(relx=0.5, rely=0.5, anchor="center")
            self._busy_panel = panel

            tk.Label(
                panel,
                text="Working…",
                font=("Segoe UI", 13, "bold"),
                fg="#ffffff",
                bg="#1e1e1e",
            ).pack(pady=(0, 4))
            tk.Label(
                panel,
                textvariable=self._busy_msg_var,
                font=("Segoe UI", 10),
                fg="#cccccc",
                bg="#1e1e1e",
                wraplength=280,
                justify="center",
            ).pack(pady=(0, 10))

            canvas = tk.Canvas(panel, width=84, height=84, bg="#1e1e1e", highlightthickness=0, bd=0)
            canvas.pack()
            self._busy_canvas = canvas
            self._busy_spin_angle = 0

            # True circular spinner: faint track ring + rotating arc
            pad = 14
            box = (pad, pad, 84 - pad, 84 - pad)
            canvas.create_oval(*box, outline="#3a3a3a", width=5)
            self._busy_arc = canvas.create_arc(
                *box,
                start=0,
                extent=78,
                style="arc",
                outline="#4CAF50",
                width=5,
            )

            def _animate_spinner():
                if not getattr(self, "_busy_panel", None) or not panel.winfo_exists():
                    return
                try:
                    self._busy_spin_angle = (getattr(self, "_busy_spin_angle", 0) + 14) % 360
                    canvas.itemconfig(self._busy_arc, start=self._busy_spin_angle)
                    self._busy_anim_id = self.after(35, _animate_spinner)
                except Exception:
                    pass

            _animate_spinner()
            def _cancel():
                self._busy_cancelled = True
                self.hide_busy(force=True)
                try:
                    # Drop sticky cloud socket if a hang is likely
                    if get_db_mode() == "supabase":
                        close_shared_supabase_conn()
                except Exception:
                    pass

            tk.Button(
                panel,
                text="Cancel",
                font=("Segoe UI", 10, "bold"),
                bg="#333333",
                fg="#ffffff",
                activebackground="#555555",
                activeforeground="#ffffff",
                relief="flat",
                padx=14,
                pady=4,
                cursor="hand2",
                command=_cancel,
            ).pack(pady=(12, 0))

            try:
                self.config(cursor="watch")
            except Exception:
                pass
            try:
                panel.lift()
                self.update_idletasks()
            except Exception:
                pass

        def hide_busy(self, force=False):
            if force:
                self._busy_depth = 0
            else:
                self._busy_depth = max(0, getattr(self, "_busy_depth", 1) - 1)
            if getattr(self, "_busy_depth", 0) > 0:
                return
            anim = getattr(self, "_busy_anim_id", None)
            if anim is not None:
                try:
                    self.after_cancel(anim)
                except Exception:
                    pass
                self._busy_anim_id = None
            panel = getattr(self, "_busy_panel", None)
            self._busy_panel = None
            if panel is not None:
                try:
                    panel.destroy()
                except Exception:
                    pass
            try:
                self.config(cursor="")
            except Exception:
                pass

        def is_busy_cancelled(self):
            return bool(getattr(self, "_busy_cancelled", False))

        def run_busy(self, message, fn, *args, **kwargs):
            self.show_busy(message)
            try:
                return fn(*args, **kwargs)
            finally:
                self.hide_busy()

        def show_app_error(self, title, error, parent=None):
            """Error popup with optional Save-to-file for support/debugging."""
            import traceback as _tb
            parent = parent or self
            if isinstance(error, BaseException):
                detail = "".join(_tb.format_exception(type(error), error, error.__traceback__))
                short = str(error)
            else:
                detail = str(error)
                short = detail

            win = tb.Toplevel(parent)
            win.title(title or "Error")
            win.geometry("520x360")
            win.transient(parent)
            try:
                win.grab_set()
            except Exception:
                pass

            tb.Label(win, text=title or "Error", font=("Segoe UI", 14, "bold"), bootstyle="danger").pack(
                anchor=W, padx=16, pady=(14, 6)
            )
            tb.Label(win, text=short, font=("Segoe UI", 10), wraplength=480, justify=LEFT).pack(
                anchor=W, padx=16, pady=(0, 8)
            )

            txt = tk.Text(win, height=10, wrap="word", font=("Consolas", 9))
            txt.pack(fill=BOTH, expand=True, padx=16, pady=6)
            txt.insert("1.0", detail)
            txt.config(state="disabled")

            btn_row = tb.Frame(win)
            btn_row.pack(fill=X, padx=16, pady=12)

            def _save_log():
                default_name = f"payroll_error_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                path = filedialog.asksaveasfilename(
                    parent=win,
                    title="Save error log",
                    defaultextension=".txt",
                    initialfile=default_name,
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                )
                if not path:
                    return
                try:
                    with open(path, "w", encoding="utf-8") as f:
                        f.write(f"App: {APP_TITLE}\n")
                        f.write(f"Time: {datetime.now().isoformat()}\n")
                        f.write(f"Mode: {get_db_mode()}\n")
                        f.write(f"Title: {title}\n\n")
                        f.write(detail)
                    messagebox.showinfo("Saved", f"Error log saved to:\n{path}", parent=win)
                except Exception as e:
                    messagebox.showerror("Save Failed", str(e), parent=win)

            tb.Button(btn_row, text="Save error log…", bootstyle="warning", command=_save_log).pack(side=LEFT)
            tb.Button(btn_row, text="Close", bootstyle="secondary", command=win.destroy).pack(side=RIGHT)

        def open_app_updates_dialog(self):
            """Open the Cloud Software Updates and Fail-Safe Recovery panel in a standalone modal window."""
            dlg = tb.Toplevel(self)
            dlg.title(self._tr("In-App Updates & Fail-Safe Recovery"))
            dlg.geometry("820x650")
            dlg.minsize(700, 520)
            try:
                dlg.transient(self)
                dlg.grab_set()
            except Exception:
                pass

            bottom_bar = tb.Frame(dlg, padding=(14, 10))
            bottom_bar.pack(side=BOTTOM, fill=X)
            tb.Button(bottom_bar, text=self._tr("Close"), bootstyle="secondary", width=12, command=dlg.destroy).pack(side=RIGHT)

            content_frame = tb.Frame(dlg)
            content_frame.pack(side=TOP, fill=BOTH, expand=True)

            self._build_app_updates_panel(content_frame, dlg)

        def _do_quick_rollback_from_login(self):
            """Quick rollback or revert handler invoked directly from the login screen."""
            bak_path = os.path.join(get_updates_dir(), "payroll_app.py.bak")
            has_bak = os.path.isfile(bak_path)
            dyn_path = get_updates_script_path()
            has_dyn = os.path.isfile(dyn_path)

            if not has_dyn and not has_bak:
                messagebox.showinfo(
                    "Factory Version",
                    "The application is already running its original built-in factory version.\nThere are no cloud updates installed to roll back.",
                    parent=self,
                )
                return

            msg = "Select a recovery option:\n\n"
            if has_bak:
                msg += "• Click [Yes] to restore the previous update backup (.bak)\n"
                msg += "• Click [No] to revert completely to the original factory built-in version\n"
                msg += "• Click [Cancel] to keep current engine"
                ans = messagebox.askyesnocancel("Revert / Rollback Engine", msg, parent=self)
                if ans is True:
                    ok, res = rollback_cloud_update(target="bak")
                    if ok:
                        ans_r = messagebox.askyesnocancel(
                            "Rollback Applied",
                            f"{res}\n\nTo apply the rollback:\n• [Yes] = Restart app automatically\n• [No] = Shutdown app cleanly now (reopen manually)\n• [Cancel] = Continue current session",
                            parent=self,
                        )
                        if ans_r is True:
                            restart_app()
                        elif ans_r is False:
                            self.shutdown_app()
                    else:
                        messagebox.showerror("Rollback Failed", res, parent=self)
                elif ans is False:
                    ok, res = rollback_cloud_update(target="factory")
                    if ok:
                        ans_r = messagebox.askyesnocancel(
                            "Reverted to Factory",
                            f"{res}\n\nTo apply the built-in version:\n• [Yes] = Restart app automatically\n• [No] = Shutdown app cleanly now (reopen manually)\n• [Cancel] = Continue current session",
                            parent=self,
                        )
                        if ans_r is True:
                            restart_app()
                        elif ans_r is False:
                            self.shutdown_app()
                    else:
                        messagebox.showerror("Revert Failed", res, parent=self)
            else:
                if messagebox.askyesno(
                    "Revert to Factory Version",
                    "No previous backup was found, but a cloud update is installed.\n\n"
                    "Would you like to revert completely to the original factory built-in version?",
                    parent=self,
                ):
                    ok, res = rollback_cloud_update(target="factory")
                    if ok:
                        ans_r = messagebox.askyesnocancel(
                            "Reverted to Factory",
                            f"{res}\n\nTo apply the built-in version:\n• [Yes] = Restart app automatically\n• [No] = Shutdown app cleanly now (reopen manually)\n• [Cancel] = Continue current session",
                            parent=self,
                        )
                        if ans_r is True:
                            restart_app()
                        elif ans_r is False:
                            self.shutdown_app()
                    else:
                        messagebox.showerror("Revert Failed", res, parent=self)

        def shutdown_app(self):
            """Cleanly shuts down the application, saving state, releasing locks, and closing."""
            try:
                self.destroy()
            except Exception:
                pass
            try:
                cleanup()
            except Exception:
                pass
            sys.exit(0)

        def _check_login_updates_bg(self):
            """Checks for cloud software updates in the background on startup and displays an install badge."""
            def _bg():
                try:
                    status, data = check_for_cloud_update()
                    r_ver = data.get("remote_version", "Latest")
                    if status == "update_available":
                        def _show():
                            try:
                                if hasattr(self, "_login_upd_badge_frame") and self._login_upd_badge_frame.winfo_exists():
                                    for child in self._login_upd_badge_frame.winfo_children():
                                        child.destroy()
                                    
                                    def _do_quick_install():
                                        btn.config(text=f"⏳ Installing v{r_ver}...", state="disabled")
                                        def _worker():
                                            ok, msg = install_cloud_update(data.get("remote_code"), data.get("remote_hash"))
                                            def _done():
                                                if ok:
                                                    if hasattr(self, "_login_upd_badge_frame") and self._login_upd_badge_frame.winfo_exists():
                                                        self._login_upd_badge_frame.grid_remove()
                                                    messagebox.showinfo(
                                                        "Update Installed Successfully 🎉",
                                                        f"Update v{r_ver} has been installed successfully!\n\n"
                                                        "The application will now close.\n"
                                                        "Simply reopen Payroll App to use the new version.",
                                                        parent=self,
                                                    )
                                                    self.shutdown_app()
                                                else:
                                                    btn.config(text=f"❌ Retry Installing v{r_ver}", state="normal")
                                                    messagebox.showerror("Update Failed", msg, parent=self)
                                            self.after(0, _done)
                                        threading.Thread(target=_worker, daemon=True).start()

                                    btn = tb.Button(
                                        self._login_upd_badge_frame,
                                        text=f"✨ Update Available: v{r_ver} — Click to Install Now",
                                        bootstyle="warning",
                                        cursor="hand2",
                                        command=_do_quick_install,
                                    )
                                    btn.pack(fill=X, pady=(6, 0))
                                    self._login_upd_badge_frame.grid()
                            except Exception:
                                pass
                        self.after(0, _show)
                    elif status == "installed_pending_restart":
                        def _show_restart():
                            try:
                                if hasattr(self, "_login_upd_badge_frame") and self._login_upd_badge_frame.winfo_exists():
                                    for child in self._login_upd_badge_frame.winfo_children():
                                        child.destroy()
                                    btn = tb.Button(
                                        self._login_upd_badge_frame,
                                        text=f"✅ Update v{r_ver} Ready — Click to Close & Reopen",
                                        bootstyle="success",
                                        cursor="hand2",
                                        command=self.shutdown_app,
                                    )
                                    btn.pack(fill=X, pady=(6, 0))
                                    self._login_upd_badge_frame.grid()
                            except Exception:
                                pass
                        self.after(0, _show_restart)
                    else:
                        def _hide():
                            try:
                                if hasattr(self, "_login_upd_badge_frame") and self._login_upd_badge_frame.winfo_exists():
                                    self._login_upd_badge_frame.grid_remove()
                            except Exception:
                                pass
                        self.after(0, _hide)
                except Exception:
                    pass
            threading.Thread(target=_bg, daemon=True).start()

        def show_login_page(self):
            self.clear_window()
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(0, weight=1)
            
            frame = tb.Frame(self, padding=40)
            frame.grid(row=0, column=0)
            frame.grid_columnconfigure(1, weight=1)
            
            # Add a fancy logo/title area
            title_frame = tb.Frame(frame)
            title_frame.grid(row=0, column=0, columnspan=2, pady=(0, 30))
            tb.Label(title_frame, text="❖", font=("Segoe UI", 36), bootstyle="primary").pack(side=TOP, pady=(0, 5))
            tb.Label(title_frame, text=APP_LOGO_TITLE, font=("Segoe UI", 26, "bold"), bootstyle="primary").pack(side=TOP)
            tb.Label(title_frame, text="Secure Data Vault", font=("Segoe UI", 12, "italic"), bootstyle="secondary").pack(side=TOP, pady=(5, 0))
            
            login_users = ["admin", "moe", "ziad"]
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT username FROM users ORDER BY username ASC")
                rows = cursor.fetchall()
                conn.close()
                if rows:
                    login_users = []
                    seen = set()
                    for r in rows:
                        if not r or not r[0]:
                            continue
                        name = plain_label(r[0])
                        key = name.lower()
                        if name and key not in seen:
                            seen.add(key)
                            login_users.append(name)
            except Exception:
                pass
            
            tb.Label(frame, text="Username:", font=("Segoe UI", 11)).grid(row=1, column=0, pady=10, padx=(0, 12), sticky=E)
            user_frame = tb.Frame(frame)
            user_frame.grid(row=1, column=1, pady=10, sticky=EW)
            self.username_entry = tb.Combobox(user_frame, width=28, font=("Segoe UI", 11), values=login_users, state="readonly")
            cached_user = "admin"
            try:
                cfg = get_db_config()
                if isinstance(cfg, dict) and cfg.get("last_selected_username"):
                    cached_user = cfg.get("last_selected_username")
            except Exception:
                pass
            if cached_user in login_users:
                self.username_entry.set(cached_user)
            elif "admin" in login_users:
                self.username_entry.set("admin")
            elif login_users:
                self.username_entry.set(login_users[0])
            self.username_entry.pack(side=LEFT, fill=X, expand=True)
            
            tb.Label(frame, text="Password:", font=("Segoe UI", 11)).grid(row=2, column=0, pady=10, padx=(0, 12), sticky=E)
            pw_frame = tb.Frame(frame)
            pw_frame.grid(row=2, column=1, pady=10, sticky=EW)
            
            self.password_entry = tb.Entry(pw_frame, show="*", width=28, font=("Segoe UI", 11))
            self.password_entry.pack(side=LEFT, fill=X, expand=True)
            self.password_entry.bind("<Return>", lambda e: self.login_on_enter())
            
            def toggle_password_visibility():
                try:
                    if self.password_entry.cget("show") == "*":
                        self.password_entry.config(show="")
                        eye_btn.config(text="Hide")
                    else:
                        self.password_entry.config(show="*")
                        eye_btn.config(text="Show")
                except Exception:
                    pass
                    
            eye_btn = tb.Button(pw_frame, text="Show", bootstyle="secondary outline", cursor="hand2", width=5, command=toggle_password_visibility)
            eye_btn.pack(side=LEFT, padx=(6, 0))
            
            tb.Button(frame, text="Login", bootstyle="primary", width=25, cursor="hand2", command=self.login).grid(row=3, column=0, columnspan=2, pady=(24, 10), ipadx=10, ipady=5)

            self._login_status = tb.Label(frame, text="", font=("Segoe UI", 10), bootstyle="secondary")
            self._login_status.grid(row=4, column=0, columnspan=2, pady=(0, 6))
            self._login_progress = tb.Progressbar(frame, mode="determinate", length=280, bootstyle="success-striped")
            self._login_progress.grid(row=5, column=0, columnspan=2, pady=(0, 10))
            self._login_progress.grid_remove()

            # Recovery & Cloud Updates Actions on Login Screen
            recov_frame = tb.Frame(frame)
            recov_frame.grid(row=6, column=0, columnspan=2, pady=(8, 0))

            tb.Button(
                recov_frame,
                text="Check / Retrieve Updates",
                bootstyle="info-outline",
                cursor="hand2",
                command=self.open_app_updates_dialog,
            ).pack(side=LEFT, padx=5)

            tb.Button(
                recov_frame,
                text="Revert / Rollback Engine",
                bootstyle="danger-outline",
                cursor="hand2",
                command=self._do_quick_rollback_from_login,
            ).pack(side=LEFT, padx=5)

            tb.Button(
                recov_frame,
                text="Shutdown App",
                bootstyle="secondary-outline",
                cursor="hand2",
                command=self.shutdown_app,
            ).pack(side=LEFT, padx=5)

            # Engine version & status indicator
            try:
                ver_info = get_active_code_info()
                ver_mode = " [Safe Mode Fallback]" if ver_info.get("is_safe_mode") else (" [Cloud Dynamic Engine]" if ver_info.get("is_dynamic") else " [Built-in Engine]")
                ver_text = f"v{ver_info.get('version', APP_VERSION)}{ver_mode}"
                tb.Label(frame, text=ver_text, font=("Segoe UI", 8), bootstyle="secondary").grid(row=7, column=0, columnspan=2, pady=(10, 0))
            except Exception:
                pass

            # Automatic Update Notification Banner on Login Screen
            self._login_upd_badge_frame = tb.Frame(frame)
            self._login_upd_badge_frame.grid(row=8, column=0, columnspan=2, pady=(4, 0))
            self._login_upd_badge_frame.grid_remove()
            self._check_login_updates_bg()

            # Focus password entry immediately for instant typing
            try:
                self.password_entry.focus_set()
            except Exception:
                pass

            # Start pre-loading data in the background BEFORE user finishes typing password
            self._start_prelogin_background_sync()

        def _start_prelogin_background_sync(self):
            """Pre-load and synchronize cloud data in the background while the user enters their password."""
            if getattr(self, "_prelogin_sync_started", False):
                return
            self._prelogin_sync_started = True
            self._prelogin_sync_event = threading.Event()
            self._prelogin_sync_result = {"ok": True, "msg": ""}

            def _bg_worker():
                try:
                    # 1. Warm up core metadata and caches
                    try:
                        self.get_db_locations()
                        self.get_db_categories()
                        self.get_shop_employee_id()
                    except Exception:
                        pass

                    # 2. In Supabase mode, synchronize local cache with cloud ahead of time
                    if get_db_mode() == "supabase":
                        ok, msg = sync_local_cache_with_cloud(backfill=False, init_schema=False)
                        self._prelogin_sync_result["ok"] = ok
                        self._prelogin_sync_result["msg"] = msg
                        try:
                            # Re-warm caches with latest synced data
                            self._cache_locations = None
                            self._cache_categories = None
                            self.get_db_locations()
                            self.get_db_categories()
                            self.get_shop_employee_id()
                        except Exception:
                            pass
                except Exception as e:
                    self._prelogin_sync_result["ok"] = False
                    self._prelogin_sync_result["msg"] = str(e)
                finally:
                    self._prelogin_sync_event.set()

            threading.Thread(target=_bg_worker, daemon=True).start()

        def login_on_enter(self):
            self.login()
            return "break"

        def login(self):
            if getattr(self, "_login_syncing", False):
                return
            username = (self.username_entry.get() or "").strip()
            password = (self.password_entry.get() or "").strip()
            if not username:
                self.show_app_error("Login Failed", "Please select a username.")
                return
            if not password:
                self.show_app_error("Login Failed", "Please enter a password.")
                return
            hashed_input = hashlib.sha256(password.encode()).hexdigest()

            self.show_busy(self._tr("Signing in…"))
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()

                def _pw_ok(stored):
                    stored = decrypt_val(stored) if stored is not None else None
                    if stored is None:
                        return False
                    stored_s = str(stored).strip()
                    return stored_s == hashed_input or stored_s == password

                cursor.execute("SELECT password FROM users WHERE username=?", (username,))
                user = cursor.fetchone()

                # Fallback: scan users client-side (handles encryption / offline mismatches)
                if not user:
                    try:
                        cursor.execute("SELECT username, password FROM users")
                        for row in cursor.fetchall() or []:
                            if not row:
                                continue
                            uname = decrypt_val(row[0]) if row[0] is not None else ""
                            if str(uname).strip().lower() == username.lower():
                                user = (row[1],)
                                break
                    except Exception:
                        pass

                if user:
                    if _pw_ok(user[0]):
                        # Normalize stored password to current hash format
                        try:
                            cursor.execute(
                                "UPDATE users SET password=? WHERE username=?",
                                (hashed_input, username),
                            )
                            commit_and_save(conn)
                        except Exception:
                            try:
                                conn.close()
                            except Exception:
                                pass
                            conn = sqlite3.connect(TEMP_DB_PATH)
                        try:
                            conn.close()
                        except Exception:
                            pass
                        self.current_user = username
                        try:
                            global CURRENT_SESSION_USER
                            CURRENT_SESSION_USER = username
                        except Exception:
                            pass
                        try:
                            save_last_selected_username(username)
                        except Exception:
                            pass
                        self.hide_busy()
                        self._finish_login_and_sync()
                        return
                    conn.close()
                    self.hide_busy()
                    self.show_app_error(
                        "Login Failed",
                        "Invalid password.\n\nTip: after a reset, use password: admin",
                    )
                else:
                    conn.close()
                    self.hide_busy()
                    self.show_app_error("Login Failed", "Invalid username.")
            except Exception as e:
                self.hide_busy()
                self.show_app_error("Login Failed", e)

        def _register_modal_popup(self, win):
            """Register a modal window to stay in front and re-surface if background is clicked."""
            if win is None:
                return
            if not hasattr(self, "_active_modal_popups"):
                self._active_modal_popups = []
            if win not in self._active_modal_popups:
                self._active_modal_popups.append(win)
            try:
                win.attributes("-topmost", True)
                win.lift()
            except Exception:
                pass

            def _on_win_destroy(e):
                if getattr(e, "widget", None) is win:
                    if hasattr(self, "_active_modal_popups") and win in self._active_modal_popups:
                        self._active_modal_popups.remove(win)
                    if hasattr(self, "_active_modal_popups") and self._active_modal_popups:
                        next_top = self._active_modal_popups[-1]
                        if self._widget_alive(next_top):
                            try:
                                next_top.attributes("-topmost", True)
                                next_top.lift()
                                next_top.focus_set()
                            except Exception:
                                pass

            win.bind("<Destroy>", _on_win_destroy, add="+")
            self._ensure_popup_click_redirection()

        def _ensure_popup_click_redirection(self):
            """Ensure clicking anywhere outside the active modal brings it back to the front."""
            if getattr(self, "_popup_interceptor_bound", False):
                return
            self._popup_interceptor_bound = True

            def _is_descendant_of(w, parent):
                cur = w
                for _ in range(32):
                    if cur is None:
                        return False
                    if cur is parent:
                        return True
                    cur = getattr(cur, "master", None)
                return False

            def _on_global_click(event):
                if not hasattr(self, "_active_modal_popups") or not self._active_modal_popups:
                    return
                # Filter out destroyed widgets
                self._active_modal_popups = [p for p in self._active_modal_popups if self._widget_alive(p)]
                if not self._active_modal_popups:
                    return
                top_popup = self._active_modal_popups[-1]
                w = getattr(event, "widget", None)
                if w and _is_descendant_of(w, top_popup):
                    return
                # User clicked outside the active smaller window: bring it immediately to the front!
                try:
                    top_popup.deiconify()
                    top_popup.lift()
                    top_popup.attributes("-topmost", True)
                    top_popup.focus_force()
                    top_popup.bell()
                except Exception:
                    pass
                return "break"

            self.bind_all("<Button-1>", _on_global_click, add="+")

        def _present_window(self, win):
            """Bring a popup to the front and keep it usable."""
            if win is None:
                return
            try:
                win.deiconify()
            except Exception:
                pass
            try:
                win.lift()
                win.focus_force()
                win.attributes("-topmost", True)
            except Exception:
                pass
            self._register_modal_popup(win)

        def _safe_grab_set(self, win):
            """Modal grab without nested grabs. Skip entirely on macOS Aqua Tk."""
            if win is None:
                return
            self._register_modal_popup(win)
            try:
                if platform.system() == "Darwin":
                    return
            except Exception:
                pass
            try:
                current = win.grab_current()
                if current is not None:
                    try:
                        if str(current) != str(win):
                            return
                    except Exception:
                        return
            except Exception:
                pass
            try:
                win.grab_set()
            except Exception:
                pass

        def _safe_grab_release(self, win):
            try:
                if win is not None and win.winfo_exists():
                    win.grab_release()
            except Exception:
                pass
            if hasattr(self, "_active_modal_popups") and win in self._active_modal_popups:
                self._active_modal_popups.remove(win)

        def _envelope_ui_open(self):
            if getattr(self, "_envelope_opening", False):
                return True
            return self._widget_alive(getattr(self, "_envelope_popup", None)) or self._widget_alive(
                getattr(self, "_expense_dialog", None)
            )

        def _open_sheet(self, parent, title, geometry):
            """Child window with top-priority placement and click-refocus protection."""
            parent = parent if self._widget_alive(parent) else self
            popup = tb.Toplevel(parent)
            popup.title(title)
            try:
                popup.geometry(geometry)
            except Exception:
                pass
            try:
                popup.transient(parent)
            except Exception:
                pass
            try:
                popup.lift()
                popup.focus_force()
                popup.attributes("-topmost", True)
            except Exception:
                pass
            self._register_modal_popup(popup)
            return popup

        def apply_and_memorize_column_widths(
            self, table_key, tree, columns, default_widths=None, min_widths=None, hidden_cols=None
        ):
            """Set proportional column widths with smart minimums, restore saved widths, and auto-save user adjustments."""
            if not self._widget_alive(tree):
                return
            
            saved_widths = get_saved_column_widths(table_key)
            hidden = set(hidden_cols or [])
            
            try:
                screen_w = tree.winfo_screenwidth() or 1280
            except Exception:
                screen_w = 1280
            
            try:
                win_w = max(950, min(screen_w - 60, self.winfo_width() if hasattr(self, "winfo_width") else 1150))
            except Exception:
                win_w = 1150
            
            visible_cols_count = max(1, len([c for c in columns if c not in hidden]))
            base_col_w = max(120, int(win_w / visible_cols_count))

            for col in columns:
                tree.heading(col, text=col)
                col_str = str(col)
                c_lower = col_str.lower()
                
                if col in hidden or (col_str in ("id", "record id", self._tr("Record ID")) and table_key == "calendar_table"):
                    tree.column(col, width=0, minwidth=0, stretch=tk.NO)
                    continue

                min_w = (min_widths or {}).get(col, (min_widths or {}).get(col_str, 50))
                
                if col_str in saved_widths:
                    w = max(min_w, saved_widths[col_str])
                elif default_widths and (col in default_widths or col_str in default_widths):
                    w = default_widths.get(col, default_widths.get(col_str, base_col_w))
                else:
                    # Smart defaults based on content requirements
                    if "cycle" in c_lower or "دورة" in c_lower:
                        w = max(base_col_w, 310)
                        min_w = max(min_w, 260)
                    elif "name" in c_lower or "اسم" in c_lower or "employee" in c_lower or "موظف" in c_lower:
                        w = max(base_col_w, 180)
                        min_w = max(min_w, 110)
                    elif "note" in c_lower or "desc" in c_lower or "ملاحظ" in c_lower or "وصف" in c_lower:
                        w = max(base_col_w, 240)
                        min_w = max(min_w, 130)
                    elif "category" in c_lower or "فئة" in c_lower:
                        w = max(base_col_w, 175)
                        min_w = max(min_w, 100)
                    elif "calculation" in c_lower or "حساب" in c_lower:
                        w = max(base_col_w, 150)
                        min_w = max(min_w, 90)
                    elif "date" in c_lower or "تاريخ" in c_lower:
                        w = max(base_col_w, 115)
                        min_w = max(min_w, 85)
                    elif "id" in c_lower:
                        w = 60
                        min_w = 40
                    elif "status" in c_lower or "حالة" in c_lower:
                        w = max(base_col_w, 110)
                        min_w = max(min_w, 75)
                    elif "rate" in c_lower or "percentage" in c_lower or "hours" in c_lower:
                        w = max(base_col_w, 115)
                        min_w = max(min_w, 70)
                    else:
                        w = max(base_col_w, 130)

                tree.column(col, width=w, minwidth=min_w, stretch=False, anchor=CENTER)

            def _save_cols(event=None):
                try:
                    save_table_column_widths(table_key, tree, columns)
                except Exception:
                    pass

            tree.bind("<ButtonRelease-1>", _save_cols, add="+")
            tree.bind("<B1-ButtonRelease>", _save_cols, add="+")
            tree.bind("<Unmap>", _save_cols, add="+")
            tree.bind("<Destroy>", _save_cols, add="+")

        def _point_inside_widget(self, widget, px, py):
            if not self._widget_alive(widget):
                return False
            try:
                wx = widget.winfo_rootx()
                wy = widget.winfo_rooty()
                ww = widget.winfo_width()
                wh = widget.winfo_height()
                return (wx <= px <= wx + ww) and (wy <= py <= wy + wh)
            except Exception:
                return False

        def _show_rev_cycles_popover(self, event=None):
            self._cancel_popover_close()
            pop = getattr(self, "rev_cycles_popover", None)
            if self._widget_alive(pop):
                try:
                    pop.deiconify()
                    pop.lift()
                    return
                except Exception:
                    pass
            
            btn = getattr(self, "btn_browse_cycles", None)
            if not self._widget_alive(btn):
                return
            
            pop = tb.Toplevel(self.tab_calendar)
            pop.overrideredirect(True)
            try:
                pop.attributes("-topmost", True)
            except Exception:
                pass
            self.rev_cycles_popover = pop
            
            try:
                self.update_idletasks()
                bx = btn.winfo_rootx()
                by = btn.winfo_rooty() + btn.winfo_height() + 4
                screen_w = self.winfo_screenwidth() or 1600
                screen_h = self.winfo_screenheight() or 900
                bw = min(880, max(680, int(screen_w * 0.55)))
                bh = min(430, max(340, min(int(screen_h * 0.45), screen_h - by - 15)))
                max_x = screen_w - bw - 10
                px = max(10, min(bx, max_x))
                pop.geometry(f"{bw}x{bh}+{px}+{by}")
            except Exception:
                pop.geometry("860x420")
            
            outer = tb.Frame(pop, padding=10, bootstyle="info", borderwidth=2, relief="ridge")
            outer.pack(fill=BOTH, expand=True)
            
            top_bar = tb.Frame(outer, padding=(6, 4))
            top_bar.pack(fill=X, side=TOP)
            
            # Year Switcher embedded inside popover
            tb.Label(top_bar, text=self._tr("Year:"), font=("Segoe UI", 12, "bold")).pack(side=LEFT, padx=(2, 4))
            tb.Button(
                top_bar,
                text="◀",
                bootstyle="outline-primary",
                width=3,
                cursor="hand2",
                command=self.prev_rev_year,
            ).pack(side=LEFT, padx=2)
            self.lbl_popover_year = tb.Label(
                top_bar,
                text=str(self.rev_cal_year),
                font=("Segoe UI", 15, "bold"),
                bootstyle="primary",
            )
            self.lbl_popover_year.pack(side=LEFT, padx=8)
            tb.Button(
                top_bar,
                text="▶",
                bootstyle="outline-primary",
                width=3,
                cursor="hand2",
                command=self.next_rev_year,
            ).pack(side=LEFT, padx=2)
            
            tb.Separator(top_bar, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=10)
            
            tb.Button(
                top_bar,
                text=f"✓ {self._tr('Select All')}",
                bootstyle="primary-outline",
                cursor="hand2",
                command=self.select_all_rev_cycles,
            ).pack(side=LEFT, padx=4)
            
            tb.Button(
                top_bar,
                text=f"✕ {self._tr('Clear Selection')}",
                bootstyle="secondary-outline",
                cursor="hand2",
                command=self.clear_rev_cycles,
            ).pack(side=LEFT, padx=4)
            
            self.lbl_popover_sel_summary = tb.Label(
                top_bar,
                text="",
                font=("Segoe UI", 11, "bold"),
                bootstyle="info",
            )
            self.lbl_popover_sel_summary.pack(side=LEFT, padx=12)
            
            tb.Button(
                top_bar,
                text="✕ " + self._tr("Close Window"),
                bootstyle="danger-outline",
                cursor="hand2",
                command=self._hide_rev_cycles_popover,
            ).pack(side=RIGHT, padx=4)
            
            grid_holder = tb.Frame(outer)
            grid_holder.pack(fill=BOTH, expand=True, pady=(4, 0))
            
            canvas = tk.Canvas(grid_holder, highlightthickness=0, borderwidth=0)
            self.rev_pop_canvas = canvas
            v_sb = tb.Scrollbar(grid_holder, orient=VERTICAL, command=canvas.yview)
            
            self.rev_cards_container = tb.Frame(canvas, padding=2)
            self.rev_cards_container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            c_win = canvas.create_window((0, 0), window=self.rev_cards_container, anchor="nw")
            
            def _configure_canvas_width(event):
                try:
                    canvas.itemconfigure(c_win, width=event.width)
                except Exception:
                    pass
            canvas.bind("<Configure>", _configure_canvas_width)
            canvas.configure(yscrollcommand=v_sb.set)
            
            v_sb.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)
            
            self._rebuild_cycle_cards()
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)
            
            def _on_pop_wheel(event):
                if not self._widget_alive(canvas):
                    return
                try:
                    if event.delta:
                        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                    elif event.num == 4:
                        canvas.yview_scroll(-1, "units")
                    elif event.num == 5:
                        canvas.yview_scroll(1, "units")
                except Exception:
                    pass
            
            for target in (pop, outer, grid_holder, canvas, self.rev_cards_container):
                target.bind("<MouseWheel>", _on_pop_wheel)
                target.bind("<Button-4>", _on_pop_wheel)
                target.bind("<Button-5>", _on_pop_wheel)
            
            pop.bind("<Enter>", lambda e: self._cancel_popover_close())
            pop.bind("<Leave>", lambda e: self._schedule_popover_close())
            outer.bind("<Enter>", lambda e: self._cancel_popover_close())
            outer.bind("<Leave>", lambda e: self._schedule_popover_close())
            grid_holder.bind("<Enter>", lambda e: self._cancel_popover_close())
            grid_holder.bind("<Leave>", lambda e: self._schedule_popover_close())
            canvas.bind("<Enter>", lambda e: self._cancel_popover_close())
            canvas.bind("<Leave>", lambda e: self._schedule_popover_close())

        def _cancel_popover_close(self):
            aid = getattr(self, "_popover_close_after_id", None)
            if aid is not None:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
                self._popover_close_after_id = None

        def _schedule_popover_close(self):
            self._cancel_popover_close()
            try:
                self._popover_close_after_id = self.after(350, self._check_and_close_popover)
            except Exception:
                pass

        def _check_and_close_popover(self):
            self._popover_close_after_id = None
            pop = getattr(self, "rev_cycles_popover", None)
            if not self._widget_alive(pop):
                return
            try:
                px, py = self.winfo_pointerxy()
                in_btn = self._point_inside_widget(getattr(self, "btn_browse_cycles", None), px, py)
                in_pop = self._point_inside_widget(pop, px, py)
                if not in_btn and not in_pop:
                    self._hide_rev_cycles_popover()
            except Exception:
                self._hide_rev_cycles_popover()

        def _hide_rev_cycles_popover(self):
            self._cancel_popover_close()
            pop = getattr(self, "rev_cycles_popover", None)
            self.rev_cycles_popover = None
            if self._widget_alive(pop):
                try:
                    pop.destroy()
                except Exception:
                    pass

        def _update_popover_sel_summary(self):
            lbl = getattr(self, "lbl_popover_sel_summary", None)
            if not self._widget_alive(lbl):
                return
            sel = getattr(self, "selected_rev_cycles", set()) or set()
            if not sel:
                lbl.config(text=f"{self._tr('All Cycles')} ({self._tr('No filter')})")
            elif len(sel) == 1:
                k = list(sel)[0]
                lbl.config(text=f"1 {self._tr('Cycle Selected')}: {cycle_label(k)}")
            elif len(sel) == 24:
                lbl.config(text=f"{self._tr('All')} 24 {self._tr('Cycles Selected')}")
            else:
                lbl.config(text=f"{len(sel)} {self._tr('Cycles Selected')}")

        def open_calendar_columns_dialog(self, parent=None):
            """Dialog to customize which columns are displayed in the Shop Earnings / Calendar table."""
            parent = parent if self._widget_alive(parent) else self
            dialog = tb.Toplevel(parent)
            dialog.title(self._tr("⚙️ Customize Table Columns"))
            dialog.geometry("520x620")
            dialog.transient(parent)
            dialog.grab_set()
            dialog.focus_set()
            
            main_f = tb.Frame(dialog, padding=20)
            main_f.pack(fill=BOTH, expand=True)
            
            tb.Label(
                main_f,
                text=self._tr("Select Columns to Display"),
                font=("Segoe UI", 14, "bold"),
                bootstyle="primary",
            ).pack(anchor=W, pady=(0, 4))
            
            tb.Label(
                main_f,
                text=self._tr("Check the columns you want visible in the Shop Earnings table. Changes are saved automatically."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=480,
                justify=LEFT,
            ).pack(anchor=W, pady=(0, 12))
            
            scroll_f = tb.Frame(main_f)
            scroll_f.pack(fill=BOTH, expand=True, pady=5)
            
            canvas = tk.Canvas(scroll_f, highlightthickness=0)
            sb = tb.Scrollbar(scroll_f, orient=VERTICAL, command=canvas.yview)
            inner = tb.Frame(canvas, padding=5)
            inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            c_win = canvas.create_window((0, 0), window=inner, anchor="nw")
            canvas.bind("<Configure>", lambda e: canvas.itemconfigure(c_win, width=e.width))
            canvas.configure(yscrollcommand=sb.set)
            
            sb.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)
            
            hidden_now = get_calendar_hidden_columns()
            check_vars = {}
            
            for col_key, desc in ALL_CALENDAR_COLUMNS:
                col_tr = self._tr(col_key)
                is_checked = (col_key not in hidden_now and col_tr not in hidden_now)
                var = tk.BooleanVar(value=is_checked)
                check_vars[col_key] = var
                
                row_f = tb.Frame(inner, padding=(4, 6))
                row_f.pack(fill=X, expand=True)
                
                cb = tb.Checkbutton(
                    row_f,
                    text=f"{col_tr}   —   {desc}",
                    variable=var,
                    bootstyle="primary-round-toggle",
                    cursor="hand2",
                )
                cb.pack(anchor=W)
            
            btn_f = tb.Frame(main_f, padding=(0, 15, 0, 0))
            btn_f.pack(fill=X, side=BOTTOM)
            
            def _save_and_apply():
                new_hidden = {"Record ID", self._tr("Record ID")}
                for col_key, var in check_vars.items():
                    if not var.get():
                        new_hidden.add(col_key)
                        new_hidden.add(self._tr(col_key))
                save_calendar_hidden_columns(new_hidden)
                self.refresh_calendar_column_visibility()
                dialog.destroy()
                messagebox.showinfo(
                    self._tr("Columns Updated"),
                    self._tr("Table columns display updated successfully."),
                    parent=self,
                )

            def _select_all_cols(val):
                for var in check_vars.values():
                    var.set(val)

            tb.Button(btn_f, text=self._tr("Save & Apply"), bootstyle="success", cursor="hand2", command=_save_and_apply).pack(side=LEFT, padx=5)
            tb.Button(btn_f, text=self._tr("Select All"), bootstyle="secondary-outline", cursor="hand2", command=lambda: _select_all_cols(True)).pack(side=LEFT, padx=5)
            tb.Button(btn_f, text=self._tr("Reset to Default"), bootstyle="warning-outline", cursor="hand2", command=lambda: (_select_all_cols(True), check_vars.get("Written Up", tk.BooleanVar()).set(False))).pack(side=LEFT, padx=5)
            tb.Button(btn_f, text=self._tr("Cancel"), bootstyle="secondary", cursor="hand2", command=dialog.destroy).pack(side=RIGHT, padx=5)

        def refresh_calendar_column_visibility(self):
            """Apply current hidden columns to tree_calendar."""
            if not self._widget_alive(getattr(self, "tree_calendar", None)):
                return
            hidden = get_calendar_hidden_columns()
            disp_cols = [c for c in self.columns if c != self._tr("Name")]
            self.apply_and_memorize_column_widths(
                "calendar_table",
                self.tree_calendar,
                disp_cols,
                hidden_cols=list(hidden),
            )

        def _bind_dialog_save_keys(self, dialog, save_fn):
            """Enter / keypad Enter runs Save while this dialog is focused."""
            dialog._save_fn = save_fn
            def _event_in_dialog(widget):
                w = widget
                for _ in range(48):
                    if w is None:
                        return False
                    if w is dialog:
                        return True
                    w = getattr(w, "master", None)
                return False

            def _on_return(event):
                try:
                    if not dialog.winfo_exists():
                        return
                    w = event.widget
                    cls = ""
                    try:
                        cls = str(w.winfo_class()).lower()
                    except Exception:
                        pass
                    if "button" in cls:
                        return
                    if not _event_in_dialog(w):
                        return
                    save_fn()
                    return "break"
                except Exception:
                    pass

            dialog.bind("<Return>", _on_return)
            dialog.bind("<KP_Enter>", _on_return)
            try:
                dialog.bind_all("<Return>", _on_return)
                dialog.bind_all("<KP_Enter>", _on_return)
            except Exception:
                pass

            def _unbind_save_keys(event=None):
                if event is not None and getattr(event, "widget", None) is not dialog:
                    return
                for seq in ("<Return>", "<KP_Enter>"):
                    try:
                        dialog.unbind_all(seq)
                    except Exception:
                        pass

            dialog.bind("<Destroy>", _unbind_save_keys, add="+")

        def _clear_topmost(self, win):
            try:
                if win.winfo_exists():
                    win.attributes("-topmost", False)
            except Exception:
                pass

        def _finish_login_and_sync(self):
            """After a valid password, transition immediately to main application with zero delay."""
            try:
                log_user_action("login", extra_summary="Logged in")
            except Exception:
                pass
            self.show_main_application()

        def _tr(self, text):
            if not hasattr(self, 'lang'):
                self.lang = 'en'
            if self.lang == 'en':
                return text
            return TRANSLATIONS.get(text, text)
            
        def toggle_language(self):
            if getattr(self, "_rebuilding_ui", False):
                return
            self.lang = 'ar' if getattr(self, 'lang', 'en') == 'en' else 'en'
            # Defer past the button click + any pending DateEntry/menu Tcl work
            self._rebuilding_ui = True

            def _rebuild():
                try:
                    self.show_main_application()
                finally:
                    self._rebuilding_ui = False

            try:
                self.after(80, _rebuild)
            except Exception:
                try:
                    _rebuild()
                except Exception:
                    self._rebuilding_ui = False

        def show_main_application(self):
            # Do not show a busy overlay around full UI teardown — clear_window()
            # destroys children and races the spinner / DateEntry Tcl cleanup.
            try:
                self.hide_busy(force=True)
            except Exception:
                pass
            self._show_main_application_body()

        def _show_main_application_body(self):
            self.is_logged_in = True
            self.last_activity_time = time.time()
            try:
                self.clear_window()
            except Exception:
                # Never abort language/login rebuild because of a destroy race
                pass
            self.grid_rowconfigure(0, weight=0)
            self.grid_columnconfigure(0, weight=0)
            
            # Header
            header = tb.Frame(self, bootstyle="primary")
            header.pack(fill=X, side=TOP)
            tb.Label(header, text=f"💈 {APP_TITLE}", font=("Segoe UI", 18, "bold"), bootstyle="inverse-primary").pack(side=LEFT, padx=20, pady=15)
            
            right_frame = tb.Frame(header, bootstyle="primary")
            right_frame.pack(side=RIGHT, padx=20, pady=15)
            
            tb.Button(right_frame, text=self._tr("📁 Employee Folders"), bootstyle="success", cursor="hand2", command=self.open_folders_window).pack(side=LEFT, padx=10)
            tb.Button(right_frame, text=self._tr("💸 Expense Reports"), bootstyle="warning", cursor="hand2", command=self.open_expenses_window).pack(side=LEFT, padx=10)
            tb.Button(right_frame, text=self._tr("📁 Shop Files"), bootstyle="info", cursor="hand2", command=self.open_shop_files_window).pack(side=LEFT, padx=10)
            tb.Button(right_frame, text=self._tr("⚙️ Settings"), bootstyle="secondary", cursor="hand2", command=self.open_settings_password_prompt).pack(side=LEFT, padx=10)
            
            username_to_show = getattr(self, "current_user", "admin")
            logged_in_text = f"{self._tr('Logged in as')} {username_to_show.capitalize()}"
            self.lbl_logged_in_user = tb.Label(right_frame, text=logged_in_text, font=("Segoe UI", 10), bootstyle="inverse-primary")
            self.lbl_logged_in_user.pack(side=LEFT, padx=10)

            if get_db_mode() == "supabase":
                self.lbl_live_sync = tb.Label(
                    right_frame,
                    text="☁️ " + self._tr("Connected"),
                    font=("Segoe UI", 9),
                    bootstyle="inverse-primary",
                )
                self.lbl_live_sync.pack(side=LEFT, padx=6)
            
            btn_text = "🌐 العربية" if getattr(self, 'lang', 'en') == 'en' else "🌐 English"
            tb.Button(right_frame, text=btn_text, bootstyle="info", cursor="hand2", command=self.toggle_language).pack(side=LEFT)
            
            self.notebook = tb.Notebook(self, bootstyle="info")
            self.notebook.pack(fill=BOTH, expand=True, padx=20, pady=20)
            
            self.tab_calendar = tb.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_calendar, text=self._tr("🟢 💈 Shop Earnings"))
            self.setup_calendar_tab()
            
            self.tab_names = tb.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_names, text=self._tr("🔵 💇‍♂️ Barbers / Stylists"))
            self.setup_names_tab()
            
            self.tab_data_entry = tb.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_data_entry, text=self._tr("🟠 ✍️ Manual Ledger"))
            self.setup_data_entry_tab()
            try:
                self.notebook.hide(self.tab_data_entry)
            except Exception:
                pass
            
            self.tab_financials = tb.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_financials, text=self._tr("🟣 📊 P&L / Financials"))
            self.setup_financials_tab()
            
            self.tab_cash_cal = tb.Frame(self.notebook, padding=10)
            self.notebook.add(self.tab_cash_cal, text=self._tr("🔴 📅 Cash Calendar"))
            self.setup_cash_calendar_tab()
            
            self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)
            self.start_live_sync()

        def open_shop_files_window(self):
            """Single-window shop document archive: pick location → browse/add files."""
            if self._widget_alive(getattr(self, "shop_files_win", None)):
                self.shop_files_win.lift()
                self.shop_files_win.focus_force()
                return

            win = tb.Toplevel(self)
            self.shop_files_win = win
            win.title(self._tr("📁 Shop Files"))
            try:
                self.update_idletasks()
                w = self.winfo_width()
                h = self.winfo_height()
                x = self.winfo_x()
                y = self.winfo_y()
                win.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                win.geometry("1000x640")
            win.grab_set()
            win.focus_force()
            self._present_window(win)

            self._shop_files_location = None
            self._shop_files_doc_map = {}

            header = tb.Frame(win)
            header.pack(fill=X, padx=28, pady=(22, 8))
            tb.Label(
                header,
                text=self._tr("📁 Shop Files"),
                font=("Segoe UI", 22, "bold"),
                bootstyle="info",
            ).pack(side=LEFT)
            tb.Button(
                header,
                text=self._tr("Close Window"),
                bootstyle="light",
                cursor="hand2",
                command=win.destroy,
            ).pack(side=RIGHT)

            self._shop_files_subtitle = tb.Label(
                win,
                text=self._tr("Open a location to manage leases, rent, permits, and other shop records."),
                font=("Segoe UI", 11),
                bootstyle="secondary",
            )
            self._shop_files_subtitle.pack(anchor=W, padx=28, pady=(0, 12))

            self._shop_files_body = tb.Frame(win)
            self._shop_files_body.pack(fill=BOTH, expand=True, padx=28, pady=(0, 24))

            self._shop_files_show_locations()

        def _shop_files_clear_body(self):
            body = getattr(self, "_shop_files_body", None)
            if not body or not body.winfo_exists():
                return
            for child in body.winfo_children():
                child.destroy()

        def _shop_files_count_docs(self, location):
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute(
                    "SELECT COUNT(*) FROM shop_documents WHERE location = ?",
                    (location,),
                )
                n = cur.fetchone()[0] or 0
                conn.close()
                return int(n)
            except Exception:
                return 0

        def _shop_files_show_locations(self):
            self._shop_files_location = None
            self._shop_files_clear_body()
            if self._widget_alive(getattr(self, "_shop_files_subtitle", None)):
                self._shop_files_subtitle.configure(
                    text=self._tr("Select which shop you want to open files for.")
                )

            body = self._shop_files_body
            intro = tb.Frame(body)
            intro.pack(fill=X, pady=(0, 16))
            tb.Label(
                intro,
                text=self._tr("Choose a shop location"),
                font=("Segoe UI", 16, "bold"),
            ).pack(anchor=W)
            tb.Label(
                intro,
                text=self._tr("Open a location to manage leases, rent, permits, and other shop records."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
            ).pack(anchor=W, pady=(4, 0))

            grid = tb.Frame(body)
            grid.pack(fill=BOTH, expand=True)

            locations = self.get_db_locations() or []
            if not locations:
                tb.Label(
                    grid,
                    text=self._tr("No locations configured. Add locations in Settings first."),
                    font=("Segoe UI", 12),
                    bootstyle="warning",
                ).pack(pady=40)
                return

            # 2-column clickable location cards
            for i, loc in enumerate(locations):
                count = self._shop_files_count_docs(loc)
                card = tb.Frame(grid, padding=18, bootstyle="secondary")
                r, c = divmod(i, 2)
                card.grid(row=r, column=c, sticky="nsew", padx=8, pady=8)
                grid.columnconfigure(c, weight=1)
                grid.rowconfigure(r, weight=0)

                title = tb.Label(card, text=f"🏪  {loc}", font=("Segoe UI", 15, "bold"), cursor="hand2")
                title.pack(anchor=W)
                meta = tb.Label(
                    card,
                    text=f"{count} {self._tr('documents')}",
                    font=("Segoe UI", 10),
                    bootstyle="secondary",
                    cursor="hand2",
                )
                meta.pack(anchor=W, pady=(6, 10))
                open_btn = tb.Button(
                    card,
                    text=self._tr("Open archive →"),
                    bootstyle="info",
                    cursor="hand2",
                    command=lambda location=loc: self._shop_files_show_archive(location),
                )
                open_btn.pack(anchor=W)

                def _open_loc(_e=None, location=loc):
                    self._shop_files_show_archive(location)

                for wdg in (card, title, meta):
                    wdg.bind("<Button-1>", _open_loc)

        def _shop_files_fetch_docs(self, location):
            rows = []
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute(
                    """
                    SELECT id, title, doc_date, description, file_path, created_at
                    FROM shop_documents
                    WHERE location = ?
                    ORDER BY COALESCE(doc_date, '') DESC, id DESC
                    """,
                    (location,),
                )
                rows = cur.fetchall() or []
                conn.close()
            except Exception:
                rows = []
            return rows

        def _shop_files_show_archive(self, location):
            self._shop_files_location = location
            self._shop_files_clear_body()
            if self._widget_alive(getattr(self, "_shop_files_subtitle", None)):
                self._shop_files_subtitle.configure(
                    text=f"{self._tr('Shop document archive')}  ·  {location}"
                )

            body = self._shop_files_body
            top = tb.Frame(body)
            top.pack(fill=X, pady=(0, 12))

            tb.Button(
                top,
                text=self._tr("← Back to locations"),
                bootstyle="secondary outline",
                cursor="hand2",
                command=self._shop_files_show_locations,
            ).pack(side=LEFT)

            tb.Button(
                top,
                text=self._tr("+ Add Document"),
                bootstyle="success",
                cursor="hand2",
                command=lambda: self._shop_files_show_add_form(location),
            ).pack(side=RIGHT)
            tb.Button(
                top,
                text=self._tr("Refresh"),
                bootstyle="secondary outline",
                cursor="hand2",
                command=lambda: self._shop_files_show_archive(location),
            ).pack(side=RIGHT, padx=8)

            tb.Label(
                body,
                text=location,
                font=("Segoe UI", 18, "bold"),
            ).pack(anchor=W, pady=(0, 4))
            tb.Label(
                body,
                text=self._tr("Keep rent leases, permits, and scanned shop papers in one place."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
            ).pack(anchor=W, pady=(0, 12))

            actions = tb.Frame(body)
            actions.pack(fill=X, pady=(0, 8))
            tb.Button(
                actions,
                text=self._tr("Open / Preview"),
                bootstyle="info",
                cursor="hand2",
                command=self._shop_files_preview_selected,
            ).pack(side=LEFT)
            tb.Button(
                actions,
                text=self._tr("Delete Document"),
                bootstyle="danger",
                cursor="hand2",
                command=self._shop_files_delete_selected,
            ).pack(side=LEFT, padx=8)

            tree_holder = tb.Frame(body)
            tree_holder.pack(fill=BOTH, expand=True)

            rows = self._shop_files_fetch_docs(location)
            self._shop_files_doc_map = {}
            self.shop_files_tree = None

            if not rows:
                empty = tb.Frame(tree_holder, padding=30)
                empty.pack(expand=True)
                tb.Label(
                    empty,
                    text=self._tr(
                        "No shop documents yet.\nAdd a scanned lease, rent receipt, or other record."
                    ),
                    font=("Segoe UI", 13),
                    bootstyle="secondary",
                    justify=CENTER,
                ).pack(pady=(40, 16))
                tb.Button(
                    empty,
                    text=self._tr("+ Add Document"),
                    bootstyle="success",
                    cursor="hand2",
                    command=lambda: self._shop_files_show_add_form(location),
                ).pack()
                return

            cols = (
                self._tr("Name"),
                self._tr("Date"),
                self._tr("Description"),
                self._tr("File"),
            )
            tree = tb.Treeview(
                tree_holder,
                columns=cols,
                show="headings",
                bootstyle="info",
                selectmode="browse",
            )
            tree.heading(cols[0], text=cols[0])
            tree.heading(cols[1], text=cols[1])
            tree.heading(cols[2], text=cols[2])
            tree.heading(cols[3], text=cols[3])
            tree.column(cols[0], width=220, minwidth=140, anchor=W)
            tree.column(cols[1], width=110, minwidth=90, anchor=CENTER)
            tree.column(cols[2], width=360, minwidth=160, anchor=W)
            tree.column(cols[3], width=180, minwidth=100, anchor=W)
            self._attach_tree_scrollbars(tree_holder, tree)
            self.shop_files_tree = tree

            for row in rows:
                doc_id, title, doc_date, description, file_path, _created = row
                fname = os.path.basename(file_path) if file_path else "—"
                iid = tree.insert(
                    "",
                    tk.END,
                    values=(
                        title or "",
                        doc_date or "",
                        description or "",
                        fname,
                    ),
                )
                self._shop_files_doc_map[iid] = {
                    "id": doc_id,
                    "title": title or "",
                    "doc_date": doc_date or "",
                    "description": description or "",
                    "file_path": file_path or "",
                }

            tree.bind("<Double-1>", lambda _e: self._shop_files_preview_selected())

        def _shop_files_selected_doc(self):
            tree = getattr(self, "shop_files_tree", None)
            if not tree or not tree.winfo_exists():
                return None
            sel = tree.selection()
            if not sel:
                return None
            return self._shop_files_doc_map.get(sel[0])

        def _shop_files_preview_selected(self):
            doc = self._shop_files_selected_doc()
            parent = getattr(self, "shop_files_win", self)
            if not doc:
                messagebox.showinfo(
                    self._tr("📁 Shop Files"),
                    self._tr("Select a document first."),
                    parent=parent,
                )
                return
            path = doc.get("file_path") or ""
            self.preview_expense_document(path, parent=parent)

        def _shop_files_delete_selected(self):
            doc = self._shop_files_selected_doc()
            parent = getattr(self, "shop_files_win", self)
            if not doc:
                messagebox.showinfo(
                    self._tr("📁 Shop Files"),
                    self._tr("Select a document first."),
                    parent=parent,
                )
                return
            if not messagebox.askyesno(
                self._tr("Delete Document"),
                self._tr("Delete this shop document and its file?"),
                parent=parent,
            ):
                return
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute("DELETE FROM shop_documents WHERE id = ?", (doc["id"],))
                commit_and_save(conn)
                conn.close()
                delete_shop_document_file(doc.get("file_path"))
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=parent)
                return
            loc = self._shop_files_location
            if loc:
                self._shop_files_show_archive(loc)

        def _shop_files_show_add_form(self, location):
            self._shop_files_clear_body()
            if self._widget_alive(getattr(self, "_shop_files_subtitle", None)):
                self._shop_files_subtitle.configure(
                    text=f"{self._tr('+ Add Document')}  ·  {location}"
                )

            body = self._shop_files_body
            top = tb.Frame(body)
            top.pack(fill=X, pady=(0, 16))
            tb.Button(
                top,
                text=self._tr("← Back to archive"),
                bootstyle="secondary outline",
                cursor="hand2",
                command=lambda: self._shop_files_show_archive(location),
            ).pack(side=LEFT)

            tb.Label(
                body,
                text=self._tr("+ Add Document"),
                font=("Segoe UI", 18, "bold"),
            ).pack(anchor=W)
            tb.Label(
                body,
                text=self._tr("Enter a name, date, short description, then attach the scanned file."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
            ).pack(anchor=W, pady=(4, 18))

            form = tb.Frame(body)
            form.pack(fill=X, anchor=N)
            pad = {"sticky": "w", "pady": 8, "padx": (0, 12)}
            ent_pad = {"sticky": "w", "pady": 8}

            tb.Label(form, text=self._tr("Document name"), font=("Segoe UI", 10, "bold")).grid(
                row=0, column=0, **pad
            )
            title_ent = tb.Entry(form, width=42)
            title_ent.grid(row=0, column=1, **ent_pad)

            tb.Label(form, text=self._tr("Document date"), font=("Segoe UI", 10, "bold")).grid(
                row=1, column=0, **pad
            )
            date_ent = tb.DateEntry(form, bootstyle="info", dateformat="%Y-%m-%d")
            date_ent.grid(row=1, column=1, **ent_pad)
            try:
                today = datetime.today().strftime("%Y-%m-%d")
                date_ent.entry.delete(0, tk.END)
                date_ent.entry.insert(0, today)
            except Exception:
                pass

            tb.Label(form, text=self._tr("Description"), font=("Segoe UI", 10, "bold")).grid(
                row=2, column=0, **pad
            )
            desc_ent = tb.Entry(form, width=42)
            desc_ent.grid(row=2, column=1, **ent_pad)

            tb.Label(form, text=self._tr("Attached file:"), font=("Segoe UI", 10, "bold")).grid(
                row=3, column=0, **pad
            )
            file_row = tb.Frame(form)
            file_row.grid(row=3, column=1, sticky="w", pady=8)
            file_var = tk.StringVar(value="")
            file_lbl = tb.Label(
                file_row,
                text=self._tr("No file chosen yet"),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=420,
            )
            file_lbl.pack(side=LEFT)

            parent_win = getattr(self, "shop_files_win", self)

            def choose_file():
                path = filedialog.askopenfilename(
                    parent=parent_win,
                    title=self._tr("Choose file…"),
                    filetypes=[
                        ("Documents & Images", "*.pdf *.png *.jpg *.jpeg *.gif *.bmp *.webp *.tif *.tiff"),
                        ("PDF", "*.pdf"),
                        ("Images", "*.png *.jpg *.jpeg *.gif *.bmp *.webp *.tif *.tiff"),
                        ("All Files", "*.*"),
                    ],
                )
                if path:
                    file_var.set(path)
                    file_lbl.configure(text=os.path.basename(path), bootstyle="info")

            tb.Button(
                file_row,
                text=self._tr("Choose file…"),
                bootstyle="info outline",
                cursor="hand2",
                command=choose_file,
            ).pack(side=LEFT, padx=(12, 0))

            hint = tb.Label(
                body,
                text=self._tr("Tip: scan the paper first, then upload the PDF or photo here."),
                font=("Segoe UI", 9),
                bootstyle="secondary",
            )
            hint.pack(anchor=W, pady=(8, 0))

            btn_row = tb.Frame(body)
            btn_row.pack(anchor=W, pady=24)

            def save_doc():
                title = title_ent.get().strip()
                try:
                    doc_date = date_ent.entry.get().strip()
                except Exception:
                    doc_date = ""
                description = desc_ent.get().strip()
                src = file_var.get().strip()
                if not title or not src:
                    messagebox.showerror(
                        "Error",
                        self._tr("A name and file are required."),
                        parent=parent_win,
                    )
                    return
                if not os.path.isfile(src):
                    messagebox.showerror(
                        "Error",
                        self._tr("Document file was not found."),
                        parent=parent_win,
                    )
                    return
                try:
                    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute(
                        """
                        INSERT INTO shop_documents
                            (location, title, doc_date, description, file_path, created_at)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        (location, title, doc_date, description, "", created_at),
                    )
                    doc_id = cur.lastrowid
                    stored = store_shop_document_file(src, location, doc_id=doc_id)
                    cur.execute(
                        "UPDATE shop_documents SET file_path = ? WHERE id = ?",
                        (stored, doc_id),
                    )
                    commit_and_save(conn)
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=parent_win)
                    return
                messagebox.showinfo(
                    self._tr("📁 Shop Files"),
                    self._tr("Shop document saved."),
                    parent=parent_win,
                )
                self._shop_files_show_archive(location)

            tb.Button(
                btn_row,
                text=self._tr("Save Document"),
                bootstyle="success",
                cursor="hand2",
                command=save_doc,
            ).pack(side=LEFT, ipadx=18, ipady=6)
            tb.Button(
                btn_row,
                text=self._tr("Cancel"),
                bootstyle="secondary",
                cursor="hand2",
                command=lambda: self._shop_files_show_archive(location),
            ).pack(side=LEFT, padx=10)

        def manual_sync_cloud(self):
            """Manual trigger for cloud sync with immediate visual feedback."""
            if getattr(self, "_manual_sync_in_progress", False):
                return
            self._manual_sync_in_progress = True

            btn = getattr(self, "btn_sync_now", None)
            lbl = getattr(self, "lbl_live_sync", None)

            if self._widget_alive(btn):
                btn.config(text="⏳ " + self._tr("Syncing…"), state="disabled")
            if self._widget_alive(lbl):
                lbl.config(text="🔄 " + self._tr("Connecting…"))

            def _bg():
                ok = False
                msg = ""
                try:
                    close_shared_supabase_conn()
                    ok, msg = sync_local_cache_with_cloud(backfill=True, init_schema=False)
                except Exception as e:
                    ok = False
                    msg = str(e)
                finally:
                    self._manual_sync_in_progress = False

                def _ui():
                    b = getattr(self, "btn_sync_now", None)
                    l = getattr(self, "lbl_live_sync", None)
                    if self._widget_alive(b):
                        b.config(text="🔄 " + self._tr("Sync Cloud"), state="normal")
                    if self._widget_alive(l):
                        if ok:
                            t_str = datetime.now().strftime("%H:%M:%S")
                            l.config(text=f"✅ {self._tr('Synced')} {t_str}")
                        else:
                            pending = offline_pending_count()
                            l.config(text=f"📴 {self._tr('Offline')} ({pending} {self._tr('pending')})")
                    if ok:
                        self._schedule_soft_ui_refresh(full=True)

                try:
                    self.after(0, _ui)
                except Exception:
                    pass

            threading.Thread(target=_bg, daemon=True).start()

        def start_live_sync(self):
            """Auto-refresh the active view from Supabase while logged in."""
            self.stop_live_sync()
            if get_db_mode() != "supabase":
                return
            self._last_sync_fingerprint = cloud_data_fingerprint()
            self._live_sync_after_id = self.after(800, self._live_sync_tick)

        def stop_live_sync(self):
            aid = getattr(self, "_live_sync_after_id", None)
            if aid is not None:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
            self._live_sync_after_id = None

        def _live_sync_tick(self):
            self._live_sync_after_id = None
            try:
                if (
                    getattr(self, "is_logged_in", False)
                    and get_db_mode() == "supabase"
                    and self.winfo_exists()
                ):
                    if _SYNC_IN_PROGRESS or getattr(self, "_rebuilding_ui", False):
                        pass
                    else:
                        def _bg():
                            status = None
                            changed = False
                            ok = False
                            msg = ""
                            after = None
                            try:
                                before = cloud_data_fingerprint()
                                ok, msg = sync_local_cache_with_cloud(backfill=False, init_schema=False)
                                after = cloud_data_fingerprint()
                                changed = after != before or after != getattr(
                                    self, "_last_sync_fingerprint", None
                                )
                                if ok:
                                    try:
                                        maybe_run_scheduled_cloud_backup()
                                    except Exception:
                                        pass
                                if ok or msg == "busy":
                                    status = cloud_sync_status_label(ok=True)
                                else:
                                    pending = offline_pending_count()
                                    if pending:
                                        status = f"☁️ {pending} change(s) waiting to upload"
                                    else:
                                        status = cloud_sync_status_label(ok=False)
                            except Exception:
                                pending = offline_pending_count()
                                status = f"📴 Offline — {pending} pending"

                            def _ui():
                                try:
                                    if self._widget_alive(getattr(self, "lbl_live_sync", None)) and status:
                                        self.lbl_live_sync.config(text=status)
                                    if ok or changed:
                                        if after is not None:
                                            self._last_sync_fingerprint = after
                                        self._schedule_soft_ui_refresh(full=changed)
                                except Exception:
                                    pass

                            try:
                                self.after(0, _ui)
                            except Exception:
                                pass

                        threading.Thread(target=_bg, daemon=True).start()
            except Exception:
                try:
                    if self._widget_alive(getattr(self, "lbl_live_sync", None)):
                        pending = offline_pending_count()
                        self.lbl_live_sync.config(
                            text=f"📴 Offline — changes saved locally ({pending} pending)"
                        )
                except Exception:
                    pass

            next_interval = 12000 if (is_supabase_offline() or offline_pending_count()) else LIVE_SYNC_INTERVAL_MS
            if (
                getattr(self, "is_logged_in", False)
                and get_db_mode() == "supabase"
                and self.winfo_exists()
            ):
                self._live_sync_after_id = self.after(
                    next_interval, self._live_sync_tick
                )

        def _schedule_soft_ui_refresh(self, full=True):
            """Queue a quiet UI refresh on idle so typing/clicks aren't blocked."""
            if getattr(self, "_soft_refresh_pending", False):
                if full:
                    self._soft_refresh_full = True
                return
            self._soft_refresh_pending = True
            self._soft_refresh_full = bool(full)
            try:
                if self._widget_alive(getattr(self, "lbl_live_sync", None)):
                    self.lbl_live_sync.config(text="☁️ Updating…")
            except Exception:
                pass

            def _run():
                do_full = bool(getattr(self, "_soft_refresh_full", False))
                self._soft_refresh_full = False
                self._soft_refresh_pending = False
                try:
                    self._refresh_active_view_from_cloud(full=do_full)
                except Exception:
                    pass
                try:
                    if self._widget_alive(getattr(self, "lbl_live_sync", None)):
                        self.lbl_live_sync.config(text=cloud_sync_status_label())
                except Exception:
                    pass

            # Yield to the event loop first, then refresh
            self.after(50, _run)

        def _refresh_active_view_from_cloud(self, full=True):
            if getattr(self, "_rebuilding_ui", False):
                return
            if not hasattr(self, "notebook"):
                return
            # Rebuilding main-window widgets while the envelope sheet is open
            # deadlocks Aqua Tk (the sheet and the whole app stop accepting clicks).
            if self._envelope_ui_open():
                return
            # Always redraw Expense Reports if that window is open so a row
            # added on the other PC appears without restarting the app.
            try:
                win = getattr(self, "expenses_win", None)
                if self._widget_alive(win) and hasattr(self, "load_expenses_data"):
                    self.load_expenses_data(quiet=True)
            except Exception:
                pass
            if not full:
                return
            try:
                if hasattr(self, "load_financials_data"):
                    self.load_financials_data(quiet=True)
            except Exception:
                pass
            try:
                if hasattr(self, "load_cash_calendar_data") and not self._envelope_ui_open():
                    self.load_cash_calendar_data(quiet=True)
            except Exception:
                pass
            try:
                cb = getattr(self, "_envelope_popup_reload", None)
                pop = getattr(self, "_envelope_popup", None)
                if callable(cb) and self._widget_alive(pop):
                    cb()
            except Exception:
                pass
            try:
                if hasattr(self, "load_calendar_data"):
                    self.load_calendar_data(quiet=True)
            except Exception:
                pass
            try:
                if hasattr(self, "load_employees"):
                    self.load_employees(quiet=True)
            except Exception:
                pass

        def setup_calendar_tab(self):
            self.rev_cal_year = datetime.today().year
            today_iso = datetime.today().strftime('%Y-%m-%d')
            cur_ck = last_completed_cycle_for_date(today_iso)
            p = parse_cycle_key(cur_ck)
            if p:
                self.rev_cal_year = p[0].year
            self.selected_rev_cycles = {cur_ck} if cur_ck else {f"{self.rev_cal_year}-01-1"}
            self.cycle_card_widgets = {}

            # Top Control Bar
            top_frame = tb.Frame(self.tab_calendar, padding=(15, 10))
            top_frame.pack(side=TOP, fill=X)
            
            # Left: Year & Selection Controls
            year_frame = tb.Labelframe(top_frame, text=self._tr("Year:"), padding=(8, 4), bootstyle="info")
            year_frame.pack(side=LEFT, fill=Y, padx=(0, 10))
            
            tb.Button(year_frame, text="◀", bootstyle="outline-primary", width=3, cursor="hand2", command=self.prev_rev_year).pack(side=LEFT, padx=3)
            self.lbl_rev_year = tb.Label(year_frame, text=str(self.rev_cal_year), font=("Segoe UI", 15, "bold"), bootstyle="primary")
            self.lbl_rev_year.pack(side=LEFT, padx=10)
            tb.Button(year_frame, text="▶", bootstyle="outline-primary", width=3, cursor="hand2", command=self.next_rev_year).pack(side=LEFT, padx=3)
            
            tb.Separator(year_frame, orient=VERTICAL).pack(side=LEFT, fill=Y, padx=8)
            
            tb.Button(year_frame, text=f"🎯 {self._tr('Current Cycle')}", bootstyle="info-outline", cursor="hand2", command=self.select_current_rev_cycle).pack(side=LEFT, padx=3)
            
            self.btn_browse_cycles = tb.Button(
                year_frame,
                text=f"📅 {self._tr('Pay Cycles')} ▾",
                bootstyle="primary-outline",
                cursor="hand2",
                command=self._show_rev_cycles_popover,
            )
            self.btn_browse_cycles.pack(side=LEFT, padx=3)
            self.btn_browse_cycles.bind("<Enter>", self._show_rev_cycles_popover)
            self.btn_browse_cycles.bind("<Leave>", lambda e: self._schedule_popover_close())

            # Middle: Employee Filter
            emp_lf = tb.Labelframe(top_frame, text=self._tr("Employee:"), padding=(8, 4), bootstyle="secondary")
            emp_lf.pack(side=LEFT, fill=Y, padx=(0, 10))
            self.cal_name_filter = tb.Combobox(emp_lf, width=18, state="readonly")
            self.cal_name_filter.set(self._tr("All"))
            self.cal_name_filter.pack(side=LEFT, padx=5, pady=2)
            self.cal_name_filter.bind("<<ComboboxSelected>>", lambda e: self.load_calendar_data(quiet=True))

            # Right: Actions
            action_lf = tb.Labelframe(top_frame, text=self._tr("Actions"), padding=(8, 4), bootstyle="primary")
            action_lf.pack(side=RIGHT, fill=Y)
            
            self.btn_missing_rate_action = tb.Button(
                action_lf,
                text="⚠️ " + self._tr("+Add Rate"),
                bootstyle="danger",
                cursor="hand2",
            )
            tb.Button(action_lf, text=self._tr("Import Excel Sales"), bootstyle="info", cursor="hand2", command=self.open_excel_import_dialog).pack(side=LEFT, padx=4)
            tb.Button(action_lf, text=self._tr("✏️ Edit"), bootstyle="warning", cursor="hand2", command=self.edit_selected_record).pack(side=LEFT, padx=4)
            tb.Button(action_lf, text=self._tr("🗑️ Delete"), bootstyle="danger", cursor="hand2", command=self.delete_selected_record).pack(side=LEFT, padx=4)
            tb.Button(action_lf, text=self._tr("⚙️ Columns"), bootstyle="secondary-outline", cursor="hand2", command=self.open_calendar_columns_dialog).pack(side=LEFT, padx=4)
            tb.Label(action_lf, text=self._tr("(Ctrl / ⌘ multi-select)"), font=("Segoe UI", 8), bootstyle="secondary").pack(side=LEFT, padx=2)

            # Bottom Summary Bar
            summary_frame = tb.Frame(self.tab_calendar, padding=(12, 8), bootstyle="primary")
            summary_frame.pack(side=BOTTOM, fill=X, pady=(8, 0))
            self.lbl_summary = tb.Label(summary_frame, text="", font=("Segoe UI", 12, "bold"), bootstyle="inverse-primary")
            self.lbl_summary.pack(side=LEFT, padx=10)

            # Records Table Container with Frozen Employee Column
            tree_frame = tb.Frame(self.tab_calendar)
            tree_frame.pack(side=TOP, fill=BOTH, expand=True, padx=15, pady=(0, 5))
            
            # 1. Left Frozen Pane: Employee Name
            frozen_frame = tb.Frame(tree_frame, width=175)
            frozen_frame.pack_propagate(False)
            frozen_frame.pack(side=LEFT, fill=Y)

            self.tree_frozen = tb.Treeview(
                frozen_frame,
                columns=(self._tr("Name"),),
                show="headings",
                bootstyle="primary",
                selectmode="extended",
            )
            self.tree_frozen.heading(self._tr("Name"), text=self._tr("Name"))
            self.tree_frozen.column(self._tr("Name"), width=170, stretch=True)
            self.tree_frozen.pack(side=TOP, fill=BOTH, expand=True)

            frozen_spacer = tb.Frame(frozen_frame, height=18)
            frozen_spacer.pack(side=BOTTOM, fill=X)

            # 2. Right Vertical Scrollbar (Synchronized for both panes)
            scroll_y = tb.Scrollbar(tree_frame, orient=VERTICAL)
            scroll_y.pack(side=RIGHT, fill=Y)

            # 3. Right Scrollable Pane: Remaining Columns
            main_table_frame = tb.Frame(tree_frame)
            main_table_frame.pack(side=LEFT, fill=BOTH, expand=True)

            scroll_x = tb.Scrollbar(main_table_frame, orient=HORIZONTAL)
            scroll_x.pack(side=BOTTOM, fill=X)

            self.columns = tuple(self._tr(c) for c in (
                "Record ID", "Date", "Cycle", "Name", "Location",
                "Service Sales", "Service Sales Calculations",
                "Service Add-on Sales", "Product Sales", "Tip",
                "Hour Rate", "Percentage", "Hours",
                "Total Calculation", "Notes", "Written Up",
            ))

            disp_cols = [c for c in self.columns if c != self._tr("Name")]
            self.tree_calendar = tb.Treeview(
                main_table_frame,
                columns=self.columns,
                displaycolumns=disp_cols,
                show="headings",
                bootstyle="primary",
                xscrollcommand=scroll_x.set,
                selectmode="extended",
            )
            scroll_x.config(command=self.tree_calendar.xview)

            # Synchronize vertical scrolling
            def _sync_yview(*args):
                try:
                    self.tree_frozen.yview(*args)
                except Exception:
                    pass
                try:
                    self.tree_calendar.yview(*args)
                except Exception:
                    pass
            scroll_y.config(command=_sync_yview)

            def _on_cal_yscroll(*args):
                scroll_y.set(*args)
                try:
                    self.tree_frozen.yview_moveto(args[0])
                except Exception:
                    pass
            self.tree_calendar.configure(yscrollcommand=_on_cal_yscroll)

            def _on_frozen_yscroll(*args):
                scroll_y.set(*args)
                try:
                    self.tree_calendar.yview_moveto(args[0])
                except Exception:
                    pass
            self.tree_frozen.configure(yscrollcommand=_on_frozen_yscroll)

            # Synchronize mousewheel scrolling across both panes
            def _on_wheel(e):
                delta = int(-1 * (e.delta / 120)) if getattr(e, "delta", 0) else (1 if getattr(e, "num", 0) == 5 else -1)
                try:
                    self.tree_frozen.yview_scroll(delta, "units")
                    self.tree_calendar.yview_scroll(delta, "units")
                except Exception:
                    pass
                return "break"

            for t in (self.tree_frozen, self.tree_calendar):
                t.bind("<MouseWheel>", _on_wheel)
                t.bind("<Button-4>", _on_wheel)
                t.bind("<Button-5>", _on_wheel)

            # Synchronize row selection between frozen column and table
            def _sync_sel_from_frozen(e=None):
                sel = self.tree_frozen.selection()
                if self.tree_calendar.selection() != sel:
                    self.tree_calendar.selection_set(sel)
                    if sel:
                        self.tree_calendar.focus(sel[0])

            def _sync_sel_from_cal(e=None):
                sel = self.tree_calendar.selection()
                if self.tree_frozen.selection() != sel:
                    self.tree_frozen.selection_set(sel)
                    if sel:
                        self.tree_frozen.focus(sel[0])

            self.tree_frozen.bind("<<TreeviewSelect>>", _sync_sel_from_frozen)
            self.tree_calendar.bind("<<TreeviewSelect>>", _sync_sel_from_cal)

            self.tree_frozen.bind("<Double-1>", lambda e: self.edit_selected_record())
            self.tree_calendar.bind("<Double-1>", lambda e: self.edit_selected_record())

            self.apply_and_memorize_column_widths(
                "calendar_table",
                self.tree_calendar,
                disp_cols,
                hidden_cols=list(get_calendar_hidden_columns()),
            )
            self.tree_calendar.pack(side=TOP, fill=BOTH, expand=True)
            self.load_calendar_data()

        def _rebuild_cycle_cards(self):
            container = getattr(self, "rev_cards_container", None)
            if not self._widget_alive(container):
                return
            for child in container.winfo_children():
                child.destroy()
            self.cycle_card_widgets = {}

            # Configure 3 uniform columns for responsive compact layout
            for c in range(3):
                container.columnconfigure(c, weight=1, uniform="cycle_col")

            year_cycles = cycles_for_year(self.rev_cal_year)
            metrics_map = getattr(self, "last_cycle_metrics", {}) or {}
            
            # Continuous 14-day cycles (26-27 per year) across 3 columns with full vertical scrolling
            for idx, ck in enumerate(year_cycles):
                row_idx = idx // 3
                col_idx = idx % 3
                bounds = cycle_bounds(ck)
                s_b = bounds[0][5:].replace('-', '/') if bounds else ""
                e_b = bounds[1][5:].replace('-', '/') if bounds else ""
                date_range_str = f"{s_b} → {e_b}"
                
                metrics = metrics_map.get(ck, {"rev": 0.0, "cnt": 0})
                rev_val = metrics["rev"]
                cnt_val = metrics["cnt"]
                
                # Card frame
                is_sel = ck in self.selected_rev_cycles
                card = tb.Frame(
                    container,
                    borderwidth=2 if is_sel else 1,
                    relief="solid" if is_sel else "groove",
                    padding=(8, 5),
                    cursor="hand2",
                    bootstyle="primary" if is_sel else "default",
                )
                card.grid(row=row_idx, column=col_idx, padx=4, pady=3, sticky="nsew")
                
                lbl_title = tb.Label(
                    card,
                    text=("✓ " if is_sel else "") + cycle_short_label(ck),
                    font=("Segoe UI", 10, "bold"),
                    bootstyle="inverse-primary" if is_sel else "primary",
                    cursor="hand2",
                )
                lbl_title.pack(anchor=W)
                
                sub_frame = tb.Frame(card, cursor="hand2")
                sub_frame.pack(fill=X, expand=True, pady=1)
                
                lbl_dates = tb.Label(
                    sub_frame,
                    text=date_range_str,
                    font=("Segoe UI", 8),
                    bootstyle="inverse-primary" if is_sel else "secondary",
                    cursor="hand2",
                )
                lbl_dates.pack(side=LEFT)
                
                lbl_cnt = tb.Label(
                    sub_frame,
                    text=f"{cnt_val} {self._tr('entries')}" if cnt_val > 0 else f"0 {self._tr('entries')}",
                    font=("Segoe UI", 8),
                    bootstyle="inverse-primary" if is_sel else "secondary",
                    cursor="hand2",
                )
                lbl_cnt.pack(side=RIGHT)
                
                lbl_rev = tb.Label(
                    card,
                    text=f"${rev_val:,.2f}",
                    font=("Segoe UI", 10, "bold"),
                    bootstyle="inverse-primary" if is_sel else ("success" if rev_val > 0 else "secondary"),
                    cursor="hand2",
                )
                lbl_rev.pack(anchor=W)
                
                def _make_click(target_ck):
                    return lambda e: self.toggle_rev_cycle(target_ck, e)
                
                def _on_card_wheel(e):
                    cv = getattr(self, "rev_pop_canvas", None)
                    if cv and self._widget_alive(cv):
                        try:
                            if e.delta:
                                cv.yview_scroll(int(-1 * (e.delta / 120)), "units")
                            elif e.num == 4:
                                cv.yview_scroll(-1, "units")
                            elif e.num == 5:
                                cv.yview_scroll(1, "units")
                        except Exception:
                            pass

                handler = _make_click(ck)
                for w in (card, lbl_title, sub_frame, lbl_dates, lbl_cnt, lbl_rev):
                    w.bind("<Button-1>", handler)
                    w.bind("<MouseWheel>", _on_card_wheel)
                    w.bind("<Button-4>", _on_card_wheel)
                    w.bind("<Button-5>", _on_card_wheel)
                    
                self.cycle_card_widgets[ck] = {
                    "frame": card,
                    "title": lbl_title,
                    "dates": lbl_dates,
                    "cnt": lbl_cnt,
                    "rev": lbl_rev,
                }

        def prev_rev_year(self):
            self.rev_cal_year -= 1
            if hasattr(self, "lbl_rev_year") and self._widget_alive(self.lbl_rev_year):
                self.lbl_rev_year.config(text=str(self.rev_cal_year))
            if hasattr(self, "lbl_popover_year") and self._widget_alive(self.lbl_popover_year):
                self.lbl_popover_year.config(text=str(self.rev_cal_year))
            self._rebuild_cycle_cards()
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def next_rev_year(self):
            self.rev_cal_year += 1
            if hasattr(self, "lbl_rev_year") and self._widget_alive(self.lbl_rev_year):
                self.lbl_rev_year.config(text=str(self.rev_cal_year))
            if hasattr(self, "lbl_popover_year") and self._widget_alive(self.lbl_popover_year):
                self.lbl_popover_year.config(text=str(self.rev_cal_year))
            self._rebuild_cycle_cards()
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def select_current_rev_cycle(self):
            today_str = datetime.today().strftime('%Y-%m-%d')
            cur_ck = last_completed_cycle_for_date(today_str)
            p = parse_cycle_key(cur_ck)
            if p and p[0].year != self.rev_cal_year:
                self.rev_cal_year = p[0].year
                if hasattr(self, "lbl_rev_year") and self._widget_alive(self.lbl_rev_year):
                    self.lbl_rev_year.config(text=str(self.rev_cal_year))
                if hasattr(self, "lbl_popover_year") and self._widget_alive(self.lbl_popover_year):
                    self.lbl_popover_year.config(text=str(self.rev_cal_year))
            self.selected_rev_cycles = {cur_ck} if cur_ck else set()
            self._rebuild_cycle_cards()
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def select_all_rev_cycles(self):
            self.selected_rev_cycles = set(cycles_for_year(self.rev_cal_year))
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def clear_rev_cycles(self):
            self.selected_rev_cycles.clear()
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def toggle_rev_cycle(self, cycle_key, event=None):
            if cycle_key in self.selected_rev_cycles:
                self.selected_rev_cycles.remove(cycle_key)
            else:
                self.selected_rev_cycles.add(cycle_key)
            
            self._update_popover_sel_summary()
            self.load_calendar_data(quiet=True)

        def stop_filter_pollers(self, which=None):
            mapping = {
                "fin": ("_poll_fin_after_id", "tab_financials"),
                "expense": ("_poll_expense_after_id", "expenses_win"),
                "ledger": ("_poll_ledger_cycle_after_id", "tab_data_entry"),
            }
            targets = [mapping[which]] if which in mapping else list(mapping.values())
            for attr, win_attr in targets:
                aid = getattr(self, attr, None)
                if aid is None:
                    continue
                cancelled = False
                if win_attr:
                    win = getattr(self, win_attr, None)
                    if win is not None:
                        try:
                            if win.winfo_exists():
                                win.after_cancel(aid)
                                cancelled = True
                        except Exception:
                            pass
                if not cancelled:
                    try:
                        self.after_cancel(aid)
                    except Exception:
                        pass
                setattr(self, attr, None)

        def _dateentry_alive(self, widget):
            try:
                if widget is None or not widget.winfo_exists():
                    return False
                entry = getattr(widget, "entry", None)
                if entry is None:
                    return False
                return bool(entry.winfo_exists())
            except Exception:
                return False

        def poll_date_changes(self):
            pass

        def set_period(self, period):
            if period == "today" or period == "week" or period == "month":
                self.select_current_rev_cycle()
            elif period == "all":
                self.select_all_rev_cycles()

        def load_calendar_data(self, quiet=False):
            if quiet:
                return self._load_calendar_data_body()
            self.show_busy(self._tr("Loading shop earnings…"))
            try:
                return self._load_calendar_data_body()
            finally:
                self.hide_busy()

        def _load_calendar_data_body(self):
            try:
                conn_d = sqlite3.connect(TEMP_DB_PATH)
                cur_d = conn_d.cursor()
                dropped = _dedupe_payroll_table(cur_d)
                if dropped:
                    conn_d.commit()
                    try:
                        schedule_cloud_push(0.2)
                    except Exception:
                        pass
                conn_d.close()
            except Exception:
                pass

            for item in self.tree_calendar.get_children():
                self.tree_calendar.delete(item)
            if hasattr(self, "tree_frozen") and self._widget_alive(self.tree_frozen):
                for item in self.tree_frozen.get_children():
                    self.tree_frozen.delete(item)
                
            emp_f = getattr(self, 'cal_name_filter', None)
            emp_val = emp_f.get() if emp_f else self._tr("All")
            
            conn = sqlite3.connect(TEMP_DB_PATH)
            col_names = getattr(self, "_cache_payroll_cols", None)
            if col_names is None:
                cursor_check = conn.cursor()
                cursor_check.execute("PRAGMA table_info(payroll_records)")
                col_names = [col[1] for col in cursor_check.fetchall()]
                self._cache_payroll_cols = col_names
            hr_col = "r.hour_rate" if "hour_rate" in col_names else "NULL"
            perc_col = "r.percentage" if "percentage" in col_names else "NULL"
            has_cycle_key = "cycle_key" in col_names
            cyc_col = "r.cycle_key" if has_cycle_key else "NULL"
            
            # Load employee name -> id mapping
            cursor2 = conn.cursor()
            cursor2.execute("SELECT id, name FROM employees")
            name_to_id = {}
            for emp_id, name in cursor2.fetchall() or []:
                plain = decrypt_val(name) if name is not None else ""
                if plain is None:
                    plain = ""
                plain = str(plain).strip()
                if plain:
                    name_to_id[plain] = emp_id
            
            # Query all records for this entire year to populate both cards and table
            query_all = f'''
                SELECT r.id, r.record_date, e.name, r.location, r.revenue, r.service_addon_sales, r.product_sales, r.tip, 
                    {hr_col}, e.hour_rate, {perc_col}, e.percentage, r.hours, r.calculation, r.notes, r.written_up, e.use_tiered_payout, r.employee_id, {cyc_col}
                FROM payroll_records r
                JOIN employees e ON r.employee_id = e.id
                WHERE 1=1
            '''
            params_all = []
            if emp_val and emp_val != self._tr("All"):
                emp_id = name_to_id.get(emp_val)
                if emp_id is None:
                    for n, i in name_to_id.items():
                        if str(n).strip().lower() == str(emp_val).strip().lower():
                            emp_id = i
                            break
                if emp_id is not None:
                    query_all += " AND r.employee_id = ?"
                    params_all.append(emp_id)
                else:
                    query_all += " AND 1=0"
                
            query_all += " ORDER BY r.record_date DESC"
            
            cursor = conn.cursor()
            cursor.execute(query_all, params_all)
            all_year_rows = cursor.fetchall() or []
            conn.close()

            # Aggregate metrics per cycle for this year's cards
            cycle_metrics = {ck: {"rev": 0.0, "cnt": 0} for ck in cycles_for_year(self.rev_cal_year)}
            
            # Map rows to their resolved cycle
            resolved_rows = []
            for row in all_year_rows:
                rec_date = row[1]
                resolved_ck = cycle_for_date(rec_date)
                resolved_rows.append((row, resolved_ck))
                if resolved_ck in cycle_metrics:
                    cycle_metrics[resolved_ck]["rev"] += to_float(row[4], 0.0)
                    cycle_metrics[resolved_ck]["cnt"] += 1

            self.last_cycle_metrics = cycle_metrics

            # Update the 24 cycle card widgets if open
            for ck, metrics in cycle_metrics.items():
                w_dict = (getattr(self, "cycle_card_widgets", None) or {}).get(ck)
                if not w_dict:
                    continue
                try:
                    if not self._widget_alive(w_dict.get("frame")):
                        continue
                    is_sel = ck in self.selected_rev_cycles
                    lbl_title = w_dict["title"]
                    lbl_rev = w_dict["rev"]
                    lbl_cnt = w_dict["cnt"]
                    card_frame = w_dict["frame"]
                    
                    lbl_title.config(
                        text=("✓ " if is_sel else "") + cycle_short_label(ck),
                        bootstyle="inverse-primary" if is_sel else "primary"
                    )
                    lbl_rev.config(
                        text=f"${metrics['rev']:,.2f}",
                        bootstyle="success" if (metrics['rev'] > 0 and not is_sel) else ("primary" if is_sel else "secondary")
                    )
                    lbl_cnt.config(
                        text=f"{metrics['cnt']} {self._tr('entries')}" if metrics['cnt'] > 0 else f"0 {self._tr('entries')}"
                    )
                    card_frame.config(
                        relief="solid" if is_sel else "groove",
                        borderwidth=2 if is_sel else 1
                    )
                except Exception:
                    pass

            # Filter rows for the table: if self.selected_rev_cycles is non-empty, filter by it; otherwise show all
            if self.selected_rev_cycles:
                table_rows = [r_tuple for r_tuple in resolved_rows if r_tuple[1] in self.selected_rev_cycles]
            else:
                table_rows = resolved_rows

            # Pre-compute employee payout details and add-on sales to apply rules:
            # 1. Below 50%: top add-on earner gets 50%, rest get 40%
            # 2. At or above 50%: if tiered checked, get their service %; if fixed >= 50%, get 50%
            # 3. Hourly employees (hour_rate > 0 and no percentage/not tiered):
            #    (hours worked * hour rate) + (100% * tip)
            emp_service_perc = {}
            emp_use_tiered = {}
            emp_is_hourly = {}
            addon_by_emp = {}
            payout_cache = {}

            for row, row_ck in table_rows:
                emp_name = row[2]
                emp_id = row[17] if len(row) > 17 else name_to_id.get(emp_name, None)
                addon_by_emp[emp_id] = addon_by_emp.get(emp_id, 0.0) + to_float(row[5], 0.0)
                if emp_id not in emp_service_perc:
                    rec_perc = row[10]
                    emp_perc = row[11]
                    rec_hr = row[8]
                    emp_hr = row[9]
                    hr_num = to_float(rec_hr if rec_hr is not None else emp_hr, 0.0)
                    use_tiered = (row[16] == 1 or row[16] == '1' or row[16] is True)
                    emp_use_tiered[emp_id] = use_tiered
                    if use_tiered:
                        if emp_id:
                            if emp_id not in payout_cache:
                                bounds = cycle_bounds(row_ck) if row_ck else (row[1], row[1])
                                payout_cache[emp_id] = self.get_employee_payout_details(emp_id, bounds[0], bounds[1])
                            _, s_perc, _, _ = payout_cache[emp_id]
                            emp_service_perc[emp_id] = to_float(s_perc, 0.0)
                        else:
                            emp_service_perc[emp_id] = 0.0
                        emp_is_hourly[emp_id] = False
                    else:
                        s_perc = to_float(rec_perc if rec_perc is not None else emp_perc, 0.0)
                        emp_service_perc[emp_id] = s_perc
                        emp_is_hourly[emp_id] = (hr_num > 0 and s_perc <= 0)

            top_below_50_emp = None
            top_below_50_amt = -1.0
            for eid, amt in addon_by_emp.items():
                if eid is None:
                    continue
                if emp_is_hourly.get(eid, False):
                    # Hourly employees do not compete for commission add-on rate
                    continue
                s_perc = emp_service_perc.get(eid, 0.0)
                is_tiered = emp_use_tiered.get(eid, False)
                # Only commission barbers below 50%
                if (is_tiered or s_perc > 0) and s_perc < 0.50:
                    if amt > top_below_50_amt:
                        top_below_50_amt = amt
                        top_below_50_emp = eid

            total_rev = 0.0
            total_addon = 0.0
            total_prod = 0.0
            total_tip = 0.0
            total_hrs = 0.0
            total_svc_calc = 0.0
            total_calc = 0.0
            missing_rate_emp_ids = {}

            for r_idx, (row, row_ck) in enumerate(table_rows):
                emp_name = row[2]
                emp_id = row[17] if len(row) > 17 else name_to_id.get(emp_name, None)

                rec_hr = row[8]
                emp_hr = row[9]
                hour_rate = to_float(rec_hr if rec_hr is not None else emp_hr, 0.0)
                product_perc = 0.00

                service_perc = emp_service_perc.get(emp_id, 0.0)
                use_tiered = emp_use_tiered.get(emp_id, False)
                is_hourly = emp_is_hourly.get(emp_id, (hour_rate > 0 and not use_tiered and service_perc <= 0))

                if use_tiered and emp_id in payout_cache:
                    _, _, product_perc, t_hr = payout_cache[emp_id]
                    product_perc = to_float(product_perc, 0.0)
                    if not rec_hr:
                        hour_rate = to_float(t_hr, hour_rate)
                else:
                    product_perc = product_percent_for_sales(to_float(row[6], 0.0))

                rev_v = to_float(row[4], 0.0)
                addon_v = to_float(row[5], 0.0)
                prod_v = to_float(row[6], 0.0)
                tip_v = to_float(row[7], 0.0)
                loc_v = row[3] if row[3] else ""
                hrs_v = to_float(row[12], 0.0)

                rate_missing = (not use_tiered and service_perc <= 0 and hour_rate <= 0)
                hours_missing = (is_hourly and hrs_v <= 0)

                if rate_missing and emp_id:
                    missing_rate_emp_ids[emp_id] = emp_name

                if is_hourly:
                    svc_calc = round(hrs_v * hour_rate, 2)
                    addon_calc = 0.0
                    tip_calc = round(tip_v * 1.0, 2)
                    calc_v = round(svc_calc + tip_calc, 2)
                elif rate_missing:
                    svc_calc = 0.0
                    addon_calc = 0.0
                    tip_calc = round(tip_v * 1.0, 2)
                    calc_v = round(tip_calc, 2)
                else:
                    svc_calc = round((rev_v * service_perc) + (hrs_v * hour_rate), 2)
                    if service_perc < 0.50:
                        if emp_id is not None and emp_id == top_below_50_emp:
                            addon_rate = 0.50
                        else:
                            addon_rate = 0.40
                    else:
                        addon_rate = service_perc
                    addon_calc = round(addon_v * addon_rate, 2)
                    prod_calc = round(prod_v * product_perc, 2)
                    tip_calc = round(tip_v * 1.0, 2)
                    calc_v = round(svc_calc + addon_calc + tip_calc, 2)

                total_rev += rev_v
                total_addon += addon_v
                total_prod += prod_v
                total_tip += tip_v
                total_hrs += hrs_v
                total_svc_calc += svc_calc
                total_calc += calc_v

                cyc_display = cycle_label(row_ck) if row_ck else ""

                if rate_missing:
                    hr_disp = "⚠️ Not Set"
                    perc_disp = "🔴 Missing Rate (+Add)"
                elif is_hourly:
                    hr_disp = f"${hour_rate:,.2f}/hr"
                    perc_disp = "Hourly (No %)"
                elif use_tiered:
                    hr_disp = f"${hour_rate:,.2f}/hr" if hour_rate > 0 else "0.00/hr"
                    perc_disp = f"{service_perc * 100:.0f}% (P: {product_perc * 100:.0f}%)"
                else:
                    hr_disp = f"${hour_rate:,.2f}/hr" if hour_rate > 0 else "0.00/hr"
                    perc_disp = f"{service_perc * 100:.1f}%"

                if is_hourly and hours_missing:
                    hrs_disp = f"⚠️ {hrs_v:.1f} (Missing)"
                else:
                    hrs_disp = f"{hrs_v:.1f}"

                if rate_missing:
                    calc_disp = f"⚠️ ${calc_v:,.2f} (+Add Rate)"
                elif is_hourly and hours_missing:
                    calc_disp = f"⚠️ ${calc_v:,.2f} (Need Hours)"
                else:
                    calc_disp = f"⭐ ${calc_v:,.2f} ⭐"

                row_tags = ()
                if rate_missing:
                    row_tags = ('missing_rate',)
                elif is_hourly and hours_missing:
                    row_tags = ('missing_hours',)

                tree_row = [
                    row[0],
                    row[1],
                    cyc_display,
                    row[2],
                    loc_v,
                    f"${rev_v:,.2f}",
                    f"${svc_calc:,.2f}",
                    f"${addon_v:,.2f}",
                    f"${prod_v:,.2f}",
                    f"${tip_v:,.2f}",
                    hr_disp,
                    perc_disp,
                    hrs_disp,
                    calc_disp,
                    row[14] if row[14] else "",
                    row[15] if row[15] else "",
                ]
                iid = f"row_{row[0]}_{r_idx}"
                self.tree_calendar.insert('', tk.END, iid=iid, values=tree_row, tags=row_tags)
                if hasattr(self, "tree_frozen") and self._widget_alive(self.tree_frozen):
                    self.tree_frozen.insert('', tk.END, iid=iid, values=(row[2],), tags=row_tags)

            self.tree_calendar.insert('', tk.END, iid='spacer_row', values=("", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""))
            if hasattr(self, "tree_frozen") and self._widget_alive(self.tree_frozen):
                self.tree_frozen.insert('', tk.END, iid='spacer_row', values=("",))

            self.tree_calendar.insert('', tk.END, iid='totals_row', values=(
                "", "", "", "=== TOTALS ===", "",
                f"${total_rev:,.2f}", f"${total_svc_calc:,.2f}",
                f"${total_addon:,.2f}", f"${total_prod:,.2f}", f"${total_tip:,.2f}",
                "", "", f"{total_hrs:.1f}",
                f"⭐ ${total_calc:,.2f} ⭐",
                "", "",
            ), tags=('totals',))
            self.tree_calendar.tag_configure('totals', background='#375a7f', foreground='white', font=('Segoe UI', 11, 'bold'))
            self.tree_calendar.tag_configure('missing_rate', foreground='#e74c3c', font=('Segoe UI', 9, 'bold'))
            self.tree_calendar.tag_configure('missing_hours', foreground='#e67e22', font=('Segoe UI', 9, 'bold'))
            if hasattr(self, "tree_frozen") and self._widget_alive(self.tree_frozen):
                self.tree_frozen.insert('', tk.END, iid='totals_row', values=("=== TOTALS ===",), tags=('totals',))
                self.tree_frozen.tag_configure('totals', background='#375a7f', foreground='white', font=('Segoe UI', 11, 'bold'))
                self.tree_frozen.tag_configure('missing_rate', foreground='#e74c3c', font=('Segoe UI', 9, 'bold'))
                self.tree_frozen.tag_configure('missing_hours', foreground='#e67e22', font=('Segoe UI', 9, 'bold'))

            btn_add_rate = getattr(self, "btn_missing_rate_action", None)
            if self._widget_alive(btn_add_rate):
                if missing_rate_emp_ids:
                    first_eid, first_ename = next(iter(missing_rate_emp_ids.items()))
                    btn_add_rate.config(
                        text=f"⚠️ {self._tr('+Add Rate for')} {first_ename}",
                        command=lambda eid=first_eid: self.open_employee_edit_by_id(eid),
                    )
                    btn_add_rate.pack(side=LEFT, padx=4)
                else:
                    btn_add_rate.pack_forget()

            # Format summary banner text
            num_sel = len(self.selected_rev_cycles)
            if num_sel == 1:
                ck_single = next(iter(self.selected_rev_cycles))
                sel_desc = f"{self._tr('Selected')}: {cycle_label_with_year(ck_single)}"
            elif num_sel > 1:
                sel_desc = f"{self._tr('Selected')}: {num_sel} {self._tr('Pay Cycles')}"
            else:
                sel_desc = f"{self._tr('Year:')} {self.rev_cal_year} ({self._tr('All Cycles')})"

            self.lbl_summary.config(
                text=f"{sel_desc}  |  {self._tr('Revenue')}: ${total_rev:,.2f}  |  {self._tr('Service Add-on Sales')}: ${total_addon:,.2f}  |  {self._tr('Tips')}: ${total_tip:,.2f}  |  {self._tr('Total Calculation')}: ${total_calc:,.2f}"
            )

        def export_excel(self):
            if hasattr(self, "tree_calendar"):
                rows = self.tree_calendar.get_children()
                if not rows or len(rows) <= 2:
                    self.load_calendar_data(quiet=True)
                    rows = self.tree_calendar.get_children()
            else:
                rows = []

            if not rows:
                messagebox.showinfo(self._tr("Export"), self._tr("No data available to export."))
                return
                
            filepath = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[
                    ("CSV Files", "*.csv"),
                    ("Excel Files", "*.xlsx"),
                    ("All Files", "*.*")
                ],
                title=self._tr("Save Revenue Data")
            )
            
            if filepath:
                try:
                    cols_to_export = list(self.columns[1:])
                    data_rows = []
                    for item in rows:
                        if item in ('spacer_row', 'totals_row'):
                            continue
                        vals = self.tree_calendar.item(item)['values']
                        if vals and len(vals) > 1:
                            data_rows.append(list(vals[1:]))
                    
                    if filepath.lower().endswith(".csv"):
                        import csv
                        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                            writer = csv.writer(f)
                            writer.writerow(cols_to_export)
                            for r in data_rows:
                                writer.writerow(r)
                    else:
                        import openpyxl
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Revenue Dashboard"
                        ws.append(cols_to_export)
                        for r in data_rows:
                            ws.append(r)
                        wb.save(filepath)
                    messagebox.showinfo(self._tr("Success"), f"{self._tr('Data successfully exported to:')}\n{filepath}")
                except Exception as e:
                    messagebox.showerror(self._tr("Error"), f"{self._tr('Could not save file:')}\n{e}")

        def delete_selected_record(self):
            selected = self.tree_calendar.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select one or more records to delete.")
                return
            
            record_ids = []
            for sel in selected:
                item_vals = self.tree_calendar.item(sel)['values']
                if item_vals:
                    record_ids.append(item_vals[0])
            
            if not record_ids:
                return 
                
            confirm_msg = "Are you sure you want to completely delete this record? This cannot be undone." if len(record_ids) == 1 else f"Are you sure you want to completely delete all {len(record_ids)} selected records?\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.\nThis cannot be undone."
            if messagebox.askyesno("Confirm Delete", confirm_msg):
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    if len(record_ids) == 1:
                        cursor.execute("DELETE FROM payroll_records WHERE id=?", (record_ids[0],))
                    else:
                        placeholders = ",".join("?" for _ in record_ids)
                        cursor.execute(f"DELETE FROM payroll_records WHERE id IN ({placeholders})", record_ids)
                    commit_and_save(conn)
                except Exception as e:
                    messagebox.showerror("Database Error", f"Failed to delete record: {e}")
                    if 'conn' in locals():
                        conn.close()
                    return
                    
                conn.close()
                self.load_calendar_data()

        def edit_selected_record(self):
            selected = self.tree_calendar.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select a record to edit.")
                return
            item = self.tree_calendar.item(selected[0])
            record_id = item['values'][0]
            if not record_id:
                return
            
            if "missing_rate" in (item.get("tags", ()) or ()):
                vals = item.get("values", [])
                emp_name = vals[3] if len(vals) > 3 else ""
                if messagebox.askyesno(
                    self._tr("Missing Rate"),
                    f"Employee '{emp_name}' does not have a percentage or hourly rate configured.\n\n"
                    "Would you like to open the employee settings now to configure their rate/percentage?",
                    parent=self
                ):
                    try:
                        conn_e = sqlite3.connect(TEMP_DB_PATH)
                        c_e = conn_e.cursor()
                        c_e.execute("SELECT employee_id FROM payroll_records WHERE id=?", (record_id,))
                        row_e = c_e.fetchone()
                        conn_e.close()
                        if row_e and row_e[0]:
                            self.open_employee_edit_by_id(row_e[0])
                            return
                    except Exception:
                        pass
            
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(payroll_records)")
            col_names = [col[1] for col in cursor.fetchall()]
            hr_col = "r.hour_rate" if "hour_rate" in col_names else "NULL"
            perc_col = "r.percentage" if "percentage" in col_names else "NULL"
            
            cursor.execute(f'''
                SELECT r.record_date, r.revenue, r.service_addon_sales, r.hours, r.notes, r.written_up, e.name, r.written_up_desc,
                    COALESCE({hr_col}, e.hour_rate), COALESCE({perc_col}, e.percentage), e.use_tiered_payout, e.id
                FROM payroll_records r
                JOIN employees e ON r.employee_id = e.id
                WHERE r.id=?
            ''', (record_id,))
            rec = cursor.fetchone()
            conn.close()
            
            if not rec:
                return
                
            self.open_edit_record_dialog(record_id, rec)

        def open_edit_record_dialog(self, record_id, rec, parent=None):
            parent_win = parent if self._widget_alive(parent) else self
            dialog = tb.Toplevel(parent_win)
            dialog.title(f"Edit Record: {rec[6]}")
            dialog.transient(parent_win)
            self._register_modal_popup(dialog)
            dialog.focus_set()
            try:
                dialog.update_idletasks()
                screen_w = dialog.winfo_screenwidth()
                screen_h = dialog.winfo_screenheight()
                dlg_w = min(540, max(460, screen_w - 40))
                dlg_h = min(680, max(500, screen_h - 90))
                x = max(0, (screen_w - dlg_w) // 2)
                y = max(10, min(25, (screen_h - dlg_h) // 4))
                dialog.geometry(f"{dlg_w}x{dlg_h}+{x}+{y}")
            except Exception:
                dialog.geometry("520x660")
            
            pad = {'padx': 15, 'pady': 8, 'sticky': E}
            ent_pad = {'padx': 15, 'pady': 8, 'sticky': W}
            
            tb.Label(dialog, text="Date:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, **pad)
            date_ent = tb.DateEntry(dialog, bootstyle="primary", dateformat='%Y-%m-%d')
            date_ent.entry.delete(0, tk.END)
            date_ent.entry.insert(0, rec[0])
            date_ent.grid(row=0, column=1, **ent_pad)
            
            tb.Label(dialog, text="Revenue (Service Sales):", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, **pad)
            rev_ent = tb.Entry(dialog, width=32)
            rev_ent.insert(0, str(rec[1]) if rec[1] else "0")
            rev_ent.grid(row=1, column=1, **ent_pad)

            tb.Label(dialog, text="Service Add-on Sales:", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, **pad)
            addon_ent = tb.Entry(dialog, width=32)
            addon_ent.insert(0, str(rec[2]) if rec[2] else "0")
            addon_ent.grid(row=2, column=1, **ent_pad)
            
            tb.Label(dialog, text="Hours:", font=("Segoe UI", 10, "bold")).grid(row=3, column=0, **pad)
            hrs_ent = tb.Entry(dialog, width=32)
            hrs_ent.insert(0, str(rec[3]) if rec[3] else "0")
            hrs_ent.grid(row=3, column=1, **ent_pad)
            
            # Hour Rate ($):
            tb.Label(dialog, text=self._tr("Hour Rate ($):"), font=("Segoe UI", 10, "bold")).grid(row=4, column=0, **pad)
            hr_ent = tb.Entry(dialog, width=32)
            hr_ent.insert(0, f"{to_float(rec[8], 0.0):.2f}")
            hr_ent.grid(row=4, column=1, **ent_pad)

            # Percentage (%):
            tb.Label(dialog, text=self._tr("Percentage (%):"), font=("Segoe UI", 10, "bold")).grid(row=5, column=0, **pad)
            perc_ent = tb.Entry(dialog, width=32)
            perc_display = to_float(rec[9], 0.0) * 100.0
            perc_ent.insert(0, f"{perc_display:.1f}")
            perc_ent.grid(row=5, column=1, **ent_pad)
            if rec[10]:  # use_tiered_payout
                perc_ent.config(state="disabled")

            tb.Label(dialog, text="Notes:", font=("Segoe UI", 10, "bold")).grid(row=6, column=0, **pad)
            notes_ent = tb.Entry(dialog, width=32)
            notes_ent.insert(0, rec[4] if rec[4] else "")
            notes_ent.grid(row=6, column=1, **ent_pad)
            
            tb.Label(dialog, text="Written Up:", font=("Segoe UI", 10, "bold")).grid(row=7, column=0, **pad)
            writ_ent = tb.Combobox(dialog, width=30, state="readonly", values=["", "Yes", "No"])
            writ_ent.set(rec[5] if rec[5] in ["Yes", "No"] else "")
            writ_ent.grid(row=7, column=1, **ent_pad)

            lbl_reason = tb.Label(dialog, text=self._tr("Write Up Reason:"), font=("Segoe UI", 10, "bold"))
            reason_ent = tb.Entry(dialog, width=32)
            if len(rec) > 7 and rec[7]:
                reason_ent.insert(0, rec[7])

            # Help Definitions
            help_msgs = {
                "Date": ("Enter the date of the work shift (YYYY-MM-DD).", "أدخل تاريخ وردية العمل (YYYY-MM-DD)."),
                "Revenue": ("Main service sales generated by the employee.", "مبيعات الخدمات الرئيسية التي حققها الموظف."),
                "Service Add-on Sales": ("Additional sales from add-on services (split at the same rate).", "المبيعات الإضافية من الخدمات التكميلية (تُقسم بنفس النسبة)."),
                "Hours": ("Total hours worked during this shift (used for hourly pay calculation).", "إجمالي الساعات التي تم عملها خلال الوردية (تُستخدم لحساب الدفع بالساعة)."),
                "Hour Rate": ("Hourly pay rate for this employee. Defaults to employee setting.", "أجر الساعة لهذا الموظف. القيمة الافتراضية مأخوذة من إعدادات الموظف."),
                "Percentage": ("Commission percentage for services. Disabled (shows Auto) if tiered payout is enabled for the employee.", "نسبة عمولة الخدمات. تظهر معطلة (تلقائي) إذا كان نظام الدفع المتدرج مفعلاً للموظف."),
                "Notes": ("Optional shift notes, details or references.", "ملاحظات اختيارية للوردية أو تفاصيل إضافية."),
                "Written Up": ("Select Yes if this shift includes a formal write-up/citation for the employee.", "اختر نعم إذا كانت هذه الوردية تتضمن إنذاراً أو مخالفة رسمية للموظف."),
                "Write Up Reason": ("Describe the reason for the write-up (required if Written Up is Yes).", "صف سبب الإنذار/المخالفة (مطلوب إذا تم تحديد نعم فوق).")
            }

            def add_dialog_help_btn(row, field_key):
                msg_en, msg_ar = help_msgs.get(field_key, ("", ""))
                btn = tb.Button(dialog, text="❓", bootstyle="link", cursor="hand2")
                btn.grid(row=row, column=2, padx=(5, 5), pady=2, sticky=W)
                msg = msg_ar if getattr(self, 'lang', 'en') == 'ar' else msg_en
                ToolTip(btn, text=msg)

            add_dialog_help_btn(0, "Date")
            add_dialog_help_btn(1, "Revenue")
            add_dialog_help_btn(2, "Service Add-on Sales")
            add_dialog_help_btn(3, "Hours")
            add_dialog_help_btn(4, "Hour Rate")
            add_dialog_help_btn(5, "Percentage")
            add_dialog_help_btn(6, "Notes")
            add_dialog_help_btn(7, "Written Up")

            help_reason_btn = tb.Button(dialog, text="❓", bootstyle="link", cursor="hand2")
            msg_en, msg_ar = help_msgs["Write Up Reason"]
            msg_reason = msg_ar if getattr(self, 'lang', 'en') == 'ar' else msg_en
            ToolTip(help_reason_btn, text=msg_reason)

            def on_writeup_changed(event=None):
                if writ_ent.get() == "Yes":
                    lbl_reason.grid(row=8, column=0, **pad)
                    reason_ent.grid(row=8, column=1, **ent_pad)
                    help_reason_btn.grid(row=8, column=2, padx=(5, 5), pady=2, sticky=W)
                else:
                    lbl_reason.grid_remove()
                    reason_ent.grid_remove()
                    help_reason_btn.grid_remove()

            writ_ent.bind("<<ComboboxSelected>>", on_writeup_changed)
            on_writeup_changed()
            
            def save_updates(event=None):
                date_val = date_ent.entry.get()
                if not date_val:
                    messagebox.showerror("Error", "Date is required.", parent=dialog)
                    return
                try:
                    datetime.strptime(date_val, '%Y-%m-%d')
                except ValueError:
                    messagebox.showerror("Error", "Invalid Date format. Use YYYY-MM-DD.", parent=dialog)
                    return
                    
                try:
                    r_amt = float(rev_ent.get() or 0)
                    addon_amt = float(addon_ent.get() or 0)
                    h_amt = float(hrs_ent.get() or 0)
                    cust_hr_val = float(hr_ent.get() or 0)
                except ValueError:
                    messagebox.showerror("Error", "Amounts/Hours must be numeric.", parent=dialog)
                    return
                    
                if r_amt < 0 or addon_amt < 0 or h_amt < 0 or cust_hr_val < 0:
                    messagebox.showerror("Error", "Amounts, hours and rates cannot be negative.", parent=dialog)
                    return

                write_up_val = writ_ent.get()
                reason_val = reason_ent.get().strip()
                if write_up_val == "Yes" and not reason_val:
                    messagebox.showerror("Validation Error", "Write Up Reason is mandatory when Written Up is Yes.", parent=dialog)
                    return
                    
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT e.use_tiered_payout, e.id, e.name 
                    FROM employees e JOIN payroll_records r ON r.employee_id = e.id 
                    WHERE r.id = ?
                ''', (record_id,))
                use_tiered, emp_id, emp_name = cursor.fetchone()
                
                if use_tiered:
                    perc_to_save = None
                    _, service_perc, product_perc, _ = self.get_employee_payout_details(emp_id, date_val, date_val)
                    calc = round(((r_amt + addon_amt) * service_perc) + (h_amt * cust_hr_val), 2)
                else:
                    try:
                        perc_to_save = float(perc_ent.get() or 0) / 100.0
                    except ValueError:
                        messagebox.showerror("Error", "Percentage must be a number.", parent=dialog)
                        conn.close()
                        return
                    calc = round(((r_amt + addon_amt) * perc_to_save) + (h_amt * cust_hr_val), 2)
                    
                try:
                    _cyc_to_save = cycle_for_date(date_val)
                    cursor.execute('''
                        UPDATE payroll_records 
                        SET record_date=?, revenue=?, service_addon_sales=?, hours=?, notes=?, written_up=?, calculation=?, written_up_desc=?, hour_rate=?, percentage=?, cycle_key=?
                        WHERE id=?
                    ''', (date_val, r_amt, addon_amt, h_amt, notes_ent.get(), write_up_val, calc, reason_val if write_up_val == "Yes" else "", cust_hr_val, perc_to_save, _cyc_to_save, record_id))
                    
                    commit_and_save(conn)
                except Exception as e:
                    messagebox.showerror("Database Error", f"Failed to update record: {e}", parent=dialog)
                    conn.close()
                    return
                
                conn.close()
                dialog.destroy()
                if self._widget_alive(parent_win):
                    try:
                        parent_win.lift()
                        parent_win.focus_set()
                    except Exception:
                        pass
                if self._widget_alive(getattr(self, "expenses_win", None)):
                    self.load_expenses_data(quiet=True)
                self.load_calendar_data()
                messagebox.showinfo("Success", f"Record updated.\nNew calculated pay: ${calc:,.2f}")
                
            def _cancel_edit():
                dialog.destroy()
                if self._widget_alive(parent_win):
                    try:
                        parent_win.lift()
                        parent_win.focus_set()
                    except Exception:
                        pass

            on_writeup_changed()
            btnf = tb.Frame(dialog)
            btnf.grid(row=9, column=0, columnspan=2, pady=30)
            tb.Button(btnf, text=self._tr("Save"), bootstyle="success", cursor="hand2", command=save_updates).pack(side=LEFT, ipadx=30, ipady=8)
            tb.Button(btnf, text=self._tr("Cancel"), bootstyle="secondary", cursor="hand2", command=_cancel_edit).pack(side=LEFT, padx=12, ipadx=24, ipady=8)
            self._bind_dialog_save_keys(dialog, save_updates)
            dialog.bind("<Escape>", lambda e: _cancel_edit())

        def setup_names_tab(self):
            header_frame = tb.Frame(self.tab_names)
            header_frame.pack(fill=X, padx=25, pady=(20, 20))
            
            tb.Label(header_frame, text=self._tr("👥 Employee Management"), font=("Segoe UI", 22, "bold"), bootstyle="primary").pack(side=LEFT)
            
            btn_frame = tb.Frame(self.tab_names)
            btn_frame.pack(fill=X, padx=25, pady=(0, 15))
            tb.Button(btn_frame, text=self._tr("+ Add New Employee"), bootstyle="success", cursor="hand2", command=lambda: self.open_employee_dialog()).pack(side=LEFT)
            tb.Button(btn_frame, text=self._tr("✏️ Edit Selected"), bootstyle="warning", cursor="hand2", command=self.edit_selected_employee).pack(side=LEFT, padx=10)
            tb.Button(btn_frame, text=self._tr("🗑️ Delete Selected"), bootstyle="danger", cursor="hand2", command=self.delete_selected_employee).pack(side=LEFT, padx=10)
            tb.Label(btn_frame, text=self._tr("(Ctrl / ⌘ + click to multi-select)"), font=("Segoe UI", 9), bootstyle="secondary").pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("📊 View Performance"), bootstyle="info", cursor="hand2", command=self.view_employee_summary).pack(side=LEFT, padx=10)
            
            columns = tuple(self._tr(c) for c in ("ID", "First Name", "Last Name", "Phone", "Email", "Hour Rate", "Percentage"))
            names_tree_holder = tb.Frame(self.tab_names)
            names_tree_holder.pack(fill=BOTH, expand=True, padx=25, pady=20)
            self.tree_names = tb.Treeview(names_tree_holder, columns=columns, show="headings", bootstyle="info", selectmode="extended")
            self.apply_and_memorize_column_widths(
                "employees_table",
                self.tree_names,
                columns,
            )
            
            self.tree_names.bind("<Double-1>", self.show_employee_action_popup)
            self._attach_tree_scrollbars(names_tree_holder, self.tree_names)
            self.load_employees()

        def rename_login_username(self, old_name, new_name, parent=None):
            """Rename a login account locally and queue the change for the other PCs."""
            global CURRENT_SESSION_USER
            old_name = plain_label(old_name).strip()
            new_name = (new_name or "").strip()
            if not old_name:
                messagebox.showerror("Error", "No user selected.", parent=parent)
                return False
            if not new_name:
                messagebox.showerror("Error", "Username cannot be empty.", parent=parent)
                return False
            if len(new_name) > 40:
                messagebox.showerror("Error", "Username is too long.", parent=parent)
                return False
            if new_name == old_name:
                messagebox.showinfo("Username", "That is already the username.", parent=parent)
                return False

            conn = sqlite3.connect(TEMP_DB_PATH)
            cur = conn.cursor()
            try:
                cur.execute("SELECT username, password FROM users")
                rows = cur.fetchall() or []
                stored_old = None
                stored_pw = None
                existing = set()
                for uname, pw in rows:
                    plain = plain_label(uname)
                    if not plain:
                        continue
                    existing.add(plain.lower())
                    if plain.lower() == old_name.lower():
                        stored_old = uname
                        stored_pw = pw
                if stored_old is None:
                    messagebox.showerror("Error", "User not found.", parent=parent)
                    return False
                if new_name.lower() in existing and new_name.lower() != old_name.lower():
                    messagebox.showerror("Error", "This username already exists.", parent=parent)
                    return False
                cur.execute(
                    "UPDATE users SET username=? WHERE username=?",
                    (new_name, stored_old),
                )
                commit_and_save(conn)
            except sqlite3.IntegrityError:
                messagebox.showerror("Error", "This username already exists.", parent=parent)
                return False
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=parent)
                return False
            finally:
                try:
                    conn.close()
                except Exception:
                    pass

            try:
                _queue_offline_op(
                    {
                        "op": "rename_key",
                        "table": "users",
                        "key": "username",
                        "old": old_name,
                        "row": {"username": new_name, "password": stored_pw},
                    }
                )
                schedule_cloud_push(0.05)
            except Exception:
                pass

            current = plain_label(getattr(self, "current_user", "") or "")
            if current.lower() == old_name.lower():
                self.current_user = new_name
                CURRENT_SESSION_USER = new_name
                try:
                    save_last_selected_username(new_name)
                except Exception:
                    pass
                try:
                    if self._widget_alive(getattr(self, "lbl_logged_in_user", None)):
                        self.lbl_logged_in_user.config(
                            text=f"{self._tr('Logged in as')} {new_name.capitalize()}"
                        )
                except Exception:
                    pass
            try:
                log_user_action(
                    "rename_user",
                    table="users",
                    record_id=new_name,
                    row={"username": new_name, "old_username": old_name},
                    extra_summary=f"Changed username from {old_name} to {new_name}",
                )
            except Exception:
                pass
            messagebox.showinfo(
                "Success",
                self._tr("Username updated. Use the new name at login."),
                parent=parent,
            )
            return True

        def change_login_username(self):
            dialog = tb.Toplevel(self)
            dialog.title(self._tr("Change Username"))
            dialog.geometry("420x260")
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()

            who = getattr(self, "current_user", DEFAULT_ADMIN_USERNAME) or DEFAULT_ADMIN_USERNAME
            tb.Label(
                dialog,
                text=self._tr("Change username"),
                font=("Segoe UI", 14, "bold"),
                bootstyle="primary",
            ).pack(pady=(18, 8))
            tb.Label(
                dialog,
                text=f"{self._tr('Current username')}:  {who}",
                font=("Segoe UI", 10),
                bootstyle="secondary",
            ).pack(anchor=W, padx=30)
            tb.Label(dialog, text=self._tr("New username:"), font=("Segoe UI", 11, "bold")).pack(
                anchor=W, padx=30, pady=(12, 4)
            )
            new_ent = tb.Entry(dialog, width=32, font=("Segoe UI", 11))
            new_ent.insert(0, who)
            new_ent.pack(padx=30, pady=4)
            new_ent.focus()
            new_ent.select_range(0, tk.END)

            def save_name(event=None):
                if self.rename_login_username(who, new_ent.get(), parent=dialog):
                    dialog.destroy()

            new_ent.bind("<Return>", save_name)
            tb.Button(
                dialog,
                text=self._tr("Update Username"),
                bootstyle="info",
                command=save_name,
            ).pack(pady=18)

        def change_login_password(self):
            dialog = tb.Toplevel(self)
            dialog.title("Change App Login Password")
            dialog.geometry("400x320")
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()
            
            tb.Label(dialog, text="Current Password:").pack(pady=(20, 5))
            current_pw_entry = tb.Entry(dialog, show="*", width=30)
            current_pw_entry.pack(pady=5)
            
            tb.Label(dialog, text="New Password:").pack(pady=(10, 5))
            new_pw_entry = tb.Entry(dialog, show="*", width=30)
            new_pw_entry.pack(pady=5)
            
            tb.Label(dialog, text="Confirm New Password:").pack(pady=(10, 5))
            confirm_pw_entry = tb.Entry(dialog, show="*", width=30)
            confirm_pw_entry.pack(pady=5)
            
            def save_password():
                current = current_pw_entry.get()
                new_pw = new_pw_entry.get()
                confirm = confirm_pw_entry.get()
                
                if new_pw != confirm:
                    messagebox.showerror("Error", "New passwords do not match.", parent=dialog)
                    return
                    
                current_user = getattr(self, "current_user", "admin")
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT password FROM users WHERE username=?", (current_user,))
                stored = cursor.fetchone()[0]
                
                hashed_current = hashlib.sha256(current.encode()).hexdigest()
                
                if stored != hashed_current and stored != current:
                    messagebox.showerror("Error", "Current password is incorrect.", parent=dialog)
                    conn.close()
                    return
                    
                hashed_new = hashlib.sha256(new_pw.encode()).hexdigest()
                cursor.execute("UPDATE users SET password=? WHERE username=?", (hashed_new, current_user))
                commit_and_save(conn)
                conn.close()
                
                messagebox.showinfo("Success", "Login password changed successfully!", parent=dialog)
                dialog.destroy()
                
            tb.Button(dialog, text="Update Password", bootstyle="success", command=save_password).pack(pady=20)

        def change_db_location(self, default_tab=None):
            return self.open_settings_dialog(default_tab="database" if default_tab != "supabase" else "supabase")

        def load_employees(self, quiet=False):
            if quiet:
                return self._load_employees_body()
            self.show_busy(self._tr("Loading employees…"))
            try:
                return self._load_employees_body()
            finally:
                self.hide_busy()

        def _load_employees_body(self):
            for item in self.tree_names.get_children():
                self.tree_names.delete(item)
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT id, first_name, last_name, phone, email, hour_rate, percentage, use_tiered_payout FROM employees")
                rows = cursor.fetchall()
            except sqlite3.OperationalError:
                try:
                    cursor.execute("SELECT id, name, '', '', '', hour_rate, percentage, 0 FROM employees")
                    rows = cursor.fetchall()
                except Exception:
                    rows = []
                    
            for row in rows:
                fmt = list(row[:7])
                use_tiered = row[7] if len(row) > 7 else 0
                fmt[1] = fmt[1] if fmt[1] else ""
                fmt[2] = fmt[2] if fmt[2] else ""
                fmt[3] = fmt[3] if fmt[3] else ""
                fmt[4] = fmt[4] if fmt[4] else ""
                rate = to_float(fmt[5], 0.0)
                perc = to_float(fmt[6], 0.0)
                fmt[5] = f"${rate:,.2f}"
                if use_tiered:
                    fmt[6] = "Tiered (35%-50%)"
                else:
                    fmt[6] = f"{perc * 100:g}%"
                self.tree_names.insert('', tk.END, values=fmt)
            conn.close()
            self.refresh_employee_dropdown()

        def delete_selected_employee(self):
            selected = self.tree_names.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select one or more employees to delete.\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.")
                return

            skipped_shop = False
            targets = []
            for sel in selected:
                item = self.tree_names.item(sel)
                vals = item.get("values") or []
                if len(vals) < 3:
                    continue
                emp_id = vals[0]
                emp_name = (str(vals[1]) + " " + str(vals[2])).strip()
                first = str(vals[1] or "").strip().lower()
                last = str(vals[2] or "").strip().lower()
                if emp_name.lower() == "shop" or first == "shop" or f"{first} {last}".strip() == "shop":
                    skipped_shop = True
                    continue
                targets.append((emp_id, emp_name))

            if skipped_shop and not targets:
                messagebox.showerror("Error", "The 'Shop' account is protected and cannot be deleted.")
                return
            if skipped_shop:
                messagebox.showinfo("Protected", "The 'Shop' account was skipped and cannot be deleted.")

            if not targets:
                return

            names_preview = ", ".join(n for _, n in targets[:5])
            if len(targets) > 5:
                names_preview += f", +{len(targets) - 5} more"

            del_win = tb.Toplevel(self)
            del_win.title("Confirm Delete")
            del_win.geometry("440x240")
            del_win.transient(self)
            del_win.grab_set()

            tb.Label(
                del_win,
                text=f"Delete {len(targets)} employee(s)?\n{names_preview}",
                font=("Segoe UI", 11, "bold"),
                justify=CENTER,
            ).pack(pady=15)
            tb.Label(del_win, text="Type 'delete' to confirm:").pack()

            entry_confirm = tb.Entry(del_win)
            entry_confirm.pack(pady=10)
            entry_confirm.focus()

            def confirm_delete(event=None):
                if entry_confirm.get().strip().lower() != "delete":
                    messagebox.showerror("Error", "You didn't type 'delete'. Canceled.", parent=del_win)
                    return
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    for emp_id, _ in targets:
                        cursor.execute("DELETE FROM employees WHERE id=?", (emp_id,))
                        cursor.execute("DELETE FROM payroll_records WHERE employee_id=?", (emp_id,))
                    commit_and_save(conn)
                    conn.close()
                    del_win.destroy()
                    self.load_employees()
                    self.load_calendar_data()
                    messagebox.showinfo("Success", f"Deleted {len(targets)} employee(s) and their payroll records.")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete: {e}", parent=del_win)

            entry_confirm.bind("<Return>", confirm_delete)
            tb.Button(del_win, text="Confirm", bootstyle="danger", command=confirm_delete).pack(pady=5)

        def view_employee_summary(self):
            selected = self.tree_names.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select an employee to view their summary.")
                return
                
            item = self.tree_names.item(selected[0])
            emp_id = item['values'][0]
            emp_name = item['values'][1]
            
            win = tb.Toplevel(self)
            win.title(f"Payroll Summary: {emp_name}")
            try:
                self.update_idletasks()
                w = self.winfo_width()
                h = self.winfo_height()
                x = self.winfo_x()
                y = self.winfo_y()
                win.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                win.geometry("1100x600")
            win.transient(self)
            win.grab_set()
            win.focus_set()
            
            tb.Label(win, text=f"Performance & Payroll Summary for {emp_name}", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(pady=15)
            
            # Row 1: Date Filter Row
            date_frame = tb.Frame(win, padding=5)
            date_frame.pack(side=TOP, fill=X, padx=25, pady=(10, 5))
            
            tb.Label(date_frame, text="From Date:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(5, 5))
            cal_from_date = tb.DateEntry(date_frame, bootstyle="info", dateformat='%Y-%m-%d')
            cal_from_date.pack(side=LEFT, padx=5)
            
            tb.Label(date_frame, text="To Date:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            cal_to_date = tb.DateEntry(date_frame, bootstyle="info", dateformat='%Y-%m-%d')
            cal_to_date.pack(side=LEFT, padx=5)
            
            def set_summary_period(period):
                today = datetime.today()
                if period == "today":
                    cal_from_date.entry.delete(0, tk.END)
                    cal_from_date.entry.insert(0, today.strftime('%Y-%m-%d'))
                    cal_to_date.entry.delete(0, tk.END)
                    cal_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
                elif period == "week":
                    start = today - timedelta(days=today.weekday())
                    cal_from_date.entry.delete(0, tk.END)
                    cal_from_date.entry.insert(0, start.strftime('%Y-%m-%d'))
                    cal_to_date.entry.delete(0, tk.END)
                    cal_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
                elif period == "month":
                    start = today.replace(day=1)
                    cal_from_date.entry.delete(0, tk.END)
                    cal_from_date.entry.insert(0, start.strftime('%Y-%m-%d'))
                    cal_to_date.entry.delete(0, tk.END)
                    cal_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
                elif period == "all":
                    cal_from_date.entry.delete(0, tk.END)
                    cal_from_date.entry.insert(0, today.replace(year=today.year - 10).strftime('%Y-%m-%d'))
                    cal_to_date.entry.delete(0, tk.END)
                    cal_to_date.entry.insert(0, today.replace(year=today.year + 10).strftime('%Y-%m-%d'))
                load_summary_data()
                
            tb.Button(date_frame, text="Today", bootstyle="outline-primary", cursor="hand2", command=lambda: set_summary_period("today")).pack(side=LEFT, padx=(20, 5))
            tb.Button(date_frame, text="This Week", bootstyle="outline-primary", cursor="hand2", command=lambda: set_summary_period("week")).pack(side=LEFT, padx=5)
            tb.Button(date_frame, text="This Month", bootstyle="outline-primary", cursor="hand2", command=lambda: set_summary_period("month")).pack(side=LEFT, padx=5)
            tb.Button(date_frame, text="All Time", bootstyle="outline-primary", cursor="hand2", command=lambda: set_summary_period("all")).pack(side=LEFT, padx=5)
            
            # Row 2: Category Filter Row (vertically stacked so it is 100% visible on standard display sizes)
            filter_row = tb.Frame(win, padding=5)
            filter_row.pack(side=TOP, fill=X, padx=25, pady=(5, 10))
            
            tb.Label(filter_row, text="Filter Category Column:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(5, 10))
            
            filter_frame = tb.Frame(filter_row)
            filter_frame.pack(side=LEFT)
            
            perf_filter_var = tk.StringVar(value="All")
            
            def on_filter_changed():
                load_summary_data()
                
            tb.Radiobutton(filter_frame, text=self._tr("Show All"), variable=perf_filter_var, value="All", bootstyle="info-toolbutton", command=on_filter_changed).pack(side=LEFT, padx=2)
            tb.Radiobutton(filter_frame, text=self._tr("Service Sales"), variable=perf_filter_var, value="Service", bootstyle="info-toolbutton", command=on_filter_changed).pack(side=LEFT, padx=2)
            tb.Radiobutton(filter_frame, text=self._tr("Product Sales"), variable=perf_filter_var, value="Product", bootstyle="info-toolbutton", command=on_filter_changed).pack(side=LEFT, padx=2)
            tb.Radiobutton(filter_frame, text=self._tr("Tips"), variable=perf_filter_var, value="Tip", bootstyle="info-toolbutton", command=on_filter_changed).pack(side=LEFT, padx=2)
            
            columns = ("Date", "Location", "Category", "Service Sales", "Service Add-on Sales", "Product Sales", "Tip", "Hours", "Calculated Pay", "Expense Amount", "Notes")
            
            summary_frame = tb.Frame(win, padding=10, bootstyle="secondary")
            summary_frame.pack(fill=X, side=BOTTOM)
            tb.Button(summary_frame, text="Close Window", bootstyle="light", cursor="hand2", command=win.destroy).pack(side=LEFT, padx=10)
            lbl_total = tb.Label(summary_frame, text="Total Earnings: $0.00", font=("Segoe UI", 14, "bold"), bootstyle="inverse-secondary")
            lbl_total.pack(side=RIGHT, padx=10)
            
            tree_frame = tb.Frame(win)
            tree_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)
            
            scroll_y = tb.Scrollbar(tree_frame, orient=VERTICAL)
            scroll_x = tb.Scrollbar(tree_frame, orient=HORIZONTAL)
            
            tree = tb.Treeview(tree_frame, columns=columns, show="headings", bootstyle="info", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            
            scroll_y.config(command=tree.yview)
            scroll_y.pack(side=RIGHT, fill=Y)
            
            scroll_x.config(command=tree.xview)
            scroll_x.pack(side=BOTTOM, fill=X)
            
            for col in columns:
                tree.heading(col, text=self._tr(col))
                tree.column(col, anchor=CENTER, width=130)
            tree.column("Notes", width=200)
                
            tree.pack(side=LEFT, fill=BOTH, expand=True)

            def load_summary_data():
                for item in tree.get_children():
                    tree.delete(item)
                    
                from_d = cal_from_date.entry.get()
                to_d = cal_to_date.entry.get()
                filter_val = perf_filter_var.get()
                
                # Fetch payroll records
                query_payroll = '''
                    SELECT r.record_date, r.location, 'Employee Revenue' AS category, r.revenue, r.service_addon_sales, r.product_sales, r.tip, r.hours, r.calculation, r.notes, e.hour_rate, e.percentage
                    FROM payroll_records r
                    JOIN employees e ON r.employee_id = e.id
                    WHERE r.employee_id=? 
                '''
                params_payroll = [emp_id]
                if from_d:
                    query_payroll += " AND r.record_date >= ?"
                    params_payroll.append(from_d)
                if to_d:
                    query_payroll += " AND r.record_date <= ?"
                    params_payroll.append(to_d)
                
                # Fetch expenses attached to this employee (salary is an expense;
                # match plaintext and leftover encrypted category values).
                query_expenses = '''
                    SELECT expense_date, category, amount, description, is_tip, tip_given
                    FROM expenses
                    WHERE employee_id=?
                '''
                params_expenses = [emp_id]
                if from_d:
                    query_expenses += " AND expense_date >= ?"
                    params_expenses.append(from_d)
                if to_d:
                    query_expenses += " AND expense_date <= ?"
                    params_expenses.append(to_d)
                    
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                
                cursor.execute(query_payroll, params_payroll)
                payroll_rows = cursor.fetchall()
                
                cursor.execute(query_expenses, params_expenses)
                expense_rows = cursor.fetchall()
                
                conn.close()
                
                # Combine rows
                combined = []
                use_tiered, service_perc, product_perc, hour_rate = self.get_employee_payout_details(emp_id, from_d, to_d)
                service_perc = to_float(service_perc, 0.0)
                product_perc = to_float(product_perc, 0.0)
                hour_rate = to_float(hour_rate, 0.0)
                for pr in payroll_rows:
                    rev_val = to_float(pr[3], 0.0)
                    addon_val = to_float(pr[4], 0.0)
                    prod_val = to_float(pr[5], 0.0)
                    tip_val = to_float(pr[6], 0.0)
                    hr_rate_val = hour_rate
                    perc_val = service_perc
                    hours_val = to_float(pr[7], 0.0)
                    
                    if use_tiered:
                        calc_val = round(((rev_val + addon_val) * service_perc) + (prod_val * product_perc) + (hours_val * hour_rate), 2)
                    else:
                        calc_val = round(((rev_val + addon_val) * perc_val) + (hours_val * hr_rate_val), 2)
                    
                    if filter_val == "Service":
                        if (not rev_val or rev_val <= 0) and (not addon_val or addon_val <= 0):
                            continue
                        prod_val = 0.0
                        tip_val = 0.0
                    elif filter_val == "Tip":
                        if not tip_val or tip_val <= 0:
                            continue
                        rev_val = 0.0
                        addon_val = 0.0
                        prod_val = 0.0
                        calc_val = 0.0
                    elif filter_val == "Product":
                        if not prod_val or prod_val <= 0:
                            continue
                        rev_val = 0.0
                        addon_val = 0.0
                        tip_val = 0.0
                        calc_val = 0.0
                        
                    combined.append((pr[0], pr[1], pr[2], rev_val, addon_val, prod_val, tip_val, pr[7], calc_val, None, pr[9], 'payroll'))
                    
                for ex in expense_rows:
                    category_name = plain_label(ex[1])
                    if category_name not in SALARY_EXPENSE_CATEGORIES and category_name != "Salary Payment":
                        continue
                    is_tip_expense = plain_label(ex[4]).lower() in ("yes", "true", "1")
                    
                    if category_name == "Salary Payment" and is_tip_expense:
                        category_name = "Salary Payment (Tip)"
                        
                    # Filter configurations:
                    if filter_val == "Service":
                        if category_name == "Salary Payment (Tip)":
                            continue
                    elif filter_val == "Tip":
                        if category_name != "Salary Payment (Tip)":
                            continue
                    elif filter_val == "Product":
                        continue

                    notes = plain_label(ex[3]) if len(ex) > 3 and ex[3] else ""
                    tip_given_val = to_float(ex[5], 0.0) if len(ex) > 5 else 0.0
                    if is_tip_expense and tip_given_val:
                        tip_note = f"{self._tr('Tip given to employee:')} ${tip_given_val:,.2f}"
                        notes = f"{notes} — {tip_note}".strip(" —") if notes else tip_note
                    combined.append((ex[0], "", category_name, None, None, None, None, None, None, ex[2], notes, 'expense'))
                    
                # Sort descending by date
                combined.sort(key=lambda x: x[0], reverse=True)
                
                total_rev = 0.0
                total_addon = 0.0
                total_prod = 0.0
                total_tip = 0.0
                total_hrs = 0.0
                total_calc = 0.0
                total_exp = 0.0
                
                for row in combined:
                    date_val, loc_val, category, rev_v, addon_v, prod_v, tip_v, hrs_v, calc_v, exp_v, notes_val, tag = row
                    
                    rev_disp = ""
                    addon_disp = ""
                    prod_disp = ""
                    tip_disp = ""
                    hrs_disp = ""
                    calc_disp = ""
                    exp_disp = ""
                    
                    if rev_v is not None:
                        rev_v = to_float(rev_v, 0.0)
                        total_rev += rev_v
                        rev_disp = f"${rev_v:,.2f}"

                    if addon_v is not None:
                        addon_v = to_float(addon_v, 0.0)
                        total_addon += addon_v
                        addon_disp = f"${addon_v:,.2f}"
                        
                    if prod_v is not None:
                        prod_v = to_float(prod_v, 0.0)
                        total_prod += prod_v
                        prod_disp = f"${prod_v:,.2f}"
                        
                    if tip_v is not None:
                        tip_v = to_float(tip_v, 0.0)
                        total_tip += tip_v
                        tip_disp = f"${tip_v:,.2f}"
                        
                    if hrs_v is not None:
                        hrs_v = to_float(hrs_v, 0.0)
                        total_hrs += hrs_v
                        hrs_disp = f"{hrs_v:.1f}"
                        
                    if calc_v is not None:
                        calc_v = to_float(calc_v, 0.0)
                        total_calc += calc_v
                        calc_disp = f"⭐ ${calc_v:,.2f} ⭐"
                        
                    if exp_v is not None:
                        exp_v = to_float(exp_v, 0.0)
                        total_exp += exp_v
                        exp_disp = f"-${exp_v:,.2f}"
                        
                    fmt = (
                        date_val,
                        loc_val,
                        self._tr(category),
                        rev_disp,
                        addon_disp,
                        prod_disp,
                        tip_disp,
                        hrs_disp,
                        calc_disp,
                        exp_disp,
                        notes_val if notes_val else ""
                    )
                    
                    tree.insert('', tk.END, values=fmt, tags=(tag,))
                
                tree.tag_configure('payroll', foreground='#00bc8c')
                tree.tag_configure('expense', foreground='#e74c3c')
                
                # Insert empty divider
                tree.insert('', tk.END, values=("", "", "", "", "", "", "", "", "", "", ""))
                
                # Insert totals row
                tree.insert('', tk.END, values=(
                    "=== TOTALS ===", 
                    "", 
                    "", 
                    f"${total_rev:,.2f}", 
                    f"${total_addon:,.2f}", 
                    f"${total_prod:,.2f}", 
                    f"${total_tip:,.2f}", 
                    f"{total_hrs:.1f}", 
                    f"⭐ ${total_calc:,.2f} ⭐", 
                    f"-${total_exp:,.2f}", 
                    ""
                ), tags=('totals',))
                tree.tag_configure('totals', foreground='#f39c12', font=('Segoe UI', 11, 'bold'))
                
                if filter_val == "Tip":
                    net_val = total_tip - total_exp
                elif filter_val == "Product":
                    net_val = total_prod - total_exp
                elif filter_val == "Service":
                    net_val = (total_rev + total_addon) - total_exp
                else: # All
                    net_val = (total_rev + total_addon + total_prod + total_tip) - total_exp
                    
                tiered_rate_str = f" | Rates: S={service_perc*100:.0f}%, P={product_perc*100:.0f}%" if use_tiered else ""
                lbl_total.config(
                    text=f"Service: ${total_rev:,.2f}  |  Add-on: ${total_addon:,.2f}  |  Product: ${total_prod:,.2f}  |  Tip: ${total_tip:,.2f}  |  Calculated: ${total_calc:,.2f}  |  Expenses: ${total_exp:,.2f}  |  Net: ${net_val:,.2f}{tiered_rate_str}"
                )

            last_state = {"from": "", "to": ""}
            def poll_summary():
                if not win.winfo_exists():
                    return
                try:
                    current_from = cal_from_date.entry.get()
                    current_to = cal_to_date.entry.get()
                    if current_from != last_state["from"] or current_to != last_state["to"]:
                        last_state["from"] = current_from
                        last_state["to"] = current_to
                        load_summary_data()
                except Exception as e:
                    print(f"Error in poll_summary: {e}")
                finally:
                    win.after(500, poll_summary)

            set_summary_period("all")
            poll_summary()

        def open_employee_edit_by_id(self, emp_id):
            if not emp_id:
                return
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT first_name, last_name, phone, email, hour_rate, percentage, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path, use_tiered_payout FROM employees WHERE id=?", (emp_id,))
                emp = cursor.fetchone()
                conn.close()
                if emp:
                    f_name, l_name, phone, email, emp_rate, emp_perc, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path, use_tiered = emp
                    f_name = decrypt_val(f_name) or ""
                    l_name = decrypt_val(l_name) or ""
                    phone = decrypt_val(phone) or ""
                    email = decrypt_val(email) or ""
                    ssn = decrypt_val(ssn) or ""
                    address = decrypt_val(address) or ""
                    start_date = decrypt_val(start_date) or ""
                    end_date = decrypt_val(end_date) or ""

                    rate_num = to_float(decrypt_val(emp_rate), 0.0)
                    emp_rate = f"{rate_num:g}" if rate_num > 0 else ""

                    perc_num = to_float(decrypt_val(emp_perc), 0.0)
                    if perc_num > 1.0:
                        emp_perc = f"{perc_num:g}"
                    elif perc_num > 0.0:
                        emp_perc = f"{perc_num * 100:g}"
                    else:
                        emp_perc = ""

                    use_tiered = use_tiered if use_tiered else 0
                    self.open_employee_dialog(emp_id, f_name, l_name, phone, email, emp_rate, emp_perc, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path, use_tiered)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open employee editor: {e}", parent=self)

        def edit_selected_employee(self):
            selected = self.tree_names.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select an employee to edit.")
                return
            item = self.tree_names.item(selected[0])
            emp_id = item['values'][0]
            self.open_employee_edit_by_id(emp_id)

        def open_employee_dialog(self, emp_id=None, f_name="", l_name="", phone="", email="", current_rate="", current_perc="", ssn="", address="", start_date="", end_date="", cv_path="", id_photo_path="", personal_photo_path="", use_tiered_payout=0):
            dialog = tb.Toplevel(self)
            dialog.title(self._tr("Edit Employee") if emp_id else self._tr("Add Employee"))
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()
            try:
                dialog.update_idletasks()
                screen_w = dialog.winfo_screenwidth()
                screen_h = dialog.winfo_screenheight()
                dlg_w = min(600, max(480, screen_w - 40))
                dlg_h = min(720, max(520, screen_h - 90))
                x = max(0, (screen_w - dlg_w) // 2)
                y = max(10, min(25, (screen_h - dlg_h) // 4))
                dialog.geometry(f"{dlg_w}x{dlg_h}+{x}+{y}")
            except Exception:
                dialog.geometry("600x680")
            
            canvas = tk.Canvas(dialog, highlightthickness=0)
            scrollbar = tb.Scrollbar(dialog, orient="vertical", command=canvas.yview)
            scrollable_frame = tb.Frame(canvas, padding=(10, 5))

            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(
                    scrollregion=canvas.bbox("all")
                )
            )

            canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            def _sync_emp_width(event):
                canvas.itemconfigure(canvas_window, width=event.width)

            canvas.bind("<Configure>", _sync_emp_width)

            def _on_emp_wheel(event):
                try:
                    delta = int(-1 * (event.delta / 120)) if getattr(event, "delta", 0) else (1 if getattr(event, "num", 0) == 5 else -1)
                    canvas.yview_scroll(delta, "units")
                except Exception:
                    pass

            def _bind_emp_wheel(w):
                try:
                    w.bind("<MouseWheel>", _on_emp_wheel, add="+")
                    w.bind("<Button-4>", _on_emp_wheel, add="+")
                    w.bind("<Button-5>", _on_emp_wheel, add="+")
                except Exception:
                    pass
                for child in w.winfo_children():
                    _bind_emp_wheel(child)

            dialog.after(100, lambda: _bind_emp_wheel(scrollable_frame))
            dialog.after(100, lambda: _bind_emp_wheel(canvas))

            canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            scrollbar.pack(side="right", fill="y")
            
            pad = {'padx': 15, 'pady': 8, 'sticky': E}
            ent_pad = {'padx': 15, 'pady': 8, 'sticky': W}
            
            tb.Label(scrollable_frame, text=self._tr("First Name:*"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0, **pad)
            entry_fname = tb.Entry(scrollable_frame, width=35)
            entry_fname.grid(row=0, column=1, **ent_pad)
            entry_fname.insert(0, f_name or "")
            
            tb.Label(scrollable_frame, text=self._tr("Last Name:"), font=("Segoe UI", 10, "bold")).grid(row=1, column=0, **pad)
            entry_lname = tb.Entry(scrollable_frame, width=35)
            entry_lname.grid(row=1, column=1, **ent_pad)
            entry_lname.insert(0, l_name or "")
            
            tb.Label(scrollable_frame, text=self._tr("Phone:"), font=("Segoe UI", 10, "bold")).grid(row=2, column=0, **pad)
            entry_phone = tb.Entry(scrollable_frame, width=35)
            entry_phone.grid(row=2, column=1, **ent_pad)
            entry_phone.insert(0, phone or "")
            
            tb.Label(scrollable_frame, text=self._tr("Email:"), font=("Segoe UI", 10, "bold")).grid(row=3, column=0, **pad)
            entry_email = tb.Entry(scrollable_frame, width=35)
            entry_email.grid(row=3, column=1, **ent_pad)
            entry_email.insert(0, email or "")

            tb.Label(scrollable_frame, text=self._tr("SSN:"), font=("Segoe UI", 10, "bold")).grid(row=4, column=0, **pad)
            entry_ssn = tb.Entry(scrollable_frame, width=35)
            entry_ssn.grid(row=4, column=1, **ent_pad)
            entry_ssn.insert(0, ssn or "")

            tb.Label(scrollable_frame, text=self._tr("Address:"), font=("Segoe UI", 10, "bold")).grid(row=5, column=0, **pad)
            entry_address = tb.Entry(scrollable_frame, width=35)
            entry_address.grid(row=5, column=1, **ent_pad)
            entry_address.insert(0, address or "")

            tb.Label(scrollable_frame, text=self._tr("Start Date:"), font=("Segoe UI", 10, "bold")).grid(row=6, column=0, **pad)
            entry_start = tb.DateEntry(scrollable_frame, bootstyle="primary", dateformat='%Y-%m-%d')
            entry_start.entry.delete(0, tk.END)
            entry_start.entry.insert(0, start_date or "")
            entry_start.grid(row=6, column=1, **ent_pad)

            tb.Label(scrollable_frame, text=self._tr("End Date:"), font=("Segoe UI", 10, "bold")).grid(row=7, column=0, **pad)
            entry_end = tb.DateEntry(scrollable_frame, bootstyle="primary", dateformat='%Y-%m-%d')
            entry_end.entry.delete(0, tk.END)
            entry_end.entry.insert(0, end_date or "")
            entry_end.grid(row=7, column=1, **ent_pad)
            
            clean_rate_str = ""
            if current_rate:
                c_rate = to_float(current_rate, 0.0)
                clean_rate_str = f"{c_rate:g}" if c_rate > 0 else ""

            clean_perc_str = ""
            if current_perc:
                c_perc = to_float(current_perc, 0.0)
                if c_perc > 1.0:
                    clean_perc_str = f"{c_perc:g}"
                elif c_perc > 0.0:
                    clean_perc_str = f"{c_perc * 100:g}"

            tb.Label(scrollable_frame, text=self._tr("Hour Rate ($):"), font=("Segoe UI", 10, "bold")).grid(row=8, column=0, **pad)
            entry_rate = tb.Entry(scrollable_frame, width=35)
            entry_rate.grid(row=8, column=1, **ent_pad)
            entry_rate.insert(0, clean_rate_str)
            
            tb.Label(scrollable_frame, text=self._tr("Percentage (1-100):"), font=("Segoe UI", 10, "bold")).grid(row=9, column=0, **pad)
            entry_perc = tb.Entry(scrollable_frame, width=35)
            entry_perc.grid(row=9, column=1, **ent_pad)
            entry_perc.insert(0, clean_perc_str)

            use_tiered_var = tk.IntVar(value=use_tiered_payout)
            tb.Label(scrollable_frame, text=self._tr("Use Tiered Payout:"), font=("Segoe UI", 10, "bold")).grid(row=10, column=0, **pad)
            chk_tiered = tb.Checkbutton(scrollable_frame, variable=use_tiered_var, text=self._tr("dynamic commission tiers + product sales split"))
            chk_tiered.grid(row=10, column=1, **ent_pad)

            def browse_file(entry):
                filepath = filedialog.askopenfilename()
                if filepath:
                    entry.delete(0, tk.END)
                    entry.insert(0, filepath)

            tb.Label(scrollable_frame, text=self._tr("CV:"), font=("Segoe UI", 10, "bold")).grid(row=11, column=0, **pad)
            cv_frame = tb.Frame(scrollable_frame)
            cv_frame.grid(row=11, column=1, **ent_pad)
            entry_cv = tb.Entry(cv_frame, width=25)
            entry_cv.pack(side=LEFT, padx=(0, 5))
            entry_cv.insert(0, cv_path or "")
            tb.Button(cv_frame, text=self._tr("Browse"), command=lambda: browse_file(entry_cv)).pack(side=LEFT)

            tb.Label(scrollable_frame, text=self._tr("ID Photo:"), font=("Segoe UI", 10, "bold")).grid(row=12, column=0, **pad)
            id_frame = tb.Frame(scrollable_frame)
            id_frame.grid(row=12, column=1, **ent_pad)
            entry_id = tb.Entry(id_frame, width=25)
            entry_id.pack(side=LEFT, padx=(0, 5))
            entry_id.insert(0, id_photo_path or "")
            tb.Button(id_frame, text=self._tr("Browse"), command=lambda: browse_file(entry_id)).pack(side=LEFT)

            tb.Label(scrollable_frame, text=self._tr("Personal Photo:"), font=("Segoe UI", 10, "bold")).grid(row=13, column=0, **pad)
            pp_frame = tb.Frame(scrollable_frame)
            pp_frame.grid(row=13, column=1, **ent_pad)
            entry_pp = tb.Entry(pp_frame, width=25)
            entry_pp.pack(side=LEFT, padx=(0, 5))
            entry_pp.insert(0, personal_photo_path or "")
            tb.Button(pp_frame, text=self._tr("Browse"), command=lambda: browse_file(entry_pp)).pack(side=LEFT)
            
            def save(event=None):
                fn = entry_fname.get().strip()
                ln = entry_lname.get().strip()
                ph = entry_phone.get().strip()
                em = entry_email.get().strip()
                rate = entry_rate.get().strip() or "0"
                perc = entry_perc.get().strip() or "0"
                s_ssn = entry_ssn.get().strip()
                s_address = entry_address.get().strip()
                s_start = entry_start.entry.get().strip()
                s_end = entry_end.entry.get().strip()
                s_cv = entry_cv.get().strip()
                s_id = entry_id.get().strip()
                s_pp = entry_pp.get().strip()
                
                if not fn:
                    messagebox.showerror("Error", self._tr("First Name is required."))
                    return
                try:
                    rate_val = to_float(rate, -1.0) if rate else 0.0
                    perc_val = to_float(perc, -1.0) if perc else 0.0
                except Exception:
                    messagebox.showerror("Error", "Hour rate and percentage must be numbers.", parent=dialog)
                    return
                    
                if rate_val < 0 or perc_val < 0:
                    messagebox.showerror("Error", "Hour rate and percentage cannot be negative.", parent=dialog)
                    return
                    
                if fn.lower() == "shop":
                    pass
                elif use_tiered_var.get() == 1:
                    pass
                else:
                    if rate_val > 0 and perc_val > 0:
                        messagebox.showerror("Error", "An employee can only have an Hour Rate OR a Percentage, not both.", parent=dialog)
                        return
                    if rate_val <= 0 and perc_val <= 0:
                        messagebox.showerror("Validation Error", "You must configure either an Hour Rate, a Percentage greater than 0, or enable 'Use Tiered Commission'.", parent=dialog)
                        return
                
                if perc_val > 1.0:
                    db_perc = perc_val / 100.0
                elif perc_val > 0.0:
                    db_perc = perc_val
                else:
                    db_perc = 0.0
                db_rate = rate_val if rate_val > 0 else 0.0
                combined_name = f"{fn} {ln}".strip()
                
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    if emp_id:
                        cursor.execute("UPDATE employees SET name=?, first_name=?, last_name=?, phone=?, email=?, hour_rate=?, percentage=?, ssn=?, address=?, start_date=?, end_date=?, cv_path=?, id_photo_path=?, personal_photo_path=?, use_tiered_payout=? WHERE id=?", 
                                    (combined_name, fn, ln, ph, em, db_rate, db_perc, s_ssn, s_address, s_start, s_end, s_cv, s_id, s_pp, use_tiered_var.get(), emp_id))
                        emp_id_to_use = emp_id
                    else:
                        cursor.execute("INSERT INTO employees (name, first_name, last_name, phone, email, hour_rate, percentage, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path, use_tiered_payout) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                                    (combined_name, fn, ln, ph, em, db_rate, db_perc, s_ssn, s_address, s_start, s_end, s_cv, s_id, s_pp, use_tiered_var.get()))
                        emp_id_to_use = cursor.lastrowid

                    new_folder = os.path.join(EMPLOYEE_FOLDERS_DIR, f"{fn}_{ln}_{emp_id_to_use}".replace(" ", "_"))
                    os.makedirs(new_folder, exist_ok=True)

                    def copy_to_folder(path):
                        if path and os.path.exists(path) and not path.startswith(new_folder):
                            try:
                                return optimize_and_save_file(path, new_folder)
                            except Exception:
                                return path
                        return path

                    final_cv = copy_to_folder(s_cv)
                    final_id = copy_to_folder(s_id)
                    final_pp = copy_to_folder(s_pp)

                    cursor.execute("UPDATE employees SET cv_path=?, id_photo_path=?, personal_photo_path=? WHERE id=?", (final_cv, final_id, final_pp, emp_id_to_use))

                    commit_and_save(conn)
                except sqlite3.IntegrityError:
                    if 'conn' in locals():
                        conn.close()
                    messagebox.showerror("Error", "An employee with this exact name already exists.")
                    return
                except Exception as e:
                    if 'conn' in locals():
                        conn.close()
                    messagebox.showerror("Database Error", f"Failed to save employee: {e}")
                    return
                
                conn.close()
                dialog.destroy()
                if self._widget_alive(getattr(self, "folders_win", None)):
                    self.load_folders()
                self.load_employees()
                self.load_calendar_data()
                    
            btnf = tb.Frame(scrollable_frame)
            btnf.grid(row=14, column=0, columnspan=2, pady=20)
            tb.Button(btnf, text=self._tr("Save Employee"), bootstyle="success", cursor="hand2", command=save).pack(side=LEFT, ipadx=25, ipady=5)
            tb.Button(btnf, text=self._tr("Cancel"), bootstyle="secondary", cursor="hand2", command=dialog.destroy).pack(side=LEFT, padx=12, ipadx=18, ipady=5)
            self._bind_dialog_save_keys(dialog, save)
            dialog.bind("<Escape>", lambda e: dialog.destroy())

        def show_employee_action_popup(self, event=None):
            selected = self.tree_names.selection()
            if not selected:
                return
            item = self.tree_names.item(selected[0])
            emp_id = item['values'][0]
            first_name = item['values'][1]
            last_name = item['values'][2]
            emp_name = f"{first_name} {last_name}".strip()

            popup = tb.Toplevel(self)
            popup.title(self._tr("Select Action"))
            popup.geometry("450x220")
            popup.transient(self)
            popup.grab_set()
            popup.focus_set()
            
            popup.update_idletasks()
            w = 450
            h = 220
            x = self.winfo_x() + (self.winfo_width() - w) // 2
            y = self.winfo_y() + (self.winfo_height() - h) // 2
            popup.geometry(f"{w}x{h}+{x}+{y}")
            
            tb.Label(popup, text=f"Choose action for: {emp_name}", font=("Segoe UI", 12, "bold"), anchor=CENTER).pack(pady=20)
            
            btn_frame = tb.Frame(popup)
            btn_frame.pack(pady=10)
            
            def handle_edit():
                popup.destroy()
                self.edit_selected_employee()
                
            def handle_performance():
                popup.destroy()
                self.view_employee_summary()
                
            tb.Button(btn_frame, text=self._tr("✏️ Edit Employee"), bootstyle="warning", width=18, command=handle_edit).pack(side=LEFT, padx=15)
            tb.Button(btn_frame, text=self._tr("📊 Performance Report"), bootstyle="info", width=18, command=handle_performance).pack(side=LEFT, padx=15)
            
            tb.Button(popup, text=self._tr("Close Window"), bootstyle="secondary outline", width=15, command=popup.destroy).pack(pady=15)

        def on_tab_changed(self, event=None):
            selected_tab = self.notebook.select()
            if not selected_tab:
                return
            tab_text = self.notebook.tab(selected_tab, "text")
            if tab_text == self._tr("💈 Shop Earnings"):
                self.load_calendar_data()
            elif tab_text == self._tr("💇‍♂️ Barbers / Stylists"):
                self.load_employees()
            elif tab_text == self._tr("✍️ Manual Ledger"):
                self.refresh_employee_dropdown()
            elif tab_text == self._tr("📊 P&L / Financials"):
                self.load_financials_data(quiet=True)
            elif tab_text == self._tr("📅 Cash Calendar"):
                if hasattr(self, 'load_cash_calendar_data'):
                    self.load_cash_calendar_data(quiet=True)

        def setup_data_entry_tab(self):
            container = tb.Frame(self.tab_data_entry)
            container.pack(expand=True)
            
            card = tb.Frame(container, padding=40)
            card.pack(expand=True)
            
            tb.Label(card, text=self._tr("✨ Add New Payroll Record"), font=("Segoe UI", 22, "bold"), bootstyle="primary").grid(row=0, column=0, columnspan=2, pady=(0, 30))
            
            self.entry_vars = {}
            pad_opt = {'padx': 20, 'pady': 12, 'sticky': E}
            ent_pad = {'padx': 20, 'pady': 12, 'sticky': W}
            
            tb.Label(card, text=self._tr("Date:"), font=("Segoe UI", 12, "bold")).grid(row=1, column=0, **pad_opt)
            self.entry_vars['Date'] = tb.DateEntry(card, bootstyle="primary", dateformat='%Y-%m-%d')
            self.entry_vars['Date'].grid(row=1, column=1, **ent_pad)

            # Cycle dropdown (row 2)
            today_str = datetime.today().strftime('%Y-%m-%d')
            init_ck = cycle_for_date(today_str)
            cycle_choices, cycle_key_map, active_lbl = get_formatted_cycle_choices(init_ck, start_from_june_2026=True)

            tb.Label(card, text=self._tr("Cycle:"), font=("Segoe UI", 12, "bold"), bootstyle="warning").grid(row=2, column=0, **pad_opt)
            cbo_cycle = tb.Combobox(card, width=38, state="readonly", values=cycle_choices, font=("Segoe UI", 11), bootstyle="warning")
            cbo_cycle.set(active_lbl)
            cbo_cycle.grid(row=2, column=1, **ent_pad)
            self.entry_vars['Cycle'] = cbo_cycle

            _cycle_manual_override = [False]
            def _on_cycle_manual_change(event=None):
                _cycle_manual_override[0] = True
            cbo_cycle.bind("<<ComboboxSelected>>", _on_cycle_manual_change)

            def _sync_cycle_from_date():
                if _cycle_manual_override[0]:
                    return
                try:
                    d_val = self.entry_vars['Date'].entry.get()
                    if d_val:
                        ck = cycle_for_date(d_val)
                        if ck:
                            new_choices, new_map, new_lbl = get_formatted_cycle_choices(ck, start_from_june_2026=True)
                            cbo_cycle["values"] = new_choices
                            cycle_key_map.update(new_map)
                            cbo_cycle.set(new_lbl)
                except Exception:
                    pass

            def _poll_ledger_cycle_sync():
                self._poll_ledger_cycle_after_id = None
                if not self._widget_alive(getattr(self, "tab_data_entry", None)):
                    return
                _sync_cycle_from_date()
                try:
                    self._poll_ledger_cycle_after_id = self.after(400, _poll_ledger_cycle_sync)
                except Exception:
                    pass
            try:
                self._poll_ledger_cycle_after_id = self.after(400, _poll_ledger_cycle_sync)
            except Exception:
                pass

            tb.Label(card, text=self._tr("Employee Name:"), font=("Segoe UI", 12, "bold")).grid(row=3, column=0, **pad_opt)
            self.cbo_employee = tb.Combobox(card, width=38, state="readonly", font=("Segoe UI", 11))
            self.cbo_employee.grid(row=3, column=1, **ent_pad)
            self.entry_vars['Employee'] = self.cbo_employee
            self.cbo_employee.bind("<<ComboboxSelected>>", self.on_employee_selected_data_entry)
            
            fields = ["Revenue", "Service Add-on Sales", "Hours", "Hour Rate", "Percentage", "Notes"]
            for i, field in enumerate(fields):
                if field == "Revenue":
                    lbl_text = "Revenue (Service Sales):"
                elif field == "Hour Rate":
                    lbl_text = "Hour Rate ($):"
                elif field == "Percentage":
                    lbl_text = "Percentage (%):"
                else:
                    lbl_text = field + ":"
                
                tb.Label(card, text=self._tr(lbl_text), font=("Segoe UI", 12, "bold")).grid(row=i+4, column=0, **pad_opt)
                ent = tb.Entry(card, width=40, font=("Segoe UI", 11))
                ent.grid(row=i+4, column=1, **ent_pad)
                self.entry_vars[field] = ent
                
            tb.Label(card, text=self._tr("Written Up:"), font=("Segoe UI", 12, "bold")).grid(row=len(fields)+4, column=0, **pad_opt)
            self.cbo_written_up = tb.Combobox(card, width=38, state="readonly", values=["", "Yes", "No"], font=("Segoe UI", 11))
            self.cbo_written_up.set("")
            self.cbo_written_up.grid(row=len(fields)+4, column=1, **ent_pad)
            self.entry_vars['Written Up'] = self.cbo_written_up

            # Conditional Write Up Reason field
            self.lbl_reason = tb.Label(card, text=self._tr("Write Up Reason:"), font=("Segoe UI", 12, "bold"))
            self.ent_reason = tb.Entry(card, width=40, font=("Segoe UI", 11))
            self.entry_vars['Write Up Reason'] = self.ent_reason

            # Help Definitions
            help_msgs = {
                "Date": ("Enter the date of the work shift (YYYY-MM-DD).", "أدخل تاريخ وردية العمل (YYYY-MM-DD)."),
                "Cycle": ("Pay cycle that owns this entry.", "دورة الدفع الخاصة بهذا السجل."),
                "Employee": ("Select the employee for this payroll record.", "اختر الموظف الخاص بسجل الرواتب هذا."),
                "Revenue": ("Main service sales generated by the employee.", "مبيعات الخدمات الرئيسية التي حققها الموظف."),
                "Service Add-on Sales": ("Additional sales from add-on services (split at the same rate).", "المبيعات الإضافية من الخدمات التكميلية (تُقسم بنفس النسبة)."),
                "Hours": ("Total hours worked during this shift (used for hourly pay calculation).", "إجمالي الساعات التي تم عملها خلال الوردية (تُستخدم لحساب الدفع بالساعة)."),
                "Hour Rate": ("Hourly pay rate for this employee. Defaults to employee setting.", "أجر الساعة لهذا الموظف. القيمة الافتراضية مأخوذة من إعدادات الموظف."),
                "Percentage": ("Commission percentage for services. Disabled (shows Auto) if tiered payout is enabled for the employee.", "نسبة عمولة الخدمات. تظهر معطلة (تلقائي) إذا كان نظام الدفع المتدرج مفعلاً للموظف."),
                "Notes": ("Optional shift notes, details or references.", "ملاحظات اختيارية للوردية أو تفاصيل إضافية."),
                "Written Up": ("Select Yes if this shift includes a formal write-up/citation for the employee.", "اختر نعم إذا كانت هذه الوردية تتضمن إنذاراً أو مخالفة رسمية للموظف."),
                "Write Up Reason": ("Describe the reason for the write-up (required if Written Up is Yes).", "صف سبب الإنذار/المخالفة (مطلوب إذا تم تحديد نعم فوق).")
            }

            def add_help_btn(row, col, field_key):
                msg_en, msg_ar = help_msgs.get(field_key, ("", ""))
                btn = tb.Button(card, text="❓", bootstyle="link", cursor="hand2")
                btn.grid(row=row, column=col, padx=(2, 5), pady=5, sticky=W)
                msg = msg_ar if getattr(self, 'lang', 'en') == 'ar' else msg_en
                ToolTip(btn, text=msg)

            add_help_btn(1, 2, "Date")
            add_help_btn(2, 2, "Cycle")
            add_help_btn(3, 2, "Employee")
            for i, field in enumerate(fields):
                add_help_btn(i+4, 2, field)
            add_help_btn(len(fields)+4, 2, "Written Up")

            self.help_reason_btn = tb.Button(card, text="❓", bootstyle="link", cursor="hand2")
            msg_en, msg_ar = help_msgs["Write Up Reason"]
            msg_reason = msg_ar if getattr(self, 'lang', 'en') == 'ar' else msg_en
            ToolTip(self.help_reason_btn, text=msg_reason)
            
            def handle_written_up_toggle(event=None):
                if self.cbo_written_up.get() == "Yes":
                    self.lbl_reason.grid(row=len(fields)+5, column=0, **pad_opt)
                    self.ent_reason.grid(row=len(fields)+5, column=1, **ent_pad)
                    self.help_reason_btn.grid(row=len(fields)+5, column=2, padx=(2, 5), pady=5, sticky=W)
                else:
                    self.lbl_reason.grid_remove()
                    self.ent_reason.grid_remove()
                    self.help_reason_btn.grid_remove()

            self.cbo_written_up.bind("<<ComboboxSelected>>", handle_written_up_toggle)
            handle_written_up_toggle()

            def clear_payroll_form():
                for field in ["Revenue", "Service Add-on Sales", "Hours", "Hour Rate", "Percentage", "Notes", "Write Up Reason"]:
                    if field in self.entry_vars:
                        if field == "Percentage":
                            self.entry_vars[field].config(state="normal")
                        self.entry_vars[field].delete(0, tk.END)
                self.entry_vars["Written Up"].set("")
                if hasattr(self, "lbl_reason"):
                    self.lbl_reason.grid_remove()
                if hasattr(self, "ent_reason"):
                    self.ent_reason.grid_remove()
                if hasattr(self, "help_reason_btn"):
                    self.help_reason_btn.grid_remove()

            btnf = tb.Frame(card)
            btnf.grid(row=len(fields)+6, column=0, columnspan=2, pady=40)
            tb.Button(btnf, text=self._tr("Calculate & Save to Database"), bootstyle="success", cursor="hand2", command=self.save_payroll_record).pack(side=LEFT, ipadx=40, ipady=12)
            tb.Button(btnf, text=self._tr("Cancel"), bootstyle="secondary", cursor="hand2", command=clear_payroll_form).pack(side=LEFT, padx=12, ipadx=24, ipady=12)
            
            self.entry_vars['Date'].entry.bind("<Return>", lambda e: self.save_payroll_record())
            self.cbo_employee.bind("<Return>", lambda e: self.save_payroll_record())
            for field in fields:
                self.entry_vars[field].bind("<Return>", lambda e: self.save_payroll_record())
            self.cbo_written_up.bind("<Return>", lambda e: self.save_payroll_record())
            self.ent_reason.bind("<Return>", lambda e: self.save_payroll_record())
            
            self.refresh_employee_dropdown()

        def on_employee_selected_data_entry(self, event=None):
            emp_name = self.cbo_employee.get()
            if not emp_name:
                return
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT hour_rate, percentage, use_tiered_payout FROM employees WHERE name=?", (emp_name,))
            row = cursor.fetchone()
            conn.close()
            if row:
                hr_rate, perc, use_tiered = row
                hr_rate = to_float(hr_rate, 0.0)
                perc = to_float(perc, 0.0)
                
                # Update Hour Rate field
                self.entry_vars['Hour Rate'].delete(0, tk.END)
                self.entry_vars['Hour Rate'].insert(0, f"{hr_rate:.2f}")
                
                # Update Percentage field
                self.entry_vars['Percentage'].config(state="normal")
                self.entry_vars['Percentage'].delete(0, tk.END)
                if use_tiered:
                    self.entry_vars['Percentage'].insert(0, "Auto")
                    self.entry_vars['Percentage'].config(state="disabled")
                else:
                    self.entry_vars['Percentage'].insert(0, f"{perc * 100:.1f}")

        def refresh_employee_dropdown(self):
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM employees")
            names = []
            for row in cursor.fetchall() or []:
                raw = row[0] if row else ""
                plain = decrypt_val(raw) if raw is not None else ""
                plain = str(plain or "").strip()
                if plain:
                    names.append(plain)
            conn.close()

            cbo = getattr(self, "cbo_employee", None)
            if self._widget_alive(cbo):
                cbo["values"] = names

            cal_f = getattr(self, "cal_name_filter", None)
            if self._widget_alive(cal_f):
                current_val = cal_f.get()
                all_text = self._tr("All")
                cal_f["values"] = [all_text] + names
                if current_val not in [all_text] + names:
                    cal_f.set(all_text)

            fin_f = getattr(self, "fin_emp_filter", None)
            if self._widget_alive(fin_f):
                current_val = fin_f.get()
                all_text = self._tr("All Employees")
                fin_f["values"] = [all_text] + names
                if current_val not in [all_text] + names:
                    fin_f.set(all_text)

            if hasattr(self, "rebuild_fin_exclude_menu"):
                try:
                    self.rebuild_fin_exclude_menu(names)
                except Exception:
                    pass

        def open_folders_window(self):
            if self._widget_alive(getattr(self, "folders_win", None)):
                self.folders_win.focus()
                return
                
            self.folders_win = tb.Toplevel(self)
            self.folders_win.title(self._tr("📁 Employee Folders"))
            try:
                self.update_idletasks()
                w = self.winfo_width()
                h = self.winfo_height()
                x = self.winfo_x()
                y = self.winfo_y()
                self.folders_win.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                self.folders_win.geometry("1100x600")
            self.folders_win.grab_set()
            self.folders_win.focus_force()
            self._present_window(self.folders_win)
            
            header_frame = tb.Frame(self.folders_win)
            header_frame.pack(fill=X, padx=25, pady=(20, 20))
            
            tb.Label(header_frame, text=self._tr("📁 Employee Folders"), font=("Segoe UI", 22, "bold"), bootstyle="primary").pack(side=LEFT)
            tb.Button(header_frame, text=self._tr("Refresh"), bootstyle="secondary outline", cursor="hand2", command=self.load_folders).pack(side=RIGHT, padx=10)
            
            columns = tuple(self._tr(c) for c in ("ID", "First Name", "Last Name", "Phone", "Files Count"))
            
            summary_frame = tb.Frame(self.folders_win, padding=10, bootstyle="secondary")
            summary_frame.pack(fill=X, side=BOTTOM, padx=25, pady=(0, 20))
            tb.Button(summary_frame, text=self._tr("Close Window"), bootstyle="light", cursor="hand2", command=self.folders_win.destroy).pack(side=LEFT)

            folders_tree_holder = tb.Frame(self.folders_win)
            folders_tree_holder.pack(fill=BOTH, expand=True, padx=25, pady=20)
            self.tree_folders = tb.Treeview(folders_tree_holder, columns=columns, show="headings", bootstyle="success")
            for col in columns:
                self.tree_folders.heading(col, text=col)
                self.tree_folders.column(col, anchor=CENTER, width=130, minwidth=90)

            self.tree_folders.bind("<Double-1>", lambda e: self.open_folder_dialog())
            self._attach_tree_scrollbars(folders_tree_holder, self.tree_folders)
            
            os.makedirs(EMPLOYEE_FOLDERS_DIR, exist_ok=True)
            self.load_folders()

        def load_folders(self):
            for item in self.tree_folders.get_children():
                self.tree_folders.delete(item)
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT id, first_name, last_name, phone FROM employees")
                rows = cursor.fetchall()
            except sqlite3.OperationalError:
                try:
                    cursor.execute("SELECT id, name, '', '' FROM employees")
                    rows = cursor.fetchall()
                except Exception:
                    rows = []
            
            for row in rows:
                emp_id = row[0]
                first = row[1] or ""
                last = row[2] or ""
                phone = row[3] or ""
                folder = os.path.join(EMPLOYEE_FOLDERS_DIR, f"{first}_{last}_{emp_id}".replace(" ", "_"))
                count = 0
                if os.path.exists(folder):
                    try:
                        count = len([name for name in os.listdir(folder) if os.path.isfile(os.path.join(folder, name))])
                    except Exception:
                        count = 0
                self.tree_folders.insert('', tk.END, values=(emp_id, first, last, phone, count))
            conn.close()

        def open_folder_dialog(self):
            selected = self.tree_folders.selection()
            if not selected:
                return
            item = self.tree_folders.item(selected[0])
            emp_id = item['values'][0]
            
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT first_name, last_name, phone, email, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path FROM employees WHERE id=?", (emp_id,))
            emp = cursor.fetchone()
            conn.close()
            
            if not emp: return
            first, last, phone, email, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path = emp
            first = first or ""
            last = last or ""
            phone = phone or ""
            email = email or ""
            ssn = ssn or ""
            address = address or ""
            start_date = start_date or ""
            end_date = end_date or ""
            cv_name = os.path.basename(cv_path) if cv_path else ""
            id_name = os.path.basename(id_photo_path) if id_photo_path else ""
            pp_name = os.path.basename(personal_photo_path) if personal_photo_path else ""
            
            folder = os.path.join(EMPLOYEE_FOLDERS_DIR, f"{first}_{last}_{emp_id}".replace(" ", "_"))
            os.makedirs(folder, exist_ok=True)
                
            win = tb.Toplevel(self)
            win.title(self._tr("Employee details and documents"))
            win.geometry("750x650")
            win.transient(self.folders_win if self._widget_alive(getattr(self, "folders_win", None)) else self)
            win.grab_set()
            win.focus_set()
            
            tb.Label(win, text=f"{self._tr('Documents')}: {first} {last}", font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(pady=10)
            
            info_container = tb.Frame(win, padding=10)
            info_container.pack(fill=X)

            info_grid = tb.Frame(info_container)
            info_grid.pack(side=LEFT, fill=BOTH, expand=True)

            tb.Label(info_grid, text=f"{self._tr('Phone:')} {phone}", font=("Segoe UI", 11)).grid(row=0, column=0, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('Email:')} {email}", font=("Segoe UI", 11)).grid(row=0, column=1, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('SSN:')} {ssn}", font=("Segoe UI", 11)).grid(row=1, column=0, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('Address:')} {address}", font=("Segoe UI", 11)).grid(row=1, column=1, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('Start Date:')} {start_date}", font=("Segoe UI", 11)).grid(row=2, column=0, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('End Date:')} {end_date}", font=("Segoe UI", 11)).grid(row=2, column=1, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('CV:')} {cv_name}", font=("Segoe UI", 11)).grid(row=3, column=0, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('ID Photo:')} {id_name}", font=("Segoe UI", 11)).grid(row=3, column=1, sticky=W, padx=15, pady=2)
            tb.Label(info_grid, text=f"{self._tr('Personal Photo:')} {pp_name}", font=("Segoe UI", 11)).grid(row=4, column=0, sticky=W, padx=15, pady=2)
            
            def edit_folder_info():
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT first_name, last_name, phone, email, hour_rate, percentage, ssn, address, start_date, end_date, cv_path, id_photo_path, personal_photo_path, use_tiered_payout FROM employees WHERE id=?", (emp_id,))
                emp_data = cursor.fetchone()
                conn.close()
                if emp_data:
                    f_name, l_name, ph, em, emp_rate, emp_perc, s_ssn, s_address, s_start, s_end, s_cv, s_id, s_pp, use_tiered = emp_data
                    f_name = decrypt_val(f_name) or ""
                    l_name = decrypt_val(l_name) or ""
                    ph = decrypt_val(ph) or ""
                    em = decrypt_val(em) or ""
                    s_ssn = decrypt_val(s_ssn) or ""
                    s_address = decrypt_val(s_address) or ""
                    s_start = decrypt_val(s_start) or ""
                    s_end = decrypt_val(s_end) or ""

                    rate_num = to_float(decrypt_val(emp_rate), 0.0)
                    emp_rate = f"{rate_num:g}" if rate_num > 0 else ""

                    perc_num = to_float(decrypt_val(emp_perc), 0.0)
                    if perc_num > 1.0:
                        emp_perc = f"{perc_num:g}"
                    elif perc_num > 0.0:
                        emp_perc = f"{perc_num * 100:g}"
                    else:
                        emp_perc = ""

                    use_tiered = use_tiered if use_tiered else 0
                    win.destroy()
                    self.open_employee_dialog(emp_id, f_name, l_name, ph, em, emp_rate, emp_perc, s_ssn, s_address, s_start, s_end, s_cv, s_id, s_pp, use_tiered)
                
            tb.Button(info_container, text=self._tr("✏️ Edit Info"), bootstyle="warning outline", command=edit_folder_info).pack(side=RIGHT, padx=15, anchor=NE)
            
            btn_frame = tb.Frame(win, padding=10)
            btn_frame.pack(fill=X)
            
            def refresh_files():
                for i in listbox.get_children():
                    listbox.delete(i)
                for f in os.listdir(folder):
                    if os.path.isfile(os.path.join(folder, f)):
                        listbox.insert('', tk.END, values=(f, "❌"))

            def upload_doc():
                filepath = filedialog.askopenfilename(title=self._tr("Select File"))
                if filepath:
                    try:
                        shutil.copy(filepath, folder)
                        refresh_files()
                        self.load_folders()
                        messagebox.showinfo("Success", "File uploaded successfully.", parent=win)
                    except Exception as e:
                        messagebox.showerror("Error", str(e), parent=win)
                        
            def open_dir():
                try:
                    if platform.system() == "Windows":
                        os.startfile(folder)
                    elif platform.system() == "Darwin":
                        subprocess.call(["open", folder])
                    else:
                        subprocess.call(["xdg-open", folder])
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=win)
                    
            def open_selected_file():
                sel = listbox.selection()
                if not sel: return
                filename = listbox.item(sel[0])['values'][0]
                filepath = os.path.join(folder, filename)
                try:
                    open_path_with_default_app(filepath)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=win)

            def on_tree_click(event):
                region = listbox.identify_region(event.x, event.y)
                if region != "cell":
                    return
                column = listbox.identify_column(event.x)
                if column == "#2":
                    item_id = listbox.identify_row(event.y)
                    if not item_id:
                        return
                    filename = listbox.item(item_id)['values'][0]
                    
                    confirm = messagebox.askyesno(
                        "Confirm File Deletion", 
                        f"Are you sure you want to delete {filename}?\nThis cannot be undone.", 
                        parent=win
                    )
                    if confirm:
                        try:
                            filepath = os.path.join(folder, filename)
                            if os.path.exists(filepath):
                                os.remove(filepath)
                            refresh_files()
                            self.load_folders()
                            messagebox.showinfo("Deleted", "File deleted successfully.", parent=win)
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to delete file:\n{e}", parent=win)

            tb.Button(btn_frame, text=self._tr("Upload Document"), bootstyle="success", command=upload_doc).pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("Open Folder"), bootstyle="info", command=open_dir).pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("Open Selected"), bootstyle="secondary", command=open_selected_file).pack(side=LEFT, padx=5)
            
            close_frame = tb.Frame(win, padding=10, bootstyle="secondary")
            close_frame.pack(fill=X, side=BOTTOM, padx=20, pady=(0, 20))
            tb.Button(close_frame, text=self._tr("Close Window"), bootstyle="light", cursor="hand2", command=win.destroy).pack(side=LEFT)
            
            docs_tree_holder = tb.Frame(win)
            docs_tree_holder.pack(fill=BOTH, expand=True, padx=20, pady=10)
            listbox = tb.Treeview(docs_tree_holder, columns=("filename", "delete_btn"), show="headings", bootstyle="primary")
            listbox.heading("filename", text=self._tr("Documents"))
            listbox.heading("delete_btn", text="")
            listbox.column("filename", anchor=W, stretch=tk.YES, minwidth=200)
            listbox.column("delete_btn", width=60, minwidth=60, anchor=CENTER, stretch=tk.NO)
            self._attach_tree_scrollbars(docs_tree_holder, listbox)
            
            listbox.bind("<Double-1>", lambda e: open_selected_file())
            listbox.bind("<Button-1>", on_tree_click)
            
            refresh_files()

        def save_payroll_record(self):
            if getattr(self, "_saving_payroll", False):
                return
            date_val = self.entry_vars['Date'].entry.get()
            emp_name = self.entry_vars['Employee'].get()
            revenue = self.entry_vars['Revenue'].get() or "0"
            addon_sales = self.entry_vars.get('Service Add-on Sales', None)
            addon_sales_val = addon_sales.get() or "0" if addon_sales else "0"
            hours = self.entry_vars['Hours'].get() or "0"
            notes = self.entry_vars['Notes'].get()
            written_up = self.entry_vars['Written Up'].get()
            write_up_reason = self.entry_vars['Write Up Reason'].get().strip()
            
            cust_hr = self.entry_vars['Hour Rate'].get() or "0"
            cust_perc = self.entry_vars['Percentage'].get() or "0"
            
            if not date_val:
                messagebox.showerror("Error", "Date is required.")
                return
            try:
                datetime.strptime(date_val, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Error", "Invalid Date format. Use YYYY-MM-DD.")
                return
                
            if not emp_name:
                messagebox.showerror("Error", "Please select an employee.")
                return
                
            try:
                revenue = float(revenue)
                addon_sales_val = float(addon_sales_val)
                hours = float(hours)
                cust_hr = float(cust_hr)
            except ValueError:
                messagebox.showerror("Error", "Revenue, Add-on Sales, Hours and Hour Rate must be numbers.")
                return
                
            if revenue < 0 or addon_sales_val < 0 or hours < 0 or cust_hr < 0:
                messagebox.showerror("Error", "Amounts and hours cannot be negative.")
                return

            if written_up == "Yes" and not write_up_reason:
                messagebox.showerror("Validation Error", "Write Up Reason is mandatory when Written Up is Yes.")
                return

            self._saving_payroll = True
            self.show_busy(self._tr("Saving payroll record…"))
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT id, hour_rate, percentage, use_tiered_payout FROM employees WHERE name=?", (emp_name,))
                emp = cursor.fetchone()
                
                if not emp:
                    conn.close()
                    messagebox.showerror("Error", "Employee not found.")
                    return
                    
                emp_id, hr_rate, perc, use_tiered = emp
                
                hour_rate_val = cust_hr
                if use_tiered:
                    perc_to_save = None
                    _, service_perc, product_perc, _ = self.get_employee_payout_details(emp_id, date_val, date_val)
                    calculation = round(((revenue + addon_sales_val) * service_perc) + (hours * hour_rate_val), 2)
                else:
                    try:
                        perc_to_save = float(cust_perc) / 100.0
                    except ValueError:
                        conn.close()
                        messagebox.showerror("Error", "Percentage must be a number.")
                        return
                    calculation = round(((revenue + addon_sales_val) * perc_to_save) + (hours * hour_rate_val), 2)
                    
                try:
                    _cyc_to_save = None
                    if 'Cycle' in self.entry_vars:
                        c_val = str(self.entry_vars['Cycle'].get() or "").strip()
                        if c_val:
                            # Try parsing as direct key or matching against standard cycle labels
                            if parse_cycle_key(c_val):
                                _cyc_to_save = c_val
                            else:
                                for yr in (datetime.today().year - 1, datetime.today().year, datetime.today().year + 1):
                                    for ck in cycles_for_year(yr):
                                        if cycle_label_with_year(ck) == c_val or cycle_label(ck) == c_val:
                                            _cyc_to_save = ck
                                            break
                                    if _cyc_to_save:
                                        break
                    if not _cyc_to_save:
                        _cyc_to_save = cycle_for_date(date_val)
                    cursor.execute('''
                        INSERT INTO payroll_records (employee_id, record_date, payment_amount, payment_type, revenue, service_addon_sales, hours, calculation, notes, written_up, written_up_desc, hour_rate, percentage, cycle_key)
                        VALUES (?, ?, NULL, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (emp_id, date_val, revenue, addon_sales_val, hours, calculation, notes, written_up, write_up_reason if written_up == "Yes" else "", hour_rate_val, perc_to_save, _cyc_to_save))
                    
                    commit_and_save(conn)
                except Exception as e:
                    conn.close()
                    messagebox.showerror("Database Error", f"Failed to save record: {e}")
                    return
                    
                conn.close()
            finally:
                self._saving_payroll = False
                self.hide_busy()
            
            messagebox.showinfo("Success", f"Record saved!\nCalculated amount: ${calculation:,.2f}")
            
            for field in ["Revenue", "Service Add-on Sales", "Hours", "Hour Rate", "Percentage", "Notes", "Write Up Reason"]:
                if field in self.entry_vars:
                    if field == 'Percentage':
                        self.entry_vars[field].config(state="normal")
                    self.entry_vars[field].delete(0, tk.END)
            self.entry_vars['Written Up'].set("")
            if hasattr(self, 'lbl_reason'):
                self.lbl_reason.grid_remove()
            if hasattr(self, 'ent_reason'):
                self.ent_reason.grid_remove()
            if hasattr(self, 'help_reason_btn'):
                self.help_reason_btn.grid_remove()
                
            self.load_calendar_data()

        def open_expenses_window(self):
            if self._widget_alive(getattr(self, "expenses_win", None)):
                self.expenses_win.focus()
                return
                
            self.expenses_win = tb.Toplevel(self)
            self.expenses_win.title(self._tr("Expense Reports"))
            try:
                self.update_idletasks()
                w = self.winfo_width()
                h = self.winfo_height()
                x = self.winfo_x()
                y = self.winfo_y()
                self.expenses_win.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                self.expenses_win.geometry("1100x600")
            self.expenses_win.focus_force()
            self._present_window(self.expenses_win)
            
            header_frame = tb.Frame(self.expenses_win)
            header_frame.pack(fill=X, padx=25, pady=(20, 10))
            
            tb.Label(header_frame, text=self._tr("💸 Expense Reports"), font=("Segoe UI", 22, "bold"), bootstyle="warning").pack(side=LEFT)
            
            btn_frame = tb.Frame(self.expenses_win)
            btn_frame.pack(fill=X, padx=25, pady=(0, 15))
            
            tb.Button(btn_frame, text=self._tr("+ Add Expense"), bootstyle="success", cursor="hand2", command=self.open_expense_dialog).pack(side=LEFT)
            tb.Button(btn_frame, text=self._tr("✏️ Edit Selected"), bootstyle="warning", cursor="hand2", command=self.edit_selected_expense).pack(side=LEFT, padx=10)
            tb.Button(btn_frame, text=self._tr("🗑️ Delete Selected"), bootstyle="danger", cursor="hand2", command=self.delete_selected_expense).pack(side=LEFT, padx=10)
            tb.Label(btn_frame, text=self._tr("(Ctrl / ⌘ + click to multi-select)"), font=("Segoe UI", 9), bootstyle="secondary").pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("💾 Export Excel"), bootstyle="info", cursor="hand2", command=self.export_expenses_excel).pack(side=LEFT, padx=10)
            tb.Button(btn_frame, text=self._tr("Refresh"), bootstyle="secondary outline", cursor="hand2", command=self.load_expenses_data).pack(side=RIGHT)
            
            filters_lf = tb.Labelframe(self.expenses_win, text=self._tr("Filters"), padding=10, bootstyle="warning")
            filters_lf.pack(fill=X, padx=25, pady=(0, 10))
            
            tb.Label(filters_lf, text=self._tr("From:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(5, 5))
            self.exp_from_date = tb.DateEntry(filters_lf, bootstyle="warning", dateformat='%Y-%m-%d')
            self.exp_from_date.pack(side=LEFT, padx=5)
            
            tb.Label(filters_lf, text=self._tr("To:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.exp_to_date = tb.DateEntry(filters_lf, bootstyle="warning", dateformat='%Y-%m-%d')
            self.exp_to_date.pack(side=LEFT, padx=5)
            
            today = datetime.today()
            from_date = today.replace(day=1).strftime('%Y-%m-%d')
            to_date = today.strftime('%Y-%m-%d')
            self.exp_from_date.entry.delete(0, tk.END)
            self.exp_from_date.entry.insert(0, from_date)
            self.exp_to_date.entry.delete(0, tk.END)
            self.exp_to_date.entry.insert(0, to_date)
            
            tb.Label(filters_lf, text=self._tr("Employee:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.exp_emp_filter = tb.Combobox(filters_lf, width=18, state="readonly")
            self.exp_emp_filter.pack(side=LEFT, padx=5)
            
            tb.Label(filters_lf, text=self._tr("Category:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.exp_cat_filter = tb.Combobox(filters_lf, width=15, state="readonly", values=["All"] + self.get_db_categories() + ["Employee Revenue"])
            self.exp_cat_filter.set("All")
            self.exp_cat_filter.pack(side=LEFT, padx=5)
            
            def change_filter_popdown_color(event=None):
                try:
                    popdown = self.exp_cat_filter.tk.call('ttk::combobox::PopdownWindow', self.exp_cat_filter)
                    listbox = popdown + '.f.l'
                    self.exp_cat_filter.tk.call(listbox, 'itemconfigure', 'end', '-background', 'green')
                    self.exp_cat_filter.tk.call(listbox, 'itemconfigure', 'end', '-foreground', 'white')
                except Exception:
                    pass
            self.exp_cat_filter.bind("<Map>", change_filter_popdown_color)
            
            tb.Label(filters_lf, text=self._tr("Status:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.exp_status_filter = tb.Combobox(filters_lf, width=12, state="readonly", values=["All", "Pending", "Approved", "Rejected"])
            self.exp_status_filter.set("All")
            self.exp_status_filter.pack(side=LEFT, padx=5)
            
            tb.Label(filters_lf, text=self._tr("Cycle:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.exp_cycle_vars = {}
            self.exp_cycle_select_all = tk.BooleanVar(value=True)
            self._exp_cycle_all_mode = True
            self._exp_cycle_submenus = []
            self.exp_cycle_summary = tk.StringVar(value=self._tr("All Cycles"))
            self.btn_exp_cycles = tb.Menubutton(
                filters_lf,
                textvariable=self.exp_cycle_summary,
                bootstyle="warning-outline",
                width=24,
                cursor="hand2",
            )
            self.btn_exp_cycles.pack(side=LEFT, padx=5)
            self.exp_cycle_menu = tk.Menu(self.btn_exp_cycles, tearoff=0)
            self.btn_exp_cycles.configure(menu=self.exp_cycle_menu)
            self.rebuild_exp_cycle_menu()
            
            self.refresh_expense_filter_employees()
            
            summary_frame = tb.Frame(self.expenses_win, padding=12, bootstyle="warning")
            summary_frame.pack(fill=X, side=BOTTOM, padx=25, pady=(10, 20))
            
            self.lbl_expenses_summary = tb.Label(summary_frame, text="", font=("Segoe UI", 14, "bold"), bootstyle="inverse-warning")
            self.lbl_expenses_summary.pack(side=RIGHT, padx=10)
            
            tb.Button(summary_frame, text=self._tr("Close Window"), bootstyle="light", cursor="hand2", command=self.expenses_win.destroy).pack(side=LEFT, padx=10)

            # Main partition container
            main_content = tb.Frame(self.expenses_win)
            main_content.pack(fill=BOTH, expand=True, padx=25, pady=10)
            
            # Left Treeview Panel
            tree_frame = tb.Frame(main_content)
            tree_frame.pack(side=LEFT, fill=BOTH, expand=True)
            
            cols = (self._tr("ID"), self._tr("Date"), self._tr("Cycle"), self._tr("Employee"), self._tr("Category"), self._tr("Payment Type"), self._tr("Amount"), self._tr("Status"))
            
            scroll_y = tb.Scrollbar(tree_frame, orient=VERTICAL)
            scroll_x = tb.Scrollbar(tree_frame, orient=HORIZONTAL)
            
            self.tree_expenses = tb.Treeview(tree_frame, columns=cols, show="headings", bootstyle="warning", yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set, selectmode="extended")
            
            scroll_y.config(command=self.tree_expenses.yview)
            scroll_y.pack(side=RIGHT, fill=Y)
            
            scroll_x.config(command=self.tree_expenses.xview)
            scroll_x.pack(side=BOTTOM, fill=X)
            
            self.apply_and_memorize_column_widths(
                "expenses_table",
                self.tree_expenses,
                cols,
            )
            
            self.tree_expenses.bind("<Double-1>", lambda e: self.edit_selected_expense())
            self.tree_expenses.tag_configure('income', foreground='#00bc8c', font=("Segoe UI", 11, "bold"))
            self.tree_expenses.tag_configure('expense', foreground='#e74c3c', font=("Segoe UI", 11))
            self.tree_expenses.pack(side=LEFT, fill=BOTH, expand=True)
            
            # Right Chart Panel
            self.chart_frame = tb.Labelframe(main_content, text=self._tr("Expense Distribution"), padding=10, bootstyle="warning")
            self.chart_frame.pack(side=RIGHT, fill=BOTH, padx=(20, 0), ipadx=10)
            
            bg_color = self.style.colors.bg if hasattr(self, 'style') and hasattr(self.style, 'colors') else "#222222"
            self.chart_canvas = tk.Canvas(self.chart_frame, width=320, height=335, highlightthickness=0, bg=bg_color)
            self.chart_canvas.pack(fill=BOTH, expand=True, pady=5)
            
            divider = tb.Frame(self.chart_frame, height=2, bootstyle="warning")
            divider.pack(fill=X, pady=5)
            
            self.revenue_chart_canvas = tk.Canvas(self.chart_frame, width=320, height=320, highlightthickness=0, bg=bg_color)
            self.revenue_chart_canvas.pack(fill=BOTH, expand=True, pady=5)
            
            self.load_expenses_data()
            self.stop_filter_pollers(which="expense")
            self.poll_expense_filter_changes()

        def refresh_expense_filter_employees(self):
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM employees")
            employees = cursor.fetchall()
            conn.close()
            
            all_txt = self._tr("All")
            gen_txt = self._tr("General/None")
            self._expense_employee_map = {all_txt: "ALL", gen_txt: "GENERAL"}
            values = [all_txt, gen_txt]
            for emp_id, emp_name in employees:
                display_str = f"{emp_name} (ID: {emp_id})"
                values.append(display_str)
                self._expense_employee_map[display_str] = emp_id
                
            if self._widget_alive(getattr(self, "exp_emp_filter", None)):
                self.exp_emp_filter['values'] = values
                self.exp_emp_filter.set(all_txt)

        def rebuild_exp_cycle_menu(self, select_all=None):
            menu = getattr(self, "exp_cycle_menu", None)
            try:
                if menu is None or not menu.winfo_exists():
                    return
            except Exception:
                return
            self._exp_cycle_updating = True
            self._exp_cycle_submenus = []
            
            # Determine reference dates & cycles
            today_dt = datetime.today()
            today_str = today_dt.strftime('%Y-%m-%d')
            curr_yr = today_dt.year
            curr_ck = cycle_for_date(today_str) or f"{curr_yr}-01-1"

            # Discover all distinct years from DB & adjacent years
            all_known_years = {curr_yr, curr_yr - 1}
            all_known_cycles = set()
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute("SELECT DISTINCT cycle_key FROM expenses WHERE cycle_key IS NOT NULL AND TRIM(cycle_key)!=''")
                for (ck,) in cur.fetchall() or []:
                    if ck:
                        s_ck = str(ck).strip()
                        all_known_cycles.add(s_ck)
                        p = parse_cycle_key(s_ck)
                        if p:
                            all_known_years.add(p[0].year if hasattr(p[0], 'year') else int(p[0]))
                cur.execute("SELECT DISTINCT cycle_key FROM payroll_records WHERE cycle_key IS NOT NULL AND TRIM(cycle_key)!=''")
                for (ck,) in cur.fetchall() or []:
                    if ck:
                        s_ck = str(ck).strip()
                        all_known_cycles.add(s_ck)
                        p = parse_cycle_key(s_ck)
                        if p:
                            all_known_years.add(p[0].year if hasattr(p[0], 'year') else int(p[0]))
                conn.close()
            except Exception:
                pass

            for y in list(all_known_years):
                for ck in cycles_for_year(y):
                    all_known_cycles.add(ck)

            sorted_years = sorted(list(all_known_years), reverse=True)

            # Build list of 8 most recent cycles descending from current cycle
            recent_cycles = []
            for off in range(0, -8, -1):
                k = add_cycles(curr_ck, off)
                if k and k not in recent_cycles:
                    recent_cycles.append(k)

            prev = set(self.get_exp_included_cycles())
            if select_all is True:
                prev = set(all_known_cycles)
                self._exp_cycle_all_mode = True
            elif select_all is False:
                prev = set()
                self._exp_cycle_all_mode = False
            elif not hasattr(self, "_exp_cycle_all_mode"):
                self._exp_cycle_all_mode = True
                prev = set(all_known_cycles)

            try:
                menu.delete(0, "end")
            except Exception:
                self._exp_cycle_updating = False
                return

            self.exp_cycle_vars = {}
            try:
                all_on = getattr(self, "_exp_cycle_all_mode", True) or (len(prev) == len(all_known_cycles))
                self.exp_cycle_select_all = tk.BooleanVar(value=all_on)
                
                # 1. Top fast-actions
                menu.add_checkbutton(
                    label=f"✓  {self._tr('All Cycles')}",
                    variable=self.exp_cycle_select_all,
                    command=self._on_exp_cycle_select_all,
                )
                menu.add_command(
                    label=f"✕  {self._tr('Clear Selection')}",
                    command=self._on_exp_cycle_clear,
                )
                menu.add_separator()

                # 2. Recent Cycles section (most recent at the top!)
                menu.add_command(
                    label=f"⏱  ─── {self._tr('Recent Cycles')} ───",
                    state="disabled",
                )
                for ck in recent_cycles:
                    lbl = cycle_label_with_year(ck)
                    var = tk.BooleanVar(value=True if all_on else (ck in prev))
                    self.exp_cycle_vars[ck] = var
                    menu.add_checkbutton(
                        label=lbl,
                        variable=var,
                        command=lambda k=ck: self._on_exp_cycle_changed(k),
                    )

                menu.add_separator()
                menu.add_command(
                    label=f"📅  ─── {self._tr('Browse by Year')} ───",
                    state="disabled",
                )

                # 3. Year Submenus (Cascades)
                for yr in sorted_years:
                    yr_cycles = cycles_for_year(yr)
                    yr_cycles.reverse()  # Latest cycles first (Dec down to Jan)
                    
                    sub = tk.Menu(menu, tearoff=0)
                    self._exp_cycle_submenus.append(sub)
                    
                    sub.add_command(
                        label=f"✓  {self._tr('Select All')} {yr}",
                        command=lambda y=yr: self._on_exp_cycle_select_year(y),
                    )
                    sub.add_separator()

                    for ck in yr_cycles:
                        lbl = cycle_label_with_year(ck)
                        if ck not in self.exp_cycle_vars:
                            var = tk.BooleanVar(value=True if all_on else (ck in prev))
                            self.exp_cycle_vars[ck] = var
                        else:
                            var = self.exp_cycle_vars[ck]
                        sub.add_checkbutton(
                            label=lbl,
                            variable=var,
                            command=lambda k=ck: self._on_exp_cycle_changed(k),
                        )

                    menu.add_cascade(
                        label=f"📅  {yr} {self._tr('Cycles')} ({len(yr_cycles)})",
                        menu=sub,
                    )

            except tk.TclError:
                self.exp_cycle_vars = {}
                self._exp_cycle_updating = False
                return

            self._update_exp_cycle_summary()
            self._exp_cycle_updating = False

        def _on_exp_cycle_select_all(self):
            if getattr(self, "_exp_cycle_updating", False):
                return
            self._exp_cycle_all_mode = True
            for var in (self.exp_cycle_vars or {}).values():
                var.set(True)
            if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                self.exp_cycle_select_all.set(True)
            self._update_exp_cycle_summary()
            self.load_expenses_data(quiet=True)

        def _on_exp_cycle_clear(self):
            if getattr(self, "_exp_cycle_updating", False):
                return
            self._exp_cycle_all_mode = False
            for var in (self.exp_cycle_vars or {}).values():
                var.set(False)
            if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                self.exp_cycle_select_all.set(False)
            self._update_exp_cycle_summary()
            self.load_expenses_data(quiet=True)

        def _on_exp_cycle_select_year(self, year):
            if getattr(self, "_exp_cycle_updating", False):
                return
            self._exp_cycle_all_mode = False
            yr_str = str(year)
            for ck, var in (self.exp_cycle_vars or {}).items():
                if ck.startswith(yr_str):
                    var.set(True)
                else:
                    var.set(False)
            if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                self.exp_cycle_select_all.set(False)
            self._update_exp_cycle_summary()
            self.load_expenses_data(quiet=True)

        def _on_exp_cycle_changed(self, cycle_key):
            if getattr(self, "_exp_cycle_updating", False):
                return
            # If all cycles were active, clicking one cycle isolates that specific cycle
            if getattr(self, "_exp_cycle_all_mode", True):
                self._exp_cycle_updating = True
                for k, var in (self.exp_cycle_vars or {}).items():
                    if k != cycle_key:
                        var.set(False)
                    else:
                        var.set(True)
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(False)
                self._exp_cycle_all_mode = False
                self._exp_cycle_updating = False

            self._update_exp_cycle_summary()
            self.load_expenses_data(quiet=True)

        def _update_exp_cycle_summary(self):
            summary = getattr(self, "exp_cycle_summary", None)
            if summary is None:
                return
            vars_map = getattr(self, "exp_cycle_vars", None) or {}
            total = len(vars_map)
            selected = self.get_exp_included_cycles()
            if getattr(self, "_exp_cycle_all_mode", True) or not vars_map or len(selected) == total:
                summary.set(self._tr("All Cycles"))
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(True)
            elif not selected:
                summary.set(self._tr("All Cycles"))
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(False)
            elif len(selected) == 1:
                summary.set(cycle_short_label(selected[0]))
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(False)
            elif len(selected) == 2:
                s1 = cycle_short_label(selected[0])
                s2 = cycle_short_label(selected[1])
                summary.set(f"{s1}, {s2}")
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(False)
            else:
                summary.set(f"{len(selected)} {self._tr('Cycles Selected')}")
                if hasattr(self, "exp_cycle_select_all") and self.exp_cycle_select_all is not None:
                    self.exp_cycle_select_all.set(False)

        def get_exp_included_cycles(self):
            vars_map = getattr(self, "exp_cycle_vars", None) or {}
            included = []
            for ck, var in vars_map.items():
                try:
                    if var.get():
                        included.append(ck)
                except Exception:
                    pass
            return included

        def _exp_cycle_included(self, cycle_key):
            if getattr(self, "_exp_cycle_all_mode", True):
                return True
            vars_map = getattr(self, "exp_cycle_vars", None) or {}
            if not vars_map:
                return True
            sel_all = getattr(self, "exp_cycle_select_all", None)
            if sel_all is not None:
                try:
                    if sel_all.get():
                        return True
                except Exception:
                    pass
            included = set(self.get_exp_included_cycles())
            if not included:
                return True
            return cycle_key in included

        def poll_expense_filter_changes(self):
            self._poll_expense_after_id = None
            if not self._widget_alive(getattr(self, "expenses_win", None)):
                return
            if not self._dateentry_alive(getattr(self, "exp_from_date", None)) or not self._dateentry_alive(
                getattr(self, "exp_to_date", None)
            ):
                return
            try:
                current_from = self.exp_from_date.entry.get()
                current_to = self.exp_to_date.entry.get()
                emp_val = self.exp_emp_filter.get()
                cat_val = self.exp_cat_filter.get()
                status_val = self.exp_status_filter.get()
                cycle_val = tuple(sorted(self.get_exp_included_cycles()))
                
                state = (current_from, current_to, emp_val, cat_val, status_val, cycle_val)
                if not hasattr(self, "_last_expense_filter_state"):
                    self._last_expense_filter_state = state
                elif self._last_expense_filter_state != state:
                    self._last_expense_filter_state = state
                    self.load_expenses_data(quiet=True)
            except tk.TclError:
                return
            except Exception as e:
                if "invalid command name" in str(e):
                    return
                print(f"Error in poll_expense_filter_changes: {e}")
            if self._widget_alive(getattr(self, "expenses_win", None)):
                self._poll_expense_after_id = self.expenses_win.after(750, self.poll_expense_filter_changes)

        def load_expenses_data(self, quiet=False):
            if quiet:
                return self._load_expenses_data_body()
            self.show_busy(self._tr("Loading expenses…"))
            try:
                return self._load_expenses_data_body()
            finally:
                self.hide_busy()

        def _load_expenses_data_body(self):
            if not self._widget_alive(getattr(self, "tree_expenses", None)):
                return
            for item in self.tree_expenses.get_children():
                self.tree_expenses.delete(item)
                
            from_d = self.exp_from_date.entry.get()
            to_d = self.exp_to_date.entry.get()
            emp_display = self.exp_emp_filter.get()
            cat_val = self.exp_cat_filter.get()
            status_val = self.exp_status_filter.get()
            
            query = '''
                SELECT ex.id, ex.expense_date, ex.employee_id, e.name, ex.category, ex.amount, ex.status, ex.description, ex.payment_type, ex.assignee_id, e2.name AS assignee_name, ex.is_tip, ex.tip_given, ex.cycle_key
                FROM expenses ex
                LEFT JOIN employees e ON ex.employee_id = e.id
                LEFT JOIN employees e2 ON ex.assignee_id = e2.id
                WHERE 1=1
            '''
            params = []
            
            if from_d:
                query += " AND ex.expense_date >= ?"
                params.append(from_d)
            if to_d:
                query += " AND ex.expense_date <= ?"
                params.append(to_d)
                
            if emp_display:
                mapped_emp = self._expense_employee_map.get(emp_display, "ALL")
                if mapped_emp == "GENERAL":
                    query += " AND ex.employee_id IS NULL"
                elif mapped_emp != "ALL":
                    query += " AND ex.employee_id = ?"
                    params.append(mapped_emp)

            query += " ORDER BY ex.expense_date DESC"
            
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            
            # Fetch individual expenses (filter category/status in Python so
            # leftover encrypted values still match Salary Payment, etc.)
            cursor.execute(query, params)
            expense_rows = cursor.fetchall()
            want_cat = plain_label(cat_val) if cat_val and cat_val not in ("All", self._tr("All")) else ""
            want_status = plain_label(status_val) if status_val and status_val not in ("All", self._tr("All")) else ""
            if want_cat or want_status:
                filtered = []
                for row in expense_rows or []:
                    cat_p = plain_label(row[4] if len(row) > 4 else "")
                    st_p = plain_label(row[6] if len(row) > 6 else "")
                    if want_cat:
                        if want_cat == "Salary Payment":
                            if cat_p not in SALARY_EXPENSE_CATEGORIES:
                                continue
                        elif cat_p != want_cat:
                            continue
                    if want_status and st_p != want_status:
                        continue
                    filtered.append(row)
                expense_rows = filtered
            
            # Fetch individual payroll revenues (if 'All' or 'Employee Revenue' is filtered)
            show_emp_rev = (cat_val == "All" or cat_val == "Employee Revenue")
            payroll_rows = []
            
            if show_emp_rev:
                query_emp_rev = '''
                    SELECT r.id, r.record_date, r.employee_id, e.name, 'Employee Revenue' AS category, r.revenue, 'Approved' AS status, r.notes, NULL AS payment_type, NULL AS assignee_id, NULL AS assignee_name
                    FROM payroll_records r
                    JOIN employees e ON r.employee_id = e.id
                    WHERE r.record_date >= ? AND r.record_date <= ?
                '''
                params_emp_rev = [from_d, to_d]
                
                if emp_display:
                    mapped_emp = self._expense_employee_map.get(emp_display, "ALL")
                    if mapped_emp == "GENERAL":
                        query_emp_rev += " AND 1=0"
                    elif mapped_emp != "ALL":
                        query_emp_rev += " AND r.employee_id = ?"
                        params_emp_rev.append(mapped_emp)
                        
                query_emp_rev += " ORDER BY r.record_date DESC"
                
                cursor.execute(query_emp_rev, params_emp_rev)
                payroll_rows = cursor.fetchall()
                
            conn.close()
            
            # Combine and sort descendently by date
            combined_rows = []
            for r in expense_rows:
                combined_rows.append(r)
            for r in payroll_rows:
                combined_rows.append(r)
                
            combined_rows.sort(key=lambda x: x[1], reverse=True)
            
            total_amt = 0.0
            total_revenue = 0.0
            
            for row in combined_rows:
                exp_id, date_val, emp_id, emp_name, category, amount, status, desc, p_type = row[:9]
                assignee_name = row[10] if len(row) > 10 else None
                amt_val = to_float(amount, 0.0)
                category = plain_label(category)
                status = plain_label(status)
                emp_name = plain_label(emp_name) or emp_name
                assignee_name = plain_label(assignee_name) or assignee_name
                desc = plain_label(desc) if desc else ""
                is_tip_flag = plain_label(row[11]) if len(row) > 11 else ""
                tip_given_val = to_float(row[12], 0.0) if len(row) > 12 else 0.0

                # Cycle label: stored cycle_key for expenses (index 13), else
                # derived from the row's date (covers revenue rows too).
                stored_cycle = row[13] if len(row) > 13 else None
                cycle_key_val = (str(stored_cycle).strip() if stored_cycle else "") or cycle_for_date(date_val)
                if not self._exp_cycle_included(cycle_key_val):
                    continue
                cycle_disp = cycle_label(cycle_key_val) if cycle_key_val else ""

                if is_tip_flag.lower() in ("yes", "true", "1") and tip_given_val:
                    tip_note = f"{self._tr('Tip given to employee:')} ${tip_given_val:,.2f}"
                    desc = f"{desc}  ·  {tip_note}".strip(" ·") if desc else tip_note
                
                if is_income_expense_category(category):
                    total_revenue += amt_val
                    formatted_amt = f"+${amt_val:,.2f}"
                    row_tag = ("income",)
                else:
                    total_amt += amt_val
                    formatted_amt = f"-${amt_val:,.2f}"
                    row_tag = ("expense",)
                    
                if category == "Cash Envelope Received" and assignee_name:
                    emp_disp = f"Shop ({self._tr('Received From')}: {assignee_name})"
                else:
                    emp_disp = emp_name if emp_name else self._tr("General/None")

                formatted_row = (
                    exp_id,
                    date_val,
                    cycle_disp,
                    emp_disp,
                    self._tr(category),
                    p_type if p_type else "",
                    formatted_amt,
                    self._tr(status)
                )
                self.tree_expenses.insert('', tk.END, values=formatted_row, tags=row_tag)
            
            net_income = total_revenue - total_amt
            self._last_total_revenue = total_revenue
            self._last_total_amt = total_amt
            
            lbl_text = f"{self._tr('Total Revenue')}: ${total_revenue:,.2f}  |  {self._tr('Total Expenses')}: ${total_amt:,.2f}  |  {self._tr('Net Income')}: ${net_income:,.2f}"
            self.lbl_expenses_summary.config(text=lbl_text)
            self.draw_expense_chart()

        def draw_expense_chart(self):
            if not self._widget_alive(getattr(self, "chart_canvas", None)):
                return
                
            self.chart_canvas.delete("all")
            
            # Aggregates parsed amounts by category directly from the tree
            cat_totals = {}
            for item in self.tree_expenses.get_children():
                values = self.tree_expenses.item(item)['values']
                if not values or len(values) < 7:
                    continue
                cat = values[4]
                raw_amt = values[6]
                
                # Exclude cash envelopes and employee revenues (both are incoming gains, not expense costs)
                if cat == self._tr("Cash Envelope Received") or cat == "Cash Envelope Received":
                    continue
                if cat == self._tr("Employee Revenue") or cat == "Employee Revenue":
                    continue
                    
                try:
                    amt_clean = str(raw_amt).replace("+", "").replace("-", "").replace("$", "").replace(",", "")
                    amt = float(amt_clean)
                    if amt > 0:
                        cat_totals[cat] = cat_totals.get(cat, 0.0) + amt
                except Exception:
                    pass
                    
            if not cat_totals:
                self.chart_canvas.create_text(
                    160, 160, 
                    text=self._tr("No expense data to display in chart"), 
                    font=("Segoe UI", 11, "italic"), 
                    fill="gray",
                    justify=tk.CENTER
                )
                return
                
            # Theme specific adaptive foreground color
            fg_color = self.style.colors.fg if hasattr(self, 'style') and hasattr(self.style, 'colors') else "#ffffff"
            # Custom color palette reflecting darkly aesthetic
            colors = ["#375a7f", "#00bc8c", "#f39c12", "#e74c3c", "#9b59b6", "#3498db", "#1abc9c", "#d35400"]
            total_sum = sum(cat_totals.values())
            
            # Circular bounding box variables for pie chart
            x1, y1, x2, y2 = 85, 35, 235, 185
            start_angle = 0
            color_idx = 0
            
            for cat, amt in cat_totals.items():
                percentage = amt / total_sum
                extent = percentage * 360
                color = colors[color_idx % len(colors)]
                color_idx += 1
                
                self.chart_canvas.create_arc(
                    x1, y1, x2, y2, 
                    start=start_angle, 
                    extent=extent, 
                    fill=color, 
                    outline="#222222", 
                    width=1
                )
                start_angle += extent
                
            # Single-column vertical Legend stack
            legend_y = 195
            legend_x = 15
            
            # Text title
            self.chart_canvas.create_text(
                160, 15, 
                text=self._tr("Expense Distribution"), 
                font=("Segoe UI", 11, "bold"), 
                fill=fg_color, 
                justify=tk.CENTER
            )
            
            for i, (cat, amt) in enumerate(cat_totals.items()):
                color = colors[i % len(colors)]
                perc_str = f"({(amt/total_sum)*100:.1f}%)"
                
                curr_x = legend_x
                curr_y = legend_y + (i * 16)
                
                # Solid color oval for category indicator
                self.chart_canvas.create_oval(
                    curr_x, curr_y + 4, 
                    curr_x + 12, curr_y + 16, 
                    fill=color, 
                    outline=""
                )
                # Localized category name with sum and percentage
                disp_text = f"{cat}: ${amt:,.0f} {perc_str}"
                self.chart_canvas.create_text(
                    curr_x + 18, curr_y + 10, 
                    text=disp_text, 
                    anchor=tk.W, 
                    font=("Segoe UI", 9, "bold"), 
                    fill=fg_color
                )
            self.draw_revenue_vs_expenses_chart()

        def draw_revenue_vs_expenses_chart(self):
            if not self._widget_alive(getattr(self, "revenue_chart_canvas", None)):
                return
                
            self.revenue_chart_canvas.delete("all")
            
            total_revenue = getattr(self, '_last_total_revenue', 0.0)
            total_expenses = getattr(self, '_last_total_amt', 0.0)
            
            # Text Title
            fg_color = self.style.colors.fg if hasattr(self, 'style') and hasattr(self.style, 'colors') else "#ffffff"
            self.revenue_chart_canvas.create_text(
                160, 15, 
                text=self._tr("Expenses vs. Revenue"), 
                font=("Segoe UI", 11, "bold"), 
                fill=fg_color, 
                justify=tk.CENTER
            )
            
            if total_revenue == 0.0 and total_expenses == 0.0:
                self.revenue_chart_canvas.create_text(
                    160, 110, 
                    text=self._tr("No financial data to display in chart"), 
                    font=("Segoe UI", 10, "italic"), 
                    fill="gray",
                    justify=tk.CENTER
                )
                return
                
            total_sum = total_revenue + total_expenses
            
            # Spherical bounding box coordinates
            x1, y1, x2, y2 = 75, 35, 245, 205
            
            rev_percentage = total_revenue / total_sum
            rev_extent = rev_percentage * 360
            
            exp_percentage = total_expenses / total_sum
            exp_extent = exp_percentage * 360
            
            # Green for Revenue: '#00bc8c'
            # Red for Expenses: '#e74c3c'
            if total_revenue > 0:
                self.revenue_chart_canvas.create_arc(
                    x1, y1, x2, y2, 
                    start=0, 
                    extent=rev_extent, 
                    fill="#00bc8c", 
                    outline="#222222", 
                    width=1
                )
            if total_expenses > 0:
                self.revenue_chart_canvas.create_arc(
                    x1, y1, x2, y2, 
                    start=rev_extent, 
                    extent=exp_extent, 
                    fill="#e74c3c", 
                    outline="#222222", 
                    width=1
                )
                
            # Single-column vertical Legend stack
            legend_y = 215
            
            # Revenue Legend (Green)
            self.revenue_chart_canvas.create_oval(
                15, legend_y + 2, 
                27, legend_y + 14, 
                fill="#00bc8c", 
                outline=""
            )
            rev_perc_str = f"({rev_percentage*100:.1f}%)"
            self.revenue_chart_canvas.create_text(
                33, legend_y + 8, 
                text=f"{self._tr('Total Revenue')}: ${total_revenue:,.2f} {rev_perc_str}", 
                anchor=tk.W, 
                font=("Segoe UI", 9, "bold"), 
                fill=fg_color
            )
            
            # Expenses Legend (Red) - stacked vertically 24px below
            self.revenue_chart_canvas.create_oval(
                15, legend_y + 26, 
                27, legend_y + 38, 
                fill="#e74c3c", 
                outline=""
            )
            exp_perc_str = f"({exp_percentage*100:.1f}%)"
            self.revenue_chart_canvas.create_text(
                33, legend_y + 32, 
                text=f"{self._tr('Total Expenses')}: ${total_expenses:,.2f} {exp_perc_str}", 
                anchor=tk.W, 
                font=("Segoe UI", 9, "bold"), 
                fill=fg_color
            )

        def open_expense_dialog(self, expense_id=None, data=None, on_save_callback=None):
            if self._widget_alive(getattr(self, "_envelope_popup", None)):
                parent_win = self._envelope_popup
            elif self._widget_alive(getattr(self, "expenses_win", None)):
                parent_win = self.expenses_win
            else:
                parent_win = self
            
            is_env = False
            if data and len(data) > 1 and is_envelope_category(data[1]):
                is_env = True
            elif self._widget_alive(getattr(self, "_envelope_popup", None)):
                is_env = True

            if is_env:
                win_title = self._tr("Edit Cash Envelope") if expense_id else self._tr("Add Cash Envelope")
            else:
                win_title = self._tr("Edit Expense") if expense_id else self._tr("Add Expense")

            dialog = self._open_sheet(
                parent_win,
                win_title,
                "560x680",
            )
            # Open centered at the top-middle of the screen
            try:
                dialog.update_idletasks()
                screen_w = dialog.winfo_screenwidth()
                screen_h = dialog.winfo_screenheight()
                dlg_w = min(560, max(460, screen_w - 40))
                dlg_h = min(720, max(520, screen_h - 90))
                x = max(0, (screen_w - dlg_w) // 2)
                y = max(10, min(25, (screen_h - dlg_h) // 4))
                dialog.geometry(f"{dlg_w}x{dlg_h}+{x}+{y}")
            except Exception:
                pass
            if platform.system() != "Darwin":
                try:
                    dialog.minsize(480, 520)
                except Exception:
                    pass
            self._expense_dialog = dialog

            # Keep Save always visible at the bottom
            footer = tb.Frame(dialog, padding=(15, 10))
            footer.pack(side=BOTTOM, fill=X)

            body = tb.Frame(dialog)
            body.pack(fill=BOTH, expand=True, padx=5, pady=5)

            canvas = tk.Canvas(body, highlightthickness=0)
            scrollbar = tb.Scrollbar(body, orient=VERTICAL, command=canvas.yview)
            form = tb.Frame(canvas, padding=(10, 5))
            form.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas_window = canvas.create_window((0, 0), window=form, anchor="nw")

            def _sync_form_width(event):
                canvas.itemconfigure(canvas_window, width=event.width)

            canvas.bind("<Configure>", _sync_form_width)
            canvas.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)

            def _on_mousewheel(event):
                try:
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except Exception:
                    pass

            def _bind_canvas_wheel(_e=None):
                canvas.bind_all("<MouseWheel>", _on_mousewheel)
                canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
                canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

            def _unbind_canvas_wheel(_e=None):
                try:
                    canvas.unbind_all("<MouseWheel>")
                    canvas.unbind_all("<Button-4>")
                    canvas.unbind_all("<Button-5>")
                except Exception:
                    pass

            if platform.system() == "Darwin":
                canvas.bind("<MouseWheel>", _on_mousewheel)
            else:
                canvas.bind("<Enter>", _bind_canvas_wheel)
                canvas.bind("<Leave>", _unbind_canvas_wheel)

            def _cleanup_expense_dialog(event=None):
                # Toplevel is in every child's bindtags, so ignore descendant Destroy.
                if event is not None and getattr(event, "widget", None) is not dialog:
                    return
                if getattr(self, "_expense_dialog", None) is dialog:
                    self._expense_dialog = None
                _unbind_canvas_wheel()
                self._safe_grab_release(dialog)
                if self._widget_alive(parent_win):
                    try:
                        parent_win.lift()
                        parent_win.focus_set()
                    except Exception:
                        pass

            dialog.bind("<Destroy>", _cleanup_expense_dialog)

            pad = {'padx': 12, 'pady': 7, 'sticky': E}
            ent_pad = {'padx': 12, 'pady': 7, 'sticky': W}
            form.columnconfigure(1, weight=1)
            
            tb.Label(form, text=self._tr("Date:"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0, **pad)
            # DateEntry opens a second grab_set calendar. That deadlocks if this
            # form was opened from the cash-envelope day window.
            from_envelope = self._widget_alive(getattr(self, "_envelope_popup", None))
            if platform.system() == "Darwin" or from_envelope:
                date_ent = tb.Entry(form, width=30)
                date_ent.entry = date_ent
            else:
                date_ent = tb.DateEntry(form, bootstyle="warning", dateformat='%Y-%m-%d')
            if data:
                date_ent.entry.delete(0, tk.END)
                date_ent.entry.insert(0, data[0])
            else:
                today = datetime.today()
                date_ent.entry.delete(0, tk.END)
                date_ent.entry.insert(0, today.strftime('%Y-%m-%d'))
            date_ent.grid(row=0, column=1, **ent_pad)

            # --- Pay cycle (auto-filled from the Date, but adjustable) --------
            tb.Label(form, text=self._tr("Cycle:"), font=("Segoe UI", 10, "bold"), bootstyle="warning").grid(row=1, column=0, **pad)
            cycle_cbo = tb.Combobox(form, width=38, state="readonly", bootstyle="warning")
            cycle_cbo.grid(row=1, column=1, **ent_pad)

            # Determine initial cycle: if existing data has a cycle_key, use it. Otherwise derive from date.
            cur_date_str = date_ent.entry.get()
            _initial_cycle = None
            if data and len(data) > 12 and data[12]:
                raw_c = str(data[12]).strip()
                if raw_c:
                    _initial_cycle = cycle_for_date(cycle_bounds(raw_c)[0] if cycle_bounds(raw_c) else raw_c) or raw_c
            if not _initial_cycle:
                _initial_cycle = cycle_for_date(cur_date_str) or cycle_for_date(datetime.today().strftime('%Y-%m-%d'))

            cycle_choices, cycle_key_map, init_lbl = get_formatted_cycle_choices(_initial_cycle, start_from_june_2026=True)
            cycle_cbo["values"] = cycle_choices
            cycle_cbo.set(init_lbl)
            try:
                cycle_cbo.current(cycle_choices.index(init_lbl))
            except Exception:
                pass

            _cycle_user_overridden = [False]
            def _on_cycle_selected(event=None):
                _cycle_user_overridden[0] = True

            cycle_cbo.bind("<<ComboboxSelected>>", _on_cycle_selected)

            last_tracked_date = [cur_date_str]

            def _sync_cycle_to_date():
                if _cycle_user_overridden[0]:
                    return
                try:
                    d_val = date_ent.entry.get()
                    if d_val:
                        ck = cycle_for_date(d_val)
                        if ck:
                            new_choices, new_map, new_lbl = get_formatted_cycle_choices(ck, start_from_june_2026=True)
                            cycle_cbo["values"] = new_choices
                            cycle_key_map.update(new_map)
                            cycle_cbo.set(new_lbl)
                            try:
                                cycle_cbo.current(new_choices.index(new_lbl))
                            except Exception:
                                pass
                except Exception:
                    pass

            def _poll_cycle_sync():
                if not dialog.winfo_exists():
                    return
                try:
                    cur = date_ent.entry.get()
                except Exception:
                    return
                if cur != last_tracked_date[0]:
                    last_tracked_date[0] = cur
                    _sync_cycle_to_date()
                try:
                    dialog.after(400, _poll_cycle_sync)
                except Exception:
                    pass

            try:
                dialog.after(400, _poll_cycle_sync)
            except Exception:
                pass

            lbl_emp = tb.Label(form, text=self._tr("Employee:"), font=("Segoe UI", 10, "bold"))
            lbl_emp.grid(row=2, column=0, **pad)
            emp_cbo = tb.Combobox(form, width=28, state="readonly")
            
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM employees")
            employees = cursor.fetchall()
            conn.close()
            
            gen_txt = self._tr("General/None")
            emp_list = [gen_txt]
            emp_id_map = {gen_txt: None}
            for e_id, e_name in employees:
                display_str = f"{e_name} (ID: {e_id})"
                emp_list.append(display_str)
                emp_id_map[display_str] = e_id
                
            emp_cbo['values'] = emp_list
            if data and data[3]:
                found = False
                for key, val in emp_id_map.items():
                    if val == data[3]:
                        emp_cbo.set(key)
                        found = True
                        break
                if not found:
                    emp_cbo.set(gen_txt)
            else:
                emp_cbo.set(gen_txt)
            emp_cbo.grid(row=2, column=1, **ent_pad)
            
            tb.Label(form, text=self._tr("Category:"), font=("Segoe UI", 10, "bold")).grid(row=3, column=0, **pad)
            cats = self.get_db_categories()
            cat_cbo = tb.Combobox(form, width=28, state="readonly", values=cats)
            
            def change_popdown_color(event=None):
                try:
                    popdown = cat_cbo.tk.call('ttk::combobox::PopdownWindow', cat_cbo)
                    listbox = popdown + '.f.l'
                    cat_cbo.tk.call(listbox, 'itemconfigure', 'end', '-background', 'green')
                    cat_cbo.tk.call(listbox, 'itemconfigure', 'end', '-foreground', 'white')
                except Exception:
                    pass
            if platform.system() != "Darwin":
                cat_cbo.bind("<Map>", change_popdown_color)
            
            if data:
                cat_cbo.set(plain_label(data[1]) or data[1])
            else:
                cat_cbo.set("Office Supplies" if "Office Supplies" in cats else (cats[0] if cats else ""))
            cat_cbo.grid(row=3, column=1, **ent_pad)
            
            # Tip Included checkbox (row 3) and tip-given amount (pops in below)
            lbl_tip = tb.Label(form, text=self._tr("Tip Included:"), font=("Segoe UI", 10, "bold"))
            tip_var = tk.IntVar(value=0)
            existing_tip = plain_label(data[10]) if data and len(data) > 10 else ""
            if existing_tip.lower() in ("yes", "true", "1"):
                tip_var.set(1)
            tip_chk = tb.Checkbutton(form, text="", variable=tip_var, bootstyle="success-round-toggle")
            lbl_tip_given = tb.Label(
                form,
                text=self._tr("How much of the tip did you give this employee?"),
                font=("Segoe UI", 10, "bold"),
                wraplength=170,
                justify=RIGHT,
            )
            tip_given_ent = tb.Entry(form, width=30)
            if data and len(data) > 11 and data[11] is not None and str(data[11]).strip() not in ("", "None", "0", "0.0"):
                try:
                    tip_given_ent.insert(0, f"{to_float(data[11], 0.0):.2f}")
                except Exception:
                    pass
            tip_given_hint = tb.Label(
                form,
                text=self._tr("How much of the tip did you give this employee?"),
                font=("Segoe UI", 8),
                bootstyle="secondary",
            )

            # Received From field (row 3)
            lbl_assignee = tb.Label(form, text=self._tr("Received From:"), font=("Segoe UI", 10, "bold"))
            assignee_cbo = tb.Combobox(form, width=28, state="readonly")
            
            assignee_list = [self._tr("General/None")]
            assignee_id_map = {self._tr("General/None"): None}
            for e_id, e_name in employees:
                if e_name.lower() != "shop":
                    display_str = f"{e_name} (ID: {e_id})"
                    assignee_list.append(display_str)
                    assignee_id_map[display_str] = e_id
            assignee_cbo['values'] = assignee_list
            if data and len(data) > 8 and data[8] and is_envelope_category(data[1] if len(data) > 1 else ""):
                found = False
                for key, val in assignee_id_map.items():
                    if val == data[8]:
                        assignee_cbo.set(key)
                        found = True
                        break
                if not found:
                    assignee_cbo.set(self._tr("General/None"))
            else:
                assignee_cbo.set(self._tr("General/None"))

            lbl_amt = tb.Label(form, text=self._tr("Amount:"), font=("Segoe UI", 10, "bold"))
            amt_ent = tb.Entry(form, width=30)
            if data and str(data[2] if len(data) > 2 else "").strip() not in ("", "None"):
                try:
                    amt_ent.insert(0, f"{to_float(data[2], 0.0):.2f}")
                except Exception:
                    amt_ent.insert(0, "")

            lbl_status = tb.Label(form, text=self._tr("Status:"), font=("Segoe UI", 10, "bold"))
            status_cbo = tb.Combobox(form, width=28, state="readonly", values=["Pending", "Approved", "Rejected"])
            if data:
                status_cbo.set(data[4])
            else:
                status_cbo.set("Pending")

            lbl_pay = tb.Label(form, text=self._tr("Payment Type:"), font=("Segoe UI", 10, "bold"))
            pts = self.get_db_payments()
            pay_type_cbo = tb.Combobox(form, width=28, state="readonly", values=pts)
            if data and len(data) > 6 and data[6]:
                pay_type_cbo.set(data[6])
            else:
                pay_type_cbo.set("Cash" if "Cash" in pts else (pts[0] if pts else ""))

            lbl_loc = tb.Label(form, text=self._tr("Location (Optional):"), font=("Segoe UI", 10, "bold"))
            locs = [""] + self.get_db_locations()
            loc_cbo = tb.Combobox(form, width=28, state="readonly", values=locs)
            if data and len(data) > 7 and data[7]:
                loc_cbo.set(data[7])
            else:
                loc_cbo.set("")

            lbl_desc = tb.Label(form, text=self._tr("Description:"), font=("Segoe UI", 10, "bold"))
            desc_ent = tb.Entry(form, width=30)
            if data:
                desc_ent.insert(0, data[5] if data[5] else "")

            lbl_doc = tb.Label(form, text=self._tr("Document / Receipt:"), font=("Segoe UI", 10, "bold"))
            doc_outer = tb.Frame(form)
            form.columnconfigure(1, weight=1)

            doc_list_holder = tb.Frame(doc_outer)
            doc_list_holder.pack(fill=X, expand=True)
            doc_paths = parse_expense_documents(data[9] if data and len(data) > 9 else None)
            removed_docs = []

            def refresh_doc_list():
                for child in doc_list_holder.winfo_children():
                    child.destroy()
                if not doc_paths:
                    empty = tb.Label(
                        doc_list_holder,
                        text=self._tr("No document attached"),
                        font=("Segoe UI", 10),
                        bootstyle="secondary",
                    )
                    empty.pack(anchor=W)
                    return
                for idx, path in enumerate(list(doc_paths)):
                    row = tb.Frame(doc_list_holder)
                    row.pack(fill=X, pady=2)
                    name = os.path.basename(path)
                    exists = bool(os.path.isfile(path) or os.path.isfile(resolve_local_doc_path(path)) or get_db_mode() == "supabase")
                    link = tb.Label(
                        row,
                        text=f"📄 {name}" + ("" if exists else f" ({self._tr('missing')})"),
                        font=("Segoe UI", 10, "underline"),
                        bootstyle="info" if exists else "danger",
                        cursor="hand2",
                    )
                    link.pack(side=LEFT, fill=X, expand=True)
                    link.bind("<Button-1>", lambda e, p=path: self.preview_expense_document(p, parent=dialog))

                    def remove_one(i=idx, p=path):
                        try:
                            if i < len(doc_paths) and doc_paths[i] == p:
                                doc_paths.pop(i)
                            elif p in doc_paths:
                                doc_paths.remove(p)
                        except Exception:
                            pass
                        removed_docs.append(p)
                        refresh_doc_list()

                    tb.Button(
                        row,
                        text="✕",
                        bootstyle="danger",
                        width=2,
                        cursor="hand2",
                        command=remove_one,
                    ).pack(side=RIGHT, padx=(6, 0))

            def browse_documents():
                filepaths = filedialog.askopenfilenames(
                    parent=dialog,
                    title=self._tr("Upload Document"),
                    filetypes=[
                        ("Documents & Images", "*.pdf *.png *.jpg *.jpeg *.gif *.bmp *.webp *.tif *.tiff"),
                        ("PDF", "*.pdf"),
                        ("Images", "*.png *.jpg *.jpeg *.gif *.bmp *.webp *.tif *.tiff"),
                        ("All Files", "*.*"),
                    ],
                )
                if filepaths:
                    for fp in filepaths:
                        if fp and fp not in doc_paths:
                            doc_paths.append(fp)
                    refresh_doc_list()

            refresh_doc_list()

            doc_btn_row = tb.Frame(form)
            tb.Button(doc_btn_row, text=self._tr("Add files"), bootstyle="info outline", command=browse_documents).pack(side=LEFT, padx=(0, 6))
            tb.Label(doc_btn_row, text=self._tr("Click filename to preview"), font=("Segoe UI", 8), bootstyle="secondary").pack(side=LEFT, padx=8)

            def layout_conditional_fields(event=None):
                cat = cat_cbo.get()
                extra = 0
                if cat == "Salary Payment":
                    lbl_tip.grid(row=4, column=0, **pad)
                    tip_chk.grid(row=4, column=1, sticky=W, padx=12, pady=7)
                    lbl_assignee.grid_remove()
                    assignee_cbo.grid_remove()
                    lbl_emp.grid(row=2, column=0, **pad)
                    emp_cbo.grid(row=2, column=1, **ent_pad)
                    if tip_var.get():
                        lbl_tip_given.grid(row=5, column=0, **pad)
                        tip_given_ent.grid(row=5, column=1, **ent_pad)
                        extra = 1
                    else:
                        lbl_tip_given.grid_remove()
                        tip_given_ent.grid_remove()
                    tip_given_hint.grid_remove()
                elif cat == "Cash Envelope Received" or is_envelope_category(cat):
                    lbl_tip.grid_remove()
                    tip_chk.grid_remove()
                    lbl_tip_given.grid_remove()
                    tip_given_ent.grid_remove()
                    tip_given_hint.grid_remove()
                    lbl_assignee.grid(row=4, column=0, **pad)
                    assignee_cbo.grid(row=4, column=1, **ent_pad)
                    lbl_emp.grid_remove()
                    emp_cbo.grid_remove()
                else:
                    lbl_tip.grid_remove()
                    tip_chk.grid_remove()
                    lbl_tip_given.grid_remove()
                    tip_given_ent.grid_remove()
                    tip_given_hint.grid_remove()
                    lbl_assignee.grid_remove()
                    assignee_cbo.grid_remove()
                    lbl_emp.grid(row=2, column=0, **pad)
                    emp_cbo.grid(row=2, column=1, **ent_pad)

                r = 5 + extra
                lbl_amt.grid(row=r, column=0, **pad)
                amt_ent.grid(row=r, column=1, **ent_pad)
                lbl_status.grid(row=r + 1, column=0, **pad)
                status_cbo.grid(row=r + 1, column=1, **ent_pad)
                lbl_pay.grid(row=r + 2, column=0, **pad)
                pay_type_cbo.grid(row=r + 2, column=1, **ent_pad)
                lbl_loc.grid(row=r + 3, column=0, **pad)
                loc_cbo.grid(row=r + 3, column=1, **ent_pad)
                lbl_desc.grid(row=r + 4, column=0, **pad)
                desc_ent.grid(row=r + 4, column=1, **ent_pad)
                lbl_doc.grid(row=r + 5, column=0, sticky=NE, padx=12, pady=7)
                doc_outer.grid(row=r + 5, column=1, sticky=EW, padx=12, pady=7)
                doc_btn_row.grid(row=r + 6, column=0, columnspan=2, sticky=EW, padx=12, pady=(2, 10))

            tip_chk.configure(command=layout_conditional_fields)
            cat_cbo.bind("<<ComboboxSelected>>", layout_conditional_fields)
            layout_conditional_fields()
            
            def save_expense():
                if getattr(dialog, "_saving", False):
                    return
                dt = normalize_iso_date(date_ent.entry.get())
                if not dt and data:
                    dt = normalize_iso_date(data[0] if data else "")
                if not dt:
                    messagebox.showerror("Error", "Date is required. Use YYYY-MM-DD.", parent=dialog)
                    return

                try:
                    raw_amt = amt_ent.get().strip()
                    if not raw_amt:
                        messagebox.showerror("Error", self._tr("Amount is required."), parent=dialog)
                        return
                    amount = float(str(raw_amt).replace(",", "").replace("$", "").strip())
                except ValueError:
                    messagebox.showerror("Error", "Amount must be numeric.", parent=dialog)
                    return
                    
                if amount < 0:
                    messagebox.showerror("Error", "Amount cannot be negative.", parent=dialog)
                    return
                    
                category = cat_cbo.get()
                if data and is_envelope_category(data[1] if len(data) > 1 else ""):
                    category = "Cash Envelope Received"
                elif is_envelope_category(category):
                    category = "Cash Envelope Received"
                if category == "Cash Envelope Received" and self.is_date_in_locked_cash_month(dt):
                    messagebox.showerror(
                        "Locked",
                        "This cash calendar month is locked. Unlock it from Cash Calendar before saving envelopes.",
                        parent=dialog,
                    )
                    return

                dialog._saving = True
                try:
                    save_btn.config(state="disabled")
                except Exception:
                    pass
                try:
                    saving_lbl.pack(side=LEFT, padx=12)
                except Exception:
                    pass

                try:
                    if category == "Cash Envelope Received":
                        emp_id = self.get_shop_employee_id()
                        assignee_val = assignee_cbo.get()
                        assignee_id = assignee_id_map.get(assignee_val)
                        is_tip = "No"
                        tip_given = 0.0
                    else:
                        emp_sel = emp_cbo.get()
                        emp_id = emp_id_map.get(emp_sel)
                        assignee_id = None
                        is_tip = "No"
                        tip_given = 0.0
                        if category == "Salary Payment" and tip_var.get():
                            is_tip = "Yes"
                            raw_tip = (tip_given_ent.get() or "").strip()
                            if not raw_tip:
                                messagebox.showerror(
                                    "Error",
                                    self._tr("Enter how much tip you gave the employee."),
                                    parent=dialog,
                                )
                                dialog._saving = False
                                try:
                                    save_btn.config(state="normal")
                                    saving_lbl.pack_forget()
                                except Exception:
                                    pass
                                return
                            try:
                                tip_given = float(str(raw_tip).replace(",", "").replace("$", "").strip())
                            except ValueError:
                                messagebox.showerror("Error", "Amount must be numeric.", parent=dialog)
                                dialog._saving = False
                                try:
                                    save_btn.config(state="normal")
                                    saving_lbl.pack_forget()
                                except Exception:
                                    pass
                                return
                            if tip_given < 0:
                                messagebox.showerror("Error", "Amount cannot be negative.", parent=dialog)
                                dialog._saving = False
                                try:
                                    save_btn.config(state="normal")
                                    saving_lbl.pack_forget()
                                except Exception:
                                    pass
                                return
                        
                    status = status_cbo.get() or "Pending"
                    description = desc_ent.get().strip()
                    pay_type = pay_type_cbo.get()
                    location = loc_cbo.get()
                    # Resolve cycle_key from combobox selection with fallback to date
                    cycle_val = str(cycle_cbo.get() or "").strip()
                    cycle_key = cycle_key_map.get(cycle_val)
                    if not cycle_key and cycle_val:
                        if parse_cycle_key(cycle_val):
                            cycle_key = cycle_val
                        else:
                            for yr in (base_yr - 1, base_yr, base_yr + 1):
                                for ck in cycles_for_year(yr):
                                    if cycle_label_with_year(ck) == cycle_val or cycle_label(ck) == cycle_val or ck in cycle_val:
                                        cycle_key = ck
                                        break
                                if cycle_key:
                                    break
                    if not cycle_key:
                        cycle_key = cycle_for_date(dt)
                    
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    saved_id = expense_id
                    if expense_id:
                        cursor.execute('''
                            UPDATE expenses 
                            SET expense_date=?, category=?, amount=?, employee_id=?, status=?, description=?, payment_type=?, location=?, is_tip=?, assignee_id=?, tip_given=?, cycle_key=?
                            WHERE id=?
                        ''', (dt, category, amount, emp_id, status, description, pay_type, location, is_tip, assignee_id, tip_given, cycle_key, expense_id))
                    else:
                        cursor.execute('''
                            INSERT INTO expenses (expense_date, category, amount, employee_id, status, description, payment_type, location, is_tip, assignee_id, tip_given, cycle_key)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (dt, category, amount, emp_id, status, description, pay_type, location, is_tip, assignee_id, tip_given, cycle_key))
                        saved_id = cursor.lastrowid
                    protect_local_expense_id(saved_id)

                    if doc_paths or removed_docs:
                        stored_paths = store_expense_documents(list(doc_paths), saved_id)
                        cursor.execute(
                            "UPDATE expenses SET document_path=? WHERE id=?",
                            (serialize_expense_documents(stored_paths), saved_id),
                        )
                        for p in removed_docs:
                            if p not in stored_paths:
                                delete_expense_document_file(p)

                    commit_and_save(conn)
                    conn.close()
                    try:
                        schedule_cloud_push(0.05)
                    except Exception:
                        pass
                    self._last_sync_fingerprint = None
                    cb = on_save_callback
                    self._safe_grab_release(dialog)
                    dialog.destroy()
                    if self._widget_alive(parent_win):
                        try:
                            parent_win.lift()
                            parent_win.focus_set()
                        except Exception:
                            pass
                    if self._widget_alive(getattr(self, "expenses_win", None)):
                        self.load_expenses_data(quiet=True)
                    try:
                        if hasattr(self, "load_financials_data"):
                            self.load_financials_data(quiet=True)
                    except Exception:
                        pass
                    if cb:
                        cb()
                except Exception as e:
                    dialog._saving = False
                    try:
                        save_btn.config(state="normal")
                        saving_lbl.pack_forget()
                    except Exception:
                        pass
                    messagebox.showerror("Error", str(e), parent=dialog)
                
            def cancel_expense():
                self._safe_grab_release(dialog)
                try:
                    dialog.destroy()
                except Exception:
                    pass
                if self._widget_alive(parent_win):
                    try:
                        parent_win.lift()
                        parent_win.focus_set()
                    except Exception:
                        pass

            btns = tb.Frame(footer)
            btns.pack()
            save_btn = tb.Button(btns, text=self._tr("Save"), bootstyle="success", command=save_expense)
            save_btn.pack(side=LEFT, ipadx=40, ipady=8)
            tb.Button(btns, text=self._tr("Cancel"), bootstyle="secondary", command=cancel_expense).pack(
                side=LEFT, padx=12, ipadx=24, ipady=8
            )
            saving_lbl = tb.Label(footer, text=self._tr("Saving…") + "  " + self._tr("Please wait, do not click…"), font=("Segoe UI", 10, "bold"), bootstyle="info")
            self._bind_dialog_save_keys(dialog, save_expense)
            dialog.bind("<Escape>", lambda e: cancel_expense())

        def preview_expense_document(self, filepath, parent=None):
            """Show an in-app preview for images, or open PDF/other files for preview."""
            parent = parent or self
            path = ensure_document_file_available(filepath)
            if not path or not os.path.isfile(path):
                messagebox.showerror("Error", "Document file was not found.", parent=parent)
                return

            ext = os.path.splitext(path)[1].lower()
            image_exts = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tif", ".tiff"}

            def do_print():
                try:
                    print_path_with_default_app(path)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not print file:\n{e}", parent=parent)

            def open_external():
                try:
                    open_path_with_default_app(path)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=parent)

            if ext in image_exts:
                preview = tb.Toplevel(parent)
                preview.title(f"{self._tr('Document Preview')}: {os.path.basename(path)}")
                preview.transient(parent)
                preview.grab_set()
                preview.focus_set()
                preview.geometry("860x700")

                tb.Label(preview, text=os.path.basename(path), font=("Segoe UI", 12, "bold")).pack(pady=(12, 6))
                canvas_holder = tb.Frame(preview)
                canvas_holder.pack(fill=BOTH, expand=True, padx=10, pady=5)
                canvas = tk.Canvas(canvas_holder, highlightthickness=0, bg="#1e1e1e")
                scroll_y = tb.Scrollbar(canvas_holder, orient=VERTICAL, command=canvas.yview)
                scroll_x = tb.Scrollbar(canvas_holder, orient=HORIZONTAL, command=canvas.xview)
                canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
                scroll_y.pack(side=RIGHT, fill=Y)
                scroll_x.pack(side=BOTTOM, fill=X)
                canvas.pack(side=LEFT, fill=BOTH, expand=True)

                photo = None
                try:
                    from PIL import Image, ImageTk
                    img = Image.open(path)
                    max_w, max_h = 820, 620
                    img.thumbnail((max_w, max_h), Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                except Exception:
                    try:
                        photo = tk.PhotoImage(file=path)
                    except Exception:
                        photo = None

                if photo is not None:
                    preview._preview_img = photo  # keep reference
                    canvas.create_image(10, 10, anchor=NW, image=photo)
                    canvas.config(scrollregion=canvas.bbox("all"))
                else:
                    tb.Label(
                        preview,
                        text="Could not render this image in-app.\nUse Open with default app.",
                        font=("Segoe UI", 11),
                        bootstyle="warning",
                    ).pack(pady=20)

                btn_row = tb.Frame(preview)
                btn_row.pack(pady=12)
                tb.Button(btn_row, text=self._tr("Print"), bootstyle="success", command=do_print).pack(side=LEFT, padx=8)
                tb.Button(btn_row, text=self._tr("Open with default app"), bootstyle="info", command=open_external).pack(side=LEFT, padx=8)
                tb.Button(btn_row, text=self._tr("Close"), bootstyle="secondary", command=preview.destroy).pack(side=LEFT, padx=8)
                return

            # PDF / other: preview dialog that opens the file
            preview = tb.Toplevel(parent)
            preview.title(f"{self._tr('Document Preview')}: {os.path.basename(path)}")
            preview.transient(parent)
            preview.grab_set()
            preview.focus_set()
            preview.geometry("440x240")
            tb.Label(preview, text=f"📄 {os.path.basename(path)}", font=("Segoe UI", 13, "bold")).pack(pady=(25, 10))
            tb.Label(
                preview,
                text=self._tr("Click below to open a preview with your default app."),
                wraplength=360,
                bootstyle="secondary",
            ).pack(pady=5)

            btn_row = tb.Frame(preview)
            btn_row.pack(pady=20)
            tb.Button(btn_row, text=self._tr("Print"), bootstyle="success", command=do_print).pack(side=LEFT, padx=8)
            tb.Button(btn_row, text=self._tr("Open with default app"), bootstyle="primary", command=open_external).pack(side=LEFT, padx=8)
            tb.Button(btn_row, text=self._tr("Close"), bootstyle="secondary", command=preview.destroy).pack(side=LEFT, padx=8)
            # Auto-open preview for PDFs so one click feels like preview
            if ext == ".pdf":
                try:
                    open_path_with_default_app(path)
                except Exception:
                    pass

        def edit_selected_expense(self):
            selected = self.tree_expenses.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select a record to edit.", parent=self.expenses_win)
                return
                
            item_vals = self.tree_expenses.item(selected[0])['values']
            exp_id = item_vals[0]
            category = item_vals[4]
            if not exp_id:
                return
                
            if category == "Employee Revenue" or category == self._tr("Employee Revenue"):
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("PRAGMA table_info(payroll_records)")
                col_names = [col[1] for col in cursor.fetchall()]
                hr_col = "r.hour_rate" if "hour_rate" in col_names else "NULL"
                perc_col = "r.percentage" if "percentage" in col_names else "NULL"
                
                cursor.execute(f'''
                    SELECT r.record_date, r.revenue, r.service_addon_sales, r.hours, r.notes, r.written_up, e.name, r.written_up_desc,
                        COALESCE({hr_col}, e.hour_rate), COALESCE({perc_col}, e.percentage), e.use_tiered_payout, e.id
                    FROM payroll_records r
                    JOIN employees e ON r.employee_id = e.id
                    WHERE r.id=?
                ''', (exp_id,))
                rec = cursor.fetchone()
                conn.close()
                
                if rec:
                    self.open_edit_record_dialog(exp_id, rec, parent=self.expenses_win)
                return
                
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute(EXPENSE_EDIT_SELECT, (exp_id,))
            row = cursor.fetchone()
            conn.close()
            
            if row:
                self.open_expense_dialog(expense_id=exp_id, data=row)

        def delete_selected_expense(self):
            selected = self.tree_expenses.selection()
            if not selected:
                messagebox.showwarning("Select", "Please select one or more records to delete.\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.", parent=self.expenses_win)
                return

            payroll_ids = []
            expense_ids = []
            locked_blocked = 0
            for sel in selected:
                item_vals = self.tree_expenses.item(sel)['values']
                if not item_vals:
                    continue
                exp_id = item_vals[0]
                exp_date = item_vals[1] if len(item_vals) > 1 else ""
                category = item_vals[4] if len(item_vals) > 4 else ""
                if not exp_id:
                    continue
                if category == "Employee Revenue" or category == self._tr("Employee Revenue"):
                    payroll_ids.append(exp_id)
                elif (category == "Cash Envelope Received" or category == self._tr("Cash Envelope Received")) and self.is_date_in_locked_cash_month(str(exp_date)):
                    locked_blocked += 1
                else:
                    expense_ids.append(exp_id)

            if locked_blocked and not payroll_ids and not expense_ids:
                messagebox.showwarning(
                    "Locked",
                    "Selected cash envelope(s) are in a locked month. Unlock the month from Cash Calendar first.",
                    parent=self.expenses_win,
                )
                return

            total = len(payroll_ids) + len(expense_ids)
            if total == 0:
                return

            confirm_msg = (
                f"Are you sure you want to delete {total} selected record(s)? This cannot be undone."
                if total > 1
                else "Are you sure you want to delete this record? This cannot be undone."
            )
            if locked_blocked:
                confirm_msg += f"\n\nNote: {locked_blocked} locked cash envelope(s) will be skipped."
            if not messagebox.askyesno("Confirm Delete", confirm_msg, parent=self.expenses_win):
                return

            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                if payroll_ids:
                    placeholders = ",".join("?" for _ in payroll_ids)
                    cursor.execute(f"DELETE FROM payroll_records WHERE id IN ({placeholders})", payroll_ids)
                if expense_ids:
                    placeholders = ",".join("?" for _ in expense_ids)
                    try:
                        cursor.execute(
                            f"SELECT document_path FROM expenses WHERE id IN ({placeholders})",
                            expense_ids,
                        )
                        docs_root = os.path.abspath(get_expense_docs_dir())
                        for (doc_path,) in cursor.fetchall() or []:
                            for p in parse_expense_documents(doc_path):
                                delete_expense_document_file(p)
                    except Exception:
                        pass
                    cursor.execute(f"DELETE FROM expenses WHERE id IN ({placeholders})", expense_ids)
                commit_and_save(conn)
                conn.close()
                self.load_expenses_data()
                if payroll_ids:
                    self.load_calendar_data()
                messagebox.showinfo("Success", f"Deleted {total} record(s).", parent=self.expenses_win)
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete: {e}", parent=self.expenses_win)

        def export_expenses_excel(self):
            rows = self.tree_expenses.get_children()
            if not rows:
                messagebox.showinfo("Export", "No data available to export.", parent=self.expenses_win)
                return
                
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
                title="Export Expenses Data",
                parent=self.expenses_win
            )
            
            if filepath:
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Expenses"
                    headers = [self._tr(c) for c in ["Date", "Cycle", "Employee", "Category", "Payment Type", "Amount", "Status"]]
                    ws.append(headers)
                    for item in rows:
                        ws.append(list(self.tree_expenses.item(item)['values'][1:]))
                    wb.save(filepath)
                    messagebox.showinfo("Success", f"Data successfully exported to:\n{filepath}", parent=self.expenses_win)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not save Excel file.\n{e}", parent=self.expenses_win)

            self.load_calendar_data()

        def open_excel_import_dialog(self):
            parent_win = self.expenses_win if self._widget_alive(getattr(self, "expenses_win", None)) else self
            filepath = filedialog.askopenfilename(
                title=self._tr("Select Excel or Numbers Sales File"),
                filetypes=[
                    ("Spreadsheets (Excel, Numbers)", "*.xlsx *.xls *.numbers"),
                    ("Excel Files", "*.xlsx *.xls"),
                    ("Numbers Files", "*.numbers"),
                    ("All Files", "*.*")
                ],
                parent=parent_win
            )
            if not filepath:
                return
                
            is_numbers = filepath.lower().endswith('.numbers')
            headers = []
            rows = []
            period_str = ""
            location = ""
            
            if is_numbers:
                try:
                    from numbers_parser import Document
                except ImportError:
                    messagebox.showerror(
                        "Library Missing",
                        "To import Apple Numbers (.numbers) files, you must install the numbers-parser library.\n\n"
                        "Please open your terminal and run:\npip install numbers-parser",
                        parent=parent_win
                    )
                    return
                try:
                    doc = Document(filepath)
                    sheet = doc.sheets[0]
                    table = sheet.tables[0]
                    rows_list = table.rows()
                    num_rows = len(rows_list)
                    
                    # cell A4 (row 4, col 1 -> index 3, 0)
                    if num_rows > 3 and len(rows_list[3]) > 0:
                        period_str = str(rows_list[3][0].value or "").strip()
                    if not period_str:
                        period_str = "Unknown Period"
                    
                    # cell C4 (row 4, col 3 -> index 3, 2)
                    if num_rows > 3 and len(rows_list[3]) > 2:
                        location = str(rows_list[3][2].value or "").strip()
                    if not location:
                        location = "Unknown Location"
                    
                    if num_rows > 5:
                        headers = [str(cell.value).strip() for cell in rows_list[5] if cell.value is not None and str(cell.value).strip()]
                        header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in rows_list[5]]
                        for r_idx in range(6, num_rows):
                            row_cells = [cell.value for cell in rows_list[r_idx]]
                            if not any(val is not None for val in row_cells):
                                continue
                            row_data = {}
                            for c_idx, val in enumerate(row_cells):
                                if c_idx < len(header_row) and header_row[c_idx]:
                                    if hasattr(val, 'strftime'):
                                        val = val.strftime('%Y-%m-%d')
                                    elif val is not None:
                                        val = str(val).strip()
                                    else:
                                        val = ""
                                    row_data[header_row[c_idx]] = val
                            rows.append(row_data)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not read Numbers file:\n{e}", parent=parent_win)
                    return
            else:
                try:
                    import openpyxl
                except ImportError:
                    messagebox.showerror(
                        "Library Missing",
                        "To import Excel files, you must install the openpyxl library.\n\n"
                        "Please open your terminal and run:\npip install openpyxl",
                        parent=parent_win
                    )
                    return
                try:
                    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
                    sheet = wb.active
                    
                    # Fetch A4 period date string
                    cell_a4_value = str(sheet.cell(row=4, column=1).value or "").strip()
                    period_str = cell_a4_value if cell_a4_value else "Unknown Period"
                    
                    # Fetch C4 location string
                    cell_c4_value = str(sheet.cell(row=4, column=3).value or "").strip()
                    location = cell_c4_value if cell_c4_value else "Unknown Location"
                    
                    header_cells = next(sheet.iter_rows(min_row=6, max_row=6))
                    headers = [str(cell.value).strip() for cell in header_cells if cell.value is not None and str(cell.value).strip()]
                    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
                    for row_cells in sheet.iter_rows(min_row=7):
                        if not any(cell.value is not None for cell in row_cells):
                            continue
                        row_data = {}
                        for i, cell in enumerate(row_cells):
                            if i < len(header_row) and header_row[i]:
                                val = cell.value
                                if hasattr(val, 'strftime'):
                                    val = val.strftime('%Y-%m-%d')
                                elif val is not None:
                                    val = str(val).strip()
                                else:
                                    val = ""
                                row_data[header_row[i]] = val
                        rows.append(row_data)
                except KeyError as ke:
                    if 'content_types' in str(ke).lower() or '[content_types].xml' in str(ke).lower():
                        messagebox.showerror(
                            "Unsupported File Format",
                            "The selected file has a format that cannot be read as an Excel (.xlsx) archive.\n\n"
                            "This usually happens if:\n"
                            "1. It is an Apple Numbers (.numbers) file.\n"
                            "2. It is a legacy Excel (.xls) file.\n\n"
                            "Please open the file and export/save it as a standard 'Excel Workbook (.xlsx)' format first, then import it.",
                            parent=parent_win
                        )
                    else:
                        messagebox.showerror("Error", f"Could not read Excel file:\n{ke}", parent=parent_win)
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Could not read Excel file:\n{e}", parent=parent_win)
                    return
                    
            if not headers:
                messagebox.showerror("Error", "Selected file has no headers on row 6.", parent=parent_win)
                return

            # Detect initial cycle from period_str or current date
            detected_cycle = None
            if period_str and period_str != "Unknown Period":
                p_dates = parse_period_dates(period_str)
                if p_dates:
                    detected_cycle = cycle_for_date(p_dates[0])
                if not detected_cycle:
                    detected_cycle = cycle_for_date(period_str)
            if not detected_cycle:
                detected_cycle = cycle_for_date(datetime.today().strftime('%Y-%m-%d'))

            cycle_choices, cycle_key_by_label, active_cycle_lbl = get_formatted_cycle_choices(detected_cycle, start_from_june_2026=True)

            dialog = tb.Toplevel(parent_win)
            dialog.title(self._tr("Import Sales Data"))
            dialog.transient(parent_win)
            dialog.grab_set()
            dialog.focus_set()
            try:
                dialog.update_idletasks()
                screen_w = dialog.winfo_screenwidth()
                screen_h = dialog.winfo_screenheight()
                dlg_w = min(640, max(520, screen_w - 40))
                dlg_h = min(700, max(540, screen_h - 90))
                x = max(0, (screen_w - dlg_w) // 2)
                y = max(10, min(25, (screen_h - dlg_h) // 4))
                dialog.geometry(f"{dlg_w}x{dlg_h}+{x}+{y}")
            except Exception:
                pass
            self._present_window(dialog)
            
            pad = {'padx': 10, 'pady': 5, 'sticky': E}
            ent_pad = {'padx': 10, 'pady': 5, 'sticky': W}
            
            tb.Label(dialog, text=self._tr("Import Sales Data"), font=("Segoe UI", 12, "bold"), bootstyle="primary").pack(pady=(12, 6), padx=15, anchor=W)

            # Step 1: Period and Target Cycle selection
            step1_lf = tb.Labelframe(dialog, text=self._tr("1. Select Target Pay Period & Cycle"), padding=(12, 6), bootstyle="info")
            step1_lf.pack(fill=X, padx=15, pady=(0, 8))

            tb.Label(step1_lf, text=self._tr("Detected Period:"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0, **pad)
            tb.Label(step1_lf, text=period_str if period_str else self._tr("None"), font=("Segoe UI", 10), bootstyle="secondary").grid(row=0, column=1, **ent_pad)

            tb.Label(step1_lf, text=self._tr("Location:"), font=("Segoe UI", 10, "bold")).grid(row=1, column=0, **pad)
            tb.Label(step1_lf, text=location if location else self._tr("None"), font=("Segoe UI", 10), bootstyle="secondary").grid(row=1, column=1, **ent_pad)

            tb.Label(step1_lf, text=self._tr("Target Cycle:"), font=("Segoe UI", 10, "bold"), bootstyle="warning").grid(row=2, column=0, **pad)
            cb_cycle = tb.Combobox(step1_lf, values=cycle_choices, width=38, state="readonly", bootstyle="warning")
            cb_cycle.set(active_cycle_lbl)
            cb_cycle.grid(row=2, column=1, **ent_pad)

            # Step 2: Match Columns
            step2_lf = tb.Labelframe(dialog, text=self._tr("2. Match Column Headers"), padding=(12, 6), bootstyle="secondary")
            step2_lf.pack(fill=X, padx=15, pady=(0, 10))

            tb.Label(step2_lf, text=self._tr("Employee Column:"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0, **pad)
            cb_emp = tb.Combobox(step2_lf, values=headers, width=32, state="readonly")
            cb_emp.grid(row=0, column=1, **ent_pad)
            
            tb.Label(step2_lf, text=self._tr("Service Sales Column:"), font=("Segoe UI", 10, "bold")).grid(row=1, column=0, **pad)
            cb_rev = tb.Combobox(step2_lf, values=headers, width=32, state="readonly")
            cb_rev.grid(row=1, column=1, **ent_pad)

            tb.Label(step2_lf, text=self._tr("Service Add-on Sales:"), font=("Segoe UI", 10, "bold")).grid(row=2, column=0, **pad)
            cb_addon = tb.Combobox(step2_lf, values=["(None)"] + headers, width=32, state="readonly")
            cb_addon.grid(row=2, column=1, **ent_pad)
            
            tb.Label(step2_lf, text=self._tr("Product Sales Column:"), font=("Segoe UI", 10, "bold")).grid(row=3, column=0, **pad)
            cb_prod = tb.Combobox(step2_lf, values=headers, width=32, state="readonly")
            cb_prod.grid(row=3, column=1, **ent_pad)
            
            tb.Label(step2_lf, text=self._tr("Tip Column:"), font=("Segoe UI", 10, "bold")).grid(row=4, column=0, **pad)
            cb_tip = tb.Combobox(step2_lf, values=headers, width=32, state="readonly")
            cb_tip.grid(row=4, column=1, **ent_pad)
            
            def _norm_header(text):
                return " ".join("".join(ch.lower() if ch.isalnum() else " " for ch in str(text or "")).split())

            emp_exact = {"service provider", "employee", "employee name", "name", "staff", "provider"}
            rev_exact = {"services sales", "service sales", "services sale", "service sale"}
            addon_exact = {"service add on sales", "service addon sales", "service add-on sales", "addon sales", "add on sales"}
            prod_exact = {"product sales", "products sales", "product sale"}
            tip_exact = {"tip", "tips", "gratuity"}

            emp_keywords = ["service provider", "employee", "staff", "provider"]
            rev_keywords = ["services sales", "service sales"]
            addon_keywords = ["service add-on", "service addon", "addon sales", "add-on sales", "add on sales"]
            prod_keywords = ["product sales", "product sale"]
            tip_keywords = ["tip", "tips", "gratuity"]

            def _pick(exact_set, keywords, already=None):
                for h in headers:
                    if already and h in already:
                        continue
                    if _norm_header(h) in exact_set:
                        return h
                for h in headers:
                    if already and h in already:
                        continue
                    hl = _norm_header(h)
                    if any(kw in hl for kw in keywords):
                        return h
                return ""

            picked_emp = _pick(emp_exact, emp_keywords)
            used = {picked_emp} if picked_emp else set()
            picked_rev = _pick(rev_exact, rev_keywords, used)
            if picked_rev:
                used.add(picked_rev)
            picked_addon = _pick(addon_exact, addon_keywords, used)
            if picked_addon:
                used.add(picked_addon)
            picked_prod = _pick(prod_exact, prod_keywords, used)
            if picked_prod:
                used.add(picked_prod)
            picked_tip = _pick(tip_exact, tip_keywords, used)

            if picked_emp:
                cb_emp.set(picked_emp)
            if picked_rev:
                cb_rev.set(picked_rev)
            if picked_addon:
                cb_addon.set(picked_addon)
            else:
                cb_addon.set("(None)")
            if picked_prod:
                cb_prod.set(picked_prod)
            if picked_tip:
                cb_tip.set(picked_tip)
            
            if not cb_emp.get() and headers:
                cb_emp.set(headers[0])
            if not cb_addon.get():
                cb_addon.set("(None)")
            
            def execute_import():
                col_emp = cb_emp.get()
                col_rev = cb_rev.get()
                col_addon = cb_addon.get()
                col_prod = cb_prod.get()
                col_tip = cb_tip.get()
                
                if not col_emp or not col_rev or not col_prod or not col_tip:
                    messagebox.showerror("Error", "All required column mappings are required.", parent=dialog)
                    return
                
                mapped_cols = [c for c in [col_emp, col_rev, col_addon, col_prod, col_tip] if c and c != "(None)"]
                if len(mapped_cols) < len(set(mapped_cols)):
                    messagebox.showerror("Error", "Each mapping must represent a unique column.", parent=dialog)
                    return

                # Resolve chosen cycle
                selected_cycle_key = cycle_key_by_label.get(cb_cycle.get()) or detected_cycle
                c_bounds = cycle_bounds(selected_cycle_key)
                cycle_start_date = c_bounds[0] if c_bounds else datetime.today().strftime('%Y-%m-%d')

                # Check for overlapping imports in this cycle and location
                try:
                    conn_chk = sqlite3.connect(TEMP_DB_PATH)
                    cur_chk = conn_chk.cursor()
                    cur_chk.execute(
                        "SELECT COUNT(*) FROM payroll_records WHERE cycle_key = ? AND LOWER(COALESCE(location, '')) = LOWER(?)",
                        (selected_cycle_key, location)
                    )
                    count_ex = (cur_chk.fetchone() or [0])[0]
                    conn_chk.close()
                    
                    if count_ex > 0:
                        if messagebox.askyesno(
                            self._tr("Overlapping Import Warning"),
                            f"Found {count_ex} existing payroll record(s) for location '{location}' in '{cycle_label_with_year(selected_cycle_key)}'.\n\n"
                            "Importing again may create duplicate entries for this cycle.\n\n"
                            "Do you want to continue anyway?",
                            parent=dialog
                        ) is False:
                            return
                except Exception:
                    pass

                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                
                imported_records = 0
                new_employees = 0
                
                try:
                    for row in rows:
                        emp_name = row.get(col_emp, "").strip()
                        raw_rev = row.get(col_rev, "").replace("$", "").replace(",", "").strip()
                        raw_addon = ""
                        if col_addon and col_addon != "(None)":
                            raw_addon = row.get(col_addon, "").replace("$", "").replace(",", "").strip()
                        raw_prod = row.get(col_prod, "").replace("$", "").replace(",", "").strip()
                        raw_tip = row.get(col_tip, "").replace("$", "").replace(",", "").strip()
                        
                        if not emp_name or not raw_rev:
                            continue
                            
                        if "total" in emp_name.lower():
                            continue
                            
                        try:
                            revenue = float(raw_rev)
                            addon_sales = float(raw_addon) if raw_addon else 0.0
                            product_sales = float(raw_prod) if raw_prod else 0.0
                            tip = float(raw_tip) if raw_tip else 0.0
                        except ValueError:
                            continue
                            
                        cursor.execute("SELECT id, percentage FROM employees WHERE name = ?", (emp_name,))
                        emp_row = cursor.fetchone()
                        
                        if emp_row:
                            emp_id, perc = emp_row
                        else:
                            parts = emp_name.split(" ", 1)
                            first_name = parts[0]
                            last_name = parts[1] if len(parts) > 1 else ""
                            cursor.execute('''
                                INSERT INTO employees (name, first_name, last_name, hour_rate, percentage)
                                VALUES (?, ?, ?, 0.0, 0.0)
                            ''', (emp_name, first_name, last_name))
                            emp_id = cursor.lastrowid
                            perc = 0.0
                            new_employees += 1
                            
                            os.makedirs(os.path.join(EMPLOYEE_FOLDERS_DIR, f"{first_name}_{last_name}_{emp_id}".replace(" ", "_")), exist_ok=True)
                            
                        # Service Add-on Sales split calculation includes both Service Revenue AND Service Add-on Sales!
                        pay = round((revenue + addon_sales) * perc, 2) if perc > 0 else 0.0
                        note_text = f"Excel Import (Cycle: {cycle_label(selected_cycle_key)}) (Period: {period_str}) | Loc: {location} | Addon: ${addon_sales:.2f} | Prod Sales: ${product_sales:.2f} | Tips: ${tip:.2f}"
                        cursor.execute('''
                            INSERT INTO payroll_records (employee_id, record_date, payment_amount, payment_type, revenue, service_addon_sales, hours, calculation, notes, written_up, location, tip, product_sales, cycle_key)
                            VALUES (?, ?, NULL, NULL, ?, ?, 0.0, ?, ?, 'No', ?, ?, ?, ?)
                        ''', (emp_id, cycle_start_date, revenue, addon_sales, pay, note_text, location, tip, product_sales, selected_cycle_key))
                        
                        imported_records += 1
                        
                    commit_and_save(conn)
                except Exception as ex:
                    conn.close()
                    messagebox.showerror("Database error", f"Import aborted due to error:\n{ex}", parent=dialog)
                    return
                    
                conn.close()
                
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM employees WHERE hour_rate = 0.0 AND percentage = 0.0")
                missing_employees = [r[0] for r in cursor.fetchall()]
                conn.close()
                
                self.refresh_expense_filter_employees()
                self.load_expenses_data()
                self.load_employees()
                self.load_calendar_data()
                try:
                    if hasattr(self, "load_financials_data"):
                        self.load_financials_data(quiet=True)
                except Exception:
                    pass
                
                messagebox.showinfo(
                    "Import Complete",
                    f"Successfully imported {imported_records} rows into '{cycle_label_with_year(selected_cycle_key)}'.\n"
                    f"Registered {new_employees} new employees.",
                    parent=dialog
                )
                
                if missing_employees:
                    missing_names_str = ", ".join(missing_employees)
                    messagebox.showwarning(
                        "Config Rates Missing",
                        f"The following employees lack configured Hour Rates or Percentages:\n\n{missing_names_str}\n\n"
                        f"We are redirecting you to configure them.",
                        parent=dialog
                    )
                    self.notebook.select(self.tab_names)
                    
                dialog.destroy()
                
            tb.Button(dialog, text=self._tr("Import Data"), bootstyle="success", command=execute_import).pack(pady=(5, 15), ipadx=25, ipady=6)

        def get_employee_payout_details(self, employee_id, from_date, to_date):
            conn = sqlite3.connect(TEMP_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT use_tiered_payout, percentage, hour_rate FROM employees WHERE id = ?", (employee_id,))
            row = cursor.fetchone()
            if not row:
                conn.close()
                return False, 0.0, 0.0, 0.0
                
            use_tiered, default_perc, hour_rate = row
            use_tiered = bool(use_tiered)
            default_perc = to_float(default_perc, 0.0)
            hour_rate = to_float(hour_rate, 0.0)
            
            if not use_tiered:
                conn.close()
                return False, default_perc, 0.0, hour_rate
                
            # Calculate totals for tiered layout (client-side — values are encrypted at rest)
            cursor.execute('''
                SELECT revenue, service_addon_sales, product_sales
                FROM payroll_records
                WHERE employee_id = ? AND record_date >= ? AND record_date <= ?
            ''', (employee_id, from_date, to_date))
            rows = cursor.fetchall() or []
            conn.close()

            total_rev = 0.0
            total_addon = 0.0
            total_prod = 0.0
            for r in rows:
                total_rev += to_float(r[0], 0.0)
                total_addon += to_float(r[1], 0.0)
                total_prod += to_float(r[2], 0.0)
            
            total_service = total_rev + total_addon
            service_perc = service_percent_for_sales(total_service)
            product_perc = product_percent_for_sales(total_prod)
            return True, service_perc, product_perc, hour_rate

        def setup_financials_tab(self):
            container = tb.Frame(self.tab_financials, padding=10)
            container.pack(fill=BOTH, expand=True)
            
            ctrl_lf = tb.Labelframe(container, text=self._tr("Dashboard Filters"), padding=10, bootstyle="info")
            ctrl_lf.pack(side=TOP, fill=X, pady=(0, 10))
            
            # Create a control grid/sections inside ctrl_lf
            filter_grid = tb.Frame(ctrl_lf)
            filter_grid.pack(fill=X, expand=True)
            
            # Row 1 of filters
            row1 = tb.Frame(filter_grid)
            row1.pack(fill=X, pady=5)
            
            tb.Label(row1, text=self._tr("From:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(5, 5))
            self.fin_from_date = tb.DateEntry(row1, bootstyle="primary", dateformat='%Y-%m-%d')
            self.fin_from_date.pack(side=LEFT, padx=5)
            
            tb.Label(row1, text=self._tr("To:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.fin_to_date = tb.DateEntry(row1, bootstyle="primary", dateformat='%Y-%m-%d')
            self.fin_to_date.pack(side=LEFT, padx=5)
            
            self.fin_inc_service = tk.BooleanVar(value=True)
            self.fin_inc_addon = tk.BooleanVar(value=True)
            self.fin_inc_product = tk.BooleanVar(value=True)
            self.fin_inc_tips = tk.BooleanVar(value=True)
            self.fin_inc_other = tk.BooleanVar(value=True)
            self.fin_inc_expenses = tk.BooleanVar(value=True)
            
            tb.Button(row1, text=self._tr("Today"), bootstyle="outline-secondary", command=lambda: self.set_fin_period("today")).pack(side=LEFT, padx=(20, 5))
            tb.Button(row1, text=self._tr("This Week"), bootstyle="outline-secondary", command=lambda: self.set_fin_period("week")).pack(side=LEFT, padx=5)
            tb.Button(row1, text=self._tr("This Month"), bootstyle="outline-secondary", command=lambda: self.set_fin_period("month")).pack(side=LEFT, padx=5)
            tb.Button(row1, text=self._tr("All Time"), bootstyle="outline-secondary", command=lambda: self.set_fin_period("all")).pack(side=LEFT, padx=5)
            
            tb.Label(row1, text=self._tr("Employee:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.fin_emp_filter = tb.Combobox(row1, width=18, state="readonly")
            self.fin_emp_filter.pack(side=LEFT, padx=5)
            self.fin_emp_filter.bind("<<ComboboxSelected>>", self.on_fin_emp_filter_changed)
            tb.Button(row1, text=self._tr("Clear"), bootstyle="secondary outline", command=self.clear_fin_employee_filter).pack(side=LEFT, padx=5)

            # Exclude dropdown with checkbox options (default: No exclude)
            tb.Label(row1, text=self._tr("Exclude:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(15, 5))
            self.fin_exclude_vars = {}
            self.fin_exclude_summary = tk.StringVar(value=self._tr("No exclude"))
            self.btn_fin_exclude = tb.Menubutton(
                row1,
                textvariable=self.fin_exclude_summary,
                bootstyle="info-outline",
                width=18,
                cursor="hand2",
            )
            self.btn_fin_exclude.pack(side=LEFT, padx=5)
            self.fin_exclude_menu = tk.Menu(self.btn_fin_exclude, tearoff=0)
            self.btn_fin_exclude.configure(menu=self.fin_exclude_menu)

            # Row 2 of filters: Specific include/exclude checkbuttons!
            row2 = tb.Frame(filter_grid)
            row2.pack(fill=X, pady=5)
            
            tb.Label(row2, text=self._tr("Include:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(5, 10))
            
            tb.Checkbutton(row2, text=self._tr("Service Sales"), variable=self.fin_inc_service, bootstyle="success", command=self.schedule_financials_refresh).pack(side=LEFT, padx=10)
            tb.Checkbutton(row2, text=self._tr("Service Add-on Sales"), variable=self.fin_inc_addon, bootstyle="success", command=self.schedule_financials_refresh).pack(side=LEFT, padx=10)
            tb.Checkbutton(row2, text=self._tr("Product Sales"), variable=self.fin_inc_product, bootstyle="success", command=self.schedule_financials_refresh).pack(side=LEFT, padx=10)
            tb.Checkbutton(row2, text=self._tr("Tips"), variable=self.fin_inc_tips, bootstyle="success", command=self.schedule_financials_refresh).pack(side=LEFT, padx=10)
            tb.Checkbutton(row2, text=self._tr("Other Earnings / Envelopes"), variable=self.fin_inc_other, bootstyle="success", command=self.schedule_financials_refresh).pack(side=LEFT, padx=10)
            tb.Checkbutton(
                row2,
                text=self._tr("Expenses"),
                variable=self.fin_inc_expenses,
                bootstyle="danger",
                command=self._on_fin_expenses_master_changed,
            ).pack(side=LEFT, padx=(10, 2))

            # Expense-category include dropdown (checked = included)
            self.fin_exp_cat_vars = {}
            self.fin_exp_select_all = tk.BooleanVar(value=True)
            self.fin_exp_cat_summary = tk.StringVar(value=self._tr("Select All"))
            self.btn_fin_exp_cats = tb.Menubutton(
                row2,
                textvariable=self.fin_exp_cat_summary,
                bootstyle="danger-outline",
                width=16,
                cursor="hand2",
            )
            self.btn_fin_exp_cats.pack(side=LEFT, padx=(0, 10))
            self.fin_exp_cat_menu = tk.Menu(self.btn_fin_exp_cats, tearoff=0)
            self.btn_fin_exp_cats.configure(menu=self.fin_exp_cat_menu)
            self.rebuild_fin_expense_cat_menu()
            
            cards_frame = tb.Frame(container)
            cards_frame.pack(side=TOP, fill=X, pady=(0, 15))
            
            self.card_earning = tb.Label(cards_frame, text=f"{self._tr('Total Earnings')}: $0.00", font=("Segoe UI", 13, "bold"), bootstyle="inverse-success", padding=10)
            self.card_earning.pack(side=LEFT, expand=True, fill=X, padx=5)
            
            self.card_expense = tb.Label(cards_frame, text=f"{self._tr('Total Expenses')}: $0.00", font=("Segoe UI", 13, "bold"), bootstyle="inverse-danger", padding=10)
            self.card_expense.pack(side=LEFT, expand=True, fill=X, padx=5)
            
            self.card_net = tb.Label(cards_frame, text=f"{self._tr('Net Profit')}: $0.00", font=("Segoe UI", 13, "bold"), bootstyle="inverse-info", padding=10)
            self.card_net.pack(side=LEFT, expand=True, fill=X, padx=5)
            
            today = datetime.today()
            earliest = get_earliest_entry_date()
            self.fin_from_date.entry.delete(0, tk.END)
            self.fin_from_date.entry.insert(0, earliest)
            self.fin_to_date.entry.delete(0, tk.END)
            self.fin_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
            
            self.log_lf = tb.Labelframe(container, padding=10, bootstyle="secondary")
            self.log_lf.pack(fill=BOTH, expand=True)
            
            right_header = tb.Frame(self.log_lf)
            right_header.pack(fill=X, pady=(0, 5))
            tb.Label(right_header, text=self._tr("Detailed Transaction Log"), font=("Segoe UI", 11, "bold"), bootstyle="secondary").pack(side=LEFT)
            
            btn_det = tb.Button(right_header, text="❓", bootstyle="link", cursor="hand2")
            btn_det.pack(side=RIGHT)
            msg_det = ("ℹ️ سجل المعاملات المفصل: يعرض قائمة بالمعاملات المالية للموظف المحدد بالتاريخ، بما في ذلك أرباح الخدمات، عمولة المنتجات، البقشيش، والمصروفات."
                    if getattr(self, 'lang', 'en') == 'ar' else
                    "ℹ️ Detailed Transaction Log: Lists chronological transactions for the selected employee, showing service earnings, product commission, tip payout, and expenses.")
            ToolTip(btn_det, text=msg_det)
            
            cols_detail = (self._tr("Date"), self._tr("Type"), self._tr("Name"), self._tr("Category"), self._tr("Amount"), self._tr("Cycle"))
            self.tree_fin_details = tb.Treeview(self.log_lf, columns=cols_detail, show="headings", bootstyle="secondary")
            self.apply_and_memorize_column_widths(
                "financials_details_table",
                self.tree_fin_details,
                cols_detail,
            )
            
            scroll_y_detail = tb.Scrollbar(self.log_lf, orient=VERTICAL, command=self.tree_fin_details.yview)
            scroll_x_detail = tb.Scrollbar(self.log_lf, orient=HORIZONTAL, command=self.tree_fin_details.xview)
            self.tree_fin_details.config(yscrollcommand=scroll_y_detail.set, xscrollcommand=scroll_x_detail.set)
            scroll_y_detail.pack(side=RIGHT, fill=Y)
            scroll_x_detail.pack(side=BOTTOM, fill=X)
            self.tree_fin_details.pack(side=LEFT, fill=BOTH, expand=True)
            self.tree_fin_details.bind("<ButtonRelease-1>", lambda e: self.open_fin_detail_item())
            
            self._last_fin_filter_state = ("", "")
            self.refresh_employee_dropdown()
            self.stop_filter_pollers(which="fin")
            self.poll_fin_filters()

        def set_fin_period(self, period):
            today = datetime.today()
            self._fin_from_user_set = True
            if period == "today":
                self.fin_from_date.entry.delete(0, tk.END)
                self.fin_from_date.entry.insert(0, today.strftime('%Y-%m-%d'))
                self.fin_to_date.entry.delete(0, tk.END)
                self.fin_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
            elif period == "week":
                start = today - timedelta(days=today.weekday())
                self.fin_from_date.entry.delete(0, tk.END)
                self.fin_from_date.entry.insert(0, start.strftime('%Y-%m-%d'))
                self.fin_to_date.entry.delete(0, tk.END)
                self.fin_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
            elif period == "month":
                start = today.replace(day=1)
                self.fin_from_date.entry.delete(0, tk.END)
                self.fin_from_date.entry.insert(0, start.strftime('%Y-%m-%d'))
                self.fin_to_date.entry.delete(0, tk.END)
                self.fin_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
            elif period == "all":
                self._fin_from_user_set = False
                earliest = get_earliest_entry_date()
                self.fin_from_date.entry.delete(0, tk.END)
                self.fin_from_date.entry.insert(0, earliest)
                self.fin_to_date.entry.delete(0, tk.END)
                self.fin_to_date.entry.insert(0, today.strftime('%Y-%m-%d'))
            self.load_financials_data(quiet=True)

        def poll_fin_filters(self):
            self._poll_fin_after_id = None
            if not self._widget_alive(getattr(self, "tab_financials", None)):
                return
            if not self._dateentry_alive(getattr(self, "fin_from_date", None)) or not self._dateentry_alive(
                getattr(self, "fin_to_date", None)
            ):
                return
            try:
                current_from = self.fin_from_date.entry.get()
                current_to = self.fin_to_date.entry.get()
                emp_val = self.fin_emp_filter.get() if hasattr(self, "fin_emp_filter") else ""
                ex_val = tuple(sorted(self.get_fin_excluded_names())) if hasattr(self, "fin_exclude_vars") else ()
                exp_cats = tuple(sorted(self.get_fin_included_expense_categories())) if hasattr(self, "fin_exp_cat_vars") else ()
                exp_on = bool(self.fin_inc_expenses.get()) if hasattr(self, "fin_inc_expenses") else True
                state = (current_from, current_to, emp_val, ex_val, exp_on, exp_cats)
                if state != getattr(self, "_last_fin_filter_state", None):
                    self._last_fin_filter_state = state
                    self.load_financials_data(quiet=True)
            except tk.TclError:
                return
            except Exception as e:
                if "invalid command name" in str(e):
                    return
                print(f"Error in poll_fin_filters: {e}")
            if self._widget_alive(getattr(self, "tab_financials", None)):
                self._poll_fin_after_id = self.tab_financials.after(750, self.poll_fin_filters)

        def schedule_financials_refresh(self, delay_ms=80):
            """Quiet, debounced P&L reload — no loading overlay for filter clicks."""
            aid = getattr(self, "_fin_refresh_after_id", None)
            if aid is not None:
                try:
                    self.after_cancel(aid)
                except Exception:
                    pass
            self._fin_refresh_after_id = self.after(delay_ms, self._run_quiet_financials_refresh)

        def _run_quiet_financials_refresh(self):
            self._fin_refresh_after_id = None
            try:
                self.load_financials_data(quiet=True)
            except Exception as e:
                print(f"Error refreshing financials: {e}")

        def load_financials_data(self, quiet=False):
            if quiet:
                return self._load_financials_data_body()
            self.show_busy(self._tr("Loading financials…"))
            try:
                return self._load_financials_data_body()
            finally:
                self.hide_busy()

        def _load_financials_data_body(self):
            if not self._widget_alive(getattr(self, "tree_fin_details", None)):
                return
                
            for item in self.tree_fin_details.get_children():
                self.tree_fin_details.delete(item)
                
            from_d = self.fin_from_date.entry.get()
            to_d = self.fin_to_date.entry.get()
            
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                
                cursor.execute("SELECT id, name FROM employees")
                employees = []
                for row in cursor.fetchall() or []:
                    if row is None or len(row) < 2:
                        continue
                    dec_name = decrypt_val(row[1]) if row[1] is not None else ""
                    dec_name = str(dec_name or "").strip()
                    employees.append((row[0], dec_name))
                
                selected_name = self.fin_emp_filter.get()
                selected_emp_id = None
                if selected_name and selected_name != self._tr("All Employees"):
                    for emp_id, emp_name in employees:
                        if emp_name == selected_name:
                            selected_emp_id = emp_id
                            break

                exclude_names = set(self.get_fin_excluded_names()) if hasattr(self, "get_fin_excluded_names") else set()
                exclude_emp_ids = set()
                for emp_id, emp_name in employees:
                    if emp_name in exclude_names:
                        exclude_emp_ids.add(emp_id)
                # If Include picks a person who is also excluded, ignore exclude for that person
                if selected_emp_id is not None and selected_emp_id in exclude_emp_ids:
                    exclude_emp_ids.discard(selected_emp_id)
                
                summary_data = {}
                for emp_id, emp_name in employees:
                    if emp_id in exclude_emp_ids:
                        continue
                    if selected_emp_id is None or emp_id == selected_emp_id:
                        summary_data[emp_id] = {
                            "name": emp_name,
                            "service_sales": 0.0,
                            "addon_sales": 0.0,
                            "product_sales": 0.0,
                            "tips": 0.0,
                            "other_earning": 0.0,
                            "expenses": 0.0
                        }
                
                shop_id = None
                for emp_id, emp_name in employees:
                    if str(emp_name).lower() == "shop":
                        shop_id = emp_id
                        break
                        
                cursor.execute('''
                    SELECT r.employee_id, r.revenue, r.service_addon_sales, r.product_sales, r.tip
                    FROM payroll_records r
                    WHERE r.record_date >= ? AND r.record_date <= ?
                ''', (from_d, to_d))
                for row in cursor.fetchall() or []:
                    if row is None or len(row) < 5:
                        continue
                    emp_id, rev, addon, prod, tip = row[0], row[1], row[2], row[3], row[4]
                    e_id = emp_id if emp_id else shop_id
                    if e_id in summary_data:
                        summary_data[e_id]["service_sales"] += to_float(rev, 0.0)
                        summary_data[e_id]["addon_sales"] += to_float(addon, 0.0)
                        summary_data[e_id]["product_sales"] += to_float(prod, 0.0)
                        summary_data[e_id]["tips"] += to_float(tip, 0.0)
                        
                cursor.execute('''
                    SELECT employee_id, category, amount
                    FROM expenses
                    WHERE expense_date >= ? AND expense_date <= ?
                ''', (from_d, to_d))
                for row in cursor.fetchall() or []:
                    if row is None or len(row) < 3:
                        continue
                    emp_id, cat, amt = row[0], row[1], row[2]
                    e_id = emp_id if emp_id else shop_id
                    amt_val = to_float(amt, 0.0)
                    cat_plain = plain_label(cat)
                    if e_id in summary_data:
                        if is_income_expense_category(cat_plain):
                            if cat_plain == "Cash Envelope Received":
                                if shop_id in summary_data:
                                    summary_data[shop_id]["other_earning"] += amt_val
                            else:
                                summary_data[e_id]["other_earning"] += amt_val
                        else:
                            if self._fin_expense_category_included(cat_plain):
                                summary_data[e_id]["expenses"] += amt_val
                            
                grand_earnings = 0.0
                grand_expenses = 0.0
                
                for emp_id, data in summary_data.items():
                    total_earn = 0.0
                    if self.fin_inc_service.get():
                        total_earn += data["service_sales"]
                    if self.fin_inc_addon.get():
                        total_earn += data["addon_sales"]
                    if self.fin_inc_product.get():
                        total_earn += data["product_sales"]
                    if self.fin_inc_tips.get():
                        total_earn += data["tips"]
                    if self.fin_inc_other.get():
                        total_earn += data["other_earning"]
                    
                    # expenses already filtered by selected categories above
                    exp_val = data["expenses"] if self.fin_inc_expenses.get() else 0.0
                    
                    grand_earnings += total_earn
                    grand_expenses += exp_val
                
                self.card_earning.config(text=f"{self._tr('Total Earnings')}: ${grand_earnings:,.2f}")
                self.card_expense.config(text=f"{self._tr('Total Expenses')}: ${grand_expenses:,.2f}")
                self.card_net.config(text=f"{self._tr('Net Profit')}: ${grand_earnings - grand_expenses:,.2f}", 
                                    bootstyle="success" if (grand_earnings - grand_expenses) >= 0 else "danger")
                
                self.reload_detailed_log(selected_emp_id, exclude_emp_ids=exclude_emp_ids, conn=conn, cursor=cursor)
                conn.close()
            except Exception as e:
                try:
                    if _is_dead_pg_error(e) or _is_connectivity_error(e):
                        get_shared_supabase_conn(force_reconnect=True)
                except Exception:
                    pass
                print(f"Error loading financials: {e}")

        def rebuild_fin_exclude_menu(self, names):
            menu = getattr(self, "fin_exclude_menu", None)
            try:
                if menu is None or not menu.winfo_exists():
                    return
            except Exception:
                return
            self._fin_exclude_updating = True
            prev = set(self.get_fin_excluded_names())
            try:
                menu.delete(0, "end")
            except Exception:
                self._fin_exclude_updating = False
                return
            self.fin_exclude_vars = {}
            try:
                for name in names or []:
                    var = tk.BooleanVar(value=(name in prev))
                    self.fin_exclude_vars[name] = var
                    menu.add_checkbutton(
                        label=name,
                        variable=var,
                        command=self._on_fin_exclude_changed,
                    )
            except tk.TclError:
                self.fin_exclude_vars = {}
                self._fin_exclude_updating = False
                return
            self._update_fin_exclude_summary()
            self._fin_exclude_updating = False

        def _fin_expense_category_choices(self):
            """Expense categories available for P&L include filter (exclude income cats)."""
            cats = []
            try:
                for c in self.get_db_categories() or []:
                    name = plain_label(c)
                    if name and not is_income_expense_category(name):
                        cats.append(name)
            except Exception:
                cats = [
                    "Travel", "Equipment", "Office Supplies", "Meals", "Software",
                    "Salary Payment", "Amazon Order", "Groceries", "Other",
                ]
            # Always treat salary as an expense, even if missing from Settings.
            cats.append("Salary Payment")
            cats.append("Salary Payment (Tip)")
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute("SELECT DISTINCT category FROM expenses")
                for row in cur.fetchall() or []:
                    name = plain_label(row[0] if row else None)
                    if name and not is_income_expense_category(name):
                        cats.append(name)
                conn.close()
            except Exception:
                pass
            # Stable unique order, salary first so it is obvious in the list
            seen = set()
            out = []
            for c in cats:
                if c not in seen:
                    seen.add(c)
                    out.append(c)
            preferred = ["Salary Payment", "Salary Payment (Tip)"]
            rest = [c for c in out if c not in preferred]
            return [c for c in preferred if c in seen] + rest

        def rebuild_fin_expense_cat_menu(self, select_all=None):
            menu = getattr(self, "fin_exp_cat_menu", None)
            try:
                if menu is None or not menu.winfo_exists():
                    return
            except Exception:
                return
            self._fin_exp_cat_updating = True
            cats = self._fin_expense_category_choices()
            prev = set(self.get_fin_included_expense_categories())
            if select_all is True:
                prev = set(cats)
            elif select_all is False:
                prev = set()
            elif not prev and getattr(self, "fin_inc_expenses", None) is not None and self.fin_inc_expenses.get():
                # First build / empty selection with Expenses on → include all
                prev = set(cats)
            try:
                menu.delete(0, "end")
            except Exception:
                self._fin_exp_cat_updating = False
                return
            self.fin_exp_cat_vars = {}
            try:
                all_on = True if select_all is True or not prev or len(prev) == len(cats) else False
                only = None
                if select_all is False:
                    all_on = False
                elif prev and len(prev) == 1 and select_all is not True:
                    all_on = False
                    only = next(iter(prev))
                self.fin_exp_select_all = tk.BooleanVar(value=all_on)
                menu.add_checkbutton(
                    label=self._tr("Select All"),
                    variable=self.fin_exp_select_all,
                    command=self._on_fin_expense_select_all,
                )
                for name in cats:
                    var = tk.BooleanVar(value=True if all_on else (name == only))
                    self.fin_exp_cat_vars[name] = var
                    menu.add_checkbutton(
                        label=name,
                        variable=var,
                        command=lambda n=name: self._on_fin_expense_cat_changed(n),
                    )
            except tk.TclError:
                self.fin_exp_cat_vars = {}
                self._fin_exp_cat_updating = False
                return
            self._update_fin_expense_cat_summary()
            self._fin_exp_cat_updating = False

        def get_fin_included_expense_categories(self):
            vars_map = getattr(self, "fin_exp_cat_vars", None) or {}
            names = []
            for name, var in vars_map.items():
                try:
                    if var.get():
                        names.append(name)
                except Exception:
                    pass
            return names

        def _fin_expense_category_included(self, cat_plain):
            if not getattr(self, "fin_inc_expenses", None) or not self.fin_inc_expenses.get():
                return False
            cat = plain_label(cat_plain)
            if not cat or is_income_expense_category(cat):
                return False
            vars_map = getattr(self, "fin_exp_cat_vars", None) or {}
            if not vars_map:
                return True
            # Select All → every real expense, including salary and any new category
            try:
                if getattr(self, "fin_exp_select_all", None) and self.fin_exp_select_all.get():
                    return True
            except Exception:
                pass
            selected = self.get_fin_included_expense_categories()
            if selected and len(selected) == len(vars_map):
                return True
            aliases = {cat}
            if cat in SALARY_EXPENSE_CATEGORIES:
                aliases.update(SALARY_EXPENSE_CATEGORIES)
            for name, var in vars_map.items():
                if name in aliases:
                    try:
                        if var.get():
                            return True
                    except Exception:
                        pass
            return False

        def _update_fin_expense_cat_summary(self):
            summary = getattr(self, "fin_exp_cat_summary", None)
            if summary is None:
                return
            vars_map = getattr(self, "fin_exp_cat_vars", None) or {}
            selected = self.get_fin_included_expense_categories()
            if not getattr(self, "fin_inc_expenses", None) or not self.fin_inc_expenses.get():
                summary.set(self._tr("None"))
            elif not vars_map or len(selected) == len(vars_map):
                summary.set(self._tr("Select All"))
            elif not selected:
                summary.set(self._tr("None"))
            elif len(selected) == 1:
                summary.set(selected[0])
            else:
                summary.set(f"{len(selected)} {self._tr('categories')}")

        def _on_fin_expense_select_all(self):
            if getattr(self, "_fin_exp_cat_updating", False):
                return
            self._fin_exp_cat_updating = True
            try:
                on = bool(self.fin_exp_select_all.get()) if getattr(self, "fin_exp_select_all", None) else True
                for var in (getattr(self, "fin_exp_cat_vars", None) or {}).values():
                    var.set(on)
                if on:
                    self.fin_inc_expenses.set(True)
            finally:
                self._fin_exp_cat_updating = False
            self._update_fin_expense_cat_summary()
            self.schedule_financials_refresh()

        def _on_fin_expense_cat_changed(self, name=None):
            if getattr(self, "_fin_exp_cat_updating", False):
                return
            self._fin_exp_cat_updating = True
            try:
                vars_map = getattr(self, "fin_exp_cat_vars", None) or {}
                if name and name in vars_map:
                    # Check if 'Select All' was previously active
                    select_all_active = False
                    if getattr(self, "fin_exp_select_all", None) and self.fin_exp_select_all.get():
                        select_all_active = True
                    
                    if select_all_active:
                        # Make only 'name' selected and uncheck all others
                        for n, var in vars_map.items():
                            var.set(n == name)
                        self.fin_exp_select_all.set(False)
                    else:
                        # Standard toggle (handled by Tkinter), update Select All state
                        all_checked = all(var.get() for var in vars_map.values())
                        if getattr(self, "fin_exp_select_all", None):
                            self.fin_exp_select_all.set(all_checked)
                    
                    any_checked = any(var.get() for var in vars_map.values())
                    self.fin_inc_expenses.set(any_checked)
            finally:
                self._fin_exp_cat_updating = False
            self._update_fin_expense_cat_summary()
            self.schedule_financials_refresh()

        def _on_fin_expenses_master_changed(self):
            if self.fin_inc_expenses.get():
                # Expenses checked → include all categories
                self.rebuild_fin_expense_cat_menu(select_all=True)
            else:
                self._update_fin_expense_cat_summary()
            self.schedule_financials_refresh()

        def _update_fin_exclude_summary(self):
            summary = getattr(self, "fin_exclude_summary", None)
            if summary is None:
                return
            names = self.get_fin_excluded_names()
            if not names:
                summary.set(self._tr("No exclude"))
            elif len(names) == 1:
                summary.set(names[0])
            else:
                summary.set(f"{len(names)} {self._tr('excluded')}")

        def _on_fin_exclude_changed(self):
            if getattr(self, "_fin_exclude_updating", False):
                return
            self._update_fin_exclude_summary()
            self.schedule_financials_refresh()

        def get_fin_excluded_names(self):
            vars_map = getattr(self, "fin_exclude_vars", None) or {}
            names = []
            for name, var in vars_map.items():
                try:
                    if var.get():
                        names.append(name)
                except Exception:
                    pass
            return names

        def on_fin_emp_filter_changed(self, event=None):
            self.schedule_financials_refresh()

        def clear_fin_employee_filter(self):
            if self._widget_alive(getattr(self, "fin_emp_filter", None)):
                self.fin_emp_filter.set(self._tr("All Employees"))
            for var in (getattr(self, "fin_exclude_vars", None) or {}).values():
                try:
                    var.set(False)
                except Exception:
                    pass
            self._update_fin_exclude_summary()
            self.schedule_financials_refresh()

        def reload_detailed_log(self, target_emp_id=None, exclude_emp_ids=None, conn=None, cursor=None):
            for item in self.tree_fin_details.get_children():
                self.tree_fin_details.delete(item)
                
            from_d = self.fin_from_date.entry.get()
            to_d = self.fin_to_date.entry.get()
            
            own_conn = conn is None
            if own_conn:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
            
            cursor.execute("SELECT id, name FROM employees")
            name_map = {}
            for row in cursor.fetchall() or []:
                if row is None or len(row) < 2:
                    continue
                dec_name = decrypt_val(row[1]) if row[1] is not None else ""
                dec_name = str(dec_name or "").strip()
                name_map[row[0]] = dec_name
            
            shop_id = None
            for e_id, name in name_map.items():
                if str(name).lower() == "shop":
                    shop_id = e_id
                    break
            
            exclude_emp_ids = set(exclude_emp_ids or [])
            detailed_rows = []
            
            # Fetch payroll details (earnings & tips)
            query_p = '''
                SELECT id, record_date, employee_id, revenue, service_addon_sales, product_sales, notes, tip
                FROM payroll_records
                WHERE record_date >= ? AND record_date <= ?
            '''
            params_p = [from_d, to_d]
            if target_emp_id is not None:
                if target_emp_id == shop_id:
                    query_p += " AND (employee_id = ? OR employee_id IS NULL)"
                else:
                    query_p += " AND employee_id = ?"
                params_p.append(target_emp_id)
            
            cursor.execute(query_p, params_p)
            for row in cursor.fetchall() or []:
                if row is None or len(row) < 8:
                    continue
                rec_id, date_val, emp_id, rev, addon, prod, notes, tip = row[:8]
                e_id = emp_id if emp_id else shop_id
                if e_id in exclude_emp_ids:
                    continue
                emp_name = name_map.get(e_id, "General/None")
                rev = to_float(rev, 0.0)
                addon = to_float(addon, 0.0)
                prod = to_float(prod, 0.0)
                tip = to_float(tip, 0.0)
                
                _cyc_p = cycle_for_date(date_val)
                if self.fin_inc_service.get() and rev > 0:
                    detailed_rows.append((date_val, self._tr("Earning"), emp_name, self._tr("Service Sales"), rev, _cyc_p, "payroll", rec_id))
                if self.fin_inc_addon.get() and addon > 0:
                    detailed_rows.append((date_val, self._tr("Earning"), emp_name, self._tr("Service Add-on Sales"), addon, _cyc_p, "payroll", rec_id))
                if self.fin_inc_product.get() and prod > 0:
                    detailed_rows.append((date_val, self._tr("Earning"), emp_name, self._tr("Product Sales"), prod, _cyc_p, "payroll", rec_id))
                if self.fin_inc_tips.get() and tip > 0:
                    detailed_rows.append((date_val, self._tr("Earning"), emp_name, self._tr("Tips"), tip, _cyc_p, "payroll", rec_id))
            
            # Fetch expenses list
            query_ex = '''
                SELECT id, expense_date, employee_id, category, amount, description, cycle_key
                FROM expenses
                WHERE expense_date >= ? AND expense_date <= ?
            '''
            params_ex = [from_d, to_d]
            if target_emp_id is not None:
                if target_emp_id == shop_id:
                    query_ex += " AND (employee_id = ? OR employee_id IS NULL)"
                else:
                    query_ex += " AND employee_id = ?"
                params_ex.append(target_emp_id)
                
            cursor.execute(query_ex, params_ex)
            for row in cursor.fetchall() or []:
                if row is None or len(row) < 6:
                    continue
                exp_id, date_val, emp_id, category, amount, desc = row[:6]
                cyc_stored = row[6] if len(row) > 6 else None
                _cyc_e = (str(cyc_stored).strip() if cyc_stored else "") or cycle_for_date(date_val)
                e_id = emp_id if emp_id else shop_id
                if e_id in exclude_emp_ids:
                    continue
                emp_name = name_map.get(e_id, "General/None")
                amount = to_float(amount, 0.0)
                cat_plain = plain_label(category)
                
                is_income = is_income_expense_category(cat_plain)
                if is_income:
                    if self.fin_inc_other.get():
                        detailed_rows.append((date_val, self._tr("Earning"), emp_name, self._tr(cat_plain), amount, _cyc_e, "expense", exp_id))
                else:
                    if self.fin_inc_expenses.get() and self._fin_expense_category_included(cat_plain):
                        detailed_rows.append((date_val, self._tr("Expense"), emp_name, self._tr(cat_plain or str(category or "")), amount, _cyc_e, "expense", exp_id))
            
            if own_conn:
                conn.close()
            
            detailed_rows.sort(key=lambda x: x[0] or "", reverse=True)
            self._fin_detail_map = {}
            
            for row in detailed_rows:
                date_val, t_type, emp_name, category, amount, cyc_key, kind, rec_id = row
                fmt_amt = f"+{amount:,.2f}" if t_type == self._tr("Earning") else f"-{amount:,.2f}"
                tag = "income" if t_type == self._tr("Earning") else "expense"
                cyc_disp = cycle_label(cyc_key) if cyc_key else ""
                iid = self.tree_fin_details.insert('', tk.END, values=(date_val, t_type, emp_name, category, fmt_amt, cyc_disp), tags=(tag,))
                self._fin_detail_map[iid] = (kind, rec_id)
                
            self.tree_fin_details.tag_configure("income", foreground="#00bc8c")
            self.tree_fin_details.tag_configure("expense", foreground="#e74c3c")

        def open_fin_detail_item(self):
            tree = getattr(self, "tree_fin_details", None)
            if not self._widget_alive(tree):
                return
            sel = tree.selection()
            if not sel:
                return
            meta = (getattr(self, "_fin_detail_map", None) or {}).get(sel[0])
            if not meta:
                return
            kind, rec_id = meta
            if kind == "expense":
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT expense_date, category, amount, employee_id, status, description, payment_type, location, assignee_id, document_path, is_tip, tip_given FROM expenses WHERE id=?",
                    (rec_id,),
                )
                row = cursor.fetchone()
                conn.close()
                if row:
                    self.open_expense_dialog(expense_id=rec_id, data=row, on_save_callback=lambda: self.load_financials_data(quiet=True))
            elif kind == "payroll":
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("PRAGMA table_info(payroll_records)")
                col_names = [col[1] for col in cursor.fetchall()]
                hr_col = "r.hour_rate" if "hour_rate" in col_names else "NULL"
                perc_col = "r.percentage" if "percentage" in col_names else "NULL"
                cursor.execute(f'''
                    SELECT r.record_date, r.revenue, r.service_addon_sales, r.hours, r.notes, r.written_up, e.name, r.written_up_desc,
                        COALESCE({hr_col}, e.hour_rate), COALESCE({perc_col}, e.percentage), e.use_tiered_payout, e.id
                    FROM payroll_records r
                    JOIN employees e ON r.employee_id = e.id
                    WHERE r.id=?
                ''', (rec_id,))
                rec = cursor.fetchone()
                conn.close()
                if rec:
                    self.open_edit_record_dialog(rec_id, rec)

        def open_settings_password_prompt(self):
            dialog = tb.Toplevel(self)
            dialog.title(self._tr("⚙️ Enter Password"))
            dialog.geometry("400x250")
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()
            
            # Center on screen
            sw = dialog.winfo_screenwidth()
            sh = dialog.winfo_screenheight()
            x = (sw - 400) // 2
            y = (sh - 250) // 2
            dialog.geometry(f"400x250+{x}+{y}")
            
            tb.Label(dialog, text=self._tr("Password:"), font=("Segoe UI", 12, "bold")).pack(pady=20)
            
            pw_frame = tb.Frame(dialog)
            pw_frame.pack(pady=10)
            
            pw_entry = tb.Entry(pw_frame, show="*", width=22, font=("Segoe UI", 11))
            pw_entry.pack(side=LEFT)
            pw_entry.focus()
            
            def toggle_settings_pw_visibility():
                try:
                    if pw_entry.cget("show") == "*":
                        pw_entry.config(show="")
                        eye_btn.config(text="Hide")
                    else:
                        pw_entry.config(show="*")
                        eye_btn.config(text="Show")
                except Exception:
                    pass
                    
            eye_btn = tb.Button(pw_frame, text="Show", bootstyle="secondary outline", cursor="hand2", width=5, command=toggle_settings_pw_visibility)
            eye_btn.pack(side=LEFT, padx=(5, 0))
            
            def check_password(event=None):
                entered_pw = pw_entry.get().strip()
                if not entered_pw:
                    return
                who = getattr(self, "current_user", None) or DEFAULT_ADMIN_USERNAME
                row = None
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute("SELECT password FROM users WHERE username=?", (who,))
                    row = cursor.fetchone()
                    if not row:
                        cursor.execute("SELECT password FROM users WHERE username=?", (DEFAULT_ADMIN_USERNAME,))
                        row = cursor.fetchone()
                    conn.close()
                except Exception as e:
                    try:
                        if _is_dead_pg_error(e) or _is_connectivity_error(e):
                            get_shared_supabase_conn(force_reconnect=True)
                            conn = sqlite3.connect(TEMP_DB_PATH)
                            cursor = conn.cursor()
                            cursor.execute("SELECT password FROM users WHERE username=?", (who,))
                            row = cursor.fetchone()
                            if not row:
                                cursor.execute("SELECT password FROM users WHERE username=?", (DEFAULT_ADMIN_USERNAME,))
                                row = cursor.fetchone()
                            conn.close()
                        else:
                            raise
                    except Exception as e2:
                        messagebox.showerror("Error", f"Could not verify password:\n{e2}", parent=dialog)
                        return
                
                if row:
                    stored_pw = decrypt_val(row[0]) if row[0] is not None else None
                    import hashlib
                    hashed_input = hashlib.sha256(entered_pw.encode('utf-8')).hexdigest()
                    if stored_pw == hashed_input or stored_pw == entered_pw:
                        dialog.destroy()
                        self.open_settings_dialog()
                        return
                
                messagebox.showerror("Error", "Invalid Password.", parent=dialog)
                
            pw_entry.bind("<Return>", check_password)
            tb.Button(dialog, text=self._tr("Unlock"), bootstyle="primary", command=check_password).pack(pady=10)

        def _build_activity_and_backup_panel(self, parent):
            """Per-user activity log and twice-daily Supabase backups."""
            try:
                top = parent.winfo_toplevel()
            except Exception:
                top = parent

            canvas = tk.Canvas(parent, highlightthickness=0, borderwidth=0)
            vscroll = tb.Scrollbar(parent, orient=VERTICAL, command=canvas.yview)
            inner = tb.Frame(canvas, padding=16)

            inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
            canvas.configure(yscrollcommand=vscroll.set)

            def _on_canvas_configure(e):
                try:
                    canvas.itemconfigure(win_id, width=e.width)
                except Exception:
                    pass
            canvas.bind("<Configure>", _on_canvas_configure)

            def _act_mousewheel(event):
                try:
                    if getattr(event, "delta", 0):
                        delta = int(-1 * (event.delta / 120))
                    elif getattr(event, "num", 0) == 5:
                        delta = 1
                    elif getattr(event, "num", 0) == 4:
                        delta = -1
                    else:
                        delta = 0
                    if delta:
                        canvas.yview_scroll(delta, "units")
                except Exception:
                    pass

            def _bind_wheel(_e=None):
                canvas.bind_all("<MouseWheel>", _act_mousewheel)
                canvas.bind_all("<Button-4>", _act_mousewheel)
                canvas.bind_all("<Button-5>", _act_mousewheel)

            def _unbind_wheel(_e=None):
                try:
                    canvas.unbind_all("<MouseWheel>")
                    canvas.unbind_all("<Button-4>")
                    canvas.unbind_all("<Button-5>")
                except Exception:
                    pass

            canvas.bind("<Enter>", _bind_wheel)
            canvas.bind("<Leave>", _unbind_wheel)
            inner.bind("<Enter>", _bind_wheel)
            inner.bind("<Leave>", _unbind_wheel)

            vscroll.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)

            # --- CSV Log Downloader from Date to Date ---
            exp_lf = tb.Labelframe(inner, text=self._tr("📥 Download Activity Logs (CSV)"), padding=12, bootstyle="primary")
            exp_lf.pack(fill=X, pady=(0, 12))

            row_dates = tb.Frame(exp_lf)
            row_dates.pack(fill=X, pady=(0, 8))

            tb.Label(row_dates, text=self._tr("From Date:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 4))
            start_dt_ent = tb.DateEntry(row_dates, bootstyle="primary", dateformat='%Y-%m-%d', width=10)
            start_dt_ent.entry.delete(0, tk.END)
            start_dt_ent.entry.insert(0, (datetime.today() - timedelta(days=30)).strftime('%Y-%m-%d'))
            start_dt_ent.pack(side=LEFT, padx=(0, 16))

            tb.Label(row_dates, text=self._tr("To Date:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 4))
            end_dt_ent = tb.DateEntry(row_dates, bootstyle="primary", dateformat='%Y-%m-%d', width=10)
            end_dt_ent.entry.delete(0, tk.END)
            end_dt_ent.entry.insert(0, datetime.today().strftime('%Y-%m-%d'))
            end_dt_ent.pack(side=LEFT)

            row_type = tb.Frame(exp_lf)
            row_type.pack(fill=X, pady=(0, 8))

            tb.Label(row_type, text=self._tr("Type:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 4))
            cbo_log_type = tb.Combobox(row_type, width=20, state="readonly", values=[self._tr("All Activity Logs"), self._tr("User Actions Only"), self._tr("Database Changes Only")])
            cbo_log_type.set(self._tr("All Activity Logs"))
            cbo_log_type.pack(side=LEFT, padx=(0, 16))
            
            def do_export_csv():
                s_date = start_dt_ent.entry.get().strip()
                e_date = end_dt_ent.entry.get().strip()
                if not s_date or not e_date:
                    messagebox.showerror("Error", self._tr("Please select valid Start and End dates."), parent=top)
                    return
                l_type = cbo_log_type.get()
                
                save_path = filedialog.asksaveasfilename(
                    parent=top,
                    title=self._tr("Save Activity Logs CSV"),
                    defaultextension=".csv",
                    filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
                    initialfile=f"activity_logs_{s_date}_to_{e_date}.csv"
                )
                if not save_path:
                    return
                    
                try:
                    import csv
                    rows_out = []
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    
                    # 1. user_action_log
                    if "User Actions" in l_type or "All" in l_type or l_type == self._tr("All Activity Logs"):
                        try:
                            cur.execute(
                                """
                                SELECT created_at, user_name, action, table_name, record_id, summary, details
                                FROM user_action_log
                                WHERE created_at >= ? AND created_at <= ?
                                ORDER BY created_at ASC
                                """,
                                (f"{s_date} 00:00:00", f"{e_date} 23:59:59")
                            )
                            for r in cur.fetchall() or []:
                                rows_out.append([
                                    r[0] or "",
                                    plain_label(r[1]) or "unknown",
                                    r[2] or "USER_ACTION",
                                    f"{r[3] or ''} #{r[4] or ''}".strip(),
                                    plain_label(r[5]) or "",
                                    plain_label(r[6]) or ""
                                ])
                        except Exception:
                            pass
                            
                    # 2. database_history_log
                    if "Database" in l_type or "All" in l_type or l_type == self._tr("All Activity Logs"):
                        try:
                            cur.execute(
                                """
                                SELECT timestamp, user_name, action, table_name, row_id, old_data, new_data
                                FROM database_history_log
                                WHERE timestamp >= ? AND timestamp <= ?
                                ORDER BY timestamp ASC
                                """,
                                (f"{s_date} 00:00:00", f"{e_date} 23:59:59")
                            )
                            for r in cur.fetchall() or []:
                                rows_out.append([
                                    r[0] or "",
                                    plain_label(r[1]) or "system",
                                    r[2] or "DB_MUTATION",
                                    f"{r[3] or ''} #{r[4] or ''}".strip(),
                                    f"Old: {plain_label(r[5]) or ''}",
                                    f"New: {plain_label(r[6]) or ''}"
                                ])
                        except Exception:
                            pass
                            
                    conn.close()
                    rows_out.sort(key=lambda x: str(x[0]))
                    
                    with open(save_path, "w", newline="", encoding="utf-8-sig") as csv_f:
                        writer = csv.writer(csv_f)
                        writer.writerow(["Timestamp", "User", "Action", "Target", "Summary / Old Data", "Details / New Data"])
                        for ro in rows_out:
                            writer.writerow(ro)
                            
                    messagebox.showinfo("Export Successful", f"Successfully exported {len(rows_out)} log entries to:\n{save_path}", parent=top)
                except Exception as e:
                    messagebox.showerror("Export Failed", str(e), parent=top)

            row_action = tb.Frame(exp_lf)
            row_action.pack(fill=X)
            tb.Button(row_action, text=self._tr("📥 Download Activity Logs (CSV)"), bootstyle="success", command=do_export_csv).pack(side=LEFT)

            log_lf = tb.Labelframe(inner, text=self._tr("Activity log"), padding=10)
            log_lf.pack(fill=X, pady=(0, 12))
            filter_row = tb.Frame(log_lf)
            filter_row.pack(fill=X, pady=(0, 8))
            tb.Label(filter_row, text=self._tr("Filter by user:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT)
            user_filter = tb.Combobox(filter_row, width=22, state="readonly")
            user_filter.pack(side=LEFT, padx=8)
            user_filter.set(self._tr("All users"))

            cols = ("When", "User", "Action")
            log_holder = tb.Frame(log_lf)
            log_holder.pack(fill=X)
            log_tree = tb.Treeview(log_holder, columns=cols, show="headings", height=8, bootstyle="info")
            log_tree.heading("When", text=self._tr("Date"))
            log_tree.heading("User", text=self._tr("User"))
            log_tree.heading("Action", text=self._tr("Action"))
            log_tree.column("When", width=150, anchor=CENTER)
            log_tree.column("User", width=110, anchor=CENTER)
            log_tree.column("Action", width=420, anchor=W)
            self._attach_tree_scrollbars(log_holder, log_tree)

            bak_lf = tb.Labelframe(inner, text=self._tr("📁 Daily Backups (Local & Supabase Cloud)"), padding=10, bootstyle="info")
            bak_lf.pack(fill=X)
            tb.Label(
                bak_lf,
                text=self._tr("Daily morning (AM) and afternoon (PM) backups are saved locally on each linked device and synced to Supabase Cloud."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=720,
                justify=LEFT,
            ).pack(anchor=W, pady=(0, 8))
            bak_cols = ("When", "Slot", "User", "Size")
            bak_holder = tb.Frame(bak_lf)
            bak_holder.pack(fill=X)
            bak_tree = tb.Treeview(bak_holder, columns=bak_cols, show="headings", height=6, bootstyle="secondary")
            bak_tree.heading("When", text=self._tr("Date & Time"))
            bak_tree.heading("Slot", text=self._tr("Type / Slot"))
            bak_tree.heading("User", text=self._tr("Device / User"))
            bak_tree.heading("Size", text=self._tr("Size"))
            bak_tree.column("When", width=150, anchor=CENTER)
            bak_tree.column("Slot", width=160, anchor=W)
            bak_tree.column("User", width=220, anchor=W)
            bak_tree.column("Size", width=90, anchor=CENTER)
            self._attach_tree_scrollbars(bak_holder, bak_tree)

            def load_logs():
                for item in log_tree.get_children():
                    log_tree.delete(item)
                want = plain_label(user_filter.get())
                all_label = self._tr("All users")
                names = [all_label]
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute(
                        """
                        CREATE TABLE IF NOT EXISTS user_action_log (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            log_uid TEXT UNIQUE,
                            created_at TEXT,
                            user_name TEXT,
                            action TEXT,
                            table_name TEXT,
                            record_id TEXT,
                            summary TEXT,
                            details TEXT
                        )
                        """
                    )
                    cur.execute(
                        "SELECT created_at, user_name, summary FROM user_action_log ORDER BY id DESC LIMIT 400"
                    )
                    rows = cur.fetchall() or []
                    conn.close()
                    seen = set()
                    for created_at, user_name, summary in rows:
                        who = plain_label(user_name) or "unknown"
                        if who.lower() not in seen:
                            seen.add(who.lower())
                            names.append(who)
                        if want and want != all_label and who.lower() != want.lower():
                            continue
                        log_tree.insert(
                            "",
                            tk.END,
                            values=(created_at or "", who, plain_label(summary) or ""),
                        )
                except Exception:
                    pass
                current = user_filter.get()
                user_filter.configure(values=names)
                if current in names:
                    user_filter.set(current)
                else:
                    user_filter.set(all_label)

            def load_backups():
                for item in bak_tree.get_children():
                    bak_tree.delete(item)
                try:
                    loading_id = bak_tree.insert("", tk.END, values=(self._tr("Loading..."), "", "", ""))
                except Exception:
                    loading_id = None

                def run():
                    try:
                        recs = list_all_backups(60)
                    except Exception:
                        recs = []

                    def populate():
                        try:
                            if loading_id and bak_tree.exists(loading_id):
                                bak_tree.delete(loading_id)
                            for rec in recs:
                                slot_label = rec.get("slot") or "Backup"
                                key = rec.get("slot_key") or ""
                                size = rec.get("size_bytes") or 0
                                try:
                                    size_txt = f"{int(size) / 1024:.0f} KB"
                                except Exception:
                                    size_txt = str(size)
                                bak_tree.insert(
                                    "",
                                    tk.END,
                                    iid=key,
                                    values=(rec.get("created_at") or "", slot_label, rec.get("created_by") or "", size_txt),
                                )
                        except Exception:
                            pass

                    try:
                        top.after(0, populate)
                    except Exception:
                        pass

                import threading
                threading.Thread(target=run, daemon=True).start()

            def do_backup_now():
                try:
                    ok, msg = create_cloud_backup(kind="manual")
                    if ok:
                        messagebox.showinfo("Backup Success", "Backup saved locally on this device and synced to Supabase Cloud.", parent=top)
                        load_backups()
                        load_logs()
                    else:
                        messagebox.showerror("Backup Failed", str(msg), parent=top)
                except Exception as e:
                    messagebox.showerror("Backup Failed", str(e), parent=top)

            def do_restore():
                sel = bak_tree.selection()
                if not sel:
                    messagebox.showwarning("Select", "Please select a backup to restore.", parent=top)
                    return
                key = sel[0]
                is_loc = str(key).startswith("local::")
                source_txt = "Local disk backup on this PC" if is_loc else "Supabase Cloud backup"
                if not messagebox.askyesno(
                    "Restore backup",
                    f"This will replace current application data with the selected {source_txt}.\n\nContinue?",
                    parent=top,
                ):
                    return
                try:
                    ok, msg = restore_cloud_backup(key)
                    if ok:
                        messagebox.showinfo("Restore", "Backup restored successfully. Refreshing views...", parent=top)
                        try:
                            self._schedule_soft_ui_refresh(full=True)
                        except Exception:
                            pass
                        load_logs()
                    else:
                        messagebox.showerror("Restore Failed", str(msg), parent=top)
                except Exception as e:
                    messagebox.showerror("Restore Failed", str(e), parent=top)

            def do_download_backup():
                sel = bak_tree.selection()
                if not sel:
                    messagebox.showwarning("Select Backup", "Please select a backup from the list to download.", parent=top)
                    return
                key = sel[0]
                payload = None
                if str(key).startswith("local::"):
                    raw_key = str(key)[7:]
                    b_dir = get_local_backups_dir()
                    json_file = os.path.join(b_dir, f"snapshot_{raw_key}.json.gz")
                    if os.path.isfile(json_file):
                        with open(json_file, "r", encoding="utf-8") as f:
                            payload = f.read()
                    if not payload:
                        enc_file = os.path.join(b_dir, f"backup_{raw_key}.enc")
                        if os.path.isfile(enc_file):
                            dest = filedialog.asksaveasfilename(
                                title="Download Backup File",
                                initialfile=f"payroll_backup_{raw_key}.enc",
                                defaultextension=".enc",
                                filetypes=[("Encrypted Database (*.enc)", "*.enc"), ("All Files (*.*)", "*.*")],
                                parent=top,
                            )
                            if dest:
                                import shutil
                                shutil.copy2(enc_file, dest)
                                messagebox.showinfo("Success", f"Backup downloaded to:\n{dest}", parent=top)
                                log_user_action("backup_download", extra_summary=f"Downloaded local backup {key} to {os.path.basename(dest)}")
                            return
                else:
                    if get_db_mode() == "supabase" and not is_supabase_offline():
                        try:
                            pg = get_shared_supabase_conn()
                            cur = pg.cursor()
                            cur.execute("SELECT payload FROM cloud_backups WHERE slot_key = %s", (key,))
                            row = cur.fetchone()
                            if row and row[0]:
                                payload = row[0]
                        except Exception as e:
                            messagebox.showerror("Error", f"Failed to fetch cloud backup: {e}", parent=top)
                            return

                if not payload:
                    messagebox.showerror("Error", "Could not retrieve backup data for the selected entry.", parent=top)
                    return

                safe_key = str(key).replace("::", "_").replace(" ", "_")
                save_path = filedialog.asksaveasfilename(
                    title="Save Backup File As",
                    initialfile=f"payroll_backup_{safe_key}.json.gz",
                    defaultextension=".json.gz",
                    filetypes=[
                        ("Compressed Backup (*.json.gz)", "*.json.gz"),
                        ("JSON Snapshot (*.json)", "*.json"),
                        ("All Files (*.*)", "*.*"),
                    ],
                    parent=top,
                )
                if not save_path:
                    return

                try:
                    snapshot = _decode_cloud_backup_payload(payload)
                    if not snapshot:
                        messagebox.showerror("Download Error", "Could not decode backup payload.", parent=top)
                        return
                    if save_path.endswith(".json"):
                        with open(save_path, "w", encoding="utf-8") as f:
                            json.dump(snapshot, f, indent=2, default=str)
                    else:
                        raw_json = json.dumps(snapshot, default=str).encode("utf-8")
                        compressed = gzip.compress(raw_json)
                        with open(save_path, "wb") as f:
                            f.write(compressed)
                    log_user_action("backup_download", extra_summary=f"Downloaded backup {key} to {os.path.basename(save_path)}")
                    messagebox.showinfo("Download Complete", f"Backup successfully saved to:\n{save_path}", parent=top)
                except Exception as exc:
                    messagebox.showerror("Download Failed", f"Failed to save backup file:\n{exc}", parent=top)

            def do_load_backup_file():
                file_path = filedialog.askopenfilename(
                    title="Select Backup File to Load & Restore",
                    filetypes=[
                        ("All Supported Backups (*.json.gz, *.json, *.enc)", "*.json.gz;*.json;*.enc"),
                        ("Compressed Snapshot (*.json.gz)", "*.json.gz"),
                        ("JSON Snapshot (*.json)", "*.json"),
                        ("Encrypted Database (*.enc)", "*.enc"),
                        ("All Files (*.*)", "*.*"),
                    ],
                    parent=top,
                )
                if not file_path:
                    return

                fn = os.path.basename(file_path)
                if not messagebox.askyesno(
                    "Confirm Restore from File",
                    f"Restoring from file will replace current application data with the contents of:\n\n{fn}\n\nAre you sure you want to proceed?",
                    icon="warning",
                    parent=top,
                ):
                    return

                try:
                    if file_path.endswith(".enc"):
                        active_enc = os.path.join(get_default_app_dir(), "payroll_data.enc")
                        import shutil
                        shutil.copy2(file_path, active_enc)
                        load_database()
                        log_user_action("backup_load_file", extra_summary=f"Restored database from file: {fn}")
                        messagebox.showinfo("Restore Success", f"Database successfully restored from:\n{fn}", parent=top)
                        try:
                            self._schedule_soft_ui_refresh(full=True)
                        except Exception:
                            pass
                        refresh_all()
                        return

                    snapshot = None
                    if file_path.endswith(".json.gz"):
                        with open(file_path, "rb") as f:
                            gz_data = f.read()
                        raw_str = gzip.decompress(gz_data).decode("utf-8")
                        snapshot = json.loads(raw_str)
                    else:
                        with open(file_path, "r", encoding="utf-8") as f:
                            content = f.read()
                        snapshot = _decode_cloud_backup_payload(content)
                        if not snapshot:
                            snapshot = json.loads(content)

                    if not snapshot or not isinstance(snapshot.get("tables"), dict):
                        messagebox.showerror("Invalid File", "The selected file does not contain valid payroll backup data.", parent=top)
                        return

                    ok, msg = restore_snapshot_dict(snapshot, source_name=f"file:{fn}")
                    if ok:
                        messagebox.showinfo("Restore Success", f"Successfully loaded and restored backup from:\n{fn}", parent=top)
                        try:
                            self._schedule_soft_ui_refresh(full=True)
                        except Exception:
                            pass
                        refresh_all()
                    else:
                        messagebox.showerror("Restore Failed", str(msg), parent=top)
                except Exception as exc:
                    messagebox.showerror("Restore Error", f"Failed to restore from file:\n{exc}", parent=top)

            def refresh_all():
                load_logs()
                load_backups()

            user_filter.bind("<<ComboboxSelected>>", lambda e: load_logs())
            btn_container = tb.Frame(bak_lf)
            btn_container.pack(fill=X, pady=(10, 0))

            b_refresh = tb.Button(btn_container, text=self._tr("🔄 Refresh"), bootstyle="secondary outline", command=refresh_all)
            b_backup = tb.Button(btn_container, text=self._tr("📥 Backup Now (Local + Cloud)"), bootstyle="success", command=do_backup_now)
            b_restore = tb.Button(btn_container, text=self._tr("Restore Selected Backup"), bootstyle="warning outline", command=do_restore)
            b_download = tb.Button(btn_container, text=self._tr("💾 Download Selected File"), bootstyle="info outline", command=do_download_backup)
            b_load = tb.Button(btn_container, text=self._tr("📂 Load Backup from Disk..."), bootstyle="primary outline", command=do_load_backup_file)

            action_buttons = [b_refresh, b_backup, b_restore, b_download, b_load]
            button_subrows = []

            def reflow_buttons(event=None):
                avail_w = btn_container.winfo_width()
                if avail_w <= 10:
                    return
                for b in action_buttons:
                    b.pack_forget()
                for f in button_subrows:
                    f.destroy()
                button_subrows.clear()

                cur_f = tb.Frame(btn_container)
                cur_f.pack(fill=X, pady=2)
                button_subrows.append(cur_f)
                cur_w = 0
                pad = 8
                for b in action_buttons:
                    req_w = b.winfo_reqwidth()
                    if cur_w > 0 and (cur_w + req_w + pad) > avail_w:
                        cur_f = tb.Frame(btn_container)
                        cur_f.pack(fill=X, pady=2)
                        button_subrows.append(cur_f)
                        cur_w = 0
                    b.pack(in_=cur_f, side=LEFT, padx=(0, pad))
                    cur_w += req_w + pad

            btn_container.bind("<Configure>", reflow_buttons)
            refresh_all()

        def _build_database_and_cloud_panel(self, parent, dialog):
            """Embeds Supabase Cloud settings with live 1GB storage meter, and local directory changer."""
            export_frame = tb.Labelframe(parent, text=self._tr("Revenue Data Export"), padding=(15, 10), bootstyle="secondary")
            export_frame.pack(side=BOTTOM, fill=X, padx=10, pady=(5, 10))

            tb.Label(
                export_frame,
                text=self._tr("Export the full table of revenue records and calculations to CSV or Excel:"),
                font=("Segoe UI", 9),
                bootstyle="secondary",
            ).pack(side=LEFT, padx=(5, 15))

            tb.Button(
                export_frame,
                text=self._tr("Export Revenue to CSV / Excel"),
                bootstyle="success",
                cursor="hand2",
                command=self.export_excel,
            ).pack(side=RIGHT, padx=5)

            db_notebook = tb.Notebook(parent, bootstyle="info")
            db_notebook.pack(side=TOP, fill=BOTH, expand=True, padx=10, pady=(10, 5))

            def _create_scrollable_subtab(parent_nb, title):
                tab = tb.Frame(parent_nb)
                parent_nb.add(tab, text=title)
                canvas = tk.Canvas(tab, highlightthickness=0)
                scrollbar = tb.Scrollbar(tab, orient="vertical", command=canvas.yview)
                scroll_content = tb.Frame(canvas, padding=15)
                scroll_content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
                win_id = canvas.create_window((0, 0), window=scroll_content, anchor="nw")
                def _on_cfg(e):
                    canvas.itemconfigure(win_id, width=e.width)
                canvas.bind("<Configure>", _on_cfg)
                canvas.configure(yscrollcommand=scrollbar.set)
                
                def _wheel(e):
                    try:
                        delta = int(-1 * (e.delta / 120)) if getattr(e, "delta", 0) else (1 if getattr(e, "num", 0) == 5 else -1)
                        canvas.yview_scroll(delta, "units")
                    except Exception:
                        pass
                def _bind_w(w):
                    try:
                        w.bind("<MouseWheel>", _wheel, add="+")
                        w.bind("<Button-4>", _wheel, add="+")
                        w.bind("<Button-5>", _wheel, add="+")
                    except Exception:
                        pass
                    for ch in w.winfo_children():
                        _bind_w(ch)
                dialog.after(120, lambda: _bind_w(scroll_content))
                dialog.after(120, lambda: _bind_w(canvas))
                
                canvas.pack(side=LEFT, fill=BOTH, expand=True)
                scrollbar.pack(side=RIGHT, fill=Y)
                return tab, scroll_content

            config_data = get_db_config()

            # --- SUB-TAB 1: SUPABASE CLOUD ---
            tab_remote_wrapper, tab_remote = _create_scrollable_subtab(db_notebook, self._tr("☁️ Supabase Cloud Database"))
            
            # --- Cloud Sync Action Bar ---
            sync_card = tb.Labelframe(tab_remote, text=self._tr("Cloud Synchronization"), padding=12, bootstyle="primary")
            sync_card.pack(fill=X, pady=(0, 10))

            self.btn_sync_now = tb.Button(
                sync_card,
                text="🔄 " + self._tr("Sync Cloud Now"),
                bootstyle="success",
                cursor="hand2",
                command=self.manual_sync_cloud,
            )
            self.btn_sync_now.pack(side=LEFT, padx=(0, 10))

            is_cloud = (get_db_mode() == "supabase")
            sync_status_txt = "☁️ " + self._tr("Connected") + " (Live Cloud Database)" if is_cloud else "💻 Local Database Mode"
            tb.Label(sync_card, text=sync_status_txt, font=("Segoe UI", 10, "bold"), bootstyle="info" if is_cloud else "secondary").pack(side=LEFT, padx=5)

            # --- Live 1 GB Free Tier Storage Meter ---
            storage_card = tb.Labelframe(tab_remote, text=self._tr("📊 Supabase Cloud Storage (1 GB Free Tier)"), padding=12, bootstyle="info")
            storage_card.pack(fill=X, pady=(5, 12))
            
            lbl_storage_txt = tb.Label(storage_card, text=self._tr("Storage Used: Checking..."), font=("Segoe UI", 11, "bold"), bootstyle="primary")
            lbl_storage_txt.pack(anchor=W, pady=(0, 4))
            
            storage_pbar = tb.Progressbar(storage_card, bootstyle="success-striped", maximum=100, value=0)
            storage_pbar.pack(fill=X, pady=(0, 6))
            
            sub_row = tb.Frame(storage_card)
            sub_row.pack(fill=X)
            lbl_storage_sub = tb.Label(sub_row, text=self._tr("Database Tables, Row History & Uploaded Documents"), font=("Segoe UI", 9), bootstyle="secondary")
            lbl_storage_sub.pack(side=LEFT)
            
            def refresh_storage_meter():
                lbl_storage_txt.config(text=self._tr("Storage Used: Calculating..."))
                def _fetch():
                    used_b, tot_b, pct, txt = get_supabase_storage_usage()
                    def _update():
                        if not dialog.winfo_exists():
                            return
                        lbl_storage_txt.config(text=f"{self._tr('Storage Used:')} {txt}")
                        storage_pbar["value"] = pct
                        if pct < 70.0:
                            storage_pbar.configure(bootstyle="success-striped")
                        elif pct < 90.0:
                            storage_pbar.configure(bootstyle="warning-striped")
                        else:
                            storage_pbar.configure(bootstyle="danger-striped")
                    try:
                        dialog.after(0, _update)
                    except Exception:
                        pass
                import threading
                threading.Thread(target=_fetch, daemon=True).start()
                
            tb.Button(sub_row, text=self._tr("🔄 Check Live Storage"), bootstyle="info outline", cursor="hand2", command=refresh_storage_meter).pack(side=RIGHT)
            try:
                dialog.after(300, refresh_storage_meter)
            except Exception:
                pass

            tb.Label(tab_remote, text=self._tr("Supabase DB Host / Project Endpoint:"), font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(5, 5))
            db_host_var = tk.StringVar(value=config_data.get("supabase_host") or "db.xxxx.supabase.co")
            db_host_entry = tb.Entry(tab_remote, textvariable=db_host_var, width=45)
            db_host_entry.pack(pady=5, fill=X)
            
            tb.Label(tab_remote, text=self._tr("Database Password:"), font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(5, 5))
            db_password_var = tk.StringVar(value=config_data.get("supabase_password") or "")
            db_password_entry = tb.Entry(tab_remote, show="*", textvariable=db_password_var, width=45)
            db_password_entry.pack(pady=5, fill=X)
            
            extra_frame = tb.Frame(tab_remote)
            extra_frame.pack(fill=X, pady=10)
            
            tb.Label(extra_frame, text=self._tr("DB Name:"), font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky=W, padx=(0, 10))
            db_name_var = tk.StringVar(value=config_data.get("supabase_database") or "postgres")
            db_name_entry = tb.Entry(extra_frame, textvariable=db_name_var, width=15)
            db_name_entry.grid(row=0, column=1, sticky=W)
            
            tb.Label(extra_frame, text=self._tr("Port:"), font=("Segoe UI", 10, "bold")).grid(row=0, column=2, sticky=W, padx=(20, 10))
            db_port_var = tk.StringVar(value=config_data.get("supabase_port") or "5432")
            db_port_entry = tb.Entry(extra_frame, textvariable=db_port_var, width=10)
            db_port_entry.grid(row=0, column=3, sticky=W)

            tb.Label(tab_remote, text=self._tr("DB Username:"), font=("Segoe UI", 10, "bold")).pack(anchor=W, pady=(5, 5))
            db_user_var = tk.StringVar(value=config_data.get("supabase_user") or "postgres")
            db_user_entry = tb.Entry(tab_remote, textvariable=db_user_var, width=45)
            db_user_entry.pack(pady=5, fill=X)

            def test_and_save_supabase():
                raw_cfg = {
                    "supabase_host": db_host_var.get().strip(),
                    "supabase_password": db_password_var.get().strip(),
                    "supabase_port": db_port_var.get().strip(),
                    "supabase_database": db_name_var.get().strip(),
                    "supabase_user": db_user_var.get().strip(),
                }
                clean_cfg = _clean_supabase_config(raw_cfg)
                host = clean_cfg["supabase_host"]
                password = clean_cfg["supabase_password"]
                port = clean_cfg["supabase_port"]
                database = clean_cfg["supabase_database"]
                username = clean_cfg["supabase_user"]

                db_host_var.set(host)
                db_port_var.set(str(port))
                db_user_var.set(username)
                db_name_var.set(database)
                
                if not host or not password:
                    messagebox.showerror("Validation Error", "Host and Password are required.", parent=dialog)
                    return
                try:
                    import pg8000.dbapi
                except ImportError:
                    messagebox.showerror("Error", "pg8000 missing. Run pip install pg8000", parent=dialog)
                    return
                try:
                    conn = pg8000.dbapi.connect(
                        host=host,
                        port=int(port),
                        user=username,
                        password=password,
                        database=database,
                        timeout=10
                    )
                    cursor = conn.cursor()
                    cursor.execute("SELECT 1")
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Connection Failed", f"Could not connect: {e}", parent=dialog)
                    return
                    
                default_dir = get_default_app_dir()
                config_file = os.path.join(default_dir, "location_config.json")
                try:
                    import json
                    new_config = get_db_config().copy()
                    new_config["mode"] = "supabase"
                    new_config["supabase_host"] = host
                    new_config["supabase_password"] = password
                    new_config["supabase_port"] = str(port)
                    new_config["supabase_database"] = database
                    new_config["supabase_user"] = username
                    with open(config_file, "w", encoding="utf-8") as f:
                        json.dump(new_config, f, indent=4)
                    refresh_storage_meter()
                    messagebox.showinfo("Success", "Supabase configuration verified and saved!\n\nPlease restart the app to activate Supabase mode.", parent=dialog)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save configuration: {e}", parent=dialog)

            def do_upload_local_to_cloud():
                raw_cfg = {
                    "supabase_host": db_host_var.get().strip(),
                    "supabase_password": db_password_var.get().strip(),
                    "supabase_port": db_port_var.get().strip(),
                    "supabase_database": db_name_var.get().strip(),
                    "supabase_user": db_user_var.get().strip(),
                }
                clean_cfg = _clean_supabase_config(raw_cfg)
                host = clean_cfg["supabase_host"]
                password = clean_cfg["supabase_password"]
                if not host or not password:
                    messagebox.showerror("Validation Error", "Host and Password are required to upload.", parent=dialog)
                    return
                status = {"text": "Starting..."}
                self.show_busy("Safely uploading local database to cloud…")
                try:
                    def _progress(t):
                        status.update(text=t)
                        try:
                            self._busy_msg_var.set(t)
                            self.update_idletasks()
                        except Exception:
                            pass
                    upload_local_database_to_supabase(progress_cb=_progress)
                    
                    # Only save mode = supabase after a 100% verified upload
                    try:
                        import json
                        default_dir = get_default_app_dir()
                        config_file = os.path.join(default_dir, "location_config.json")
                        new_config = get_db_config().copy()
                        new_config["mode"] = "supabase"
                        new_config["supabase_host"] = host
                        new_config["supabase_password"] = password
                        new_config["supabase_port"] = str(clean_cfg["supabase_port"])
                        new_config["supabase_database"] = clean_cfg["supabase_database"]
                        new_config["supabase_user"] = clean_cfg["supabase_user"]
                        with open(config_file, "w", encoding="utf-8") as f:
                            json.dump(new_config, f, indent=4)
                    except Exception:
                        pass
                    
                    self.hide_busy()
                    refresh_storage_meter()
                    messagebox.showinfo(
                        "Upload Complete",
                        f"Local database successfully uploaded to Supabase.\n\nStatus: {status.get('text')}\n\n"
                        "Restart this app, then on the 2nd device enter the same Supabase settings and restart.",
                        parent=dialog,
                    )
                except Exception as e:
                    self.hide_busy()
                    # Safe fallback: Ensure local mode stays active
                    try:
                        import json
                        default_dir = get_default_app_dir()
                        config_file = os.path.join(default_dir, "location_config.json")
                        cfg = get_db_config().copy()
                        cfg["mode"] = "local"
                        with open(config_file, "w", encoding="utf-8") as f:
                            json.dump(cfg, f, indent=4)
                    except Exception:
                        pass
                    messagebox.showwarning(
                        "Upload Incomplete - Local Data Safe",
                        f"Cloud upload encountered an issue:\n{e}\n\n"
                        "Your local database is completely intact and was NOT modified.\n"
                        "The application will continue running in Local Mode with all your data.",
                        parent=dialog,
                    )

            def do_cleanup_cloud():
                if not messagebox.askyesno(
                    "Clean Up Cloud",
                    "Remove duplicate Shop/config rows and clear sync history noise on Supabase?",
                    parent=dialog,
                ):
                    return
                self.show_busy("Cleaning cloud duplicates…")
                try:
                    if self.is_busy_cancelled():
                        return
                    cleanup_supabase_duplicates()
                    self.hide_busy()
                    refresh_storage_meter()
                    messagebox.showinfo("Cleanup Complete", "Cloud duplicates cleaned.", parent=dialog)
                except Exception as e:
                    self.hide_busy()
                    self.show_app_error("Cleanup Failed", e, parent=dialog)
            
            def do_sync_files():
                self.show_busy("Syncing documents to cloud…")
                try:
                    ok, msg = sync_all_local_files_to_cloud()
                    self.hide_busy()
                    refresh_storage_meter()
                    if ok:
                        messagebox.showinfo("Files Synced", str(msg), parent=dialog)
                    else:
                        messagebox.showerror("Sync Failed", str(msg), parent=dialog)
                except Exception as e:
                    self.hide_busy()
                    messagebox.showerror("Sync Error", str(e), parent=dialog)

            btn_row = tb.Frame(tab_remote)
            btn_row.pack(pady=10, fill=X)
            tb.Button(btn_row, text=self._tr("Verify & Save Configuration"), bootstyle="success", command=test_and_save_supabase).pack(side=LEFT, padx=(0, 8))
            tb.Button(btn_row, text="Upload Local → Cloud", bootstyle="warning", command=do_upload_local_to_cloud).pack(side=LEFT, padx=(0, 8))
            tb.Button(btn_row, text="📁 Sync Local Files → Cloud", bootstyle="info", command=do_sync_files).pack(side=LEFT, padx=(0, 8))
            tb.Button(btn_row, text="Clean Up Cloud Duplicates", bootstyle="secondary", command=do_cleanup_cloud).pack(side=LEFT)

            # --- Multi-Device First-Time Integration Guide ---
            guide_card = tb.Labelframe(tab_remote, text=self._tr("📖 First-Time Cloud Setup & Multi-Device Integration Guide"), padding=14, bootstyle="info")
            guide_card.pack(fill=X, pady=(15, 10))

            is_ar = (getattr(self, 'lang', 'en') == 'ar')
            if is_ar:
                guide_lines = (
                    "📋 خطوات ربط ومزامنة جهاز جديد للمرة الأولى:\n\n"
                    "1️⃣ الجهاز الرئيسي (الجهاز الأول الذي يحتوي على البيانات الأصلية):\n"
                    "   • أدخل بيانات Supabase: المضيف (Host)، كلمة مرور قاعدة البيانات، المنفذ (6543)، واسم المستخدم.\n"
                    "   • اضغط على 'التحقق وحفظ الإعدادات' للتأكد من الاتصال بنجاح.\n"
                    "   • اضغط على 'Upload Local → Cloud' لرفع كامل سجلات الرواتب والموظفين والمصروفات إلى السحابة.\n"
                    "   • اضغط على '📁 Sync Local Files → Cloud' لرفع مستندات وصور الموظفين.\n\n"
                    "2️⃣ الجهاز الثاني (الكمبيوتر المحمول أو الجهاز الإضافي):\n"
                    "   • افتح نفس هذه النافذة (إعدادات قاعدة البيانات والسحابة) على الجهاز الجديد.\n"
                    "   • أدخل نفس إعدادات وبيانات Supabase تماماً واضغط 'التحقق وحفظ الإعدادات'.\n"
                    "   • اضغط على زر '🔄 مزامنة السحابة الآن' لتحميل كامل قاعدة البيانات والملفات تلقائياً للجهاز.\n\n"
                    "3️⃣ الاستخدام اليومي والمزامنة التلقائية:\n"
                    "   • تتم مزامنة التغييرات تلقائياً في الخلفية بين كل الأجهزة المرتبطة.\n"
                    "   • يمكنك الضغط على '🔄 مزامنة السحابة الآن' في أي وقت لإجراء مزامنة فورية وتحديث كل الجداول."
                )
            else:
                guide_lines = (
                    "📋 Step-by-Step Multi-Device Integration Guide:\n\n"
                    "1️⃣ Primary PC (Source of Truth with Existing Data):\n"
                    "   • Enter your Supabase credentials: Host, Database Password, Port (6543), Database (postgres), and User.\n"
                    "   • Click 'Verify & Save Configuration' to confirm the cloud connection succeeds.\n"
                    "   • Click 'Upload Local → Cloud' to push all existing payroll records, employees, and expenses to Supabase.\n"
                    "   • Click '📁 Sync Local Files → Cloud' to upload employee documents and photos.\n\n"
                    "2️⃣ Secondary Device (Laptop or Second Workstation):\n"
                    "   • Open this exact Database & Cloud Settings tab on the second device.\n"
                    "   • Enter the identical Supabase credentials and click 'Verify & Save Configuration'.\n"
                    "   • Click '🔄 Sync Cloud Now' at the top of this tab to download the entire database and documents.\n\n"
                    "3️⃣ Daily Continuous Synchronization:\n"
                    "   • All changes synchronize automatically in the background between active devices.\n"
                    "   • You can click '🔄 Sync Cloud Now' at any time to force an instant refresh across all tabs."
                )

            lbl_guide = tb.Label(
                guide_card,
                text=guide_lines,
                font=("Segoe UI", 9),
                justify=RIGHT if is_ar else LEFT,
                wraplength=640,
            )
            lbl_guide.pack(fill=X, padx=5, pady=5)

            # --- SUB-TAB 2: LOCAL DIR ---
            tab_local_wrapper, tab_local = _create_scrollable_subtab(db_notebook, self._tr("📁 Local Storage Directory"))
            tb.Label(tab_local, text=self._tr("Current storage directory:"), font=("Segoe UI", 10, "bold")).pack(pady=(10, 5))
            current_dir = get_default_app_dir()
            saved_dir = config_data.get("custom_db_directory") or current_dir
            lbl_current = tb.Label(tab_local, text=saved_dir, font=("Segoe UI", 9), wraplength=500, bootstyle="secondary")
            lbl_current.pack(pady=5, padx=20)
            selected_path = tk.StringVar(value=saved_dir)
            
            def browse_folder():
                folder = filedialog.askdirectory(parent=dialog, initialdir=saved_dir, title="Select Local/Sync Directory")
                if folder:
                    selected_path.set(folder)
                    lbl_new.config(text=folder)
            
            tb.Button(tab_local, text=self._tr("📁 Select New Directory"), bootstyle="info outline", command=browse_folder).pack(pady=10)
            tb.Label(tab_local, text=self._tr("New target directory:"), font=("Segoe UI", 10, "bold")).pack(pady=(10, 5))
            lbl_new = tb.Label(tab_local, text=saved_dir, font=("Segoe UI", 9, "italic"), wraplength=500, bootstyle="success")
            lbl_new.pack(pady=5, padx=20)
            
            copy_files_var = tk.BooleanVar(value=True)
            chk_copy = tb.Checkbutton(tab_local, text=self._tr("Copy database & employee folders if not present"), variable=copy_files_var, bootstyle="success-round-toggle")
            chk_copy.pack(pady=10)
            
            def save_local_config():
                new_dir = selected_path.get()
                if not new_dir or not os.path.exists(new_dir):
                    messagebox.showerror("Error", "Selected directory does not exist.", parent=dialog)
                    return
                default_dir = get_default_app_dir()
                real_new = os.path.abspath(new_dir)
                config_file = os.path.join(default_dir, "location_config.json")
                try:
                    import json
                    new_config = {
                        "mode": "local",
                        "custom_db_directory": real_new if real_new != os.path.abspath(default_dir) else None
                    }
                    with open(config_file, "w", encoding="utf-8") as f:
                        json.dump(new_config, f, indent=4)
                    messagebox.showinfo("Restart Required", "Storage directory successfully updated!\n\nPlease restart the app to apply the changes.", parent=dialog)
                    dialog.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save: {e}", parent=dialog)
            
            def reset_local_default():
                selected_path.set(get_default_app_dir())
                lbl_new.config(text=get_default_app_dir())
                
            local_btn_frame = tb.Frame(tab_local)
            local_btn_frame.pack(pady=15)
            tb.Button(local_btn_frame, text=self._tr("Reset to Default"), bootstyle="secondary", command=reset_local_default).pack(side=LEFT, padx=10)
            tb.Button(local_btn_frame, text=self._tr("Save Configuration"), bootstyle="success", command=save_local_config).pack(side=LEFT, padx=10)

        def _build_app_updates_panel(self, parent, dialog):
            """Settings panel for cloud in-app updates, fail-safe recovery, and Supabase audit history."""
            top = dialog
            
            canvas = tk.Canvas(parent, highlightthickness=0, borderwidth=0)
            vscroll = tb.Scrollbar(parent, orient=VERTICAL, command=canvas.yview)
            inner = tb.Frame(canvas, padding=16)
            
            inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            win_id = canvas.create_window((0, 0), window=inner, anchor="nw")
            canvas.configure(yscrollcommand=vscroll.set)
            
            def _on_canvas_configure(e):
                try:
                    canvas.itemconfigure(win_id, width=e.width)
                except Exception:
                    pass
            canvas.bind("<Configure>", _on_canvas_configure)
            
            vscroll.pack(side=RIGHT, fill=Y)
            canvas.pack(side=LEFT, fill=BOTH, expand=True)

            # --- 1. CURRENT ENGINE STATUS CARD ---
            info = get_active_code_info()
            status_lf = tb.Labelframe(inner, text=self._tr("🚀 Current Engine & Version Status"), padding=14, bootstyle="info")
            status_lf.pack(fill=X, pady=(0, 14))
            
            status_grid = tb.Frame(status_lf)
            status_grid.pack(fill=X)
            
            if info["is_safe_mode"]:
                engine_txt = "⚠️ Built-in Factory Engine (Safe Mode Fallback)"
                engine_style = "warning"
            elif info["is_dynamic"]:
                engine_txt = "🟢 Cloud-Updated Engine"
                engine_style = "success"
            else:
                engine_txt = "📦 Built-in Factory Engine"
                engine_style = "info"
                
            tb.Label(status_grid, text="Active Engine:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky=W, pady=3, padx=(0, 12))
            tb.Label(status_grid, text=engine_txt, font=("Segoe UI", 10, "bold"), bootstyle=engine_style).grid(row=0, column=1, sticky=W, pady=3)
            
            tb.Label(status_grid, text="Active Version:", font=("Segoe UI", 10, "bold")).grid(row=1, column=0, sticky=W, pady=3, padx=(0, 12))
            build_date_disp = info.get('build_date') or APP_BUILD_DATE
            tb.Label(status_grid, text=f"v{info['version']} (Build Date: {build_date_disp})", font=("Segoe UI", 10)).grid(row=1, column=1, sticky=W, pady=3)

            tb.Label(status_grid, text="System Status:", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky=W, pady=3, padx=(0, 12))
            tb.Label(status_grid, text="✅ System Verified & Ready", font=("Segoe UI", 10), bootstyle="success").grid(row=2, column=1, sticky=W, pady=3)

            # Safe mode / Crash diagnostics box
            safe_flag = get_safe_mode_flag_path()
            crash_log = get_last_crash_log_path()
            if os.path.isfile(safe_flag) or info["is_safe_mode"]:
                safe_box = tb.Frame(status_lf, bootstyle="danger", padding=10)
                safe_box.pack(fill=X, pady=(10, 0))
                
                safe_msg = "⚠️ A downloaded update crashed on launch. The app automatically fell back to the safe built-in engine.\nYou can check for a cloud update or roll back."
                if os.path.isfile(crash_log):
                    try:
                        with open(crash_log, "r", encoding="utf-8") as lf:
                            first_lines = "".join([lf.readline() for _ in range(3)]).strip()
                            if first_lines:
                                safe_msg += f"\n\nCrash summary:\n{first_lines}"
                    except Exception:
                        pass
                tb.Label(safe_box, text=safe_msg, bootstyle="inverse-danger", font=("Segoe UI", 9), wraplength=650, justify=LEFT).pack(anchor=W, pady=(0, 8))
                
                def _view_crash_log():
                    if not os.path.isfile(crash_log):
                        messagebox.showinfo("Crash Log", "No crash log file found.", parent=top)
                        return
                    try:
                        with open(crash_log, "r", encoding="utf-8") as f:
                            content = f.read()
                    except Exception as ce:
                        content = f"Error reading crash log: {ce}"
                    c_dlg = tb.Toplevel(top)
                    c_dlg.title("Last Crash Diagnostics & Traceback")
                    c_dlg.geometry("700x500")
                    c_txt = tk.Text(c_dlg, wrap=WORD, font=("Consolas", 10))
                    c_txt.pack(fill=BOTH, expand=True, padx=10, pady=10)
                    c_txt.insert("1.0", content)
                    c_txt.configure(state="disabled")
                    tb.Button(c_dlg, text="Close", bootstyle="secondary", command=c_dlg.destroy).pack(pady=(0, 10))

                def _clear_safe_mode():
                    if os.path.isfile(safe_flag):
                        try:
                            os.remove(safe_flag)
                        except Exception:
                            pass
                    messagebox.showinfo("Safe Mode Cleared", "Safe mode flag cleared. The app will attempt to run the update on next restart.", parent=top)
                    safe_box.pack_forget()

                btn_c_row = tb.Frame(safe_box, bootstyle="danger")
                btn_c_row.pack(anchor=W)
                tb.Button(btn_c_row, text="📋 View Full Crash Report", bootstyle="light", command=_view_crash_log).pack(side=LEFT, padx=(0, 8))
                tb.Button(btn_c_row, text="🧹 Clear Safe Mode Flag", bootstyle="warning", command=_clear_safe_mode).pack(side=LEFT)

            # --- 2. CLOUD UPDATE & CHECK ACTIONS ---
            upd_lf = tb.Labelframe(inner, text=self._tr("🔄 Cloud Software Updates"), padding=14, bootstyle="primary")
            upd_lf.pack(fill=X, pady=(0, 14))

            tb.Label(
                upd_lf,
                text=self._tr("Check for and install the latest official software updates directly over the cloud:"),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=680,
                justify=LEFT,
            ).pack(anchor=W, pady=(0, 10))

            url_row = tb.Frame(upd_lf)
            url_row.pack(fill=X, pady=(0, 10))
            tb.Label(url_row, text="Update Server URL:", font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 10))
            ent_url = tb.Entry(url_row, font=("Segoe UI", 9))
            ent_url.pack(side=LEFT, fill=X, expand=True, padx=(0, 8))
            ent_url.insert(0, get_custom_update_server_url())

            def _save_url():
                set_custom_update_server_url(ent_url.get().strip())
                messagebox.showinfo("Saved", "Update server URL configuration saved.", parent=top)
            tb.Button(url_row, text="💾 Save URL", bootstyle="secondary outline", command=_save_url).pack(side=LEFT)

            lbl_check_status = tb.Label(upd_lf, text="Ready to check for updates.", font=("Segoe UI", 10, "italic"), bootstyle="secondary")
            lbl_check_status.pack(anchor=W, pady=(0, 10))

            btn_act_row = tb.Frame(upd_lf)
            btn_act_row.pack(fill=X, pady=(0, 8))

            cached_update = {"data": None}

            def _do_check_update():
                lbl_check_status.configure(text="⏳ Connecting to cloud update server and checking for updates...", bootstyle="info")
                btn_check.configure(state="disabled")

                def run_bg():
                    status, data = check_for_cloud_update()
                    def update_ui():
                        btn_check.configure(state="normal")
                        if status == "error":
                            lbl_check_status.configure(text=f"❌ Check failed: {data.get('error')}", bootstyle="danger")
                            messagebox.showerror("Update Check Failed", f"Could not fetch update from server:\n\n{data.get('error')}", parent=top)
                        elif status == "up_to_date":
                            lbl_check_status.configure(text=f"✅ You are running the latest version! (Code Hash: {data.get('remote_hash', '')[:8]})", bootstyle="success")
                            messagebox.showinfo("Up to Date", f"Your application is completely up to date!\n\nVersion: v{data.get('remote_version')}\nCode Hash: {data.get('remote_hash', '')[:12]}", parent=top)
                        elif status == "update_available":
                            cached_update["data"] = data
                            size_kb = int(data.get("size_bytes", 0)) / 1024
                            lbl_check_status.configure(
                                text=f"✨ Update Available! New Version: v{data.get('remote_version')} ({size_kb:.0f} KB). Ready to install.",
                                bootstyle="warning"
                            )
                            if messagebox.askyesno(
                                "Update Available ✨",
                                f"A new version is available!\n\n"
                                f"• New Version: v{data.get('remote_version')}\n"
                                f"• Current Version: v{info.get('version')}\n"
                                f"• Download Size: {size_kb:.0f} KB\n\n"
                                f"Would you like to download and install this update now?",
                                parent=top
                            ):
                                _do_install_update(data)
                    try:
                        top.after(0, update_ui)
                    except Exception:
                        pass

                import threading
                threading.Thread(target=run_bg, daemon=True).start()

            def _do_install_update(data=None):
                if not data:
                    data = cached_update.get("data")
                if not data or not data.get("remote_code"):
                    _do_check_update()
                    return

                code = data.get("remote_code")
                r_hash = data.get("remote_hash")
                lbl_check_status.configure(text="⏳ Validating syntax and installing update...", bootstyle="info")
                
                ok, msg = install_cloud_update(code, r_hash)
                if ok:
                    lbl_check_status.configure(text="✅ Update installed successfully!", bootstyle="success")
                    load_update_logs()
                    if messagebox.askyesno(
                        "Update Installed 🎉",
                        f"{msg}\n\nUpdate installed successfully!\n\n"
                        "• Click [Yes] to Close Application Cleanly (Recommended — just reopen when ready)\n"
                        "• Click [No] to Restart Automatically",
                        parent=top,
                    ):
                        dialog.destroy()
                        self.shutdown_app()
                    else:
                        dialog.destroy()
                        restart_app()
                else:
                    lbl_check_status.configure(text=f"❌ Installation aborted: {msg}", bootstyle="danger")
                    messagebox.showerror("Installation Failed", msg, parent=top)

            def _do_rollback_bak():
                bak_path = os.path.join(get_updates_dir(), "payroll_app.py.bak")
                if not os.path.isfile(bak_path):
                    messagebox.showwarning("Rollback", "No previous update backup (.bak) was found on this computer.", parent=top)
                    return
                mtime = datetime.fromtimestamp(os.path.getmtime(bak_path)).strftime("%Y-%m-%d %H:%M:%S")
                if not messagebox.askyesno(
                    "Roll Back Update ⏮️",
                    f"This will restore the previous update backup from:\n{mtime}\n\nContinue?",
                    parent=top,
                ):
                    return
                ok, msg = rollback_cloud_update(target="bak")
                if ok:
                    load_update_logs()
                    ans_r = messagebox.askyesnocancel(
                        "Rollback Success ⏮️",
                        f"{msg}\n\nTo apply the rollback:\n• [Yes] = Restart app automatically\n• [No] = Shutdown app cleanly now (reopen manually)\n• [Cancel] = Continue current session",
                        parent=top,
                    )
                    if ans_r is True:
                        dialog.destroy()
                        restart_app()
                    elif ans_r is False:
                        dialog.destroy()
                        self.shutdown_app()
                else:
                    messagebox.showerror("Rollback Failed", msg, parent=top)

            def _do_revert_factory():
                if not messagebox.askyesno(
                    "Revert to Factory Built-in Version 🏭",
                    "This will remove the downloaded dynamic update and return the application to its original built-in code engine.\n\nContinue?",
                    parent=top,
                ):
                    return
                ok, msg = rollback_cloud_update(target="factory")
                if ok:
                    load_update_logs()
                    ans_r = messagebox.askyesnocancel(
                        "Reverted to Factory 🏭",
                        f"{msg}\n\nTo apply the built-in version:\n• [Yes] = Restart app automatically\n• [No] = Shutdown app cleanly now (reopen manually)\n• [Cancel] = Continue current session",
                        parent=top,
                    )
                    if ans_r is True:
                        dialog.destroy()
                        restart_app()
                    elif ans_r is False:
                        dialog.destroy()
                        self.shutdown_app()
                else:
                    messagebox.showerror("Revert Failed", msg, parent=top)

            btn_check = tb.Button(btn_act_row, text=self._tr("🔄 Check for Updates Now"), bootstyle="primary", command=_do_check_update)
            btn_check.pack(side=LEFT, padx=(0, 8))

            tb.Button(btn_act_row, text=self._tr("⬇️ Install Update Now"), bootstyle="success", command=lambda: _do_install_update()).pack(side=LEFT, padx=(0, 8))
            tb.Button(btn_act_row, text=self._tr("⏮️ Roll Back to Previous Backup"), bootstyle="warning outline", command=_do_rollback_bak).pack(side=LEFT, padx=(0, 8))
            tb.Button(btn_act_row, text=self._tr("🏭 Revert to Factory Built-in"), bootstyle="danger outline", command=_do_revert_factory).pack(side=LEFT)

            # --- 3. UPDATE AUDIT & CLOUD LOGS ---
            hist_lf = tb.Labelframe(inner, text=self._tr("📋 Supabase Update & Recovery Audit History"), padding=12, bootstyle="secondary")
            hist_lf.pack(fill=BOTH, expand=True, pady=(0, 8))

            hist_cols = ("When", "User", "Action", "Summary")
            hist_holder = tb.Frame(hist_lf)
            hist_holder.pack(fill=BOTH, expand=True)
            hist_tree = tb.Treeview(hist_holder, columns=hist_cols, show="headings", height=7, bootstyle="secondary")
            hist_tree.heading("When", text=self._tr("Date & Time"))
            hist_tree.heading("User", text=self._tr("Device / User"))
            hist_tree.heading("Action", text=self._tr("Action Type"))
            hist_tree.heading("Summary", text=self._tr("Details"))
            hist_tree.column("When", width=140, anchor=CENTER)
            hist_tree.column("User", width=150, anchor=W)
            hist_tree.column("Action", width=150, anchor=CENTER)
            hist_tree.column("Summary", width=300, anchor=W)
            self._attach_tree_scrollbars(hist_holder, hist_tree)

            def load_update_logs():
                for item in hist_tree.get_children():
                    hist_tree.delete(item)
                
                rows_found = []
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute(
                        """
                        SELECT created_at, user_name, action, summary
                        FROM user_action_log
                        WHERE action LIKE 'app_%' OR action LIKE 'update_%'
                        ORDER BY id DESC LIMIT 50
                        """
                    )
                    for r in cur.fetchall() or []:
                        rows_found.append((r[0] or "", plain_label(r[1]) or "", r[2] or "", r[3] or ""))
                    conn.close()
                except Exception:
                    pass

                if get_db_mode() == "supabase" and not is_supabase_offline():
                    try:
                        pg = get_shared_supabase_conn()
                        cur = pg.cursor()
                        cur.execute(
                            """
                            SELECT created_at, user_name, action, summary
                            FROM user_action_log
                            WHERE action LIKE 'app_%' OR action LIKE 'update_%'
                            ORDER BY created_at DESC LIMIT 50
                            """
                        )
                        for r in cur.fetchall() or []:
                            rows_found.append((r[0] or "", plain_label(r[1]) or "", r[2] or "", r[3] or ""))
                    except Exception:
                        pass

                seen = set()
                final_rows = []
                for rf in rows_found:
                    sig = (rf[0], rf[1], rf[2], rf[3])
                    if sig not in seen:
                        seen.add(sig)
                        final_rows.append(rf)
                final_rows.sort(key=lambda x: str(x[0]), reverse=True)

                for r in final_rows[:40]:
                    act = r[2]
                    act_label = {
                        "app_update": "🚀 Update Installed",
                        "app_rollback": "⏮️ Rolled Back",
                        "update_crash": "⚠️ Crash / Safe Mode",
                        "app_update_check": "🔍 Update Checked",
                        "update_syntax_error": "❌ Syntax Rejected",
                    }.get(act, act)
                    hist_tree.insert("", tk.END, values=(r[0], r[1], act_label, r[3]))

            tb.Button(hist_lf, text=self._tr("🔄 Refresh History Log"), bootstyle="secondary outline", command=load_update_logs).pack(anchor=W, pady=(8, 0))
            load_update_logs()

        def open_settings_dialog(self, default_tab=None):
            dialog = tb.Toplevel(self)
            dialog.title(self._tr("⚙️ Config & Database Settings Panel"))
            try:
                self.update_idletasks()
                w = max(860, self.winfo_width())
                h = max(700, self.winfo_height())
                x = self.winfo_x()
                y = self.winfo_y()
                dialog.geometry(f"{w}x{h}+{x}+{y}")
            except Exception:
                dialog.geometry("860x700")
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()
            
            notebook = tb.Notebook(dialog, bootstyle="info")
            footer = tb.Frame(dialog)
            footer.pack(side=BOTTOM, fill=X, padx=20, pady=(6, 14))
            tb.Button(
                footer,
                text=self._tr("Close Window"),
                bootstyle="secondary",
                cursor="hand2",
                command=dialog.destroy,
            ).pack(side=RIGHT)
            notebook.pack(fill=BOTH, expand=True, padx=20, pady=(16, 4))
            
            # Helper tab builder (Dynamic list managers)
            def create_tab_editor(tab_parent, db_table, label_text):
                frame = tb.Frame(tab_parent, padding=20)
                frame.pack(fill=BOTH, expand=True)
                
                list_frame = tb.Frame(frame)
                list_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 10))
                
                scroll = tb.Scrollbar(list_frame)
                scroll.pack(side=RIGHT, fill=Y)
                
                item_listbox = tk.Listbox(list_frame, yscrollcommand=scroll.set, font=("Segoe UI", 11))
                item_listbox.pack(side=LEFT, fill=BOTH, expand=True)
                scroll.config(command=item_listbox.yview)
                
                def load_items():
                    item_listbox.delete(0, tk.END)
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    cursor.execute(f"SELECT name FROM {db_table} ORDER BY name ASC")
                    for row in cursor.fetchall():
                        item_listbox.insert(tk.END, row[0])
                    conn.close()
                    
                load_items()
                
                ctrl_frame = tb.Frame(frame)
                ctrl_frame.pack(side=RIGHT, fill=Y, padx=(10, 0))
                
                tb.Label(ctrl_frame, text=label_text, font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(0, 10))
                new_ent = tb.Entry(ctrl_frame, width=25, font=("Segoe UI", 11))
                new_ent.pack(pady=5)
                
                def add_item():
                    val = new_ent.get().strip()
                    if not val:
                        return
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    try:
                        cursor.execute(f"INSERT INTO {db_table} (name) VALUES (?)", (val,))
                        commit_and_save(conn)
                        new_ent.delete(0, tk.END)
                        self.invalidate_config_caches()
                        load_items()
                    except sqlite3.IntegrityError:
                        messagebox.showerror("Error", "This item already exists.", parent=dialog)
                    conn.close()
                    
                tb.Button(ctrl_frame, text="+ Add", bootstyle="success", command=add_item).pack(fill=X, pady=5)
                
                def delete_item():
                    sel = item_listbox.curselection()
                    if not sel:
                        messagebox.showwarning("Select", "Please select an item to delete.", parent=dialog)
                        return
                    val = item_listbox.get(sel[0])
                    if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{val}'?", parent=dialog):
                        conn = sqlite3.connect(TEMP_DB_PATH)
                        cursor = conn.cursor()
                        cursor.execute(f"DELETE FROM {db_table} WHERE name=?", (val,))
                        commit_and_save(conn)
                        conn.close()
                        self.invalidate_config_caches()
                        load_items()
                        
                tb.Button(ctrl_frame, text="🗑️ Delete Selected", bootstyle="danger", command=delete_item).pack(fill=X, pady=20)
                
            tab_loc = tb.Frame(notebook)
            notebook.add(tab_loc, text=self._tr("📍 Locations"))
            create_tab_editor(tab_loc, "config_locations", self._tr("Location Name:"))
            
            tab_cat = tb.Frame(notebook)
            notebook.add(tab_cat, text=self._tr("🏷️ Categories"))
            create_tab_editor(tab_cat, "config_categories", self._tr("Category Name:"))
            
            tab_pay = tb.Frame(notebook)
            notebook.add(tab_pay, text=self._tr("💳 Payment Types"))
            create_tab_editor(tab_pay, "config_payments", self._tr("Payment Type:"))
            
            tab_acc = tb.Frame(notebook, padding=30)
            notebook.add(tab_acc, text=self._tr("🔑 Account"))
            
            tb.Label(tab_acc, text=self._tr("Change username"), font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 8), anchor=W)
            who_now = getattr(self, "current_user", DEFAULT_ADMIN_USERNAME) or DEFAULT_ADMIN_USERNAME
            lbl_current_uname = tb.Label(
                tab_acc,
                text=f"{self._tr('Current username')}:  {who_now}",
                font=("Segoe UI", 10),
                bootstyle="secondary",
            )
            lbl_current_uname.pack(anchor=W, pady=(0, 6))
            tb.Label(tab_acc, text=self._tr("New username:"), font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=5)
            new_uname_ent = tb.Entry(tab_acc, width=30)
            new_uname_ent.insert(0, who_now)
            new_uname_ent.pack(pady=5, anchor=W)

            def change_username():
                old = getattr(self, "current_user", DEFAULT_ADMIN_USERNAME) or DEFAULT_ADMIN_USERNAME
                new = (new_uname_ent.get() or "").strip()
                if self.rename_login_username(old, new, parent=dialog):
                    lbl_current_uname.config(text=f"{self._tr('Current username')}:  {new}")
                    new_uname_ent.delete(0, tk.END)
                    new_uname_ent.insert(0, new)
                    try:
                        load_users()
                    except Exception:
                        pass

            tb.Button(tab_acc, text=self._tr("Update Username"), bootstyle="info", command=change_username).pack(anchor=W, pady=(4, 24))

            tb.Label(tab_acc, text=self._tr("Change your login password"), font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(0, 20), anchor=W)
            
            tb.Label(tab_acc, text=self._tr("Current Password:"), font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=5)
            curr_pw = tb.Entry(tab_acc, show="*", width=30)
            curr_pw.pack(pady=5)
            
            tb.Label(tab_acc, text=self._tr("New Password:"), font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=5)
            new_pw = tb.Entry(tab_acc, show="*", width=30)
            new_pw.pack(pady=5)
            
            def change_pass():
                c_val = curr_pw.get().strip()
                n_val = new_pw.get().strip()
                if not c_val or not n_val:
                    messagebox.showerror("Error", "Fields cannot be empty.", parent=dialog)
                    return
                import hashlib
                c_hash = hashlib.sha256(c_val.encode('utf-8')).hexdigest()
                n_hash = hashlib.sha256(n_val.encode('utf-8')).hexdigest()
                who = getattr(self, "current_user", DEFAULT_ADMIN_USERNAME) or DEFAULT_ADMIN_USERNAME
                
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT password FROM users WHERE username=?", (who,))
                row = cursor.fetchone()
                stored = row[0] if row else None
                stored = decrypt_val(stored) if stored is not None else None
                
                if stored == c_hash or stored == c_val:
                    cursor.execute("UPDATE users SET password=? WHERE username=?", (n_hash, who))
                    commit_and_save(conn)
                    conn.close()
                    messagebox.showinfo("Success", "Password updated successfully.", parent=dialog)
                    curr_pw.delete(0, tk.END)
                    new_pw.delete(0, tk.END)
                else:
                    conn.close()
                    messagebox.showerror("Error", "Incorrect current password.", parent=dialog)
                    
            tb.Button(tab_acc, text=self._tr("Update Password"), bootstyle="success", command=change_pass).pack(pady=20)

            tab_users = tb.Frame(notebook, padding=20)
            notebook.add(tab_users, text=self._tr("👥 Users"))
            tb.Label(tab_users, text=self._tr("Admin users (synced across PCs)"), font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(anchor=W, pady=(0, 8))
            tb.Label(
                tab_users,
                text="New users start with password  admin  until they change it.",
                font=("Segoe UI", 10),
                bootstyle="secondary",
            ).pack(anchor=W, pady=(0, 10))
            users_holder = tb.Frame(tab_users)
            users_holder.pack(fill=BOTH, expand=True)
            users_list = tk.Listbox(users_holder, font=("Segoe UI", 11), height=8)
            users_list.pack(side=LEFT, fill=BOTH, expand=True)
            users_scroll = tb.Scrollbar(users_holder, command=users_list.yview)
            users_scroll.pack(side=RIGHT, fill=Y)
            users_list.config(yscrollcommand=users_scroll.set)

            def load_users():
                users_list.delete(0, tk.END)
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute("SELECT username FROM users ORDER BY username ASC")
                    for row in cur.fetchall() or []:
                        name = plain_label(row[0]) if row and row[0] else ""
                        if name:
                            users_list.insert(tk.END, name)
                    conn.close()
                except Exception:
                    pass

            add_row = tb.Frame(tab_users)
            add_row.pack(fill=X, pady=10)
            tb.Label(add_row, text=self._tr("New username"), font=("Segoe UI", 10, "bold")).pack(side=LEFT)
            new_user_ent = tb.Entry(add_row, width=22)
            new_user_ent.pack(side=LEFT, padx=8)

            def add_user():
                uname = new_user_ent.get().strip()
                if not uname:
                    return
                hashed = hashlib.sha256(b"admin").hexdigest()
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute("INSERT INTO users (username, password) VALUES (?, ?)", (uname, hashed))
                    commit_and_save(conn)
                    conn.close()
                    new_user_ent.delete(0, tk.END)
                    load_users()
                    messagebox.showinfo(
                        "Users",
                        self._tr("Username added. Default password is admin until they change it."),
                        parent=dialog,
                    )
                except sqlite3.IntegrityError:
                    messagebox.showerror("Error", "This username already exists.", parent=dialog)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            def rename_user():
                sel = users_list.curselection()
                if not sel:
                    messagebox.showwarning("Select", "Please select a user to rename.", parent=dialog)
                    return
                old = users_list.get(sel[0])
                new = (rename_user_ent.get() or "").strip()
                if self.rename_login_username(old, new, parent=dialog):
                    rename_user_ent.delete(0, tk.END)
                    load_users()
                    who = getattr(self, "current_user", "") or ""
                    try:
                        lbl_current_uname.config(text=f"{self._tr('Current username')}:  {who}")
                        new_uname_ent.delete(0, tk.END)
                        new_uname_ent.insert(0, who)
                    except Exception:
                        pass

            def delete_user():
                sel = users_list.curselection()
                if not sel:
                    return
                uname = users_list.get(sel[0])
                if str(uname).strip().lower() == "admin":
                    messagebox.showerror("Error", self._tr("Cannot delete the admin account."), parent=dialog)
                    return
                if not messagebox.askyesno("Confirm", f"Delete user '{uname}'?", parent=dialog):
                    return
                conn = sqlite3.connect(TEMP_DB_PATH)
                cur = conn.cursor()
                cur.execute("DELETE FROM users WHERE username=?", (uname,))
                if cur.rowcount == 0:
                    cur.execute("SELECT username FROM users")
                    for row in cur.fetchall() or []:
                        if plain_label(row[0]).lower() == str(uname).strip().lower():
                            cur.execute("DELETE FROM users WHERE username=?", (row[0],))
                            break
                commit_and_save(conn)
                conn.close()
                load_users()

            tb.Button(add_row, text=self._tr("Add User"), bootstyle="success", command=add_user).pack(side=LEFT, padx=6)

            rename_row = tb.Frame(tab_users)
            rename_row.pack(fill=X, pady=(0, 8))
            tb.Label(rename_row, text=self._tr("New username:"), font=("Segoe UI", 10, "bold")).pack(side=LEFT)
            rename_user_ent = tb.Entry(rename_row, width=22)
            rename_user_ent.pack(side=LEFT, padx=8)
            tb.Button(
                rename_row,
                text=self._tr("Rename Selected"),
                bootstyle="info",
                command=rename_user,
            ).pack(side=LEFT, padx=6)

            def _fill_rename(_e=None):
                sel = users_list.curselection()
                if not sel:
                    return
                rename_user_ent.delete(0, tk.END)
                rename_user_ent.insert(0, users_list.get(sel[0]))

            users_list.bind("<<ListboxSelect>>", _fill_rename)
            tb.Button(tab_users, text=self._tr("🗑️ Delete Selected"), bootstyle="danger outline", command=delete_user).pack(anchor=W, pady=6)
            load_users()

            tab_act = tb.Frame(notebook)
            notebook.add(tab_act, text=self._tr("Activity & Backups"))
            self._build_activity_and_backup_panel(tab_act)

            tab_tools = tb.Frame(notebook)
            notebook.add(tab_tools, text=self._tr("Commissions"))

            comm_canvas = tk.Canvas(tab_tools, highlightthickness=0, borderwidth=0)
            comm_scroll = tb.Scrollbar(tab_tools, orient=VERTICAL, command=comm_canvas.yview)
            comm_inner = tb.Frame(comm_canvas, padding=24)
            comm_inner.bind(
                "<Configure>",
                lambda e: comm_canvas.configure(scrollregion=comm_canvas.bbox("all")),
            )
            comm_win = comm_canvas.create_window((0, 0), window=comm_inner, anchor="nw")
            comm_canvas.configure(yscrollcommand=comm_scroll.set)

            def _comm_canvas_width(event):
                try:
                    comm_canvas.itemconfigure(comm_win, width=event.width)
                except Exception:
                    pass

            comm_canvas.bind("<Configure>", _comm_canvas_width)

            def _comm_mousewheel(event):
                try:
                    comm_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except Exception:
                    pass

            def _bind_comm_wheel(_e=None):
                comm_canvas.bind_all("<MouseWheel>", _comm_mousewheel)
                comm_canvas.bind_all("<Button-4>", lambda e: comm_canvas.yview_scroll(-1, "units"))
                comm_canvas.bind_all("<Button-5>", lambda e: comm_canvas.yview_scroll(1, "units"))

            def _unbind_comm_wheel(_e=None):
                try:
                    comm_canvas.unbind_all("<MouseWheel>")
                    comm_canvas.unbind_all("<Button-4>")
                    comm_canvas.unbind_all("<Button-5>")
                except Exception:
                    pass

            comm_canvas.bind("<Enter>", _bind_comm_wheel)
            comm_canvas.bind("<Leave>", _unbind_comm_wheel)
            comm_inner.bind("<Enter>", _bind_comm_wheel)
            comm_inner.bind("<Leave>", _unbind_comm_wheel)

            comm_scroll.pack(side=RIGHT, fill=Y)
            comm_canvas.pack(side=LEFT, fill=BOTH, expand=True)

            tb.Label(comm_inner, text=self._tr("Commissions"), font=("Segoe UI", 16, "bold"), bootstyle="primary").pack(pady=(0, 8), anchor=W)
            tb.Label(
                comm_inner,
                text=self._tr("From sales () → Tosales()  =  commission %. Dollar signs are optional."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=720,
                justify=LEFT,
            ).pack(anchor=W, pady=(0, 18))

            def _build_tier_editor(parent, title, kind):
                box = tb.Labelframe(parent, text=title, padding=16, bootstyle="info")
                box.pack(fill=BOTH, expand=True, pady=(0, 22))
                cols = (self._tr("From sales ($)"), self._tr("To sales ($)"), self._tr("Commission %"))
                tree_hold = tb.Frame(box)
                tree_hold.pack(fill=BOTH, expand=True, pady=(4, 12))
                tree = tb.Treeview(tree_hold, columns=cols, show="headings", height=9, bootstyle="info")
                tree_y = tb.Scrollbar(tree_hold, orient=VERTICAL, command=tree.yview)
                tree.configure(yscrollcommand=tree_y.set)
                for c in cols:
                    tree.heading(c, text=c)
                    tree.column(c, width=180, anchor=CENTER, minwidth=140)
                tree.pack(side=LEFT, fill=BOTH, expand=True)
                tree_y.pack(side=RIGHT, fill=Y)

                def load_ui():
                    for iid in tree.get_children():
                        tree.delete(iid)
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    try:
                        cur.execute(
                            "SELECT id, from_sales, to_sales, percentage FROM payout_tiers "
                            "WHERE LOWER(COALESCE(kind, 'service')) = ? ORDER BY from_sales ASC",
                            (kind,),
                        )
                        rows = cur.fetchall() or []
                    except Exception:
                        try:
                            cur.execute(
                                "SELECT id, from_sales, to_sales, percentage FROM payout_tiers ORDER BY from_sales ASC"
                            )
                            rows = cur.fetchall() or []
                        except Exception:
                            rows = []
                    conn.close()

                    if not rows and kind == "service":
                        rows = [
                            (0, 0.0, 2500.0, 35.0),
                            (0, 2500.01, 3500.0, 38.0),
                            (0, 3500.01, 4500.0, 40.0),
                            (0, 4500.01, 6000.0, 45.0),
                            (0, 6000.01, 7500.0, 48.0),
                            (0, 7500.01, 9999999.0, 50.0),
                        ]
                    elif not rows and kind == "product":
                        rows = [
                            (0, 0.0, 150.0, 0.0),
                            (0, 150.01, 250.0, 15.0),
                            (0, 250.01, 9999999.0, 20.0),
                        ]

                    for tid, a, b, p in rows:
                        a_val = to_float(a, 0.0)
                        b_val = to_float(b, 9999999.0)
                        p_val = to_float(p, 0.0)
                        iid = f"tier_{tid}" if tid else None
                        tree.insert("", tk.END, iid=iid, values=(f"{a_val:,.2f}", f"{b_val:,.2f}", f"{p_val:.2f}"))

                form = tb.Frame(box)
                form.pack(fill=X, pady=(8, 10))
                e_from = tb.Entry(form, width=16)
                e_to = tb.Entry(form, width=16)
                e_pct = tb.Entry(form, width=12)
                tb.Label(form, text=self._tr("From sales ($)"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(0, 6))
                e_from.pack(side=LEFT, padx=6)
                tb.Label(form, text=self._tr("To sales ($)"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(14, 6))
                e_to.pack(side=LEFT, padx=6)
                tb.Label(form, text=self._tr("Commission %"), font=("Segoe UI", 10, "bold")).pack(side=LEFT, padx=(14, 6))
                e_pct.pack(side=LEFT, padx=6)

                def _fill_inputs(event=None):
                    sel = tree.selection()
                    if not sel:
                        return
                    vals = tree.item(sel[0], "values")
                    e_from.delete(0, tk.END)
                    e_from.insert(0, vals[0].replace(",", ""))
                    e_to.delete(0, tk.END)
                    e_to.insert(0, vals[1].replace(",", ""))
                    e_pct.delete(0, tk.END)
                    e_pct.insert(0, vals[2].replace(",", ""))

                tree.bind("<<TreeviewSelect>>", _fill_inputs)

                def check_overlap(from_val, to_val, exclude_id=None):
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    try:
                        if exclude_id is not None:
                            cur.execute(
                                "SELECT from_sales, to_sales FROM payout_tiers WHERE LOWER(COALESCE(kind, 'service')) = ? AND id != ? "
                                "AND from_sales <= ? AND to_sales >= ?",
                                (kind.lower(), exclude_id, to_val, from_val)
                            )
                        else:
                            cur.execute(
                                "SELECT from_sales, to_sales FROM payout_tiers WHERE LOWER(COALESCE(kind, 'service')) = ? "
                                "AND from_sales <= ? AND to_sales >= ?",
                                (kind.lower(), to_val, from_val)
                            )
                        row = cur.fetchone()
                        conn.close()
                        return row is not None
                    except Exception:
                        try:
                            if exclude_id is not None:
                                cur.execute(
                                    "SELECT from_sales, to_sales FROM payout_tiers WHERE id != ? "
                                    "AND from_sales <= ? AND to_sales >= ?",
                                    (exclude_id, to_val, from_val)
                                )
                            else:
                                cur.execute(
                                    "SELECT from_sales, to_sales FROM payout_tiers "
                                    "WHERE from_sales <= ? AND to_sales >= ?",
                                    (to_val, from_val)
                                )
                            row = cur.fetchone()
                            conn.close()
                            return row is not None
                        except Exception:
                            conn.close()
                            return False

                def add_tier():
                    raw_a = (e_from.get() or "").strip()
                    raw_b = (e_to.get() or "").strip()
                    raw_p = (e_pct.get() or "").strip()
                    if not raw_a or not raw_b or not raw_p:
                        messagebox.showerror("Error", "From, To, and % are required. You can include $.", parent=dialog)
                        return
                    a = to_float(raw_a, 0.0)
                    b = to_float(raw_b, 0.0)
                    p = to_float(raw_p, 0.0)
                    
                    if check_overlap(a, b):
                        messagebox.showerror("Error", "This range overlaps with an existing commission tier.", parent=dialog)
                        return

                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    try:
                        cur.execute(
                            "INSERT INTO payout_tiers (from_sales, to_sales, percentage, kind) VALUES (?, ?, ?, ?)",
                            (a, b, p, kind),
                        )
                    except Exception:
                        cur.execute(
                            "INSERT INTO payout_tiers (from_sales, to_sales, percentage) VALUES (?, ?, ?)",
                            (a, b, p),
                        )
                    commit_and_save(conn)
                    conn.close()
                    e_from.delete(0, tk.END)
                    e_to.delete(0, tk.END)
                    e_pct.delete(0, tk.END)
                    load_ui()

                def update_tier():
                    sel = tree.selection()
                    if not sel:
                        messagebox.showerror("Error", "Please select a tier to update from the table first.", parent=dialog)
                        return
                    iid = sel[0]

                    raw_a = (e_from.get() or "").strip()
                    raw_b = (e_to.get() or "").strip()
                    raw_p = (e_pct.get() or "").strip()
                    if not raw_a or not raw_b or not raw_p:
                        messagebox.showerror("Error", "From, To, and % are required. You can include $.", parent=dialog)
                        return
                    new_a = to_float(raw_a, 0.0)
                    new_b = to_float(raw_b, 0.0)
                    new_p = to_float(raw_p, 0.0)

                    exclude_id = None
                    if iid and iid.startswith("tier_"):
                        exclude_id = int(iid.split("_")[1])

                    if check_overlap(new_a, new_b, exclude_id):
                        messagebox.showerror("Error", "This range overlaps with an existing commission tier.", parent=dialog)
                        return

                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    if iid and iid.startswith("tier_"):
                        tier_id = int(iid.split("_")[1])
                        cur.execute(
                            "UPDATE payout_tiers SET from_sales=?, to_sales=?, percentage=? WHERE id=?",
                            (new_a, new_b, new_p, tier_id),
                        )
                    else:
                        try:
                            cur.execute(
                                "INSERT INTO payout_tiers (from_sales, to_sales, percentage, kind) VALUES (?, ?, ?, ?)",
                                (new_a, new_b, new_p, kind),
                            )
                        except Exception:
                            cur.execute(
                                "INSERT INTO payout_tiers (from_sales, to_sales, percentage) VALUES (?, ?, ?)",
                                (new_a, new_b, new_p),
                            )
                    commit_and_save(conn)
                    conn.close()
                    e_from.delete(0, tk.END)
                    e_to.delete(0, tk.END)
                    e_pct.delete(0, tk.END)
                    load_ui()

                def delete_tier():
                    sel = tree.selection()
                    if not sel:
                        return
                    iid = sel[0]
                    vals = tree.item(iid, "values")
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    if iid and iid.startswith("tier_"):
                        tier_id = int(iid.split("_")[1])
                        cur.execute("DELETE FROM payout_tiers WHERE id=?", (tier_id,))
                    else:
                        try:
                            cur.execute(
                                "DELETE FROM payout_tiers WHERE from_sales=? AND to_sales=? AND percentage=? "
                                "AND LOWER(COALESCE(kind, 'service'))=?",
                                (to_float(vals[0]), to_float(vals[1]), to_float(vals[2]), kind),
                            )
                        except Exception:
                            cur.execute(
                                "DELETE FROM payout_tiers WHERE from_sales=? AND to_sales=? AND percentage=?",
                                (to_float(vals[0]), to_float(vals[1]), to_float(vals[2])),
                            )
                    commit_and_save(conn)
                    conn.close()
                    load_ui()

                tb.Button(form, text="+ Add", bootstyle="success", command=add_tier).pack(side=LEFT, padx=14)
                tb.Button(form, text=self._tr("Update"), bootstyle="info", command=update_tier).pack(side=LEFT, padx=14)
                tb.Button(box, text=self._tr("🗑️ Delete Selected"), bootstyle="danger outline", command=delete_tier).pack(anchor=W, pady=(8, 4))
                load_ui()

            _build_tier_editor(comm_inner, self._tr("Service sales commissions"), "service")
            _build_tier_editor(comm_inner, self._tr("Product sales commissions"), "product")
            dialog.bind("<Destroy>", lambda e: _unbind_comm_wheel())

            tab_cols = tb.Frame(notebook, padding=24)
            notebook.add(tab_cols, text=self._tr("📊 Table Columns"))
            
            tb.Label(tab_cols, text=self._tr("Select Columns to Display"), font=("Segoe UI", 15, "bold"), bootstyle="primary").pack(anchor=W, pady=(0, 6))
            tb.Label(
                tab_cols,
                text=self._tr("Check the columns you want visible in the Shop Earnings table. Changes are saved automatically."),
                font=("Segoe UI", 10),
                bootstyle="secondary",
                wraplength=720,
                justify=LEFT,
            ).pack(anchor=W, pady=(0, 15))
            
            cols_scroll_f = tb.Frame(tab_cols)
            cols_scroll_f.pack(fill=BOTH, expand=True)
            
            cols_canvas = tk.Canvas(cols_scroll_f, highlightthickness=0)
            cols_sb = tb.Scrollbar(cols_scroll_f, orient=VERTICAL, command=cols_canvas.yview)
            cols_inner = tb.Frame(cols_canvas, padding=10)
            cols_inner.bind("<Configure>", lambda e: cols_canvas.configure(scrollregion=cols_canvas.bbox("all")))
            cols_c_win = cols_canvas.create_window((0, 0), window=cols_inner, anchor="nw")
            cols_canvas.bind("<Configure>", lambda e: cols_canvas.itemconfigure(cols_c_win, width=e.width))
            cols_canvas.configure(yscrollcommand=cols_sb.set)
            
            cols_sb.pack(side=RIGHT, fill=Y)
            cols_canvas.pack(side=LEFT, fill=BOTH, expand=True)
            
            settings_hidden_now = get_calendar_hidden_columns()
            settings_check_vars = {}
            
            for col_key, desc in ALL_CALENDAR_COLUMNS:
                col_tr = self._tr(col_key)
                is_checked = (col_key not in settings_hidden_now and col_tr not in settings_hidden_now)
                var = tk.BooleanVar(value=is_checked)
                settings_check_vars[col_key] = var
                
                row_f = tb.Frame(cols_inner, padding=(4, 6))
                row_f.pack(fill=X, expand=True)
                
                cb = tb.Checkbutton(
                    row_f,
                    text=f"{col_tr}   —   {desc}",
                    variable=var,
                    bootstyle="primary-round-toggle",
                    cursor="hand2",
                )
                cb.pack(anchor=W)
            
            cols_btn_f = tb.Frame(tab_cols, padding=(0, 15, 0, 0))
            cols_btn_f.pack(fill=X, side=BOTTOM)
            
            def _settings_save_cols():
                new_hidden = {"Record ID", self._tr("Record ID")}
                for col_key, var in settings_check_vars.items():
                    if not var.get():
                        new_hidden.add(col_key)
                        new_hidden.add(self._tr(col_key))
                save_calendar_hidden_columns(new_hidden)
                self.refresh_calendar_column_visibility()
                messagebox.showinfo(
                    self._tr("Columns Updated"),
                    self._tr("Table columns display updated successfully."),
                    parent=dialog,
                )

            def _settings_select_all_cols(val):
                for var in settings_check_vars.values():
                    var.set(val)

            tb.Button(cols_btn_f, text=self._tr("Save & Apply"), bootstyle="success", cursor="hand2", command=_settings_save_cols).pack(side=LEFT, padx=5)
            tb.Button(cols_btn_f, text=self._tr("Select All"), bootstyle="secondary-outline", cursor="hand2", command=lambda: _settings_select_all_cols(True)).pack(side=LEFT, padx=5)
            tb.Button(cols_btn_f, text=self._tr("Reset to Default"), bootstyle="warning-outline", cursor="hand2", command=lambda: (_settings_select_all_cols(True), settings_check_vars.get("Written Up", tk.BooleanVar()).set(False))).pack(side=LEFT, padx=5)

            # --- LAST TAB: DATABASE & CLOUD SYNC ---
            tab_db = tb.Frame(notebook)
            notebook.add(tab_db, text=self._tr("🗄️ Database & Cloud"))
            self._build_database_and_cloud_panel(tab_db, dialog)

            # --- SOFTWARE UPDATES TAB ---
            tab_update = tb.Frame(notebook)
            notebook.add(tab_update, text=self._tr("🚀 App Updates"))
            self._build_app_updates_panel(tab_update, dialog)

            if default_tab in ("database", "supabase", "db"):
                notebook.select(tab_db)
            elif default_tab in ("update", "updates", "github"):
                notebook.select(tab_update)
            elif default_tab in ("activity", "logs", "backups"):
                notebook.select(tab_act)
            elif default_tab == "locations":
                notebook.select(tab_loc)
            elif default_tab == "categories":
                notebook.select(tab_cat)
            elif default_tab == "payments":
                notebook.select(tab_pay)
            elif default_tab == "account":
                notebook.select(tab_acc)
            elif default_tab == "users":
                notebook.select(tab_users)
            elif default_tab in ("commissions", "tiers"):
                notebook.select(tab_tools)
            elif default_tab == "columns":
                notebook.select(tab_cols)

            self._present_window(dialog)

        def setup_cash_calendar_tab(self):
            import calendar
            self.cash_cal_year = datetime.today().year
            self.cash_cal_month = datetime.today().month
            
            self.cash_cal_container = tb.Frame(self.tab_cash_cal, padding=10)
            self.cash_cal_container.pack(fill=BOTH, expand=True)
            
            ctrl_frame = tb.Frame(self.cash_cal_container)
            ctrl_frame.pack(fill=X, pady=(10, 20))
            
            tb.Button(ctrl_frame, text="◀ Prev", bootstyle="outline-primary", command=self.prev_cash_month).pack(side=LEFT, padx=10)
            self.lbl_cash_month_year = tb.Label(ctrl_frame, text="", font=("Segoe UI", 16, "bold"))
            self.lbl_cash_month_year.pack(side=LEFT, padx=20)
            tb.Button(ctrl_frame, text="Next ▶", bootstyle="outline-primary", command=self.next_cash_month).pack(side=LEFT, padx=10)
            self.btn_cash_month_lock = tb.Button(
                ctrl_frame,
                text=self._tr("🔒 Lock Month"),
                bootstyle="danger",
                cursor="hand2",
                command=self.open_cash_month_lock_dialog,
            )
            self.btn_cash_month_lock.pack(side=LEFT, padx=15)
            tb.Button(
                ctrl_frame,
                text=self._tr("Delete Month Envelopes"),
                bootstyle="danger outline",
                cursor="hand2",
                command=self.delete_month_envelopes,
            ).pack(side=LEFT, padx=8)
            
            self.lbl_cash_cal_summary = tb.Label(ctrl_frame, text="", font=("Segoe UI", 12, "italic"), bootstyle="secondary")
            self.lbl_cash_cal_summary.pack(side=RIGHT, padx=20)

            self.calendar_grid_frame = tb.Frame(self.cash_cal_container)
            self.calendar_grid_frame.pack(fill=BOTH, expand=True)
            
            self.load_cash_calendar_data()

        def cash_year_month(self, year=None, month=None):
            y = self.cash_cal_year if year is None else year
            m = self.cash_cal_month if month is None else month
            return f"{int(y):04d}-{int(m):02d}"

        def is_cash_month_locked(self, year_month=None):
            ym = year_month or self.cash_year_month()
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM cash_month_locks WHERE year_month=?", (ym,))
                row = cursor.fetchone()
                conn.close()
                return bool(row)
            except Exception:
                return False

        def is_date_in_locked_cash_month(self, date_str):
            if not date_str:
                return False
            try:
                ym = str(date_str).strip()[:7]
                if len(ym) < 7 or ym[4] != "-":
                    return False
                return self.is_cash_month_locked(ym)
            except Exception:
                return False

        def verify_user_password(self, password):
            entered = (password or "").strip()
            if not entered:
                return False
            username = getattr(self, "current_user", None) or DEFAULT_ADMIN_USERNAME
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT password FROM users WHERE username=?", (username,))
                row = cursor.fetchone()
                if not row:
                    cursor.execute("SELECT password FROM users WHERE username=?", (DEFAULT_ADMIN_USERNAME,))
                    row = cursor.fetchone()
                conn.close()
                if not row:
                    return False
                stored = row[0]
                hashed = hashlib.sha256(entered.encode("utf-8")).hexdigest()
                return stored == hashed or stored == entered
            except Exception:
                return False

        def update_cash_lock_button(self):
            btn = getattr(self, "btn_cash_month_lock", None)
            if not btn:
                return
            try:
                if not btn.winfo_exists():
                    return
            except Exception:
                return
            if self.is_cash_month_locked():
                btn.config(text=self._tr("🔓 Unlock Month"), bootstyle="warning")
            else:
                btn.config(text=self._tr("🔒 Lock Month"), bootstyle="danger")

        def delete_month_envelopes(self):
            if self.is_cash_month_locked():
                messagebox.showwarning("Locked", "Unlock this month before deleting envelope data.", parent=self)
                return
            dialog = tb.Toplevel(self)
            dialog.title(self._tr("Delete Month Envelopes"))
            dialog.geometry("440x240")
            dialog.transient(self)
            dialog.grab_set()
            self._present_window(dialog)
            ym = self.cash_year_month()
            tb.Label(
                dialog,
                text=self._tr("This will permanently delete all cash envelopes for this month."),
                wraplength=380,
                font=("Segoe UI", 11),
            ).pack(pady=(20, 8), padx=16)
            tb.Label(dialog, text=f"{ym}", font=("Segoe UI", 12, "bold")).pack()
            tb.Label(dialog, text="Password:", font=("Segoe UI", 10, "bold")).pack(anchor=W, padx=30, pady=(12, 4))
            pw = tb.Entry(dialog, show="*", width=28)
            pw.pack(padx=30)

            def do_delete():
                if not self.verify_user_password(pw.get()):
                    messagebox.showerror("Error", "Incorrect password.", parent=dialog)
                    return
                start = f"{ym}-01"
                import calendar as _cal
                _, nd = _cal.monthrange(self.cash_cal_year, self.cash_cal_month)
                end = f"{ym}-{nd:02d}"
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT id, category, expense_date FROM expenses WHERE expense_date >= ? AND expense_date <= ?",
                        (start, end),
                    )
                    ids = []
                    for exp_id, cat, _dt in cur.fetchall() or []:
                        cat_p = str(decrypt_val(cat) if cat is not None else "").strip()
                        if cat_p == "Cash Envelope Received":
                            ids.append(exp_id)
                    if ids:
                        placeholders = ",".join("?" for _ in ids)
                        cur.execute(f"DELETE FROM expenses WHERE id IN ({placeholders})", ids)
                        commit_and_save(conn)
                    conn.close()
                    dialog.destroy()
                    self.load_cash_calendar_data(quiet=True)
                    messagebox.showinfo("Deleted", f"Removed {len(ids)} envelope(s) for {ym}.", parent=self)
                except Exception as e:
                    messagebox.showerror("Error", str(e), parent=dialog)

            tb.Button(dialog, text=self._tr("Delete Month Envelopes"), bootstyle="danger", command=do_delete).pack(pady=16)
            tb.Button(dialog, text=self._tr("Cancel"), bootstyle="secondary", command=dialog.destroy).pack()

        def open_cash_month_lock_dialog(self):
            import calendar
            ym = self.cash_year_month()
            locked = self.is_cash_month_locked(ym)
            month_label = f"{self._tr(calendar.month_name[self.cash_cal_month])} {self.cash_cal_year}"

            dialog = tb.Toplevel(self)
            dialog.title(self._tr("Unlock Month") if locked else self._tr("Lock Month"))
            dialog.geometry("480x340")
            dialog.transient(self)
            dialog.grab_set()
            dialog.focus_set()

            if locked:
                explain = (
                    f"{month_label} is currently locked.\n\n"
                    "Unlocking will allow adding, editing, approving, and deleting "
                    "cash envelopes for this month again.\n\n"
                    "Enter your password and click Confirm to unlock."
                )
                confirm_label = self._tr("Confirm Unlock")
                boot = "warning"
            else:
                explain = (
                    f"You are about to lock in {month_label}.\n\n"
                    "Once locked, cash envelopes for this month cannot be added, "
                    "edited, approved, or deleted until the month is unlocked.\n\n"
                    "Enter your password and click Confirm to lock this month in."
                )
                confirm_label = self._tr("Confirm Lock")
                boot = "danger"

            tb.Label(dialog, text=explain, font=("Segoe UI", 11), wraplength=430, justify=LEFT).pack(padx=20, pady=(20, 10))
            tb.Label(dialog, text=self._tr("Password:"), font=("Segoe UI", 10, "bold")).pack(anchor=W, padx=20)
            pw_frame = tb.Frame(dialog)
            pw_frame.pack(fill=X, padx=20, pady=8)
            pw_entry = tb.Entry(pw_frame, show="*", width=28, font=("Segoe UI", 11))
            pw_entry.pack(side=LEFT)
            pw_entry.focus()

            def toggle_pw():
                try:
                    if pw_entry.cget("show") == "*":
                        pw_entry.config(show="")
                        eye_btn.config(text="Hide")
                    else:
                        pw_entry.config(show="*")
                        eye_btn.config(text="Show")
                except Exception:
                    pass

            eye_btn = tb.Button(pw_frame, text="Show", bootstyle="secondary outline", cursor="hand2", width=5, command=toggle_pw)
            eye_btn.pack(side=LEFT, padx=(5, 0))

            def do_confirm(event=None):
                if not self.verify_user_password(pw_entry.get()):
                    messagebox.showerror("Error", "Invalid password.", parent=dialog)
                    return
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    cursor = conn.cursor()
                    if locked:
                        cursor.execute("DELETE FROM cash_month_locks WHERE year_month=?", (ym,))
                    else:
                        cursor.execute("DELETE FROM cash_month_locks WHERE year_month=?", (ym,))
                        cursor.execute(
                            "INSERT INTO cash_month_locks (year_month, locked_by, locked_at) VALUES (?, ?, ?)",
                            (ym, getattr(self, "current_user", DEFAULT_ADMIN_USERNAME), datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                        )
                    commit_and_save(conn)
                    conn.close()
                    dialog.destroy()
                    self.load_cash_calendar_data(quiet=True)
                    messagebox.showinfo(
                        "Success",
                        f"{month_label} has been unlocked." if locked else f"{month_label} has been locked in.",
                    )
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to update lock: {e}", parent=dialog)

            pw_entry.bind("<Return>", do_confirm)
            btn_row = tb.Frame(dialog)
            btn_row.pack(pady=20)
            tb.Button(btn_row, text=confirm_label, bootstyle=boot, command=do_confirm).pack(side=LEFT, padx=8)
            tb.Button(btn_row, text=self._tr("Cancel"), bootstyle="secondary", command=dialog.destroy).pack(side=LEFT, padx=8)

        def prev_cash_month(self):
            self.cash_cal_month -= 1
            if self.cash_cal_month < 1:
                self.cash_cal_month = 12
                self.cash_cal_year -= 1
            self.load_cash_calendar_data(quiet=True)
            
        def next_cash_month(self):
            self.cash_cal_month += 1
            if self.cash_cal_month > 12:
                self.cash_cal_month = 1
                self.cash_cal_year += 1
            self.load_cash_calendar_data(quiet=True)

        def load_cash_calendar_data(self, quiet=False):
            if getattr(self, "_cash_cal_loading", False):
                return
            if self._envelope_ui_open():
                self._cash_cal_reload_when_popup_closes = True
                return
            if quiet:
                return self._load_cash_calendar_data_body()
            self.show_busy(self._tr("Loading cash calendar…"))
            try:
                return self._load_cash_calendar_data_body()
            finally:
                self.hide_busy()

        def _load_cash_calendar_data_body(self):
            import calendar
            if getattr(self, "_cash_cal_loading", False):
                return
            if self._envelope_ui_open():
                self._cash_cal_reload_when_popup_closes = True
                return
            self._cash_cal_loading = True
            try:
                return self._load_cash_calendar_data_body_inner()
            finally:
                self._cash_cal_loading = False

        def _load_cash_calendar_data_body_inner(self):
            import calendar
            year = self.cash_cal_year
            month = self.cash_cal_month
            
            month_name = calendar.month_name[month]
            locked = self.is_cash_month_locked()
            title = f"{self._tr(month_name)} {year}"
            if locked:
                title += f"  ({self._tr('LOCKED')})"
            start_date = f"{year}-{month:02d}-01"
            _, num_days = calendar.monthrange(year, month)
            end_date = f"{year}-{month:02d}-{num_days:02d}"
            
            rows = []
            try:
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                # Fetch by date only, then filter category in Python so plaintext and
                # legacy encrypted category values both show up.
                cursor.execute('''
                    SELECT ex.id, ex.expense_date, ex.amount, ex.status, e2.name AS assignee_name,
                        ex.description, ex.location, ex.category
                    FROM expenses ex
                    LEFT JOIN employees e2 ON ex.assignee_id = e2.id
                ''')
                rows = cursor.fetchall() or []
                conn.close()
            except Exception as e:
                try:
                    self.lbl_cash_cal_summary.config(text=f"Error loading calendar: {e}")
                except Exception:
                    pass
                return

            try:
                self.lbl_cash_month_year.config(text=title)
                self.update_cash_lock_button()
            except Exception:
                pass
            
            envelopes_by_day = {}
            notes_by_day = {}
            total_amt = 0.0
            total_approved = 0.0
            total_pending = 0.0
            for row in rows:
                try:
                    exp_id, dt_str, amt, status, assignee, desc, loc, category = row
                    cat = plain_label(category)
                    if not is_envelope_category(cat):
                        continue
                    iso = normalize_iso_date(dt_str)
                    if not iso:
                        continue
                    try:
                        dt = datetime.strptime(iso, "%Y-%m-%d")
                    except ValueError:
                        continue
                    if dt.year != year or dt.month != month:
                        continue
                    day = dt.day
                    amt_f = to_float(amt, 0.0)
                    status_s = str(decrypt_val(status) if status is not None else "").strip()
                    desc_s = str(decrypt_val(desc) if desc is not None else "").strip()
                    if desc_s:
                        notes_by_day.setdefault(day, []).append(desc_s)

                    if day not in envelopes_by_day:
                        envelopes_by_day[day] = []
                    envelopes_by_day[day].append({
                        "id": exp_id,
                        "amount": amt_f,
                        "status": status_s,
                        "assignee": assignee if assignee else "Unassigned",
                        "description": desc_s,
                        "location": loc if loc else ""
                    })
                    total_amt += amt_f
                    if status_s == "Approved":
                        total_approved += amt_f
                    else:
                        total_pending += amt_f
                except Exception:
                    continue

            # Also check payroll_records for any notes on days of this month
            try:
                conn_notes = sqlite3.connect(TEMP_DB_PATH)
                cur_n = conn_notes.cursor()
                cur_n.execute("SELECT record_date, notes FROM payroll_records WHERE notes IS NOT NULL AND TRIM(notes) != ''")
                for r_date, r_notes in cur_n.fetchall() or []:
                    p_note = str(decrypt_val(r_notes) if r_notes is not None else "").strip()
                    if p_note:
                        iso = normalize_iso_date(decrypt_val(r_date) if r_date is not None else "")
                        if not iso:
                            iso = normalize_iso_date(r_date)
                        if iso and iso.startswith(f"{year}-{month:02d}-"):
                            try:
                                d_num = int(iso.split("-")[2])
                                if p_note not in notes_by_day.setdefault(d_num, []):
                                    notes_by_day[d_num].append(p_note)
                            except Exception:
                                pass
                conn_notes.close()
            except Exception:
                pass
            
            self.lbl_cash_cal_summary.config(
                text=f"Total: ${total_amt:,.2f}  |  Approved: ${total_approved:,.2f}  |  Pending/Not Approved: ${total_pending:,.2f}"
                + (f"  |  {self._tr('LOCKED')}" if locked else "")
            )
            
            # Configured shop locations — each day needs one envelope per location.
            try:
                required_locations = [
                    str(decrypt_val(n) if n is not None else "").strip()
                    for n in (self.get_db_locations() or [])
                ]
                required_locations = [n for n in required_locations if n]
            except Exception:
                required_locations = []

            try:
                for widget in self.calendar_grid_frame.winfo_children():
                    try:
                        widget.unbind("<Button-1>")
                    except Exception:
                        pass
                    for child in widget.winfo_children():
                        try:
                            child.unbind("<Button-1>")
                        except Exception:
                            pass
                    widget.destroy()
            except Exception:
                return
            
            headers = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
            for col, h in enumerate(headers):
                lbl = tb.Label(self.calendar_grid_frame, text=self._tr(h), font=("Segoe UI", 11, "bold"), anchor=CENTER, bootstyle="secondary")
                lbl.grid(row=0, column=col, sticky="nsew", padx=2, pady=5)
                self.calendar_grid_frame.columnconfigure(col, weight=1, uniform="equal")
                
            cal_first_weekday, _ = calendar.monthrange(year, month)
            start_col = (cal_first_weekday + 1) % 7
            
            current_row = 1
            current_col = start_col
            
            self.calendar_grid_frame.rowconfigure(0, weight=0)
            for r in range(1, 7):
                self.calendar_grid_frame.rowconfigure(r, weight=1, uniform="equal")
                
            for c in range(start_col):
                lbl = tb.Label(self.calendar_grid_frame, text="", bootstyle="light")
                lbl.grid(row=current_row, column=c, sticky="nsew", padx=2, pady=2)
                
            for day in range(1, num_days + 1):
                cell_frame = tb.Frame(self.calendar_grid_frame, borderwidth=1, relief="solid")
                cell_frame.grid(row=current_row, column=current_col, sticky="nsew", padx=2, pady=2)
                
                day_envelopes = envelopes_by_day.get(day, [])
                day_notes = notes_by_day.get(day, [])
                num_notes = len(day_notes)

                # Format day header and note mark with star:
                # One note -> ⭐ 1 Note; Two notes -> ⭐⭐ 2 Notes
                if num_notes == 1:
                    day_hdr_text = f"{day} ⭐"
                    note_tag_text = "⭐ 1 Note"
                elif num_notes == 2:
                    day_hdr_text = f"{day} ⭐⭐"
                    note_tag_text = "⭐⭐ 2 Notes"
                elif num_notes > 2:
                    day_hdr_text = f"{day} ⭐⭐⭐"
                    note_tag_text = f"⭐⭐⭐ {num_notes} Notes"
                else:
                    day_hdr_text = str(day)
                    note_tag_text = ""
                
                cell_hdr = tb.Frame(cell_frame)
                cell_hdr.pack(fill=X, padx=4, pady=2)
                lbl_day = tb.Label(cell_hdr, text=day_hdr_text, font=("Segoe UI", 11, "bold"), anchor="w")
                lbl_day.pack(side=LEFT)
                
                if day_envelopes:
                    day_total = sum(to_float(env["amount"], 0.0) for env in day_envelopes)
                    present_locs = set()
                    for env in day_envelopes:
                        loc = str(decrypt_val(env.get("location")) if env.get("location") else "").strip()
                        if loc:
                            present_locs.add(loc)
                    missing_locs = [loc for loc in required_locations if loc not in present_locs]
                    locations_complete = (not required_locations) or (len(missing_locs) == 0)
                    all_approved = all(
                        str(env.get("status") or "").strip() == "Approved" for env in day_envelopes
                    )
                    # Green only when every location has an envelope AND all are approved.
                    # Red if a location is missing (e.g. only 1 of 2) or any envelope is pending.
                    complete_and_approved = locations_complete and all_approved and len(day_envelopes) > 0

                    if complete_and_approved:
                        cell_frame.config(bootstyle="success")
                        cell_hdr.config(bootstyle="success")
                        lbl_day.config(bootstyle="inverse-success")
                        loc_count = f"{len(present_locs)}/{len(required_locations)}" if required_locations else str(len(day_envelopes))
                        lbl_amount = tb.Label(cell_frame, text=f"${day_total:,.2f}", font=("Segoe UI", 11, "bold"), bootstyle="inverse-success")
                        lbl_amount.pack(pady=3)
                        lbl_count = tb.Label(
                            cell_frame,
                            text=f"{loc_count} {self._tr('Approved')}",
                            font=("Segoe UI", 8),
                            bootstyle="inverse-success",
                        )
                        lbl_count.pack()
                        if note_tag_text:
                            lbl_note_tag = tb.Label(
                                cell_frame,
                                text=note_tag_text,
                                font=("Segoe UI", 8, "bold"),
                                bootstyle="inverse-success",
                            )
                            lbl_note_tag.pack(pady=(2, 0))
                    else:
                        cell_frame.config(bootstyle="danger")
                        cell_hdr.config(bootstyle="danger")
                        lbl_day.config(bootstyle="inverse-danger")

                        if missing_locs:
                            branch_text = f"{self._tr('Missing')}: {', '.join(missing_locs)}"
                        else:
                            issue_locations = sorted(list(set(
                                str(env.get("location") or "").strip()
                                for env in day_envelopes
                                if str(env.get("status") or "").strip() != "Approved" and env.get("location")
                            )))
                            branch_text = ", ".join(issue_locations) if issue_locations else self._tr("Pending")

                        lbl_branch = tb.Label(cell_hdr, text=branch_text, font=("Segoe UI", 8, "bold"), bootstyle="inverse-danger", anchor=E)
                        lbl_branch.pack(side=RIGHT, padx=(4, 0))

                        lbl_amount = tb.Label(cell_frame, text=f"${day_total:,.2f}", font=("Segoe UI", 11, "bold"), bootstyle="inverse-danger")
                        lbl_amount.pack(pady=(2, 2))
                        if required_locations:
                            status_txt = f"{len(present_locs)}/{len(required_locations)} {self._tr('locations')}"
                        else:
                            status_txt = f"{len(day_envelopes)} {self._tr('Pending')}"
                        lbl_count = tb.Label(cell_frame, text=status_txt, font=("Segoe UI", 8), bootstyle="inverse-danger")
                        lbl_count.pack()
                        if note_tag_text:
                            lbl_note_tag = tb.Label(
                                cell_frame,
                                text=note_tag_text,
                                font=("Segoe UI", 8, "bold"),
                                bootstyle="inverse-danger",
                            )
                            lbl_note_tag.pack(pady=(2, 0))
                else:
                    cell_frame.config(bootstyle="light")
                    cell_hdr.config(bootstyle="light")
                    lbl_day.config(bootstyle="secondary")
                    if note_tag_text:
                        lbl_note_tag = tb.Label(
                            cell_frame,
                            text=note_tag_text,
                            font=("Segoe UI", 8, "bold"),
                            bootstyle="warning",
                        )
                        lbl_note_tag.pack(pady=4)
                    
                click_date = f"{year}-{month:02d}-{day:02d}"
                
                def make_bind(widget, dt_str=click_date):
                    widget.bind("<Button-1>", lambda e, d=dt_str: self._on_cash_day_click(d))
                    widget.config(cursor="hand2")
                    
                make_bind(cell_frame)
                make_bind(cell_hdr)
                make_bind(lbl_day)
                for w in cell_frame.winfo_children():
                    make_bind(w)
                    for sub_w in w.winfo_children():
                        make_bind(sub_w)
                
                current_col += 1
                if current_col > 6:
                    current_col = 0
                    current_row += 1
                    
            if current_col != 0:
                for c in range(current_col, 7):
                    lbl = tb.Label(self.calendar_grid_frame, text="", bootstyle="light")
                    lbl.grid(row=current_row, column=c, sticky="nsew", padx=2, pady=2)


        def _attach_tree_scrollbars(self, parent, tree):
            """Attach vertical + horizontal scrollbars.
            `tree` must be a child of `parent` (an empty Frame that holds the table).
            """
            scroll_y = tb.Scrollbar(parent, orient=VERTICAL, command=tree.yview)
            scroll_x = tb.Scrollbar(parent, orient=HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            scroll_y.pack(side=RIGHT, fill=Y)
            scroll_x.pack(side=BOTTOM, fill=X)
            tree.pack(side=LEFT, fill=BOTH, expand=True)

        def _on_cash_day_click(self, date_str):
            """Open the day popup once per click (cells bind the frame and labels)."""
            self._envelope_opening = True
            prev = getattr(self, "_cash_day_click_after", None)
            if prev is not None:
                try:
                    self.after_cancel(prev)
                except Exception:
                    pass
            self._cash_day_click_after = self.after(
                20, lambda d=date_str: self.open_day_envelopes_popup(d)
            )

        def open_day_envelopes_popup(self, date_str):
            self._cash_day_click_after = None
            existing = getattr(self, "_envelope_popup", None)
            if self._widget_alive(existing):
                if getattr(self, "_envelope_popup_date", None) == date_str:
                    try:
                        existing.lift()
                        existing.focus_set()
                    except Exception:
                        pass
                    self._envelope_opening = False
                    return
                self._envelope_popup = None
                self._envelope_popup_reload = None
                self._envelope_popup_date = None
                try:
                    self._safe_grab_release(existing)
                except Exception:
                    pass
                try:
                    existing.destroy()
                except Exception:
                    pass

            popup = self._open_sheet(self, f"Cash Envelopes: {date_str}", "980x680")
            try:
                popup.update_idletasks()
                screen_w = popup.winfo_screenwidth()
                screen_h = popup.winfo_screenheight()
                pop_w = min(980, max(720, screen_w - 40))
                pop_h = min(680, max(520, screen_h - 90))
                x = max(0, (screen_w - pop_w) // 2)
                y = max(10, min(30, (screen_h - pop_h) // 5))
                popup.geometry(f"{pop_w}x{pop_h}+{x}+{y}")
            except Exception:
                pass
            self._envelope_popup = popup
            self._envelope_popup_date = date_str
            self._envelope_opening = False
            
            tb.Label(popup, text=f"Cash Envelopes for {date_str}", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=(15, 5))
            loc_hint = tb.Label(popup, text="", font=("Segoe UI", 10), bootstyle="secondary")
            loc_hint.pack(pady=(0, 8))
            
            # Pack buttons first at bottom so they never get pushed off-screen
            btn_frame = tb.Frame(popup)
            btn_frame.pack(side=BOTTOM, fill=X, padx=20, pady=15)

            cols = (
                self._tr("ID"),
                self._tr("Received From"),
                self._tr("Amount"),
                self._tr("Status"),
                self._tr("Location"),
                self._tr("Description"),
            )
            tree_holder = tb.Frame(popup)
            tree_holder.pack(fill=BOTH, expand=True, padx=20, pady=10)
            tree = tb.Treeview(tree_holder, columns=cols, show="headings", bootstyle="primary", height=12, selectmode="extended")
            self.apply_and_memorize_column_widths(
                "cash_calendar_day_table",
                tree,
                cols,
                hidden_cols=[self._tr("ID")],
            )

            self._attach_tree_scrollbars(tree_holder, tree)
            
            def load_day_data():
                if not self._widget_alive(popup):
                    return
                try:
                    for item in tree.get_children():
                        tree.delete(item)
                except Exception:
                    return
                try:
                    if get_db_mode() == "supabase":
                        enable_local_first_mode()
                    required_locations = [
                        str(decrypt_val(n) if n is not None else "").strip()
                        for n in (self.get_db_locations() or [])
                    ]
                    required_locations = [n for n in required_locations if n]

                    day = str(date_str)[:10]
                    conn = sqlite3.connect(TEMP_DB_PATH, timeout=3)
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT ex.id, e.name, ex.amount, ex.status, ex.location, ex.description, ex.category, ex.expense_date
                        FROM expenses ex
                        LEFT JOIN employees e ON ex.assignee_id = e.id
                        WHERE CAST(ex.expense_date AS TEXT) = ?
                           OR CAST(ex.expense_date AS TEXT) LIKE ?
                           OR CAST(ex.expense_date AS TEXT) LIKE 'enc:%'
                           OR CAST(ex.expense_date AS TEXT) LIKE 'denc:%'
                    ''', (day, day + "%"))
                    present_locs = set()
                    for row in cursor.fetchall() or []:
                        row_vals = list(row)
                        raw_dt = row_vals[7] if len(row_vals) > 7 else ""
                        iso = normalize_iso_date(decrypt_val(raw_dt) if raw_dt else "")
                        if iso != day:
                            iso = normalize_iso_date(raw_dt)
                        if iso != day:
                            continue
                        cat = plain_label(row_vals[6])
                        if not is_envelope_category(cat):
                            continue
                        row_vals[1] = row_vals[1] if row_vals[1] else "General/None"
                        amt = to_float(row_vals[2], 0.0)
                        row_vals[2] = f"${amt:,.2f}"
                        row_vals[3] = row_vals[3] if row_vals[3] else ""
                        loc = str(decrypt_val(row_vals[4]) if row_vals[4] is not None else "").strip()
                        row_vals[4] = loc
                        if loc:
                            present_locs.add(loc)
                        row_vals[5] = row_vals[5] if row_vals[5] else ""
                        tree.insert('', tk.END, values=row_vals[:6])
                    conn.close()

                    missing = [loc for loc in required_locations if loc not in present_locs]
                    if required_locations:
                        if missing:
                            loc_hint.config(
                                text=f"{self._tr('Need one envelope per location')}: {', '.join(required_locations)}  |  {self._tr('Missing')}: {', '.join(missing)}",
                                bootstyle="danger",
                            )
                        else:
                            loc_hint.config(
                                text=f"{self._tr('All locations have envelopes')}: {', '.join(required_locations)}",
                                bootstyle="success",
                            )
                    else:
                        loc_hint.config(text="")
                except Exception as e:
                    try:
                        loc_hint.config(text=f"Could not load envelopes: {e}", bootstyle="danger")
                    except Exception:
                        pass
                
            self._envelope_popup_reload = load_day_data
            popup.after(30, load_day_data)

            def _after_save():
                try:
                    load_day_data()
                except Exception:
                    pass
                self._cash_cal_reload_when_popup_closes = True

            def _on_popup_close():
                if getattr(self, "_envelope_popup", None) is popup:
                    self._envelope_popup = None
                    self._envelope_popup_reload = None
                    self._envelope_popup_date = None
                self._safe_grab_release(popup)
                try:
                    popup.destroy()
                except Exception:
                    pass
                self._cash_cal_reload_when_popup_closes = False
                def _refresh_cal():
                    try:
                        self.load_cash_calendar_data(quiet=True)
                    except Exception:
                        pass
                self.after(80, _refresh_cal)

            popup.protocol("WM_DELETE_WINDOW", _on_popup_close)
            month_locked = self.is_date_in_locked_cash_month(date_str)
            if month_locked:
                tb.Label(
                    popup,
                    text=self._tr("This month is locked. Unlock it from Cash Calendar to make changes."),
                    font=("Segoe UI", 10, "bold"),
                    bootstyle="warning",
                ).pack(pady=(0, 5))
            
            def on_tree_double_click(event):
                selected = tree.selection()
                if not selected:
                    return
                if self.is_date_in_locked_cash_month(date_str):
                    messagebox.showwarning(
                        "Locked",
                        "This month is locked. Unlock it from Cash Calendar before editing envelopes.",
                        parent=popup,
                    )
                    return
                exp_id = tree.item(selected[0])['values'][0]
                conn = sqlite3.connect(TEMP_DB_PATH)
                cursor = conn.cursor()
                cursor.execute(EXPENSE_EDIT_SELECT, (exp_id,))
                row = cursor.fetchone()
                conn.close()
                if row:
                    self.open_expense_dialog(
                        expense_id=exp_id, 
                        data=row, 
                        on_save_callback=_after_save,
                    )
            
            tree.bind("<Double-1>", on_tree_double_click)
            
            def _selected_envelope_ids():
                ids = []
                for sel in tree.selection():
                    vals = tree.item(sel).get("values") or []
                    if vals and vals[0]:
                        ids.append(vals[0])
                return ids

            def approve_envelope():
                if self.is_date_in_locked_cash_month(date_str):
                    messagebox.showwarning("Locked", "This month is locked. Unlock it before approving envelopes.", parent=popup)
                    return
                ids = _selected_envelope_ids()
                if not ids:
                    messagebox.showwarning("Select", "Please select one or more envelopes to approve.\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.", parent=popup)
                    return
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    placeholders = ",".join("?" for _ in ids)
                    conn.execute(f"UPDATE expenses SET status = 'Approved' WHERE id IN ({placeholders})", ids)
                    commit_and_save(conn)
                    conn.close()
                    load_day_data()
                    self.load_cash_calendar_data(quiet=True)
                    self.load_financials_data(quiet=True)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed: {e}", parent=popup)
                    
            def disapprove_envelope():
                if self.is_date_in_locked_cash_month(date_str):
                    messagebox.showwarning("Locked", "This month is locked. Unlock it before changing envelope status.", parent=popup)
                    return
                ids = _selected_envelope_ids()
                if not ids:
                    messagebox.showwarning("Select", "Please select one or more envelopes to mark Pending.\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.", parent=popup)
                    return
                try:
                    conn = sqlite3.connect(TEMP_DB_PATH)
                    placeholders = ",".join("?" for _ in ids)
                    conn.execute(f"UPDATE expenses SET status = 'Pending' WHERE id IN ({placeholders})", ids)
                    commit_and_save(conn)
                    conn.close()
                    load_day_data()
                    self.load_cash_calendar_data(quiet=True)
                    self.load_financials_data(quiet=True)
                except Exception as e:
                    messagebox.showerror("Error", f"Failed: {e}", parent=popup)

            def add_new_envelope():
                if self.is_date_in_locked_cash_month(date_str):
                    messagebox.showwarning("Locked", "This month is locked. Unlock it before adding envelopes.", parent=popup)
                    return
                shop_id = self.get_shop_employee_id()
                mock_data = (date_str, "Cash Envelope Received", "", shop_id, "Pending", "", "Cash", "", None)
                self.open_expense_dialog(expense_id=None, data=mock_data, on_save_callback=_after_save)

            def delete_envelope():
                if self.is_date_in_locked_cash_month(date_str):
                    messagebox.showwarning("Locked", "This month is locked. Unlock it before deleting envelopes.", parent=popup)
                    return
                ids = _selected_envelope_ids()
                if not ids:
                    messagebox.showwarning("Select", "Please select one or more envelopes to delete.\nTip: hold Ctrl (Windows) or ⌘ (Mac) to multi-select.", parent=popup)
                    return
                confirm_msg = (
                    f"Are you sure you want to delete {len(ids)} selected cash envelope(s)?"
                    if len(ids) > 1
                    else "Are you sure you want to delete this cash envelope?"
                )
                if messagebox.askyesno("Confirm Delete", confirm_msg, parent=popup):
                    try:
                        conn = sqlite3.connect(TEMP_DB_PATH)
                        placeholders = ",".join("?" for _ in ids)
                        try:
                            cur = conn.cursor()
                            cur.execute(
                                f"SELECT document_path FROM expenses WHERE id IN ({placeholders})",
                                ids,
                            )
                            for (doc_path,) in cur.fetchall() or []:
                                for p in parse_expense_documents(doc_path):
                                    delete_expense_document_file(p)
                        except Exception:
                            pass
                        conn.execute(f"DELETE FROM expenses WHERE id IN ({placeholders})", ids)
                        commit_and_save(conn)
                        conn.close()
                        load_day_data()
                        self.load_cash_calendar_data(quiet=True)
                        self.load_financials_data(quiet=True)
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed: {e}", parent=popup)

            add_btn = tb.Button(btn_frame, text=self._tr("✅ Approve Selected"), bootstyle="success", command=approve_envelope)
            add_btn.pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("❌ Set Pending"), bootstyle="danger", command=disapprove_envelope).pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("+ Add Envelope"), bootstyle="primary", command=add_new_envelope).pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("🗑️ Delete"), bootstyle="secondary outline", command=delete_envelope).pack(side=LEFT, padx=5)
            tb.Button(btn_frame, text=self._tr("Close Window"), bootstyle="light", command=_on_popup_close).pack(side=RIGHT, padx=5)
            if month_locked:
                for child in btn_frame.winfo_children():
                    try:
                        if child.cget("text") != self._tr("Close Window"):
                            child.configure(state="disabled")
                    except Exception:
                        pass



def get_machine_id():
    """Generates a stable hardware-based ID for the current machine."""
    system_info = f"{platform.node()}-{uuid.getnode()}"
    return hashlib.sha256(system_info.encode()).hexdigest()[:16].upper()

def check_machine_license():
    """Verifies that the app is running on the authorized computer."""
    machine_id = get_machine_id()
    
    if AUTHORIZED_MACHINE_ID == "ANY":
        return True
        
    if AUTHORIZED_MACHINE_ID == "":
        root = tk.Tk()
        root.withdraw()
        root.clipboard_clear()
        root.clipboard_append(machine_id)
        root.update()
        messagebox.showinfo(
            "Machine ID Generated", 
            f"This application needs to be authorized for this computer.\n\n"
            f"Your Machine ID is:\n{machine_id}\n\n"
            f"(It has been copied to your clipboard.)\n\n"
            f"Please send this ID to the developer to receive your personalized app."
        )
        root.destroy()
        return False
        
    if machine_id != AUTHORIZED_MACHINE_ID:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            "License Error", 
            f"This application is not authorized to run on this computer.\n\n"
            f"Current Machine ID: {machine_id}\n\n"
            f"Please contact the developer for a fresh app."
        )
        root.destroy()
        return False
        
    return True


if __name__ == "__main__":
    if not HAS_DEPS:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Dependency Missing", "Please install required package:\n\nOpen your terminal and run:\npip install ttkbootstrap cryptography")
        sys.exit(1)
        
    if not check_machine_license():
        sys.exit(1)

    # Dynamic update bootstrap (runs downloaded update if present, or safely falls back)
    if _check_and_run_dynamic_update():
        sys.exit(0)

    app = PayrollApp()
    app.mainloop()


